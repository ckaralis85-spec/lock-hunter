"""
Lock Hunter — AI-powered lock sourcing desktop app
- Live marketplace searches via Claude API (web search tool)
- Full LPU Belt Explorer catalog (~900 locks) synced from lpubelts.com
- Lock thumbnails pulled from each lock's LPU page (Flickr CC media)
- SQLite database, search log, new/used + shipping filters
- Two tabs: Search and Locks (browse by owned / wishlist / all + belt)

Install: "Install LockHunter.bat"  ->  dist/LockHunter.exe (+ Desktop shortcut)
Dev run: python lock_hunter.py
"""

# Version scheme (MAJOR.MINOR.PATCH): each update bumps PATCH by 1, but PATCH
# and MINOR only run 1-9. So the sequence rolls over like this:
#   ... 3.1.8 -> 3.1.9 -> 3.2.1 -> 3.2.2 ... 3.9.9 -> 4.1.1 -> 4.1.2 ...
# i.e. after x.N.9 go to x.(N+1).1, and after x.9.9 go to (x+1).1.1.
VERSION = "4.7.4"




# ---------------------------------------------------------------------------
# Update check (GitHub Releases).
#   >>> Set GITHUB_REPO to "<owner>/<repo>" for your published repository,
#   >>> e.g. "ferf/lock-hunter".
# Once a day the app asks the GitHub API for the latest release; if its tag
# (e.g. "v4.7.1") is newer than this build, it offers to open the releases
# page so the user can download the new LockHunter.exe. Until GITHUB_REPO is
# filled in, the update check is silently inert (no errors, no false alerts).
GITHUB_REPO = "ckaralis85-spec/lock-hunter"
GITHUB_RELEASES_PAGE = "https://github.com/%s/releases/latest" % GITHUB_REPO
GITHUB_LATEST_API = "https://api.github.com/repos/%s/releases/latest" % GITHUB_REPO
GITHUB_ISSUES_URL = "https://github.com/%s/issues" % GITHUB_REPO

_BROWSER_UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
               "AppleWebKit/537.36 (KHTML, like Gecko) "
               "Chrome/124.0 Safari/537.36")

# Marketplace-probe timeout as a (connect, read) pair. The 12s connect makes
# a dead/black-holed host release its pool worker in 12s instead of squatting
# on it for the old flat 30s; the read timeout stays at the full 30s, so any
# site that actually responds gets exactly as much time to answer as before —
# a pure tail-killer with NO quality trade. Applies ONLY to the ~45 no-key
# marketplace probes and the direct eBay page scrape — infrastructure calls
# (LPU catalog sync, Firestore, Bazaar feed, eBay API, the Claude call,
# update check) keep their own timeouts.
PROBE_TIMEOUT = (12, 30)

import json
import os
import platform
import re
import sqlite3
import sys
import threading
import traceback
import queue
import datetime
import webbrowser
import urllib.parse
import concurrent.futures

# Headless build helper: `python lock_hunter.py --extract-icon [out.ico]` writes
# the embedded LPU icon to a file and exits, WITHOUT importing tkinter or
# starting the GUI. "Install LockHunter.bat" uses this so the built .exe gets the
# proper file icon even when no external assets folder is present.
if __name__ == "__main__" and len(sys.argv) > 1 and sys.argv[1] == "--extract-icon":
    import base64 as _b64, zlib as _zlib
    _out = sys.argv[2] if len(sys.argv) > 2 else "_built_icon.ico"
    try:
        _src = open(os.path.abspath(__file__), encoding="utf-8").read()
        _m = re.search(r"_ICO_B64 = \((.*?)\)", _src, re.S)
        _blob = "".join(re.findall(r'"([^"]*)"', _m.group(1)))
        with open(_out, "wb") as _f:
            _f.write(_zlib.decompress(_b64.b64decode(_blob)))
        sys.exit(0)
    except Exception as _ex:
        sys.stderr.write(f"icon extract failed: {_ex}\n")
        sys.exit(1)

import tkinter as tk
from tkinter import ttk, messagebox, filedialog

try:
    import requests
except ImportError:
    raise SystemExit("Missing dependency: pip install requests")

try:
    from PIL import Image, ImageTk
    HAVE_PIL = True
except ImportError:
    HAVE_PIL = False   # app still works, just no thumbnails

def _version_tuple(v):
    """'3.2.6' -> (3,2,6) for comparison; tolerant of stray text and of
    2-part versions like '4.5' (-> (4,5,0)) so the kill-switch min_version
    can be written either way."""
    m = re.search(r"(\d+)\.(\d+)\.(\d+)", str(v))
    if m:
        return tuple(int(x) for x in m.groups())
    m = re.search(r"(\d+)\.(\d+)", str(v))
    if m:
        return (int(m.group(1)), int(m.group(2)), 0)
    m = re.search(r"(\d+)", str(v))
    return (int(m.group(1)), 0, 0) if m else (0, 0, 0)

def _latest_from_text(text):
    """Find every 'LockHunter V#.#.#' (or a bare x.y.z) in a blob of text and
    return (version_str, tuple) of the newest, or (None, None). Handles
    \\u0020-style escaping so it also works on embedded JSON."""
    cleaned = (text.replace("\\u0020", " ").replace("\\/", "/")
                   .replace("\\u002e", ".").replace("&nbsp;", " "))
    found = re.findall(r"[Ll]ock\s*[Hh]unter[\s_%20]*[Vv]?\.?\s*(\d+\.\d+\.\d+)",
                       cleaned)
    if not found:
        # a version file may just contain the bare number / JSON
        found = re.findall(r"\b(\d+\.\d+\.\d+)\b", cleaned)
    if not found:
        return None, None
    best = max(found, key=_version_tuple)
    return best, _version_tuple(best)

def find_latest_version(status_cb=lambda s: None):
    """Return (version_str, tuple, status) for the newest published GitHub
    release. status is one of:
      "ok"    - a release was found (version_str/tuple are set)
      "none"  - GitHub was reached but there's no release yet, or GITHUB_REPO
                isn't configured (treat as: user already has the latest)
      "error" - couldn't reach GitHub at all (network problem)

    Reads the latest release's tag (e.g. "v4.7.1") from the GitHub API and
    parses the version out of it."""
    if not GITHUB_REPO or GITHUB_REPO.startswith("OWNER/"):
        return None, None, "none"   # repo not filled in yet -> stay silent
    try:
        status_cb("Checking for updates…")
        r = requests.get(GITHUB_LATEST_API, headers={
            "Accept": "application/vnd.github+json",
            "User-Agent": "LockHunter",
            "X-GitHub-Api-Version": "2022-11-28",
        }, timeout=30)
        if r.status_code == 200:
            tag = str((r.json() or {}).get("tag_name") or "").strip()
            # Tags look like "v4.7.1" — the version is glued to the leading
            # "v", so match without a word boundary (unlike _latest_from_text,
            # which is for scraping free text).
            m = re.search(r"(\d+\.\d+\.\d+)", tag)
            if m:
                return m.group(1), _version_tuple(m.group(1)), "ok"
            return None, None, "none"
        if r.status_code == 404:
            # repo reached but no releases published yet
            return None, None, "none"
        return None, None, "error"
    except Exception as ex:
        status_cb(f"Update check: {ex}")
        return None, None, "error"

def bump_version(v):
    """Return the next version string following the 1-9 rollover scheme above."""
    major, minor, patch = (int(x) for x in v.split("."))
    patch += 1
    if patch > 9:
        patch = 1
        minor += 1
        if minor > 9:
            minor = 1
            major += 1
    return f"{major}.{minor}.{patch}"

def resource_path(*parts):
    """Path to a bundled resource. Tries several locations so it works whether
    the app is run from source, frozen by PyInstaller (--onefile unpacks data
    to sys._MEIPASS), or the assets sit next to the .exe / script:
      1. sys._MEIPASS         (PyInstaller onefile temp-extract dir)
      2. dir of the .exe      (sys.executable, when frozen)
      3. dir of this .py file (running from source)
      4. current working dir
      5. ~/.lockhunter/assets  (materialized from embedded data — always works)
    Returns the first path that exists; if none do, returns the best guess so
    callers' os.path.exists() checks still behave sensibly."""
    bases = []
    meipass = getattr(sys, "_MEIPASS", None)
    if meipass:
        bases.append(meipass)
    if getattr(sys, "frozen", False):
        bases.append(os.path.dirname(os.path.abspath(sys.executable)))
    bases.append(os.path.dirname(os.path.abspath(__file__)))
    bases.append(os.getcwd())
    tried = []
    for base in bases:
        cand = os.path.join(base, *parts)
        tried.append(cand)
        if os.path.exists(cand):
            return cand
    # Last resort: materialize the icon from data embedded in this file, so the
    # window/header icon works even with NO assets folder present anywhere.
    embedded = _materialize_embedded_asset(*parts)
    if embedded and os.path.exists(embedded):
        return embedded
    # nothing existed — return the first candidate (stable default)
    return tried[0] if tried else os.path.join(*parts)

def _materialize_embedded_asset(*parts):
    """If the requested asset is one we carry embedded (base64+zlib at the end
    of this file), write it once into ~/.lockhunter/assets/ and return its path.
    Returns None for assets we don't embed."""
    if not parts:
        return None
    name = parts[-1]
    blob = _EMBEDDED_ASSETS.get(name)
    if not blob:
        return None
    try:
        import base64, zlib
        out_dir = os.path.join(os.path.expanduser("~"), ".lockhunter", "assets")
        os.makedirs(out_dir, exist_ok=True)
        out = os.path.join(out_dir, name)
        if not os.path.exists(out) or os.path.getsize(out) == 0:
            data = zlib.decompress(base64.b64decode(blob))
            with open(out, "wb") as f:
                f.write(data)
        return out
    except Exception:
        return None

def _refresh_stable_icons():
    """Keep the STABLE icon copies under ~/.lockhunter/assets in sync with the
    icons this build ships, and maintain a CONTENT-HASHED copy of the .ico
    (icon_<md5-8>.ico). Desktop shortcuts point their IconLocation at the
    hashed name because Windows caches shortcut icons by PATH: rewriting the
    same old path keeps showing the stale cached image no matter how many
    times the bytes change, while a path that changes with the icon's content
    has no cache entry and always renders fresh. The legacy lpu_icon.ico path
    is refreshed too, for old shortcuts that still aim at it. When anything
    changed and this is the frozen Windows exe, existing "Lock Hunter.lnk"
    shortcuts are silently re-pointed at this exe + the new hashed icon.
    Never raises. Returns True if any file was rewritten."""
    changed = False
    try:
        import base64, hashlib, zlib
        out_dir = os.path.join(os.path.expanduser("~"), ".lockhunter", "assets")
        os.makedirs(out_dir, exist_ok=True)

        def _shipped(name):
            """Bytes of the icon THIS build carries (assets folder / PyInstaller
            bundle first, embedded blob as fallback). Never the stable copy
            itself — that's the thing being refreshed."""
            src = resource_path("assets", name)
            if os.path.abspath(src) != os.path.abspath(
                    os.path.join(out_dir, name)):
                try:
                    if os.path.exists(src):
                        with open(src, "rb") as f:
                            return f.read()
                except Exception:
                    pass
            blob = _EMBEDDED_ASSETS.get(name)
            if blob:
                try:
                    return zlib.decompress(base64.b64decode(blob))
                except Exception:
                    pass
            return None

        ico_data = _shipped("lpu_icon.ico")
        for name, data in (("lpu_icon.ico", ico_data),
                           ("lpu_icon_256.png", _shipped("lpu_icon_256.png"))):
            if not data:
                continue
            dst = os.path.join(out_dir, name)
            try:
                cur = None
                if os.path.exists(dst):
                    with open(dst, "rb") as f:
                        cur = f.read()
                if cur != data:
                    with open(dst, "wb") as f:
                        f.write(data)
                    changed = True
            except OSError:
                pass

        hashed_path = None
        if ico_data:
            h = hashlib.md5(ico_data).hexdigest()[:8]
            hashed_path = os.path.join(out_dir, "icon_%s.ico" % h)
            try:
                if not os.path.exists(hashed_path):
                    with open(hashed_path, "wb") as f:
                        f.write(ico_data)
                    changed = True
                for fn in os.listdir(out_dir):
                    if (fn.startswith("icon_") and fn.endswith(".ico")
                            and fn != os.path.basename(hashed_path)):
                        try:
                            os.remove(os.path.join(out_dir, fn))
                        except OSError:
                            pass
            except OSError:
                pass

        if changed and os.name == "nt":
            # One-shot nudge so Windows drops its cached (old) icon.
            try:
                import subprocess
                subprocess.run(["ie4uinit.exe", "-show"], check=False,
                               creationflags=0x08000000)  # CREATE_NO_WINDOW
            except Exception:
                pass
        if changed and hashed_path:
            _refresh_desktop_shortcuts(hashed_path)
    except Exception:
        pass
    return changed

def _refresh_desktop_shortcuts(icon_path):
    """Re-point every EXISTING "Lock Hunter.lnk" desktop shortcut at THIS
    running exe with the given (content-hashed) icon. Only acts when running
    as the frozen Windows exe; the PowerShell work runs in a background
    thread so startup never waits on it. Existing shortcuts only — creating
    one is the installer's / build script's job. Never raises."""
    try:
        if os.name != "nt" or not getattr(sys, "frozen", False):
            return
        def q(s):   # single-quote for PowerShell ('' escapes a quote)
            return str(s).replace("'", "''")
        exe = os.path.abspath(sys.executable)
        wd = os.path.dirname(exe)
        ps = (
            "$ErrorActionPreference='SilentlyContinue';"
            "$w=New-Object -ComObject WScript.Shell;"
            "$dirs=@([Environment]::GetFolderPath('Desktop'),"
            "\"$env:USERPROFILE\\Desktop\","
            "\"$env:USERPROFILE\\OneDrive\\Desktop\","
            "\"$env:PUBLIC\\Desktop\") | Select-Object -Unique;"
            "foreach($d in $dirs){"
            "$p=Join-Path $d 'Lock Hunter.lnk';"
            "if(Test-Path $p){"
            "$s=$w.CreateShortcut($p);"
            "$s.TargetPath='" + q(exe) + "';"
            "$s.WorkingDirectory='" + q(wd) + "';"
            "$s.IconLocation='" + q(icon_path) + ",0';"
            "$s.Save()}}"
        )
        def _run():
            try:
                import subprocess
                subprocess.run(["powershell", "-NoProfile", "-Command", ps],
                               check=False, timeout=30,
                               creationflags=0x08000000)  # CREATE_NO_WINDOW
            except Exception:
                pass
        threading.Thread(target=_run, daemon=True).start()
    except Exception:
        pass

APP_DIR = os.path.join(os.path.expanduser("~"), ".lockhunter")
DB_PATH = os.path.join(APP_DIR, "lockhunter.db")
LOG_PATH = os.path.join(APP_DIR, "lockhunter.log")
CFG_PATH = os.path.join(APP_DIR, "config.json")
IMG_DIR = os.path.join(APP_DIR, "images")

API_URL = "https://api.anthropic.com/v1/messages"

API_VERSION = "2023-06-01"
MODEL = "claude-sonnet-4-6"
SECONDHAND_SITES = (
    "the LPU Lock Bazaar (https://lpulocks.com/#/lockbazaar — community "
    "marketplace, all listings are member-sold used locks), "
    "eBay, DBA.dk, Tradera, Blocket, Lauritz, Catawiki, Facebook Marketplace, "
    "Gumtree, Leboncoin, Marktplaats, Finn.no, Kleinanzeigen, ShopGoodwill, "
    "Etsy (vintage), LiveAuctioneers, Invaluable, Delcampe, Ricardo.ch, Todocoleccion"
)

# Lock brands mapped to their country of origin / primary market. Used to steer
# searches toward the marketplaces and language where a given lock is most
# likely to appear secondhand. Keys are lowercase brand tokens matched against
# the start of the lock name. This is guidance for the AI, not a hard filter.
BRAND_ORIGINS = {
    "abloy": "Finland", "sento": "Finland",
    "assa": "Sweden", "ruko": "Denmark/Sweden", "trioving": "Norway",
    "assa abloy": "Sweden",
    "bks": "Germany", "dom": "Germany", "abus": "Germany", "burg": "Germany",
    "burg-wächter": "Germany", "ces": "Germany", "evva": "Austria",
    "winkhaus": "Germany", "wilka": "Germany", "zeiss": "Germany",
    "cawi": "Germany", "wittkopp": "Germany", "carl wittkopp": "Germany",
    "mauer": "Germany", "kromer safe": "Germany",
    "ikon": "Germany", "kaba": "Switzerland/Germany", "dorma kaba": "Switzerland/Germany",
    "gege": "Austria", "kromer": "Germany", "zi ikon": "Germany",
    "mottura": "Italy", "cisa": "Italy", "iseo": "Italy", "securemme": "Italy",
    "viro": "Italy", "mia": "Italy", "potent": "Italy",
    "fichet": "France", "bricard": "France", "vachette": "France",
    "pollux": "France", "heracles": "France", "muel": "France",
    "mul-t-lock": "Israel", "multlock": "Israel", "rav bariah": "Israel",
    "lips": "Netherlands", "nemef": "Netherlands", "oxloc": "Netherlands",
    "lockwood": "Australia", "gainsborough": "Australia", "brava": "Australia",
    "whitco": "Australia",
    "medeco": "USA", "sargent": "USA", "corbin": "USA", "russwin": "USA",
    "corbin russwin": "USA", "schlage": "USA", "kwikset": "USA",
    "american lock": "USA", "master lock": "USA", "best": "USA", "arrow": "USA",
    "s&g": "USA", "sargent and greenleaf": "USA", "yale": "USA/UK",
    "chubb": "UK", "union": "UK", "era": "UK", "squire": "UK", "ingersoll": "UK",
    "banham": "UK", "asec": "UK",
    "goal": "Japan", "miwa": "Japan", "west": "Japan", "showa": "Japan",
    "guli": "China", "wafangdian": "China",
    "border": "Russia", "kerberos": "Russia", "guardian": "Russia",
    "titan": "Russia/Ukraine",
    "robur": "Sweden", "rosengrens": "Sweden",
    "tesa": "Spain", "azbe": "Spain", "fac": "Spain", "ucem": "Spain",
    "lince": "Spain",
    "lob": "Poland", "lobix": "Poland", "gerda": "Poland", "bordo": "Poland",
    "metalplast": "Poland", "zamki": "Poland",
    "abus pl": "Poland",
    "portugal": "Portugal", "jma": "Spain", "stv": "Portugal",
    "kale": "Turkey", "kale kilit": "Turkey", "srannaz": "Turkey",
    "rostex": "Czechia", "tokoz": "Czechia", "fab": "Czechia",
    "hefele": "Slovakia",
    "vagner": "Croatia",
}

# Country -> (native-language marketplaces + language hint) for localized search.
COUNTRY_MARKETS = {
    "Finland": "Finnish sites Tori.fi, Huuto.net (search in Finnish, e.g. 'lukko')",
    "Sweden": "Swedish sites Tradera, Blocket, Marketplace (search in Swedish, 'lås')",
    "Norway": "Norwegian site Finn.no (search in Norwegian, 'lås')",
    "Denmark": "Danish sites DBA.dk, Lauritz (search in Danish, 'lås')",
    "Denmark/Sweden": "DBA.dk, Tradera, Blocket (Danish/Swedish, 'lås')",
    "Germany": "German sites eBay.de, Kleinanzeigen.de (search in German, 'Schloss', 'Zylinder', 'Schließzylinder')",
    "Austria": "eBay.at, willhaben.at, Kleinanzeigen.de (German, 'Schloss', 'Zylinder')",
    "Switzerland/Germany": "Ricardo.ch, tutti.ch, eBay.de, Kleinanzeigen.de (German/French)",
    "Italy": "Italian sites eBay.it, Subito.it (search in Italian, 'serratura', 'cilindro')",
    "France": "French sites Leboncoin, eBay.fr (search in French, 'serrure', 'cylindre', 'barillet')",
    "Netherlands": "Dutch site Marktplaats.nl (search in Dutch, 'slot', 'cilinder')",
    "Australia": "Australian Gumtree, eBay.com.au, Facebook Marketplace (English)",
    "USA": "eBay.com, US Facebook Marketplace, ShopGoodwill, Craigslist, LiveAuctioneers (English)",
    "USA/UK": "eBay.com and eBay.co.uk (English)",
    "UK": "UK sites eBay.co.uk, Gumtree (English)",
    "Japan": "Japanese sites Yahoo! Auctions Japan (auctions.yahoo.co.jp), Mercari (search in Japanese, '錠' / 'ロック' / brand in katakana)",
    "China": "AliExpress, Taobao (Chinese)",
    "Russia": "Avito.ru, Meshok.net (search in Russian, 'замок', 'цилиндр')",
    "Russia/Ukraine": "Avito.ru, OLX.ua (Russian/Ukrainian, 'замок')",
    "Israel": "Yad2 (yad2.co.il), eBay (search in Hebrew, 'מנעול')",
    "Spain": "Spanish sites Wallapop, Milanuncios, eBay.es (search in Spanish, 'cerradura', 'cilindro', 'bombín')",
    "Poland": "Polish sites OLX.pl, Allegro (search in Polish, 'wkładka', 'zamek', 'bębenkowa')",
    "Portugal": "Portuguese sites OLX.pt, CustoJusto (search in Portuguese, 'canhão', 'fechadura')",
}

def _split_catalog_names(name):
    """LPU catalog entries sometimes list several distinct locks in one entry,
    separated by '/' (e.g. 'Anbo X18u / Lex X18 / GLK X18 / Apec XS'). These
    are separate locks and must be searched individually — no seller lists all
    of them in one title. Returns the component names (deduped, order
    preserved); a normal single-name entry returns just [name].

    Splits on slashes used as name separators, but not on dimensions like
    '30/30' or '30/10': a purely numeric fragment is reattached to the
    previous name.

    Whenever a split actually happens, the UNTOUCHED full name is kept as the
    FIRST search as well. The splitter can't always tell a separator slash
    from a model slash (real catalog cases: 'ABUS 74/40 LOTO' is ONE lock
    whose size has a suffix word, and '(UCF/UCH/UCS)' lists variants inside
    parentheses), so searching the full name too guarantees the entry is
    always hunted exactly as written, with the split parts as extras."""
    if not name or "/" not in name:
        return [name]
    rough = re.split(r"\s*/\s*", name)
    parts = [seg.strip() for seg in rough if seg.strip()]
    cleaned = []
    for p in parts:
        if re.fullmatch(r"\d+", p) and cleaned:
            # numeric fragment = a size (e.g. the '30' in '30/30'); reattach
            cleaned[-1] = cleaned[-1] + "/" + p
        else:
            cleaned.append(p)
    seen, out = set(), []
    for p in cleaned:
        k = p.lower()
        if p and k not in seen:
            seen.add(k)
            out.append(p)
    # keep the full original name as the first search whenever it differs
    # from what the split produced (i.e. an actual split happened)
    full = re.sub(r"\s+", " ", name.strip())
    if full and full.lower() not in {o.lower() for o in out}:
        out.insert(0, full)
    return out or [name]


def _broaden_lock_name(name):
    """A broader form of a lock name with descriptive qualifiers stripped, so
    an over-specific catalog entry also gets searched by its core brand+model.
    Removes:
      - pin-count descriptors: '7 pin', '6-pin', '5 pins', '5 pin lock',
        '7 pin cylinder'  (the count word, plus an optional trailing
        'lock'/'cylinder')
      - parenthetical NUMERIC codes such as '(70091)'

    Examples:
      'Dejo 7 pin'               -> 'Dejo'
      'Mauer Variator A (70091)' -> 'Mauer Variator A'

    Returns '' when nothing containing a letter is left, so a bare code or a
    lone '7 pin' never becomes a junk query."""
    s = name or ""
    # parenthetical codes: only parens whose contents include a digit, so a
    # numeric code '(70091)' is dropped but a word note '(euro profile)' isn't
    s = re.sub(r"\(\s*[^)]*\d[^)]*\)", " ", s)
    # pin-count descriptors, with an optional trailing 'lock'/'cylinder'
    s = re.sub(r"\b\d+\s*-?\s*pins?\b(?:\s+(?:locks?|cylinders?|cyl))?",
               " ", s, flags=re.I)
    s = re.sub(r"\s+", " ", s).strip(" -/,")
    if not re.search(r"[A-Za-z]", s):
        return ""
    return s


# Layer-2 catalog overrides (from Ferf's reviewed spreadsheet). LPU packs
# several distinct locks into one entry in inconsistent ways; these are the
# entries where the correct search terms can't be derived automatically, so
# they're specified by hand. Keyed by the FOLDED catalog name (the app builds
# names identically, so this matches). Value = list of search terms, or None
# to skip the entry entirely (unsearchable placeholders).
_CATALOG_OVERRIDES = {
    'abus mylock / t65al': ['ABUS MyLock', 'ABUS T65AL'],
    'assa 500 / assa flexcore/flexcore plus': ['ASSA 500', 'ASSA Flexcore', 'ASSA Flexcore Plus'],
    'avocet abs / era professional cylinder / federal lock u-systems (ucf/uch/ucs/uus)-3100 / thirard federal s / thirard federal 2': ['Avocet ABS', 'ERA Professional Cylinder', 'Federal Lock U-systems 3100', 'Thirard Federal S', 'Thirard Federal 2'],
    'bison/hyt/lays/qlsy chain key lock': ['Bison Chain Key Lock', 'HYT Chain Key Lock', 'Lays Chain Key Lock', 'Qlsy Chain Key Lock'],
    'ccl sesamee 900 series': ['CCL Sesamee 900'],
    'eagle "supr-security" (with shutter)': ['Eagle "Supr-Security"'],
    'eclipse (any model)': ['Eclipse'],
    'evva gpi/als': ['EVVA GPI', 'EVVA ALS'],
    'generic/unknown 1 or 2 lever cabinet lock': ['1 Lever Cabinet lock', '2 lever cabinet lock'],
    'generic/unknown 3 lever cabinet lock': ['3 Lever Cabinet lock'],
    'generic/unknown 3 lever mortice lock': ['3 Lever mortice lock'],
    'generic/unknown 4 lever curtained lock': ['4 lever curtained lock', '4 lever lock'],
    'generic/unknown 4 lever uncurtained lock': ['4 lever uncurtained lock', '4 lever lock'],
    'generic/unknown 6 or 7 lever lock': ['6 Lever lock', '7 lever lock'],
    'generic/unranked lock cylinder': None,
    'kasp 160 series diskus': ['Kasp 160 Diskus'],
    'kawaha (any) / magmaus (any)': ['Kawaha', 'Magmaus'],
    'lips keso (door cylinder) / lips octro': ['LIPS Keso', 'LIPS Octro'],
    'master lock pro series': ['Master Lock Pro'],
    'miwa/anker 3800': ['MIWA 3800', 'Anker 3800'],
    's&g 4544 / s&g 4440 series': ['S&G 4544', 'S&G 4440'],
    'schlage original commercial/residential': ['Schlage Original Commercial', 'Schlage Original Residential'],
    'wilka pr100 series / wilka th6 / wilka si6': ['Wilka PR100', 'Wilka TH6', 'Wilka SI6'],
    'yale y110 series brass padlock': ['Yale Y110 brass padlock'],
    'yale y120 series brass padlock': ['Yale Y120 brass padlock'],
    'yuema 750 series / forte enigma': ['Yuema 750', 'Forte Enigma'],
}


def _search_terms(name):
    """Every query to run for one catalog lock name:
      1. the name AS WRITTEN — plus, for slash entries, each individual lock
         (see _split_catalog_names); and
      2. a BROADENED variant of each individual lock, with pin-count
         descriptors and numeric parenthetical codes removed (see
         _broaden_lock_name) — e.g. 'Dejo 7 pin' also searches 'Dejo', and
         'Mauer Variator A (70091)' also searches 'Mauer Variator A'.

    Deduped, order preserved, with the exact/original terms first so precise
    matches lead and the broader ones follow.

    A hand-curated override (from the reviewed spreadsheet) wins outright for the
    handful of catalog entries LPU labels ambiguously: it returns the exact
    search terms, or [] to skip an unsearchable placeholder entirely."""
    _ov = _CATALOG_OVERRIDES.get(_fold(name))
    if _ov is not None:
        return list(_ov)
    if _ov is None and _fold(name) in _CATALOG_OVERRIDES:
        return []      # explicit skip (value was None)
    parts = _split_catalog_names(name)
    out, seen = [], set()

    def add(t):
        t = re.sub(r"\s+", " ", (t or "").strip())
        k = t.lower()
        if t and k not in seen:
            seen.add(k)
            out.append(t)

    for p in parts:                       # originals, exactly as written
        add(p)
    # broaden the INDIVIDUAL locks only — never the joined full slash-name,
    # which _split_catalog_names keeps at parts[0] when a split happened.
    comps = parts[1:] if len(parts) > 1 else parts
    for p in comps:
        b = _broaden_lock_name(p)
        if b and b.lower() != p.lower():
            add(b)
    return out


def _origin_for_lock(lock_name):
    """Best-guess country of origin for a lock name, matched by brand prefix.
    Returns a country string (a key of COUNTRY_MARKETS) or ''. Longer brand
    keys are tried first so 'assa abloy' beats 'assa', 'dorma kaba' beats 'kaba'."""
    n = (lock_name or "").strip().lower()
    for brand in sorted(BRAND_ORIGINS, key=len, reverse=True):
        if n.startswith(brand):
            return BRAND_ORIGINS[brand]
    # also allow the brand to appear as a standalone word anywhere in the name
    tokens = set(re.split(r"[^a-z0-9&]+", n))
    for brand in sorted(BRAND_ORIGINS, key=len, reverse=True):
        if " " not in brand and brand in tokens:
            return BRAND_ORIGINS[brand]
    return ""


# ------------------------------------------------- direct marketplace search
# General web search barely indexes live marketplace listings (short-lived
# auctions rarely make it into the index before they end). So rather than only
# *searching about* the marketplaces, we hit their own on-site search URLs:
#  - these templates are handed to the AI to FETCH directly (web_fetch tool)
#  - and the app itself probes eBay's regional search pages deterministically.
MARKET_SEARCH_URLS = {
    "Finland": ["https://www.tori.fi/koko_suomi?q={q}",
                "https://www.huuto.net/haku?words={q}"],
    "Sweden": ["https://www.tradera.com/search?q={q}",
               "https://www.blocket.se/annonser/hela_sverige?q={q}"],
    "Norway": ["https://www.finn.no/bap/forsale/search.html?q={q}"],
    "Denmark": ["https://www.dba.dk/soeg/?soeg={q}"],
    "Denmark/Sweden": ["https://www.dba.dk/soeg/?soeg={q}",
                       "https://www.tradera.com/search?q={q}"],
    "Germany": ["https://www.kleinanzeigen.de/s-suchanfrage.html?keywords={q}",
                "https://www.ebay.de/sch/i.html?_nkw={q}"],
    "Austria": ["https://www.willhaben.at/iad/kaufen-und-verkaufen/marktplatz?keyword={q}",
                "https://www.ebay.de/sch/i.html?_nkw={q}"],
    "Switzerland/Germany": ["https://www.ricardo.ch/de/s/{q}",
                            "https://www.ebay.de/sch/i.html?_nkw={q}"],
    "Italy": ["https://www.subito.it/annunci-italia/vendita/usato/?q={q}",
              "https://www.ebay.it/sch/i.html?_nkw={q}"],
    "France": ["https://www.leboncoin.fr/recherche?text={q}",
               "https://www.ebay.fr/sch/i.html?_nkw={q}"],
    "Netherlands": ["https://www.marktplaats.nl/q/{q}/"],
    "UK": ["https://www.ebay.co.uk/sch/i.html?_nkw={q}",
           "https://www.gumtree.com/search?search_category=all&q={q}"],
    "USA": ["https://www.ebay.com/sch/i.html?_nkw={q}",
            "https://shopgoodwill.com/categories/listing?st={q}"],
    "USA/UK": ["https://www.ebay.com/sch/i.html?_nkw={q}",
               "https://www.ebay.co.uk/sch/i.html?_nkw={q}"],
    "Australia": ["https://www.ebay.com.au/sch/i.html?_nkw={q}"],
    "Japan": ["https://auctions.yahoo.co.jp/search/search?p={q}",
              "https://jp.mercari.com/search?keyword={q}"],
    "Spain": ["https://es.wallapop.com/app/search?keywords={q}",
              "https://www.ebay.es/sch/i.html?_nkw={q}"],
    "Russia": ["https://www.avito.ru/all?q={q}"],
    "Israel": ["https://www.ebay.com/sch/i.html?_nkw={q}"],
}
GLOBAL_MARKET_URLS = [
    "https://www.ebay.com/sch/i.html?_nkw={q}",
    "https://www.ebay.co.uk/sch/i.html?_nkw={q}",
    "https://www.catawiki.com/en/s?q={q}",
    "https://www.delcampe.net/en_US/collectibles/search?term={q}",
    "https://shopgoodwill.com/categories/listing?st={q}",
]

def _market_urls(lock_name, origin, deep=False):
    """Ready-to-fetch marketplace search URLs for this lock: the origin
    country's venues first, then the global secondhand tier. Never filtered
    to auction-format only — fixed-price (Buy It Now) secondhand listings are
    a huge share of rare-lock sales and must always be visible."""
    q = urllib.parse.quote_plus(lock_name.strip())
    urls = []
    for t in MARKET_SEARCH_URLS.get(origin, []):
        u = t.format(q=q)
        if u not in urls:
            urls.append(u)
    for t in GLOBAL_MARKET_URLS:
        u = t.format(q=q)
        if u not in urls:
            urls.append(u)
    return urls[:12 if deep else 7]

# eBay regional domain that matches a lock's origin country (probed first).
ORIGIN_EBAY = {
    "Germany": "www.ebay.de", "Austria": "www.ebay.at",
    "Switzerland/Germany": "www.ebay.ch", "Italy": "www.ebay.it",
    "France": "www.ebay.fr", "Spain": "www.ebay.es",
    "Netherlands": "www.ebay.nl", "UK": "www.ebay.co.uk",
    "USA": "www.ebay.com", "USA/UK": "www.ebay.com",
    "Australia": "www.ebay.com.au",
}
_EBAY_GLOBAL = ["www.ebay.com", "www.ebay.co.uk", "www.ebay.de"]

def _parse_ebay_search(html):
    """Extract (title, price, item_url) tuples from an eBay search-results
    page. Tries the classic 's-item' card markup first; if that yields
    nothing (eBay ships newer layouts too), falls back to a looser pass that
    keys off /itm/ links directly."""
    items = _parse_ebay_search_strict(html)
    return items if items else _parse_ebay_search_loose(html)

def _parse_ebay_search_loose(html):
    """Layout-robust fallback. eBay ships several search-result card designs
    (classic 's-item', newer 'su-card', mobile), and the /itm/ anchor wraps a
    large block of nested markup — so instead of reading the anchor's own text
    we find every /itm/<id> link and pull a title + price from the card markup
    AROUND it, trying the title spots eBay actually uses in order: a heading
    span, s-item__title, the newer su-styled-text primary line, the product
    image alt, then aria-label. Item ids are global, so the link is rebuilt
    canonically from the id."""
    out, seen = [], set()
    title_pats = (
        r'role="heading"[^>]*>\s*(?:<[^>]+>\s*)*([^<]{3,180})',
        r's-item__title[^"]*"[^>]*>\s*(?:<span[^>]*>\s*)*([^<]{3,180})',
        r'su-styled-text[^"]*(?:primary|bold)[^"]*"[^>]*>\s*([^<]{3,180})',
        r'<img[^>]+alt="([^"]{3,180})"',
        r'aria-label="([^"]{3,180})"',
    )
    for m in re.finditer(r'/itm/(?:[^"/?\s]+/)?(\d{9,})', html):
        iid = m.group(1)
        if iid in seen:
            continue
        window = html[max(0, m.start() - 250): m.start() + 2600]
        title = ""
        for pat in title_pats:
            tm = re.search(pat, window, re.I | re.S)
            if tm:
                title = tm.group(1)
                break
        title = re.sub(r"\s+", " ", title).strip()
        title = re.sub(r"^New Listing\s*", "", title, flags=re.I)
        title = title.replace("&amp;", "&").replace("&nbsp;", " ").strip()
        low = title.lower()
        if not title or low in ("shop on ebay", "new listing") \
                or low.startswith("opens in a"):
            continue
        pm = re.search(r'((?:US\s*\$|C\s*\$|AU\s*\$|EUR|GBP|£|€|\$)'
                       r'\s?\d[\d.,]*)', window)
        price = re.sub(r"\s+", " ", pm.group(1)).strip() if pm else ""
        seen.add(iid)
        out.append((title, price, f"https://www.ebay.com/itm/{iid}"))
        if len(out) >= 120:
            break
    return out

def _parse_ebay_search_strict(html):
    """Classic 's-item' card markup: per-item chunks so fields can't bleed
    across listings; skips the 'Shop on eBay' placeholder card."""
    out = []
    for chunk in re.split(r'<li[^>]*class="[^"]*s-item', html)[1:]:
        lm = re.search(r'href="(?:(?:https?:)?//www\.ebay\.[a-z.]+)?'
                       r'/itm/(?:[^"/?]+/)?(\d+)[^"]*"', chunk)
        tm = re.search(r's-item__title[^>]*>(.{0,500}?)</(?:div|h3)>', chunk,
                       re.S)
        pm = re.search(r's-item__price[^>]*>(.{0,160}?)</span>', chunk, re.S)
        if not (lm and tm):
            continue
        def _txt(frag):
            return re.sub(r"\s+", " ",
                          re.sub(r"<[^>]+>", " ", frag or "")).strip()
        title = re.sub(r"^New Listing\s*", "", _txt(tm.group(1)), flags=re.I)
        if not title or title.lower() == "shop on ebay":
            continue
        url = f"https://www.ebay.com/itm/{lm.group(1)}"
        price = _txt(pm.group(1)).replace("&nbsp;", " ").strip() if pm else ""
        out.append((title.replace("&amp;", "&"), price, url))
    return out

def _fold(s):
    """Accent- and case-insensitive text for MATCHING only ('Sémag' ->
    'semag'). Display always keeps the real accented name. Letters that do NOT
    decompose to ASCII under NFKD (ø, æ, ł, ß, ...) are mapped explicitly
    first — otherwise NFKD+ascii would strip them entirely and Scandinavian /
    Polish words like 'nøkkel' or 'kłódka' would fold to 'nkkel' / 'kodka' and
    never match."""
    s = str(s or "")
    for a, b in (("ø", "o"), ("Ø", "o"), ("æ", "ae"), ("Æ", "ae"),
                 ("œ", "oe"), ("Œ", "oe"), ("ß", "ss"), ("ł", "l"),
                 ("Ł", "l"), ("đ", "d"), ("Đ", "d"), ("ð", "d"), ("Ð", "d"),
                 ("þ", "th"), ("Þ", "th"), ("ı", "i")):
        if a in s:
            s = s.replace(a, b)
    import unicodedata
    s = unicodedata.normalize("NFKD", s)
    return s.encode("ascii", "ignore").decode("ascii").lower()

def _log_diag(msg):
    """Persist a scrape diagnostic to the saved log (lockhunter.log) so a
    zero-result eBay/marketplace probe can be diagnosed after the fact —
    per-probe status_cb messages only reach the live status bar, not the file."""
    try:
        log(msg)
    except Exception:
        pass


def _ebay_session():
    """A persistent session for eBay scraping. curl_cffi (real Chrome TLS
    fingerprint) when bundled — plain-requests TLS is a big reason eBay 403s —
    else a requests session carrying browser headers. Cookies persist across
    calls on the returned session, which is what the homepage warm-up needs."""
    try:
        from curl_cffi import requests as _cffi
        return _cffi.Session(impersonate="chrome"), True
    except Exception:
        s = requests.Session()
        s.headers.update(_SCRAPE_HEADERS)
        return s, False


def _ebay_get(sess, url, dom):
    """Fetch an eBay search URL on a warmed session, with headers that look
    like a same-site navigation from the homepage (Referer + Sec-Fetch), which
    matters to eBay's bot check. Returns (status, text)."""
    hdrs = {
        "Referer": f"https://{dom}/",
        "Accept": ("text/html,application/xhtml+xml,application/xml;q=0.9,"
                   "image/avif,image/webp,*/*;q=0.8"),
        "Accept-Language": "en-US,en;q=0.9,de;q=0.7,fr;q=0.6",
        "Sec-Fetch-Dest": "document", "Sec-Fetch-Mode": "navigate",
        "Sec-Fetch-Site": "same-origin", "Sec-Fetch-User": "?1",
        "Upgrade-Insecure-Requests": "1",
    }
    r = sess.get(url, headers=hdrs, timeout=_FB_TIMEOUT)
    st = getattr(r, "status_code", 0) or 0
    try:
        text = r.text or ""
    except Exception:
        text = ""
    return st, text


def search_ebay_direct(lock_name, status_cb, deep=False, origin=""):
    """Probe eBay's own search pages directly (origin-country domain first,
    then global ones) and return listing dicts whose titles match every word
    of the lock name. Deterministic — doesn't depend on the AI or on search
    engines having indexed the listing. Failures degrade to []."""
    domains = []
    od = ORIGIN_EBAY.get(origin)
    if od:
        domains.append(od)
    for d in _EBAY_GLOBAL:
        if d not in domains:
            domains.append(d)
    domains = domains[:5 if deep else 3]
    # Query variants (brand aliases like CAWI->Wittkopp, size forms, a native
    # lock word) fire on the FIRST/origin domain — the one most likely to hold
    # the listing — while the other domains get the exact term only, so recall
    # goes up without multiplying requests across every country.
    variants = _query_variants(lock_name, origin, deep)
    plan = []          # (domain, query_string, is_variant)
    for i, dom in enumerate(domains):
        if i == 0:
            for v in variants:
                plan.append((dom, v, v != lock_name))
        else:
            plan.append((dom, lock_name, False))
    tokens = _lock_tokens(lock_name)
    out, seen = [], set()
    # eBay 403s a cold /sch/ request that carries no session cookies. A real
    # browser gets those cookies by loading eBay first, so we do the same:
    # ONE persistent session (browser-TLS via curl_cffi when available), warmed
    # up on each domain's homepage before its first search. Cookies from the
    # warm-up ride along on the search requests.
    sess, _tls = _ebay_session()
    warmed = set()
    import time as _time
    for dom, qstr, is_var in plan:
        q = urllib.parse.quote_plus(qstr.strip())
        url = (f"https://{dom}/sch/i.html?_nkw={q}"
               f"&_ipg={240 if deep else 120}")
        try:
            status_cb(f"Probing {dom} directly"
                      + (f" ({qstr})…" if is_var else "…"))
            if dom not in warmed:
                warmed.add(dom)
                try:
                    sess.get(f"https://{dom}/", timeout=_FB_TIMEOUT)
                    _time.sleep(0.4)
                except Exception:
                    pass
            st, text = _ebay_get(sess, url, dom)
            # Persistent diagnostic so a zero-result eBay is DEBUGGABLE from the
            # saved log: status, page size, whether the HTML even carries /itm/
            # links (if not, it's a wall or a JS-only shell, not a parser bug),
            # and whether it looks bot-walled.
            kb = len(text) // 1024
            has_itm = "/itm/" in text
            blocked = _looks_blocked(st, text)
            diag = (f"eBay diag {dom} '{qstr}': HTTP {st or 'none'}, {kb}KB, "
                    f"itm_links={'yes' if has_itm else 'NO'}, "
                    f"blocked={'yes' if blocked else 'no'}")
            if st != 200:
                _log_diag(diag)
                status_cb(f"  {dom}: HTTP {st or 'no response'} "
                          "on search page")
                continue
            items = _parse_ebay_search(text)
            _log_diag(diag + f", parsed={len(items)}")
            # If eBay served listings (itm_links=yes) but the parser got none,
            # dump a cleaned sample of the card markup around the first /itm/
            # link so the exact structure is visible in the saved log.
            if not items and has_itm:
                _im = re.search(r'/itm/\d{9,}', text)
                if _im:
                    _s = text[max(0, _im.start() - 400): _im.start() + 700]
                    _s = re.sub(r"\s+", " ", _s.replace("<", " <"))
                    _log_diag("eBay markup sample: " + _s[:900])
            matched = 0
            for title, price, link in items:
                if tokens and not _match_tokens(title, tokens):
                    continue
                if (_LOCK_CONTEXT_FILTER and not _search_has_lock_word(lock_name)
                        and not _has_lock_context(title)):
                    continue
                if link in seen:
                    continue
                seen.add(link)
                matched += 1
                out.append({
                    "title": title, "price": price, "currency": "",
                    "condition": "used",
                    "site": "eBay" if dom == "www.ebay.com"
                            else "eBay (" + dom.replace("www.ebay.", "") + ")",
                    "url": link, "location": "",
                    "shipping": "unknown",
                    "notes": "found via direct eBay search",
                    "preverified": True,
                })
            status_cb(f"  {dom}: {len(items)} result(s) parsed, "
                      f"{matched} matched")
        except Exception as ex:
            status_cb(f"  {dom}: probe error {ex}")
            continue
    return out


# ============================================================================
# eBay Browse API (official, no scraping) — one keyset covers all marketplaces
# ----------------------------------------------------------------------------
# Far more reliable than parsing HTML: no bot-walls, structured fields, and a
# single OAuth app credential (Client ID + Secret) queries EVERY eBay country
# via the X-EBAY-C-MARKETPLACE-ID header. Credentials come from the environment
# (EBAY_CLIENT_ID / EBAY_CLIENT_SECRET); if unset, this module is simply
# inert and the app falls back to the scrape probe. Nothing is hardcoded.
EBAY_OAUTH_URL = "https://api.ebay.com/identity/v1/oauth2/token"
EBAY_BROWSE_URL = "https://api.ebay.com/buy/browse/v1/item_summary/search"
EBAY_SCOPE = "https://api.ebay.com/oauth/api_scope"

# origin country -> eBay marketplace id (the API's equivalent of the domain).
ORIGIN_EBAY_MKT = {
    "Germany": "EBAY_DE", "Austria": "EBAY_AT", "Switzerland/Germany": "EBAY_CH",
    "Italy": "EBAY_IT", "France": "EBAY_FR", "Spain": "EBAY_ES",
    "Netherlands": "EBAY_NL", "UK": "EBAY_GB", "USA": "EBAY_US",
    "USA/UK": "EBAY_US", "Australia": "EBAY_AU", "Ireland": "EBAY_IE",
    "Belgium": "EBAY_BE", "Poland": "EBAY_PL",
}
_EBAY_MKT_GLOBAL = ["EBAY_US", "EBAY_GB", "EBAY_DE"]
# marketplace id -> currency + a human suffix for the site label
_EBAY_MKT_META = {
    "EBAY_US": ("USD", ""), "EBAY_GB": ("GBP", "co.uk"), "EBAY_DE": ("EUR", "de"),
    "EBAY_FR": ("EUR", "fr"), "EBAY_IT": ("EUR", "it"), "EBAY_ES": ("EUR", "es"),
    "EBAY_AT": ("EUR", "at"), "EBAY_CH": ("CHF", "ch"), "EBAY_NL": ("EUR", "nl"),
    "EBAY_AU": ("AUD", "com.au"), "EBAY_IE": ("EUR", "ie"),
    "EBAY_BE": ("EUR", "be"), "EBAY_PL": ("PLN", "pl"),
}

_ebay_token_cache = {"token": None, "exp": 0}

def _ebay_credentials():
    cid = os.environ.get("EBAY_CLIENT_ID", "").strip()
    secret = os.environ.get("EBAY_CLIENT_SECRET", "").strip()
    return (cid, secret) if (cid and secret) else (None, None)

def ebay_api_enabled():
    return _ebay_credentials()[0] is not None

def _ebay_token(status_cb=lambda s: None):
    """Fetch/cache an application OAuth token (client-credentials grant).
    Cached until ~1 min before expiry. Returns None if creds missing or the
    token call fails (caller then just skips the API)."""
    import time as _t
    cid, secret = _ebay_credentials()
    if not cid:
        return None
    now = _t.time()
    if _ebay_token_cache["token"] and now < _ebay_token_cache["exp"] - 60:
        return _ebay_token_cache["token"]
    try:
        import base64
        basic = base64.b64encode(f"{cid}:{secret}".encode()).decode()
        r = requests.post(
            EBAY_OAUTH_URL,
            headers={"Authorization": f"Basic {basic}",
                     "Content-Type": "application/x-www-form-urlencoded"},
            data={"grant_type": "client_credentials", "scope": EBAY_SCOPE},
            timeout=30)
        if r.status_code != 200:
            status_cb(f"  eBay API auth failed: HTTP {r.status_code}")
            return None
        j = r.json()
        tok = j.get("access_token")
        if not tok:
            return None
        _ebay_token_cache["token"] = tok
        _ebay_token_cache["exp"] = now + int(j.get("expires_in", 7200))
        return tok
    except Exception as ex:
        status_cb(f"  eBay API auth error: {ex}")
        return None

def _parse_ebay_api_items(payload, marketplace):
    """Turn a Browse API item_summary/search JSON response into listing tuples
    (title, price_str, url, condition, location)."""
    out = []
    for it in (payload or {}).get("itemSummaries", []) or []:
        title = (it.get("title") or "").strip()
        url = it.get("itemWebUrl") or it.get("itemHref") or ""
        price = ""
        pr = it.get("price") or {}
        if pr.get("value"):
            cur = pr.get("currency") or _EBAY_MKT_META.get(marketplace, ("", ""))[0]
            price = f"{pr['value']} {cur}".strip()
        cond = (it.get("condition") or "").strip().lower() or "used"
        loc = ""
        il = it.get("itemLocation") or {}
        loc = il.get("country") or ""
        img = (it.get("image") or {}).get("imageUrl") or ""
        if not img:
            ti = it.get("thumbnailImages") or []
            if ti:
                img = (ti[0] or {}).get("imageUrl") or ""
        if title and url:
            out.append((title, price, url.split("?")[0], cond, loc, img))
    return out

def search_ebay_api(lock_name, status_cb, deep=False, origin=""):
    """Search eBay via the official Browse API across the origin marketplace
    first, then the global ones. Returns listing dicts shaped exactly like the
    scrape probe (so results merge/dedupe cleanly). Inert if no credentials."""
    tok = _ebay_token(status_cb)
    if not tok:
        return []
    markets = []
    om = ORIGIN_EBAY_MKT.get(origin)
    if om:
        markets.append(om)
    for m in _EBAY_MKT_GLOBAL:
        if m not in markets:
            markets.append(m)
    markets = markets[:5 if deep else 3]
    tokens = _lock_tokens(lock_name)
    limit = 100 if deep else 50
    out, seen = [], set()
    for mkt in markets:
        try:
            status_cb(f"eBay API: searching {mkt}…")
            r = requests.get(
                EBAY_BROWSE_URL,
                headers={"Authorization": f"Bearer {tok}",
                         "X-EBAY-C-MARKETPLACE-ID": mkt,
                         "Content-Type": "application/json"},
                params={"q": lock_name.strip(), "limit": str(limit)},
                timeout=30)
            if r.status_code == 401:
                # token expired mid-run — refresh once and retry this market
                _ebay_token_cache["token"] = None
                tok = _ebay_token(status_cb)
                if not tok:
                    break
                r = requests.get(
                    EBAY_BROWSE_URL,
                    headers={"Authorization": f"Bearer {tok}",
                             "X-EBAY-C-MARKETPLACE-ID": mkt},
                    params={"q": lock_name.strip(), "limit": str(limit)},
                    timeout=30)
            if r.status_code != 200:
                status_cb(f"  {mkt}: HTTP {r.status_code}")
                continue
            items = _parse_ebay_api_items(r.json(), mkt)
            suffix = _EBAY_MKT_META.get(mkt, ("", ""))[1]
            matched = 0
            for title, price, url, cond, loc, img in items:
                if tokens and not _match_tokens(title, tokens):
                    continue
                if (_LOCK_CONTEXT_FILTER and not _search_has_lock_word(lock_name)
                        and not _has_lock_context(title)):
                    continue
                if url in seen:
                    continue
                seen.add(url)
                matched += 1
                out.append({
                    "title": title, "price": price, "currency": "",
                    "condition": cond,
                    "site": "eBay" if not suffix else f"eBay ({suffix})",
                    "url": url, "location": loc, "shipping": "unknown",
                    "notes": "found via eBay API", "image": img,
                    "preverified": True,
                })
            status_cb(f"  {mkt}: {len(items)} item(s), {matched} matched")
        except Exception as ex:
            status_cb(f"  {mkt}: eBay API error {ex}")
            continue
    return out


# ============================================================================
# Additional no-API-key marketplace probes (Catawiki, Leboncoin, Kleinanzeigen)
# ----------------------------------------------------------------------------
# Same idea as the eBay probe: fetch the site's own search-results page and
# parse listings ourselves — zero API credits. These sites are JS-heavy and
# defended, so each parser is best-effort: it prefers structured data embedded
# in the page (JSON-LD, __NEXT_DATA__, itemprop microdata), falls back to
# listing-anchor scraping, and ALWAYS degrades to [] on trouble. They may need
# maintenance when a site changes its markup.
_SCRAPE_HEADERS = {
    "User-Agent": _BROWSER_UA,
    "Accept": ("text/html,application/xhtml+xml,application/xml;q=0.9,"
               "image/avif,image/webp,*/*;q=0.8"),
    "Accept-Language": "en-US,en;q=0.9,de;q=0.7,fr;q=0.7,*;q=0.5",
    "Referer": "https://www.google.com/",
    "Sec-Fetch-Dest": "document", "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "cross-site", "Upgrade-Insecure-Requests": "1",
}

_BLOCK_PHRASES = (
    "pardon our interruption", "are you a human", "are you a robot",
    "captcha", "just a moment", "attention required", "access denied",
    "unusual traffic", "request blocked", "px-captcha",
    "challenge-platform", "enable javascript and cookies",
    "checking your browser",
)

def _looks_blocked(status, text):
    """Heuristic: is this response a bot-wall/challenge rather than results?
    Challenge pages are small; real search pages are big — so the phrase test
    only applies to short responses, which keeps the word 'captcha' in the
    footer of a real 100 KB results page from false-positiving."""
    if status in (403, 407, 429, 503, 509):
        return True
    t = text or ""
    if len(t) < 25000:
        low = t[:8000].lower()
        return any(p in low for p in _BLOCK_PHRASES)
    return False

def _scrape_transports():
    """Ordered fetch transports for scraping: browser-TLS (curl_cffi,
    impersonating Chrome's TLS fingerprint) first when bundled — many
    marketplaces bot-wall plain-requests TLS no matter how good the headers
    are — then standard requests. Without curl_cffi, a second plain slot
    still gives transient failures one retry."""
    out = []
    try:
        from curl_cffi import requests as _cffi

        def _tls_get(url, headers, timeout, _c=_cffi):
            t = timeout[-1] if isinstance(timeout, (tuple, list)) else timeout
            s = _c.Session(impersonate="chrome")
            try:
                return s.get(url, headers=headers, timeout=t or 30)
            finally:
                try:
                    s.close()
                except Exception:
                    pass
        out.append(("browser-TLS", _tls_get))
    except Exception:
        pass

    def _std_get(url, headers, timeout):
        r = requests.get(url, headers=headers, timeout=timeout,
                         allow_redirects=True)
        try:
            ct = (r.headers.get("content-type") or "").lower()
            if "charset" not in ct and \
                    (r.encoding or "").lower() in ("", "iso-8859-1"):
                r.encoding = r.apparent_encoding or "utf-8"
        except Exception:
            pass
        return r
    out.append(("standard", _std_get))
    if len(out) == 1:
        out.append(("standard retry", _std_get))
    return out

def _scrape_get(url, status_cb=lambda s: None, label="", timeout=None,
                headers=None, attempts=2):
    """Hardened GET used by every scraping probe. Browser-TLS transport when
    available, realistic headers, bot-wall detection, and one retry on the
    alternate transport with a short jittered pause. Returns (status, text);
    (0, "") only when every attempt errored. A blocked-everywhere fetch still
    returns the last (status, text) so callers keep their own keep/drop
    policy (the verifier treats walls as keep-unverified, for instance)."""
    import time as _time
    import random as _random
    hdrs = dict(_SCRAPE_HEADERS)
    if headers:
        hdrs.update(headers)
    to = timeout if timeout is not None else PROBE_TIMEOUT
    transports = _scrape_transports()[:max(1, attempts)]
    last = (0, "")
    for i, (tname, getter) in enumerate(transports):
        if i:
            _time.sleep(0.5 + _random.random() * 0.7)
        try:
            r = getter(url, hdrs, to)
        except Exception as ex:
            status_cb(f"  {label or url[:40]}: {tname} fetch error {ex}")
            continue
        st = getattr(r, "status_code", 0) or 0
        try:
            text = r.text or ""
        except Exception:
            text = ""
        last = (st, text)
        if not _looks_blocked(st, text):
            return last
        if i + 1 < len(transports):
            status_cb(f"  {label or url[:40]}: bot-wall via {tname} "
                      f"(HTTP {st}) — retrying with {transports[i+1][0]}…")
            continue
        status_cb(f"  {label or url[:40]}: bot-wall on every transport "
                  f"(HTTP {st})")
    return last

_MP_UA = {
    "User-Agent": _BROWSER_UA,
    "Accept-Language": "en;q=0.9,fr;q=0.8,de;q=0.8,*;q=0.5",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

def _mp_price_near(text, idx, window=240):
    """Find a price-looking token near a position in text."""
    seg = text[idx:idx + window]
    m = re.search(r"((?:€|EUR|£|GBP|\$|US\s*\$)\s?\d[\d.,]*)", seg)
    if m:
        return re.sub(r"\s+", " ", m.group(1)).strip()
    m = re.search(r"(\d[\d.,]*)\s?(?:€|EUR|£|\$)", seg)
    return (m.group(0).strip() if m else "")

def _jsonld_products(html):
    """Yield product-ish dicts from any <script type=ld+json> blocks."""
    out = []
    for m in re.finditer(
            r'<script[^>]+type=["\']application/ld\+json["\'][^>]*>(.*?)</script>',
            html, re.S | re.I):
        raw = m.group(1).strip()
        try:
            data = json.loads(raw)
        except Exception:
            continue
        stack = [data]
        while stack:
            node = stack.pop()
            if isinstance(node, list):
                stack.extend(node)
            elif isinstance(node, dict):
                t = node.get("@type", "")
                t = " ".join(t) if isinstance(t, list) else str(t)
                if "Product" in t or "Offer" in t or "Vehicle" in t:
                    out.append(node)
                for v in node.values():
                    if isinstance(v, (dict, list)):
                        stack.append(v)
    return out

def _offer_price(node):
    off = node.get("offers")
    if isinstance(off, list) and off:
        off = off[0]
    if isinstance(off, dict):
        p = off.get("price") or off.get("lowPrice") or ""
        cur = off.get("priceCurrency") or ""
        if p:
            return f"{p} {cur}".strip()
    return ""

def _norm_sizes(s):
    """Collapse the many ways sellers write a lock size to ONE canonical form
    so matching and querying line up. 83/45, 83-45, 83x45, '83 45' and 8345 all
    become '8345'. Runs on FOLDED text. Only joins a digit group to the NEXT
    group when doing so is unambiguous (both are pure digit runs separated by a
    single size delimiter), so real separate numbers aren't fused."""
    s = re.sub(r'(?<=\d)\s*[/xX\u00d7\u2715-]\s*(?=\d)', '', s)   # 83/45 -> 8345
    return s

# Brand aliases: every set is a group of names that mean the SAME maker/line.
# Used two ways — at match time (any alias in the title satisfies that brand
# token) and at query time (each alias becomes an extra search variant). Keep
# entries FOLDED (lowercase, no accents) since that's how they're compared.
_BRAND_ALIASES = [
    {"mul-t-lock", "mult-lock", "multlock", "mtl", "mul t lock"},
    {"cawi", "wittkopp", "carl wittkopp"},
    {"s&g", "s and g", "sargent and greenleaf", "sargent greenleaf", "sng"},
    {"burg-wachter", "burg wachter", "burgwachter", "burg"},
    {"zeiss ikon", "zi ikon", "zeiss", "zi"},
    {"assa abloy", "assa"},
    {"dorma kaba", "dormakaba", "kaba"},
    {"corbin russwin", "corbin", "russwin"},
    {"abloy protec", "abloy protec2", "protec2", "protec 2"},
    {"emhart", "emha"},
]
_ALIAS_INDEX = {}
for _grp in _BRAND_ALIASES:
    for _a in _grp:
        _ALIAS_INDEX.setdefault(_a, set()).update(_grp)


def _alias_variants(term):
    """Yield query variants of a term where a recognized brand alias is
    swapped for its siblings (e.g. 'Mul-T-Lock C10' -> 'MTL C10'). The longest
    matching alias phrase wins so multi-word brands substitute cleanly. Capped
    to keep request counts sane."""
    low = _fold(term)
    out = []
    for alias in sorted(_ALIAS_INDEX, key=len, reverse=True):
        if alias and alias in low:
            for sib in _ALIAS_INDEX[alias]:
                if sib == alias:
                    continue
                variant = re.sub(re.escape(alias), sib, low, count=1)
                variant = re.sub(r"\s+", " ", variant).strip()
                if variant and variant not in out:
                    out.append(variant)
            break
    return out[:4]


_ALIAS_CANON = []      # (alias_phrase, canonical_single_token), longest first
for _grp in _BRAND_ALIASES:
    _alnum = [a for a in _grp if a.isalnum()]
    _canon = min(_alnum or _grp, key=len)
    _canon = re.sub(r"[^a-z0-9]+", "", _canon) or "x"
    for _al in _grp:
        _ALIAS_CANON.append((_al, _canon))
_ALIAS_CANON.sort(key=lambda p: len(p[0]), reverse=True)


def _canon_aliases(folded):
    """Replace any brand-alias spelling with one canonical token so multi-word
    brands (Mul-T-Lock -> mtl, Sargent and Greenleaf -> sng) survive being
    split into tokens. Boundary-guarded so a short canon can't match inside an
    unrelated word."""
    for alias, canon in _ALIAS_CANON:
        folded = re.sub(r'(?<![a-z0-9])' + re.escape(alias) + r'(?![a-z0-9])',
                        canon, folded)
    return folded


def _lock_tokens(name):
    """Canonical match tokens for a lock name: accent-folded, brand-aliases
    canonicalized, sizes glued (83/45 -> 8345), split on non-alphanumerics,
    single chars dropped. The one place token lists are made, so every probe
    matches identically."""
    prepared = _norm_sizes(_canon_aliases(_fold(name)))
    return [t for t in re.split(r"[^a-z0-9]+", prepared) if len(t) > 1]


def _match_tokens(title, tokens):
    """True if EVERY search token appears in the title, size-tolerant and
    alias-aware: the title is put through the same fold -> canon-alias ->
    size-glue pipeline as the tokens, so '8345' matches '83/45' and a title
    saying 'MTL' satisfies a 'Mul-T-Lock' token. Exact-model discipline holds —
    each token must still be accounted for."""
    if not title or not tokens:
        return False
    tl = _norm_sizes(_canon_aliases(_fold(title)))
    for tok in tokens:
        tk = _norm_sizes(_canon_aliases(tok))
        if tk in tl or tok in tl:
            continue
        return False
    return True


def _query_variants(term, origin="", deep=False):
    """The set of search strings to actually fire for one component: the term
    as written, brand-alias swaps, and (on origin-native sites) a native-
    language 'brand + lock word' form so German/French/etc. sellers' titles
    surface. Deduped (case-insensitively) and capped. The exact term is always
    first so precise matches lead."""
    variants = [term]
    seen = {_fold(term)}
    for v in _alias_variants(term):
        if v not in seen:
            variants.append(v)
            seen.add(v)
    # size-glued form: 'ABUS 83/45' also searches 'abus 8345' (eBay treats the
    # slash form and the glued form as different queries, so both are needed).
    ns = _norm_sizes(_fold(term))
    if ns != _fold(term) and ns not in seen:
        variants.append(ns)
        seen.add(ns)
    native = _NATIVE_LOCK_WORD.get(origin)
    if native:
        brand = term.split()[0] if term.split() else term
        cand = f"{brand} {native}"
        if _fold(cand) not in seen:
            variants.append(cand)
            seen.add(_fold(cand))
    cap = 5 if deep else 4
    return variants[:cap]


# One representative native lock word per origin, for query variants on that
# origin's own marketplaces (sellers there rarely write English "lock").
_NATIVE_LOCK_WORD = {
    "Germany": "schloss", "Austria": "schloss",
    "Switzerland/Germany": "schloss",
    "France": "serrure", "Italy": "serratura", "Spain": "cerradura",
    "Netherlands": "slot", "Poland": "zamek",
    "Sweden": "las", "Denmark/Sweden": "las", "Norway": "las",
    "Finland": "lukko",
}

# Lock-context words across the languages of the marketplaces we probe. A
# listing whose title contains at least one of these is very likely an actual
# lock/key item — used to filter out same-name non-lock junk (e.g. a "Sol"
# lock model vs. "Sol" sunglasses). Folded (accent/case-stripped) for matching.
_LOCK_CONTEXT_WORDS = {
    # English
    "lock", "locks", "padlock", "padlocks", "cylinder", "cylinders", "deadbolt",
    "latch", "key", "keys", "keyed", "keyway", "lockset", "mortise", "mortice",
    "rim lock", "cam lock", "keycard", "locking", "locksmith", "safe",
    # French
    "serrure", "serrures", "cadenas", "cle", "cles", "clef", "clefs", "verrou",
    "barillet", "cylindre", "gache", "goupille",
    # German
    "schloss", "schloesser", "vorhangeschloss", "vorhangschloss", "zylinder",
    "schluessel", "schliesszylinder", "riegel", "schliessanlage",
    # Spanish / Portuguese
    "cerradura", "cerraduras", "candado", "candados", "llave", "llaves",
    "bombin", "bombillo", "cilindro", "cadeado", "fechadura", "chave", "canhao",
    "trinco",
    # Italian
    "serratura", "lucchetto", "chiave", "chiavi", "cilindro", "catenaccio",
    # Dutch
    "slot", "sloten", "hangslot", "cilinder", "sleutel", "grendel",
    # Nordic
    "las", "hanglas", "nyckel", "nokkel", "sylinter", "lukko", "avain",
    "sylinteri", "cylinder",
    # Nordic additions: Danish/Norwegian key + padlock forms (as they fold),
    # Finnish padlock, Danish "security lock/cylinder"
    "nogle", "haengelas", "hengelas", "hanglaas", "riippulukko",
    "sikkerhedslas", "sikkerhedscylinder", "hengelaas", "kombinasjonslas",
    # Polish / Czech / Slovak / others
    "zamek", "klodka", "klucz", "wkladka", "wkladki", "bebenkowa", "zamok",
    "kljuc", "brava", "lokot", "kilit", "anahtar", "zamki", "cilindar",
    # Turkish
    "kilidi", "silindir",
}
# a few short ones need word-boundary care to avoid matching inside other words
_LOCK_WORDS_STRICT = {"key", "las", "slot", "safe", "cle", "cles", "brava"}

# Cyrillic lock words (Bulgarian/Russian/etc.). _fold() strips Cyrillic to
# nothing, so these are matched against the RAW title instead. Helps listings
# that pair a Latin brand ("Ruko") with a Cyrillic description.
_LOCK_WORDS_CYRILLIC = (
    "ключалка", "катинар", "цилиндър", "цилиндр", "секрет", "брава",
    "замок", "замка", "патрон", "заключв", "ключ",
)

def _has_lock_context(title):
    """True if the title looks like an actual lock/key listing (contains a
    lock-context word in any supported language)."""
    raw = str(title or "").lower()
    for w in _LOCK_WORDS_CYRILLIC:      # Cyrillic first (fold would drop it)
        if w in raw:
            return True
    tl = _fold(title)
    if not tl:
        return False
    words = set(re.split(r"[^a-z0-9]+", tl))
    for w in _LOCK_CONTEXT_WORDS:
        if " " in w or "-" in w:
            if w.replace("-", " ") in tl.replace("-", " "):
                return True
        elif w in _LOCK_WORDS_STRICT:
            if w in words:          # exact word only (avoid 'key' in 'monkey')
                return True
        elif w in words or w in tl:
            return True
    return False

# Whether the lock-context filter is active. On by default; the UI can toggle
# it (a checkbox) so the user can loosen filtering if a terse-titled real
# listing gets dropped.
_LOCK_CONTEXT_FILTER = True

def _search_has_lock_word(lock_name):
    """If the user's own search term already includes a lock word (e.g. they
    typed 'Abloy padlock'), the token match already constrains to locks, so we
    skip the extra context filter to avoid over-dropping."""
    return _has_lock_context(lock_name)

def _generic_probe(site_name, url, parse_fn, lock_name, status_cb, note):
    """Fetch one search URL, run its parser, token-filter, shape into listing
    dicts. Best-effort; returns [] on any failure."""
    tokens = _lock_tokens(lock_name)
    out, seen = [], set()
    try:
        status_cb(f"Probing {site_name} directly…")
        st, text = _scrape_get(url, status_cb, label=site_name,
                               headers=_MP_UA)
        if st != 200:
            status_cb(f"  {site_name}: HTTP {st or 'no response'} "
                      "on search page")
            return []
        items = parse_fn(text) or []
        matched = 0
        for title, price, link in items:
            if not link or link in seen:
                continue
            if tokens and not _match_tokens(title, tokens):
                continue
            # Lock-context filter: drop same-name non-lock items (e.g. a "Sol"
            # lock model vs. "Sol" sunglasses) when enabled. Skipped if the
            # search term itself already contains a lock word (then the tokens
            # do the work and we don't want to double-filter).
            if (_LOCK_CONTEXT_FILTER and not _search_has_lock_word(lock_name)
                    and not _has_lock_context(title)):
                continue
            seen.add(link)
            matched += 1
            out.append({
                "title": title.strip(), "price": price, "currency": "",
                "condition": "used", "site": site_name, "url": link,
                "location": "", "shipping": "unknown", "notes": note,
                "preverified": True,
            })
        status_cb(f"  {site_name}: {len(items)} parsed, {matched} matched")
    except Exception as ex:
        status_cb(f"  {site_name}: probe error {ex}")
        return []
    return out

# ---- Catawiki (auctions; global) ------------------------------------------
def _parse_catawiki(html):
    out = []
    # Catawiki renders lots as <a href="/en/l/12345-...">; titles in the anchor
    # or an adjacent heading. Prefer JSON-LD when present.
    for node in _jsonld_products(html):
        name = node.get("name") or ""
        urlp = node.get("url") or ""
        if name and urlp:
            out.append((name, _offer_price(node), urlp if urlp.startswith("http")
                        else "https://www.catawiki.com" + urlp))
    if out:
        return out
    for m in re.finditer(r'<a[^>]+href="(/[a-z]{2}/l/(\d+)[^"]*)"[^>]*>(.*?)</a>',
                         html, re.S | re.I):
        href, _id, inner = m.group(1), m.group(2), m.group(3)
        title = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", inner)).strip()
        # aria-label usually holds the real lot title; prefer it when the
        # visible anchor text is empty or too short to be a title.
        am = re.search(r'aria-label="([^"]{5,200})"', m.group(0))
        if am and len(am.group(1).strip()) > len(title):
            title = am.group(1).strip()
        if not title:
            continue
        price = _mp_price_near(html, m.end())
        out.append((title, price, "https://www.catawiki.com" + href))
        if len(out) >= 50:
            break
    return out

def search_catawiki_direct(lock_name, status_cb, deep=False):
    q = urllib.parse.quote_plus(lock_name.strip())
    url = f"https://www.catawiki.com/en/s?q={q}"
    return _generic_probe("Catawiki", url, _parse_catawiki, lock_name,
                          status_cb, "found via direct Catawiki search")

# ---- Leboncoin (France) ---------------------------------------------------
def _parse_leboncoin(html):
    out = []
    # Leboncoin is a Next.js app; listings live in __NEXT_DATA__ as JSON with
    # list_id + subject + price. Fall back to /<id>.htm anchors.
    m = re.search(r'<script id="__NEXT_DATA__"[^>]*>(.*?)</script>', html, re.S)
    if m:
        try:
            data = json.loads(m.group(1))
            stack = [data]
            while stack:
                node = stack.pop()
                if isinstance(node, list):
                    stack.extend(node)
                elif isinstance(node, dict):
                    if node.get("list_id") and (node.get("subject") or node.get("subject_text")):
                        lid = node["list_id"]
                        title = node.get("subject") or node.get("subject_text") or ""
                        price = ""
                        pv = node.get("price")
                        if isinstance(pv, list) and pv:
                            price = f"{pv[0]} EUR"
                        elif isinstance(pv, (int, float, str)) and pv:
                            price = f"{pv} EUR"
                        url = node.get("url") or f"https://www.leboncoin.fr/ad/{lid}"
                        if not url.startswith("http"):
                            url = "https://www.leboncoin.fr" + url
                        out.append((title, price, url))
                    for v in node.values():
                        if isinstance(v, (dict, list)):
                            stack.append(v)
        except Exception:
            pass
    if out:
        # de-dupe by url
        seen, uniq = set(), []
        for t in out:
            if t[2] not in seen:
                seen.add(t[2]); uniq.append(t)
        return uniq
    for m in re.finditer(r'<a[^>]+href="(https://www\.leboncoin\.fr/(?:ad|[a-z_]+)/(\d+)[^"]*)"[^>]*>(.*?)</a>',
                         html, re.S | re.I):
        href, _id, inner = m.group(1), m.group(2), m.group(3)
        title = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", inner)).strip()
        if title:
            out.append((title, _mp_price_near(html, m.end()), href.split("?")[0]))
        if len(out) >= 50:
            break
    return out

def search_leboncoin_direct(lock_name, status_cb, deep=False):
    q = urllib.parse.quote_plus(lock_name.strip())
    url = f"https://www.leboncoin.fr/recherche?text={q}"
    return _generic_probe("Leboncoin", url, _parse_leboncoin, lock_name,
                          status_cb, "found via direct Leboncoin search")

# ---- Kleinanzeigen (Germany) ----------------------------------------------
def _parse_kleinanzeigen(html):
    out = []
    # Kleinanzeigen server-renders article cards: <article ... data-href="/s-anzeige/..">
    # with an <h2> title and a price <p class="aditem-main--middle--price...">.
    for m in re.finditer(r'<article[^>]*\bdata-href="(/s-anzeige/[^"]+)"(.*?)</article>',
                         html, re.S | re.I):
        href, inner = m.group(1), m.group(2)
        tm = re.search(r'<h2[^>]*>(.*?)</h2>', inner, re.S)
        title = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", tm.group(1))).strip() if tm else ""
        am = re.search(r'aria-label="([^"]{5,200})"', inner)
        if am and len(am.group(1).strip()) > len(title):
            title = am.group(1).strip()
        if not title:
            continue
        pm = re.search(r'price[^>]*>(.*?)<', inner, re.S | re.I)
        price = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", pm.group(1))).strip() if pm else ""
        url = "https://www.kleinanzeigen.de" + href.split("?")[0]
        out.append((title, price, url))
        if len(out) >= 50:
            break
    if out:
        return out
    for m in re.finditer(r'<a[^>]+href="(/s-anzeige/[^"]+)"[^>]*>(.*?)</a>', html, re.S | re.I):
        href, inner = m.group(1), m.group(2)
        title = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", inner)).strip()
        if title:
            out.append((title, _mp_price_near(html, m.end()),
                        "https://www.kleinanzeigen.de" + href.split("?")[0]))
        if len(out) >= 50:
            break
    return out

def search_kleinanzeigen_direct(lock_name, status_cb, deep=False):
    q = urllib.parse.quote_plus(lock_name.strip())
    url = f"https://www.kleinanzeigen.de/s-suchanfrage.html?keywords={q}"
    return _generic_probe("Kleinanzeigen", url, _parse_kleinanzeigen, lock_name,
                          status_cb, "found via direct Kleinanzeigen search")

# Which extra probes to run for a given origin. Catawiki is global; Leboncoin
# is France-centric; Kleinanzeigen is Germany/Austria/Switzerland-centric —
# but all three run regardless (rare locks travel), origin just affects order.
# ---- Delcampe (collectibles; global) --------------------------------------
def _parse_delcampe(html):
    out = []
    # Delcampe item links look like /en_US/collectibles/.../<id>.html ; titles
    # in the anchor or its img alt, price in an adjacent .price/.item-price.
    for node in _jsonld_products(html):
        name = node.get("name") or ""
        urlp = node.get("url") or ""
        if name and urlp:
            out.append((name, _offer_price(node),
                        urlp if urlp.startswith("http")
                        else "https://www.delcampe.net" + urlp))
    if out:
        return out
    for m in re.finditer(
            r'<a[^>]+href="(https://www\.delcampe\.net/[^"]*?/(\d+)\.html)"[^>]*>(.*?)</a>',
            html, re.S | re.I):
        href, _id, inner = m.group(1), m.group(2), m.group(3)
        title = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", inner)).strip()
        am = re.search(r'(?:aria-label|title|alt)="([^"]{5,200})"', m.group(0))
        if am and len(am.group(1).strip()) > len(title):
            title = am.group(1).strip()
        if not title:
            continue
        out.append((title, _mp_price_near(html, m.end()), href.split("?")[0]))
        if len(out) >= 50:
            break
    return out

def search_delcampe_direct(lock_name, status_cb, deep=False):
    q = urllib.parse.quote_plus(lock_name.strip())
    url = f"https://www.delcampe.net/en_US/collectibles/search?term={q}"
    return _generic_probe("Delcampe", url, _parse_delcampe, lock_name,
                          status_cb, "found via direct Delcampe search")

# ---- Ricardo (Switzerland) ------------------------------------------------
def _parse_ricardo(html):
    out = []
    # Ricardo is a Next.js app: articles in __NEXT_DATA__ with id/title/
    # buyNowPrice, and item URLs like /de/a/<slug>-<id>/.
    m = re.search(r'<script id="__NEXT_DATA__"[^>]*>(.*?)</script>', html, re.S)
    if m:
        try:
            data = json.loads(m.group(1))
            stack, seen = [data], set()
            while stack:
                node = stack.pop()
                if isinstance(node, list):
                    stack.extend(node)
                elif isinstance(node, dict):
                    aid = node.get("id") or node.get("articleId")
                    title = node.get("title") or node.get("name") or ""
                    if aid and title and isinstance(title, str):
                        price = ""
                        for key in ("buyNowPrice", "startPrice", "price",
                                    "currentBidPrice"):
                            pv = node.get(key)
                            if isinstance(pv, dict):
                                pv = pv.get("amount") or pv.get("value")
                            if pv:
                                price = f"{pv} CHF"; break
                        slug = re.sub(r"[^a-z0-9]+", "-", _fold(title)).strip("-")[:60]
                        url = f"https://www.ricardo.ch/de/a/{slug}-{aid}/"
                        if url not in seen:
                            seen.add(url)
                            out.append((title, price, url))
                    for v in node.values():
                        if isinstance(v, (dict, list)):
                            stack.append(v)
        except Exception:
            pass
    if out:
        return out
    for m in re.finditer(r'<a[^>]+href="(/[a-z]{2}/a/[^"]*?-(\d+)/?)"[^>]*>(.*?)</a>',
                         html, re.S | re.I):
        href, _id, inner = m.group(1), m.group(2), m.group(3)
        title = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", inner)).strip()
        am = re.search(r'(?:aria-label|title)="([^"]{5,200})"', m.group(0))
        if am and len(am.group(1).strip()) > len(title):
            title = am.group(1).strip()
        if not title:
            continue
        out.append((title, _mp_price_near(html, m.end()),
                    "https://www.ricardo.ch" + href))
        if len(out) >= 50:
            break
    return out

def search_ricardo_direct(lock_name, status_cb, deep=False):
    q = urllib.parse.quote_plus(lock_name.strip())
    url = f"https://www.ricardo.ch/de/s/{q}"
    return _generic_probe("Ricardo", url, _parse_ricardo, lock_name,
                          status_cb, "found via direct Ricardo search")

# ---- Marktplaats (Netherlands) --------------------------------------------
def _parse_marktplaats(html):
    out = []
    # Marktplaats embeds results in a __NEXT_DATA__-like JSON and also renders
    # <a href="/v/.../m<id>-..."> anchors with a title span + price.
    for m in re.finditer(r'<script[^>]*>(.*?"itemId".*?)</script>', html, re.S):
        try:
            data = json.loads(re.sub(r'^[^{]*', '', m.group(1)))
        except Exception:
            continue
        stack = [data]
        while stack:
            node = stack.pop()
            if isinstance(node, list):
                stack.extend(node)
            elif isinstance(node, dict):
                iid = node.get("itemId") or node.get("id")
                title = node.get("title") or ""
                if iid and title and isinstance(title, str):
                    price = ""
                    pv = node.get("priceInfo") or node.get("price")
                    if isinstance(pv, dict):
                        cents = pv.get("priceCents") or pv.get("amount")
                        if isinstance(cents, (int, float)):
                            price = f"{cents/100:.2f} EUR"
                    url = node.get("vipUrl") or node.get("url") or ""
                    if url and not url.startswith("http"):
                        url = "https://www.marktplaats.nl" + url
                    if url:
                        out.append((title, price, url))
                for v in node.values():
                    if isinstance(v, (dict, list)):
                        stack.append(v)
        if out:
            break
    if out:
        seen, uniq = set(), []
        for t in out:
            if t[2] not in seen:
                seen.add(t[2]); uniq.append(t)
        return uniq
    for m in re.finditer(r'<a[^>]+href="(/v/[^"]*?/m?(\d+)[^"]*)"[^>]*>(.*?)</a>',
                         html, re.S | re.I):
        href, _id, inner = m.group(1), m.group(2), m.group(3)
        title = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", inner)).strip()
        am = re.search(r'(?:aria-label|title)="([^"]{5,200})"', m.group(0))
        if am and len(am.group(1).strip()) > len(title):
            title = am.group(1).strip()
        if not title:
            continue
        out.append((title, _mp_price_near(html, m.end()),
                    "https://www.marktplaats.nl" + href.split("?")[0]))
        if len(out) >= 50:
            break
    return out

def search_marktplaats_direct(lock_name, status_cb, deep=False):
    q = urllib.parse.quote_plus(lock_name.strip())
    url = f"https://www.marktplaats.nl/q/{q}/"
    return _generic_probe("Marktplaats", url, _parse_marktplaats, lock_name,
                          status_cb, "found via direct Marktplaats search")

# ---- Tradera (Sweden) -----------------------------------------------------
def _parse_tradera(html):
    out = []
    # Tradera exposes JSON-LD ItemList and renders /item/<cat>/<id>/<slug>.
    for node in _jsonld_products(html):
        name = node.get("name") or ""
        urlp = node.get("url") or ""
        if name and urlp:
            out.append((name, _offer_price(node),
                        urlp if urlp.startswith("http")
                        else "https://www.tradera.com" + urlp))
    if out:
        return out
    for m in re.finditer(r'<a[^>]+href="(/item/\d+/(\d+)/[^"]*)"[^>]*>(.*?)</a>',
                         html, re.S | re.I):
        href, _id, inner = m.group(1), m.group(2), m.group(3)
        title = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", inner)).strip()
        am = re.search(r'(?:aria-label|title)="([^"]{5,200})"', m.group(0))
        if am and len(am.group(1).strip()) > len(title):
            title = am.group(1).strip()
        if not title:
            continue
        out.append((title, _mp_price_near(html, m.end()),
                    "https://www.tradera.com" + href.split("?")[0]))
        if len(out) >= 50:
            break
    return out

def search_tradera_direct(lock_name, status_cb, deep=False):
    q = urllib.parse.quote_plus(lock_name.strip())
    url = f"https://www.tradera.com/search?q={q}"
    return _generic_probe("Tradera", url, _parse_tradera, lock_name,
                          status_cb, "found via direct Tradera search")

# ---- Subito (Italy) -------------------------------------------------------
def _parse_subito(html):
    out = []
    # Subito renders <a href=".../<slug>-<id>.htm"> cards; titles in an <h2>,
    # price in a span[class*=price]. JSON-LD sometimes present.
    for node in _jsonld_products(html):
        name = node.get("name") or ""
        urlp = node.get("url") or ""
        if name and urlp and urlp.startswith("http"):
            out.append((name, _offer_price(node), urlp))
    if out:
        return out
    for m in re.finditer(r'<a[^>]+href="(https://www\.subito\.it/[^"]*?-(\d+)\.htm)"[^>]*>(.*?)</a>',
                         html, re.S | re.I):
        href, _id, inner = m.group(1), m.group(2), m.group(3)
        tm = re.search(r'<h2[^>]*>(.*?)</h2>', inner, re.S)
        title = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ",
                       tm.group(1) if tm else inner)).strip()
        am = re.search(r'(?:aria-label|title)="([^"]{5,200})"', m.group(0))
        if am and len(am.group(1).strip()) > len(title):
            title = am.group(1).strip()
        if not title:
            continue
        pm = re.search(r'price[^>]*>(.*?)<', inner, re.S | re.I)
        price = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", pm.group(1))).strip() if pm else ""
        if not price:
            price = _mp_price_near(html, m.end())
        out.append((title, price, href.split("?")[0]))
        if len(out) >= 50:
            break
    return out

def search_subito_direct(lock_name, status_cb, deep=False):
    q = urllib.parse.quote_plus(lock_name.strip())
    url = f"https://www.subito.it/annunci-italia/vendita/usato/?q={q}"
    return _generic_probe("Subito", url, _parse_subito, lock_name,
                          status_cb, "found via direct Subito search")

# ---- Blocket (Sweden) -----------------------------------------------------
def _parse_blocket(html):
    out = []
    # Blocket is a Next.js app; ads carry an id + subject + price, and item
    # URLs look like /annons/.../<id>. Try JSON blobs then anchors.
    m = re.search(r'<script id="__NEXT_DATA__"[^>]*>(.*?)</script>', html, re.S)
    if m:
        try:
            data = json.loads(m.group(1))
            stack, seen = [data], set()
            while stack:
                node = stack.pop()
                if isinstance(node, list):
                    stack.extend(node)
                elif isinstance(node, dict):
                    aid = node.get("ad_id") or node.get("id")
                    subj = node.get("subject") or node.get("heading") or ""
                    if aid and subj and isinstance(subj, str):
                        price = ""
                        pv = node.get("price")
                        if isinstance(pv, dict):
                            pv = pv.get("value") or pv.get("amount")
                        if pv:
                            price = f"{pv} SEK"
                        url = node.get("share_url") or node.get("url") or \
                            f"https://www.blocket.se/annons/{aid}"
                        if not url.startswith("http"):
                            url = "https://www.blocket.se" + url
                        if url not in seen:
                            seen.add(url)
                            out.append((subj, price, url))
                    for v in node.values():
                        if isinstance(v, (dict, list)):
                            stack.append(v)
        except Exception:
            pass
    if out:
        return out
    for m in re.finditer(r'<a[^>]+href="(https://www\.blocket\.se/annons/[^"]*?(\d+))"[^>]*>(.*?)</a>',
                         html, re.S | re.I):
        href, _id, inner = m.group(1), m.group(2), m.group(3)
        title = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", inner)).strip()
        am = re.search(r'(?:aria-label|title)="([^"]{5,200})"', m.group(0))
        if am and len(am.group(1).strip()) > len(title):
            title = am.group(1).strip()
        if not title:
            continue
        out.append((title, _mp_price_near(html, m.end()), href.split("?")[0]))
        if len(out) >= 50:
            break
    return out

def search_blocket_direct(lock_name, status_cb, deep=False):
    q = urllib.parse.quote_plus(lock_name.strip())
    url = f"https://www.blocket.se/annonser/hela_sverige?q={q}"
    return _generic_probe("Blocket", url, _parse_blocket, lock_name,
                          status_cb, "found via direct Blocket search")

# ---- Willhaben (Austria) --------------------------------------------------
def _parse_willhaben(html):
    out = []
    # Willhaben embeds results in a __NEXT_DATA__ blob with heading + price +
    # a seo url; item pages end in -<id>/.
    m = re.search(r'<script id="__NEXT_DATA__"[^>]*>(.*?)</script>', html, re.S)
    if m:
        try:
            data = json.loads(m.group(1))
            stack, seen = [data], set()
            while stack:
                node = stack.pop()
                if isinstance(node, list):
                    stack.extend(node)
                elif isinstance(node, dict):
                    # willhaben stores attributes as a list of {name,values}
                    attrs = node.get("attributes")
                    if isinstance(attrs, dict):
                        attrs = attrs.get("attribute")
                    if isinstance(attrs, list) and node.get("id"):
                        amap = {}
                        for a in attrs:
                            if isinstance(a, dict) and a.get("name"):
                                vals = a.get("values") or a.get("value")
                                if isinstance(vals, list) and vals:
                                    amap[a["name"]] = vals[0]
                                elif vals:
                                    amap[a["name"]] = vals
                        title = amap.get("HEADING") or ""
                        seo = amap.get("SEO_URL") or ""
                        price = amap.get("PRICE_FOR_DISPLAY") or amap.get("PRICE") or ""
                        if price and not re.search(r"[€A-Za-z]", str(price)):
                            price = f"{price} EUR"
                        if title and seo:
                            url = seo if seo.startswith("http") else \
                                "https://www.willhaben.at/iad/" + seo.lstrip("/")
                            if url not in seen:
                                seen.add(url)
                                out.append((title, str(price), url))
                    for v in node.values():
                        if isinstance(v, (dict, list)):
                            stack.append(v)
        except Exception:
            pass
    if out:
        return out
    for m in re.finditer(r'<a[^>]+href="(/iad/[^"]*?-(\d+)/?)"[^>]*>(.*?)</a>',
                         html, re.S | re.I):
        href, _id, inner = m.group(1), m.group(2), m.group(3)
        title = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", inner)).strip()
        am = re.search(r'(?:aria-label|title)="([^"]{5,200})"', m.group(0))
        if am and len(am.group(1).strip()) > len(title):
            title = am.group(1).strip()
        if not title:
            continue
        out.append((title, _mp_price_near(html, m.end()),
                    "https://www.willhaben.at" + href.split("?")[0]))
        if len(out) >= 50:
            break
    return out

def search_willhaben_direct(lock_name, status_cb, deep=False):
    q = urllib.parse.quote_plus(lock_name.strip())
    url = f"https://www.willhaben.at/iad/kaufen-und-verkaufen/marktplatz?keyword={q}"
    return _generic_probe("Willhaben", url, _parse_willhaben, lock_name,
                          status_cb, "found via direct Willhaben search")

# ---- Etsy (vintage/handmade; global) --------------------------------------
def _parse_etsy(html):
    out = []
    # Etsy exposes JSON-LD ItemList of products, and renders /listing/<id>/slug
    # anchors otherwise.
    for node in _jsonld_products(html):
        name = node.get("name") or ""
        urlp = node.get("url") or ""
        if name and urlp and "/listing/" in urlp:
            out.append((name, _offer_price(node), urlp.split("?")[0]))
    if out:
        return out
    seen = set()
    for m in re.finditer(r'<a[^>]+href="(https://www\.etsy\.com/listing/(\d+)/[^"]*)"[^>]*>(.*?)</a>',
                         html, re.S | re.I):
        href, lid, inner = m.group(1), m.group(2), m.group(3)
        if lid in seen:
            continue
        title = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", inner)).strip()
        am = re.search(r'(?:aria-label|title)="([^"]{5,200})"', m.group(0))
        if am and len(am.group(1).strip()) > len(title):
            title = am.group(1).strip()
        if not title:
            continue
        seen.add(lid)
        out.append((title, _mp_price_near(html, m.end()), href.split("?")[0]))
        if len(out) >= 50:
            break
    return out

def search_etsy_direct(lock_name, status_cb, deep=False):
    q = urllib.parse.quote_plus(lock_name.strip())
    # bias toward vintage/collectible where old locks live
    url = f"https://www.etsy.com/search?q={q}"
    return _generic_probe("Etsy", url, _parse_etsy, lock_name,
                          status_cb, "found via direct Etsy search")

# ---- 2dehands (Belgium) — same platform family as Marktplaats -------------
def _parse_2dehands(html):
    out = []
    for m in re.finditer(r'<script[^>]*>(.*?"itemId".*?)</script>', html, re.S):
        try:
            data = json.loads(re.sub(r'^[^{]*', '', m.group(1)))
        except Exception:
            continue
        stack = [data]
        while stack:
            node = stack.pop()
            if isinstance(node, list):
                stack.extend(node)
            elif isinstance(node, dict):
                iid = node.get("itemId") or node.get("id")
                title = node.get("title") or ""
                if iid and title and isinstance(title, str):
                    price = ""
                    pv = node.get("priceInfo") or node.get("price")
                    if isinstance(pv, dict):
                        cents = pv.get("priceCents") or pv.get("amount")
                        if isinstance(cents, (int, float)):
                            price = f"{cents/100:.2f} EUR"
                    url = node.get("vipUrl") or node.get("url") or ""
                    if url and not url.startswith("http"):
                        url = "https://www.2dehands.be" + url
                    if url:
                        out.append((title, price, url))
                for v in node.values():
                    if isinstance(v, (dict, list)):
                        stack.append(v)
        if out:
            break
    if out:
        seen, uniq = set(), []
        for t in out:
            if t[2] not in seen:
                seen.add(t[2]); uniq.append(t)
        return uniq
    for m in re.finditer(r'<a[^>]+href="(/v/[^"]*?/m?(\d+)[^"]*)"[^>]*>(.*?)</a>',
                         html, re.S | re.I):
        href, _id, inner = m.group(1), m.group(2), m.group(3)
        title = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", inner)).strip()
        am = re.search(r'(?:aria-label|title)="([^"]{5,200})"', m.group(0))
        if am and len(am.group(1).strip()) > len(title):
            title = am.group(1).strip()
        if not title:
            continue
        out.append((title, _mp_price_near(html, m.end()),
                    "https://www.2dehands.be" + href.split("?")[0]))
        if len(out) >= 50:
            break
    return out

def search_2dehands_direct(lock_name, status_cb, deep=False):
    q = urllib.parse.quote_plus(lock_name.strip())
    url = f"https://www.2dehands.be/q/{q}/"
    return _generic_probe("2dehands", url, _parse_2dehands, lock_name,
                          status_cb, "found via direct 2dehands search")

# ---- Buyee (proxy for Yahoo Auctions Japan + Mercari; English) ------------
# Buyee is an international proxy that resells Japanese listings in English —
# easier to search AND buy than the native JP sites, so we probe it directly.
def _parse_buyee(html):
    out = []
    # Buyee item links look like /item/yahoo/auction/<id> or /item/mercari/...
    for node in _jsonld_products(html):
        name = node.get("name") or ""
        urlp = node.get("url") or ""
        if name and urlp:
            out.append((name, _offer_price(node),
                        urlp if urlp.startswith("http")
                        else "https://buyee.jp" + urlp))
    if out:
        return out
    seen = set()
    for m in re.finditer(r'<a[^>]+href="(/item/[a-z]+/[a-z0-9]+/([A-Za-z0-9]+))"[^>]*>(.*?)</a>',
                         html, re.S | re.I):
        href, iid, inner = m.group(1), m.group(2), m.group(3)
        if iid in seen:
            continue
        title = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", inner)).strip()
        am = re.search(r'(?:aria-label|title|alt)="([^"]{5,200})"', m.group(0))
        if am and len(am.group(1).strip()) > len(title):
            title = am.group(1).strip()
        if not title:
            continue
        seen.add(iid)
        price = _mp_price_near(html, m.end())
        if not price:
            pm = re.search(r'(¥|JPY|\$|US\s*\$)\s?[\d,]+', html[m.end():m.end()+300])
            price = pm.group(0).strip() if pm else ""
        out.append((title, price, "https://buyee.jp" + href.split("?")[0]))
        if len(out) >= 50:
            break
    return out

def search_buyee_direct(lock_name, status_cb, deep=False):
    q = urllib.parse.quote_plus(lock_name.strip())
    url = f"https://buyee.jp/item/search/query/{q}"
    return _generic_probe("Buyee (JP)", url, _parse_buyee, lock_name,
                          status_cb, "found via direct Buyee search")

# ---- Yahoo Auctions Japan (native) ----------------------------------------
def _parse_yahoo_jp(html):
    out = []
    # Auction item URLs: auctions.yahoo.co.jp/jp/auction/<id>. Titles in the
    # anchor; prices shown with ¥. Many listings are Japanese-only, so token
    # matching will naturally drop non-Latin ones — that's expected.
    seen = set()
    for m in re.finditer(r'<a[^>]+href="(https://auctions\.yahoo\.co\.jp/jp/auction/([a-z0-9]+))"[^>]*>(.*?)</a>',
                         html, re.S | re.I):
        href, iid, inner = m.group(1), m.group(2), m.group(3)
        if iid in seen:
            continue
        title = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", inner)).strip()
        am = re.search(r'(?:aria-label|title)="([^"]{5,200})"', m.group(0))
        if am and len(am.group(1).strip()) > len(title):
            title = am.group(1).strip()
        if not title:
            continue
        seen.add(iid)
        pm = re.search(r'¥\s?[\d,]+', html[m.end():m.end()+300])
        price = pm.group(0).strip() if pm else ""
        out.append((title, price, href.split("?")[0]))
        if len(out) >= 50:
            break
    return out

def search_yahoo_jp_direct(lock_name, status_cb, deep=False):
    q = urllib.parse.quote_plus(lock_name.strip())
    url = f"https://auctions.yahoo.co.jp/search/search?p={q}"
    return _generic_probe("Yahoo Auctions JP", url, _parse_yahoo_jp, lock_name,
                          status_cb, "found via direct Yahoo Auctions JP search")

# ---- Mercari Japan (native) -----------------------------------------------
def _parse_mercari_jp(html):
    out = []
    # Mercari JP item URLs: jp.mercari.com/item/<id>. Heavily JS-rendered; try
    # JSON-LD and item anchors. Japanese-only titles will be filtered by tokens.
    for node in _jsonld_products(html):
        name = node.get("name") or ""
        urlp = node.get("url") or ""
        if name and urlp and "/item/" in urlp:
            out.append((name, _offer_price(node), urlp.split("?")[0]))
    if out:
        return out
    seen = set()
    for m in re.finditer(r'<a[^>]+href="(/item/(m?\d+))"[^>]*>(.*?)</a>',
                         html, re.S | re.I):
        href, iid, inner = m.group(1), m.group(2), m.group(3)
        if iid in seen:
            continue
        title = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", inner)).strip()
        am = re.search(r'(?:aria-label|title|alt)="([^"]{5,200})"', m.group(0))
        if am and len(am.group(1).strip()) > len(title):
            title = am.group(1).strip()
        if not title:
            continue
        seen.add(iid)
        pm = re.search(r'(¥|JPY)\s?[\d,]+', html[m.end():m.end()+300])
        price = pm.group(0).strip() if pm else ""
        out.append((title, price, "https://jp.mercari.com" + href.split("?")[0]))
        if len(out) >= 50:
            break
    return out

def search_mercari_jp_direct(lock_name, status_cb, deep=False):
    q = urllib.parse.quote_plus(lock_name.strip())
    url = f"https://jp.mercari.com/search?keyword={q}"
    return _generic_probe("Mercari JP", url, _parse_mercari_jp, lock_name,
                          status_cb, "found via direct Mercari JP search")

# ---- OLX (Poland, Portugal, Ukraine, …) -----------------------------------
# OLX runs the same platform across countries on separate domains. It
# server-renders listing cards with a data-cy="l-card" wrapper, a title in an
# <h6>/<h4>, a price element, and item links like /d/oferta/<slug>-ID<id>.html.
_OLX_DOMAINS = {
    "Poland": "www.olx.pl", "Portugal": "www.olx.pt", "Ukraine": "www.olx.ua",
    "Russia/Ukraine": "www.olx.ua",
    "Romania": "www.olx.ro", "Bulgaria": "www.olx.bg",
    "Kazakhstan": "www.olx.kz", "Uzbekistan": "www.olx.uz",
}
_OLX_CUR = {"www.olx.pl": "PLN", "www.olx.pt": "EUR", "www.olx.ua": "UAH",
            "www.olx.ro": "RON", "www.olx.bg": "BGN", "www.olx.kz": "KZT",
            "www.olx.uz": "UZS"}

def _parse_olx(html, domain):
    out, seen = [], set()
    cur = _OLX_CUR.get(domain, "")
    # Preferred: per-card blocks so title & price can't bleed across listings.
    for cm in re.finditer(r'<div[^>]+data-cy="l-card"[^>]*>(.*?)</div>\s*</div>\s*</div>',
                          html, re.S | re.I):
        card = cm.group(1)
        lm = re.search(r'href="(/d/oferta/[^"]*?-ID([A-Za-z0-9]+)\.html[^"]*)"', card)
        if not lm:
            continue
        iid = lm.group(2)
        if iid in seen:
            continue
        tm = re.search(r'<h[46][^>]*>(.*?)</h[46]>', card, re.S)
        title = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", tm.group(1))).strip() if tm else ""
        if not title:
            am = re.search(r'(?:aria-label|title)="([^"]{5,200})"', card)
            title = am.group(1).strip() if am else ""
        if not title:
            continue
        pm = re.search(r'data-testid="ad-price"[^>]*>(.*?)<', card, re.S | re.I)
        if not pm:
            pm = re.search(r'([\d\s.,]+\s*(?:zł|€|₴|PLN|EUR|грн))', card)
        price = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", pm.group(1))).strip() if pm else ""
        if price and cur and not re.search(r"[a-zA-Zł€₴]", price):
            price = f"{price} {cur}"
        seen.add(iid)
        url = lm.group(1)
        out.append((title, price, f"https://{domain}" + url.split("?")[0]))
        if len(out) >= 50:
            break
    if out:
        return out
    # Fallback: bare offer anchors anywhere on the page.
    for m in re.finditer(r'<a[^>]+href="(/d/oferta/[^"]*?-ID([A-Za-z0-9]+)\.html[^"]*)"[^>]*>(.*?)</a>',
                         html, re.S | re.I):
        href, iid, inner = m.group(1), m.group(2), m.group(3)
        if iid in seen:
            continue
        title = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", inner)).strip()
        am = re.search(r'(?:aria-label|title)="([^"]{5,200})"', m.group(0))
        if am and len(am.group(1).strip()) > len(title):
            title = am.group(1).strip()
        if not title:
            continue
        seen.add(iid)
        price = _mp_price_near(html, m.end())
        if price and cur and not re.search(r"[a-zA-Zł€₴]", price):
            price = f"{price} {cur}"
        out.append((title, price, f"https://{domain}" + href.split("?")[0]))
        if len(out) >= 50:
            break
    return out

def search_olx_direct(lock_name, status_cb, deep=False, origin=""):
    """Search OLX on the origin country's domain when known (olx.pl for Polish
    brands, olx.pt for Portuguese, olx.ua for Ukrainian). For other origins,
    default to olx.pl since that's where most collectible European lock
    listings on OLX appear."""
    domain = _OLX_DOMAINS.get(origin, "www.olx.pl")
    q = urllib.parse.quote_plus(lock_name.strip())
    url = f"https://{domain}/oferty/q-{q}/"
    return _generic_probe(f"OLX ({domain.replace('www.olx.', '')})", url,
                          lambda h: _parse_olx(h, domain), lock_name,
                          status_cb, "found via direct OLX search")

# ---- helpers reused by the parsers below -----------------------------------
def _anchor_scan(html, href_re, base, note_price_window=400):
    """Generic anchor-based scan: find listing links matching href_re, take the
    anchor text or aria-label as the title, and a nearby price token. Returns
    (title, price, absolute_url) tuples. Defensive; used as a fallback."""
    out, seen = [], set()
    for m in re.finditer(href_re, html, re.S | re.I):
        href = m.group(1)
        key = m.group(2) if m.lastindex and m.lastindex >= 2 else href
        if key in seen:
            continue
        inner = m.group(m.lastindex) if m.lastindex else ""
        title = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", inner)).strip()
        am = re.search(r'(?:aria-label|title|alt)="([^"]{5,200})"', m.group(0))
        if am and len(am.group(1).strip()) > len(title):
            title = am.group(1).strip()
        if not title:
            continue
        seen.add(key)
        url = href if href.startswith("http") else base + href
        out.append((title, _mp_price_near(html, m.end(), note_price_window),
                    url.split("?")[0]))
        if len(out) >= 50:
            break
    return out

# ---- Tori.fi (Finland) — home of Abloy stock ------------------------------
def _parse_tori(html):
    out = []
    for node in _jsonld_products(html):
        name = node.get("name") or ""
        urlp = node.get("url") or ""
        if name and urlp:
            out.append((name, _offer_price(node),
                        urlp if urlp.startswith("http") else "https://www.tori.fi" + urlp))
    if out:
        return out
    return _anchor_scan(
        html, r'<a[^>]+href="(https://www\.tori\.fi/[^"]*?/(\d+)[^"]*)"[^>]*>(.*?)</a>',
        "https://www.tori.fi")

def search_tori_direct(lock_name, status_cb, deep=False):
    q = urllib.parse.quote_plus(lock_name.strip())
    url = f"https://www.tori.fi/koko_suomi?q={q}"
    return _generic_probe("Tori.fi (FI)", url, _parse_tori, lock_name,
                          status_cb, "found via direct Tori.fi search")

# ---- Huuto.net (Finland auctions) -----------------------------------------
def _parse_huuto(html):
    return _anchor_scan(
        html, r'<a[^>]+href="(/kohde/[^"]*?/(\d+))"[^>]*>(.*?)</a>',
        "https://www.huuto.net")

def search_huuto_direct(lock_name, status_cb, deep=False):
    q = urllib.parse.quote_plus(lock_name.strip())
    url = f"https://www.huuto.net/haku?words={q}"
    return _generic_probe("Huuto.net (FI)", url, _parse_huuto, lock_name,
                          status_cb, "found via direct Huuto.net search")

# ---- Gumtree (UK and AU share the platform) -------------------------------
def _parse_gumtree(html, domain):
    out = []
    for node in _jsonld_products(html):
        name = node.get("name") or ""
        urlp = node.get("url") or ""
        if name and urlp:
            out.append((name, _offer_price(node),
                        urlp if urlp.startswith("http") else f"https://{domain}" + urlp))
    if out:
        return out
    return _anchor_scan(
        html, r'<a[^>]+href="(/p/[^"]*?/(\d+))"[^>]*>(.*?)</a>',
        f"https://{domain}")

def search_gumtree_uk_direct(lock_name, status_cb, deep=False):
    q = urllib.parse.quote_plus(lock_name.strip())
    url = f"https://www.gumtree.com/search?search_scope=false&q={q}"
    return _generic_probe("Gumtree (UK)", url,
                          lambda h: _parse_gumtree(h, "www.gumtree.com"),
                          lock_name, status_cb,
                          "found via direct Gumtree UK search")

def search_gumtree_au_direct(lock_name, status_cb, deep=False):
    q = urllib.parse.quote_plus(lock_name.strip())
    url = f"https://www.gumtree.com.au/s-search?keyword={q}"
    return _generic_probe("Gumtree (AU)", url,
                          lambda h: _parse_gumtree(h, "www.gumtree.com.au"),
                          lock_name, status_cb,
                          "found via direct Gumtree AU search")

# ---- Wallapop (Spain) -----------------------------------------------------
def _parse_wallapop(html):
    out = []
    for node in _jsonld_products(html):
        name = node.get("name") or ""
        urlp = node.get("url") or ""
        if name and urlp and "/item/" in urlp:
            out.append((name, _offer_price(node), urlp.split("?")[0]))
    if out:
        return out
    return _anchor_scan(
        html, r'<a[^>]+href="(https://(?:es\.)?wallapop\.com/item/([a-z0-9-]+))"[^>]*>(.*?)</a>',
        "https://es.wallapop.com")

def search_wallapop_direct(lock_name, status_cb, deep=False):
    q = urllib.parse.quote_plus(lock_name.strip())
    url = f"https://es.wallapop.com/app/search?keywords={q}"
    return _generic_probe("Wallapop (ES)", url, _parse_wallapop, lock_name,
                          status_cb, "found via direct Wallapop search")

# ---- Milanuncios (Spain) --------------------------------------------------
def _parse_milanuncios(html):
    return _anchor_scan(
        html, r'<a[^>]+href="(/[a-z-]+/[^"]*?-(\d+)\.htm)"[^>]*>(.*?)</a>',
        "https://www.milanuncios.com")

def search_milanuncios_direct(lock_name, status_cb, deep=False):
    q = urllib.parse.quote_plus(lock_name.strip())
    url = f"https://www.milanuncios.com/anuncios/?s={q}"
    return _generic_probe("Milanuncios (ES)", url, _parse_milanuncios,
                          lock_name, status_cb,
                          "found via direct Milanuncios search")

# ---- Finn.no (Norway) -----------------------------------------------------
def _parse_finn(html):
    out = []
    for node in _jsonld_products(html):
        name = node.get("name") or ""
        urlp = node.get("url") or ""
        if name and urlp:
            out.append((name, _offer_price(node),
                        urlp if urlp.startswith("http") else "https://www.finn.no" + urlp))
    if out:
        return out
    return _anchor_scan(
        html, r'<a[^>]+href="(https://www\.finn\.no/[^"]*?finnkode=(\d+)[^"]*)"[^>]*>(.*?)</a>',
        "https://www.finn.no")

def search_finn_direct(lock_name, status_cb, deep=False):
    q = urllib.parse.quote_plus(lock_name.strip())
    url = f"https://www.finn.no/bap/forsale/search.html?q={q}"
    return _generic_probe("Finn.no (NO)", url, _parse_finn, lock_name,
                          status_cb, "found via direct Finn.no search")

# ---- ShopGoodwill (USA auctions) ------------------------------------------
def _parse_shopgoodwill(html):
    out = []
    for node in _jsonld_products(html):
        name = node.get("name") or ""
        urlp = node.get("url") or ""
        if name and urlp:
            out.append((name, _offer_price(node),
                        urlp if urlp.startswith("http") else "https://shopgoodwill.com" + urlp))
    if out:
        return out
    return _anchor_scan(
        html, r'<a[^>]+href="(/item/(\d+))"[^>]*>(.*?)</a>',
        "https://shopgoodwill.com")

def search_shopgoodwill_direct(lock_name, status_cb, deep=False):
    q = urllib.parse.quote_plus(lock_name.strip())
    url = f"https://shopgoodwill.com/categories/listing?st={q}"
    return _generic_probe("ShopGoodwill (US)", url, _parse_shopgoodwill,
                          lock_name, status_cb,
                          "found via direct ShopGoodwill search")

# ---- GovDeals (USA government surplus auctions) ----------------------------
def _parse_govdeals(html):
    return _anchor_scan(
        html, r'<a[^>]+href="(/[^"]*?asset[^"]*?/(\d+)[^"]*)"[^>]*>(.*?)</a>',
        "https://www.govdeals.com")

def search_govdeals_direct(lock_name, status_cb, deep=False):
    q = urllib.parse.quote_plus(lock_name.strip())
    url = f"https://www.govdeals.com/index.cfm?fa=Main.AdvSearchResults&kWord={q}"
    return _generic_probe("GovDeals (US)", url, _parse_govdeals, lock_name,
                          status_cb, "found via direct GovDeals search")

# ---- LiveAuctioneers (global specialty/estate auctions) --------------------
def _parse_liveauctioneers(html):
    out = []
    for node in _jsonld_products(html):
        name = node.get("name") or ""
        urlp = node.get("url") or ""
        if name and urlp:
            out.append((name, _offer_price(node),
                        urlp if urlp.startswith("http") else "https://www.liveauctioneers.com" + urlp))
    if out:
        return out
    return _anchor_scan(
        html, r'<a[^>]+href="(/item/(\d+)[^"]*)"[^>]*>(.*?)</a>',
        "https://www.liveauctioneers.com")

def search_liveauctioneers_direct(lock_name, status_cb, deep=False):
    q = urllib.parse.quote_plus(lock_name.strip())
    url = f"https://www.liveauctioneers.com/search/?keyword={q}"
    return _generic_probe("LiveAuctioneers", url, _parse_liveauctioneers,
                          lock_name, status_cb,
                          "found via direct LiveAuctioneers search")

# ---- Yad2 (Israel) — Hebrew, best-effort (cross-script matching limited) ---
def _parse_yad2(html):
    return _anchor_scan(
        html, r'<a[^>]+href="(/item/([a-z0-9]+))"[^>]*>(.*?)</a>',
        "https://www.yad2.co.il")

def search_yad2_direct(lock_name, status_cb, deep=False):
    q = urllib.parse.quote_plus(lock_name.strip())
    url = f"https://www.yad2.co.il/products/search?query={q}"
    return _generic_probe("Yad2 (IL)", url, _parse_yad2, lock_name,
                          status_cb, "found via direct Yad2 search")

# ---- Avito (Russia) — Cyrillic, best-effort, heavily bot-defended ----------
def _parse_avito(html):
    return _anchor_scan(
        html, r'<a[^>]+href="(/[^"]*?_(\d{6,})[^"]*)"[^>]*>(.*?)</a>',
        "https://www.avito.ru")

def search_avito_direct(lock_name, status_cb, deep=False):
    q = urllib.parse.quote_plus(lock_name.strip())
    url = f"https://www.avito.ru/all?q={q}"
    return _generic_probe("Avito (RU)", url, _parse_avito, lock_name,
                          status_cb, "found via direct Avito search")

# ---- Taobao/Xianyu (China) — best-effort, very hostile to automation -------
def _parse_taobao(html):
    return _anchor_scan(
        html, r'<a[^>]+href="(//item\.taobao\.com/[^"]*?id=(\d+)[^"]*)"[^>]*>(.*?)</a>',
        "https:")

def search_taobao_direct(lock_name, status_cb, deep=False):
    q = urllib.parse.quote_plus(lock_name.strip())
    url = f"https://s.taobao.com/search?q={q}"
    return _generic_probe("Taobao (CN)", url, _parse_taobao, lock_name,
                          status_cb, "found via direct Taobao search")

# ---- A batch of national marketplaces (mostly Latin-script). Each follows
#      the same pattern: try JSON-LD, then an anchor scan; degrade to []. ----
def _ld_or_anchor(html, base, href_re):
    out = []
    for node in _jsonld_products(html):
        name = node.get("name") or ""
        urlp = node.get("url") or ""
        if name and urlp:
            out.append((name, _offer_price(node),
                        urlp if urlp.startswith("http") else base + urlp))
    if out:
        return out
    return _anchor_scan(html, href_re, base)

# ---- Bland.is (Iceland) ---------------------------------------------------
def search_bland_direct(lock_name, status_cb, deep=False):
    q = urllib.parse.quote_plus(lock_name.strip())
    url = f"https://bland.is/leit/?q={q}"
    return _generic_probe(
        "Bland.is (IS)", url,
        lambda h: _anchor_scan(h, r'<a[^>]+href="(/[^"]*?/(\d{4,})[^"]*)"[^>]*>(.*?)</a>',
                               "https://bland.is"),
        lock_name, status_cb, "found via direct Bland.is search")

# ---- DoneDeal (Ireland) ---------------------------------------------------
def search_donedeal_direct(lock_name, status_cb, deep=False):
    q = urllib.parse.quote_plus(lock_name.strip())
    url = f"https://www.donedeal.ie/all?words={q}"
    return _generic_probe(
        "DoneDeal (IE)", url,
        lambda h: _ld_or_anchor(h, "https://www.donedeal.ie",
                                r'<a[^>]+href="(/[a-z0-9-]+/[^"]*?/(\d{5,}))"[^>]*>(.*?)</a>'),
        lock_name, status_cb, "found via direct DoneDeal search")

# ---- DBA.dk (Denmark) -----------------------------------------------------
def search_dba_direct(lock_name, status_cb, deep=False):
    q = urllib.parse.quote_plus(lock_name.strip())
    url = f"https://www.dba.dk/soeg/?soeg={q}"
    return _generic_probe(
        "DBA.dk (DK)", url,
        lambda h: _ld_or_anchor(h, "https://www.dba.dk",
                                r'<a[^>]+href="(https://www\.dba\.dk/[^"]*?id-(\d+)[^"]*)"[^>]*>(.*?)</a>'),
        lock_name, status_cb, "found via direct DBA.dk search")

# ---- Osta.ee (Estonia) ----------------------------------------------------
def search_osta_direct(lock_name, status_cb, deep=False):
    q = urllib.parse.quote_plus(lock_name.strip())
    url = f"https://www.osta.ee/otsi?q={q}"
    return _generic_probe(
        "Osta.ee (EE)", url,
        lambda h: _ld_or_anchor(h, "https://www.osta.ee",
                                r'<a[^>]+href="(/[^"]*?-(\d{5,})[^"]*)"[^>]*>(.*?)</a>'),
        lock_name, status_cb, "found via direct Osta.ee search")

# ---- SS.com (Latvia) ------------------------------------------------------
def search_sscom_direct(lock_name, status_cb, deep=False):
    q = urllib.parse.quote_plus(lock_name.strip())
    url = f"https://www.ss.com/en/search/?q={q}"
    return _generic_probe(
        "SS.com (LV)", url,
        lambda h: _anchor_scan(h, r'<a[^>]+href="(/en/[^"]*?/[a-z0-9]+/([a-z0-9]+)\.html)"[^>]*>(.*?)</a>',
                               "https://www.ss.com"),
        lock_name, status_cb, "found via direct SS.com search")

# ---- Skelbiu (Lithuania) --------------------------------------------------
def search_skelbiu_direct(lock_name, status_cb, deep=False):
    q = urllib.parse.quote_plus(lock_name.strip())
    url = f"https://www.skelbiu.lt/skelbimai/?keywords={q}"
    return _generic_probe(
        "Skelbiu (LT)", url,
        lambda h: _anchor_scan(h, r'<a[^>]+href="(/skelbimai/[^"]*?-(\d{5,})\.html)"[^>]*>(.*?)</a>',
                               "https://www.skelbiu.lt"),
        lock_name, status_cb, "found via direct Skelbiu search")

# ---- Kufar (Belarus) ------------------------------------------------------
def search_kufar_direct(lock_name, status_cb, deep=False):
    q = urllib.parse.quote_plus(lock_name.strip())
    url = f"https://www.kufar.by/l?query={q}"
    return _generic_probe(
        "Kufar (BY)", url,
        lambda h: _ld_or_anchor(h, "https://www.kufar.by",
                                r'<a[^>]+href="(https://www\.kufar\.by/[^"]*?/(\d{6,})[^"]*)"[^>]*>(.*?)</a>'),
        lock_name, status_cb, "found via direct Kufar search")

# ---- Bazoš (Czech .cz + Slovak .sk) ---------------------------------------
def _search_bazos(lock_name, status_cb, tld):
    q = urllib.parse.quote_plus(lock_name.strip())
    url = f"https://www.bazos.{tld}/search.php?hledat={q}"
    base = f"https://www.bazos.{tld}"
    return _generic_probe(
        f"Bazoš ({tld.upper()})", url,
        lambda h: _anchor_scan(h, r'<a[^>]+href="(/inzerat/(\d+)/[^"]*)"[^>]*>(.*?)</a>', base),
        lock_name, status_cb, f"found via direct Bazoš.{tld} search")

def search_bazos_cz_direct(lock_name, status_cb, deep=False):
    return _search_bazos(lock_name, status_cb, "cz")

def search_bazos_sk_direct(lock_name, status_cb, deep=False):
    return _search_bazos(lock_name, status_cb, "sk")

# ---- Jófogás (Hungary) ----------------------------------------------------
def search_jofogas_direct(lock_name, status_cb, deep=False):
    q = urllib.parse.quote_plus(lock_name.strip())
    url = f"https://www.jofogas.hu/magyarorszag?q={q}"
    return _generic_probe(
        "Jófogás (HU)", url,
        lambda h: _ld_or_anchor(h, "https://www.jofogas.hu",
                                r'<a[^>]+href="(https://www\.jofogas\.hu/[^"]*?(\d{6,})[^"]*)"[^>]*>(.*?)</a>'),
        lock_name, status_cb, "found via direct Jófogás search")

# ---- Njuškalo (Croatia/Slovenia) ------------------------------------------
def search_njuskalo_direct(lock_name, status_cb, deep=False):
    q = urllib.parse.quote_plus(lock_name.strip())
    url = f"https://www.njuskalo.hr/?trazi={q}"
    return _generic_probe(
        "Njuškalo (HR)", url,
        lambda h: _ld_or_anchor(h, "https://www.njuskalo.hr",
                                r'<a[^>]+href="(/[^"]*?-oglas-(\d{6,})[^"]*)"[^>]*>(.*?)</a>'),
        lock_name, status_cb, "found via direct Njuškalo search")

# ---- KupujemProdajem (Serbia) ---------------------------------------------
def search_kupujem_direct(lock_name, status_cb, deep=False):
    q = urllib.parse.quote_plus(lock_name.strip())
    url = f"https://www.kupujemprodajem.com/pretraga?keywords={q}"
    return _generic_probe(
        "KupujemProdajem (RS)", url,
        lambda h: _anchor_scan(h, r'<a[^>]+href="(/[^"]*?/oglas/(\d{6,})[^"]*)"[^>]*>(.*?)</a>',
                               "https://www.kupujemprodajem.com"),
        lock_name, status_cb, "found via direct KupujemProdajem search")

# ---- Maltapark (Malta) — ad id lives in the query string, keep it ---------
def _parse_maltapark(html):
    out, seen = [], set()
    for m in re.finditer(r'<a[^>]+href="([^"]*?adid=(\d+)[^"]*)"[^>]*>(.*?)</a>',
                         html, re.S | re.I):
        href, aid, inner = m.group(1), m.group(2), m.group(3)
        if aid in seen:
            continue
        title = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", inner)).strip()
        am = re.search(r'(?:aria-label|title|alt)="([^"]{5,200})"', m.group(0))
        if am and len(am.group(1).strip()) > len(title):
            title = am.group(1).strip()
        if not title:
            continue
        seen.add(aid)
        # keep the FULL href incl. query (that's where adid lives); only drop
        # a trailing tracking fragment after '#'
        url = href.split("#")[0]
        if not url.startswith("http"):
            url = "https://www.maltapark.com" + (url if url.startswith("/") else "/" + url)
        out.append((title, _mp_price_near(html, m.end()), url))
        if len(out) >= 50:
            break
    return out

def search_maltapark_direct(lock_name, status_cb, deep=False):
    q = urllib.parse.quote_plus(lock_name.strip())
    url = f"https://www.maltapark.com/searchresult.asp?keyword={q}"
    return _generic_probe("Maltapark (MT)", url, _parse_maltapark,
                          lock_name, status_cb,
                          "found via direct Maltapark search")

# ---- Sahibinden (Turkey) — Turkish is Latin, so matching works ------------
def search_sahibinden_direct(lock_name, status_cb, deep=False):
    q = urllib.parse.quote_plus(lock_name.strip())
    url = f"https://www.sahibinden.com/kelime-ile-arama?query_text={q}"
    return _generic_probe(
        "Sahibinden (TR)", url,
        lambda h: _ld_or_anchor(h, "https://www.sahibinden.com",
                                r'<a[^>]+href="(/ilan/[^"]*?-(\d{6,})[^"]*)"[^>]*>(.*?)</a>'),
        lock_name, status_cb, "found via direct Sahibinden search")

# ---- Bazaraki (Cyprus) ----------------------------------------------------
def search_bazaraki_direct(lock_name, status_cb, deep=False):
    q = urllib.parse.quote_plus(lock_name.strip())
    url = f"https://www.bazaraki.com/search/?q={q}"
    return _generic_probe(
        "Bazaraki (CY)", url,
        lambda h: _ld_or_anchor(h, "https://www.bazaraki.com",
                                r'<a[^>]+href="(/adv/(\d+)_[^"]*)"[^>]*>(.*?)</a>'),
        lock_name, status_cb, "found via direct Bazaraki search")

# Origin country -> the native marketplace(s) most likely to carry the lock,
# run first (ordering only; all probes still run — rare locks travel).
_ORIGIN_NATIVE = {
    "France": ["Leboncoin"],
    "Germany": ["Kleinanzeigen"], "Austria": ["Willhaben", "Kleinanzeigen"],
    "Switzerland/Germany": ["Ricardo", "Kleinanzeigen"],
    "Italy": ["Subito"],
    "Netherlands": ["Marktplaats"],
    "Belgium": ["2dehands", "Marktplaats"],
    "Sweden": ["Tradera", "Blocket"],
    "Denmark/Sweden": ["Tradera", "Blocket"],
    "Japan": ["Buyee (JP)", "Yahoo Auctions JP", "Mercari JP"],
    "Poland": ["OLX (pl)"],
    "Portugal": ["OLX (pt)"],
    "Russia/Ukraine": ["OLX (ua)"],
    "Finland": ["Tori.fi (FI)", "Huuto.net (FI)"],
    "Norway": ["Finn.no (NO)"],
    "Spain": ["Wallapop (ES)", "Milanuncios (ES)"],
    "UK": ["Gumtree (UK)"],
    "USA": ["ShopGoodwill (US)", "GovDeals (US)", "LiveAuctioneers"],
    "USA/UK": ["ShopGoodwill (US)", "Gumtree (UK)"],
    "Australia": ["Gumtree (AU)"],
    "Israel": ["Yad2 (IL)"],
    "Russia": ["Avito (RU)"],
    "China": ["Taobao (CN)"],
    "Iceland": ["Bland.is (IS)"],
    "Ireland": ["DoneDeal (IE)"],
    "Denmark": ["DBA.dk (DK)"],
    "Estonia": ["Osta.ee (EE)"],
    "Latvia": ["SS.com (LV)"],
    "Lithuania": ["Skelbiu (LT)"],
    "Belarus": ["Kufar (BY)"],
    "Czechia": ["Bazoš (CZ)"],
    "Slovakia": ["Bazoš (SK)"],
    "Hungary": ["Jófogás (HU)"],
    "Croatia": ["Njuškalo (HR)"],
    "Slovenia": ["Njuškalo (HR)"],
    "Serbia": ["KupujemProdajem (RS)"],
    "Malta": ["Maltapark (MT)"],
    "Turkey": ["Sahibinden (TR)"],
    "Cyprus": ["Bazaraki (CY)"],
    "Romania": ["OLX (ro)"],
    "Bulgaria": ["OLX (bg)"],
}

# ---- Facebook Marketplace (opt-in, single-lock searches only) -------------
# Facebook has NO public API, but it EMBEDS public Marketplace listing data as
# JSON inside <script type="application/json"> blobs that it serves to
# logged-out requests. So we can read it with a plain GET — no login, no
# browser engine, nothing stored — and walk that JSON for listing nodes.
# Two hard rules keep this from getting an IP rate-limited by Facebook:
#   * it runs SEQUENTIALLY and THROTTLED (never in the parallel probe pool), and
#   * it runs for SINGLE-LOCK searches only (never the Hunt-Wishlist batch).
# It is best-effort: Facebook rotates its page shape from time to time, so if a
# parse yields nothing the probe just returns [] like any other site would.
# Marketplace is location-scoped, so we sweep a short list of major-city
# markets. Edit _FB_MARKETS to taste — fewer markets = faster + smaller
# footprint. (City slugs are Facebook's own; adjust if a market stops
# resolving.)
_FB_MARKETS = [
    ("USA", "nyc"), ("Germany", "berlin"), ("France", "paris"),
    ("Netherlands", "amsterdam"), ("Poland", "warsaw"),
    ("Sweden", "stockholm"), ("Denmark", "copenhagen"),
    ("Bulgaria", "sofia"), ("UK", "london"), ("Italy", "milan"),
    ("Finland", "helsinki"), ("Switzerland", "zurich"),
]

# Origin-aware market selection: instead of hitting all markets every time,
# pick the lock's likely-origin market (+ neighbours) plus a small always-on
# core. Fewer, more relevant markets = better coverage of the RIGHT places AND
# fewer requests per search (lighter on Facebook's rate limit). Keys are the
# country strings _origin_for_lock returns; values are _FB_MARKETS labels.
_FB_ORIGIN_PREF = {
    "Finland": ["Finland", "Sweden", "Denmark"],
    "Sweden": ["Sweden", "Denmark", "Finland"],
    "Norway": ["Denmark", "Sweden"],
    "Denmark": ["Denmark", "Sweden", "Germany"],
    "Germany": ["Germany", "Netherlands", "Poland"],
    "Austria": ["Germany", "Switzerland", "Italy"],
    "Switzerland": ["Switzerland", "Germany", "France", "Italy"],
    "Italy": ["Italy", "Switzerland", "France"],
    "France": ["France", "Netherlands", "Switzerland"],
    "Netherlands": ["Netherlands", "Germany", "France"],
    "Poland": ["Poland", "Germany"],
    "Spain": ["France", "Italy"],
    "Portugal": ["France", "Italy"],
    "UK": ["UK", "France", "Netherlands"],
    "USA": ["USA", "UK"],
    "Israel": ["Germany", "France", "UK"],
    "Russia": ["Poland", "Bulgaria"],
    "Ukraine": ["Poland", "Bulgaria"],
    "Czechia": ["Poland", "Germany"],
    "Slovakia": ["Poland", "Germany"],
    "Croatia": ["Italy", "Germany"],
    "Turkey": ["Bulgaria", "Germany"],
    "Australia": ["UK", "USA"],
    "Japan": ["USA", "UK"],
    "China": ["USA", "UK"],
}
_FB_CORE = ["USA", "Germany", "UK"]   # huge markets — anything can surface here
_FB_MAX_MARKETS = 6


def _fb_markets_for(lock_name):
    """Origin-aware pick of Facebook markets for a lock: its likely-origin
    market (+ neighbours) plus the always-on core, capped at _FB_MAX_MARKETS.
    Unknown origin falls back to the busiest general markets. Returns a list of
    (country_label, city_slug) from _FB_MARKETS."""
    origin = _origin_for_lock(lock_name)
    labels = []
    for part in re.split(r"[/,]", origin or ""):
        for lbl in _FB_ORIGIN_PREF.get(part.strip(), []):
            if lbl not in labels:
                labels.append(lbl)
    if not labels:   # unknown brand/origin -> busiest general markets
        labels = ["USA", "Germany", "UK", "France", "Italy", "Netherlands"]
    for lbl in _FB_CORE:
        if lbl not in labels:
            labels.append(lbl)
    labels = labels[:_FB_MAX_MARKETS]
    bylabel = {c: (c, s) for c, s in _FB_MARKETS}
    return [bylabel[l] for l in labels if l in bylabel]
_FB_HEADERS = {
    "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                   "AppleWebKit/537.36 (KHTML, like Gecko) "
                   "Chrome/126.0.0.0 Safari/537.36"),
    "Accept": ("text/html,application/xhtml+xml,application/xml;q=0.9,"
               "image/avif,image/webp,*/*;q=0.8"),
    "Accept-Language": "en-US,en;q=0.9",
    "Sec-Fetch-Dest": "document", "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "none", "Upgrade-Insecure-Requests": "1",
}
_FB_TIMEOUT = 30
# Self-imposed rate limit for the Facebook probe. Facebook's no-login ceiling
# is ~30-60 requests/hour per IP. With origin-aware market selection a sweep is
# now ~6 markets + a warm-up (~7 requests), so 3 sweeps/hour (~21 requests)
# stays comfortably under. Rolling 60-minute window, persisted so restarting
# the app can't reset it.
_FB_MAX_PER_HOUR = 3
_FB_WINDOW_MIN = 60


def _fb_session():
    """Return (session, using_curl_cffi). Facebook frequently serves a login
    wall to plain-`requests` traffic because urllib3's TLS fingerprint is
    recognizably not-a-browser — perfect headers don't fix that. curl_cffi
    impersonates a real Chrome TLS fingerprint, which is how the no-login
    scrapers actually get the public SSR JSON back. If curl_cffi isn't
    installed we fall back to requests (the probe then simply tends to return
    nothing rather than erroring)."""
    try:
        from curl_cffi import requests as _cffi
        return _cffi.Session(impersonate="chrome"), True
    except Exception:
        s = requests.Session()
        s.headers.update(_FB_HEADERS)
        return s, False


def _fb_walk_listings(node, out):
    """Recursively collect Marketplace listing nodes from Facebook's embedded
    JSON. A listing is any dict carrying a 'marketplace_listing_title' plus an
    'id'. Keying on that field (rather than a fixed path) is what lets this
    survive Facebook reshuffling the surrounding structure."""
    if isinstance(node, dict):
        if node.get("marketplace_listing_title") and node.get("id"):
            out.append(node)
        for v in node.values():
            _fb_walk_listings(v, out)
    elif isinstance(node, list):
        for v in node:
            _fb_walk_listings(v, out)


def _fb_photo(n):
    """Best-effort primary photo URL from a Facebook listing JSON node."""
    for k in ("primary_listing_photo", "primaryListingPhoto"):
        img = ((n.get(k) or {}).get("image") or {}).get("uri")
        if img:
            return img
    for k in ("listing_photos", "listingPhotos"):
        lp = n.get(k) or []
        if isinstance(lp, list) and lp:
            img = ((lp[0] or {}).get("image") or {}).get("uri")
            if img:
                return img
    return ""


def _fb_extract(html):
    """Pull (title, price, url) tuples out of the JSON blobs embedded in a
    logged-out Facebook Marketplace search page. Best-effort / defensive."""
    found = {}
    for m in re.finditer(
            r'<script[^>]+type="application/json"[^>]*>(.*?)</script>',
            html, re.S):
        blob = m.group(1)
        if "marketplace_listing_title" not in blob:
            continue
        try:
            data = json.loads(blob)
        except Exception:
            continue
        nodes = []
        _fb_walk_listings(data, nodes)
        for n in nodes:
            lid = str(n.get("id") or "").strip()
            if not lid or lid in found:
                continue
            title = str(n.get("marketplace_listing_title") or "").strip()
            if not title:
                continue
            price = ""
            lp = n.get("listing_price") or {}
            if isinstance(lp, dict):
                price = (lp.get("formatted_amount")
                         or lp.get("formatted_amount_zeros_stripped") or "")
                if not price and lp.get("amount"):
                    price = (f"{lp.get('amount')} "
                             f"{lp.get('currency', '')}").strip()
            found[lid] = (title, price,
                          f"https://www.facebook.com/marketplace/item/{lid}/",
                          _fb_photo(n))
    return list(found.values())


def search_facebook_marketplace(lock_name, status_cb, throttle=2.5):
    """Opt-in, single-lock-only Facebook Marketplace probe. Sweeps a short list
    of major-city markets SEQUENTIALLY with a delay between each (Facebook rate-
    limits bursts from one IP). No login or credentials — it reads the public
    JSON embedded in the logged-out search page. Best-effort; returns [] on
    trouble. Result dicts match the other marketplace probes."""
    import time
    tokens = _lock_tokens(lock_name)
    q = urllib.parse.quote_plus(lock_name.strip())
    out, seen = [], set()
    try:
        sess, tls_ok = _fb_session()
        if not tls_ok:
            status_cb("  Facebook: curl_cffi not installed — Facebook may "
                      "return no results (see note in requirements.txt)")
        # Warm-up: visit /marketplace/ once so Facebook sees a browser that
        # browsed first, and so the session picks up any cookies.
        try:
            sess.get("https://www.facebook.com/marketplace/",
                     timeout=_FB_TIMEOUT)
        except Exception:
            pass
        markets = _fb_markets_for(lock_name)
        for i, (country, city) in enumerate(markets):
            if i:
                time.sleep(throttle)   # space requests to stay under FB's limit
            url = (f"https://www.facebook.com/marketplace/{city}/search/"
                   f"?query={q}")
            try:
                status_cb(f"Probing Facebook Marketplace ({country})…")
                r = sess.get(url, timeout=_FB_TIMEOUT)
                if r.status_code != 200:
                    status_cb(f"  FB {country}: HTTP {r.status_code}")
                    continue
                items = _fb_extract(r.text)
                if not items:
                    _low = (r.text or "")[:4000].lower()
                    if "login" in _low or _looks_blocked(r.status_code,
                                                         r.text):
                        status_cb(f"  FB {country}: got a login/anti-bot "
                                  "page instead of listings")
                matched = 0
                for title, price, link, img in items:
                    if link in seen:
                        continue
                    if tokens and not _match_tokens(title, tokens):
                        continue
                    if (_LOCK_CONTEXT_FILTER
                            and not _search_has_lock_word(lock_name)
                            and not _has_lock_context(title)):
                        continue
                    seen.add(link)
                    matched += 1
                    out.append({
                        "title": title, "price": price, "currency": "",
                        "condition": "used",
                        "site": f"Facebook Marketplace ({country})",
                        "url": link, "location": country,
                        "shipping": "unknown",
                        "notes": "found via direct Facebook Marketplace search",
                        "image": img, "preverified": True,
                    })
                status_cb(
                    f"  FB {country}: {len(items)} parsed, {matched} matched")
            except Exception as ex:
                status_cb(f"  FB {country}: {ex}")
                continue
    except Exception as ex:
        status_cb(f"  Facebook Marketplace: {ex}")
    return out


def run_extra_marketplace_probes(lock_name, status_cb, origin="", deep=False):
    """Run every no-API-key marketplace probe and merge the results, de-duped
    by url. Origin-native sites go first (cosmetic). Any individual site
    failing never affects the others — each degrades to []."""
    by_name = {
        "Catawiki": search_catawiki_direct,
        "Leboncoin": search_leboncoin_direct,
        "Kleinanzeigen": search_kleinanzeigen_direct,
        "Delcampe": search_delcampe_direct,
        "Ricardo": search_ricardo_direct,
        "Marktplaats": search_marktplaats_direct,
        "Tradera": search_tradera_direct,
        "Subito": search_subito_direct,
        "Blocket": search_blocket_direct,
        "Willhaben": search_willhaben_direct,
        "Etsy": search_etsy_direct,
        "2dehands": search_2dehands_direct,
        "Buyee (JP)": search_buyee_direct,
        "Yahoo Auctions JP": search_yahoo_jp_direct,
        "Mercari JP": search_mercari_jp_direct,
        "Tori.fi (FI)": search_tori_direct,
        "Huuto.net (FI)": search_huuto_direct,
        "Gumtree (UK)": search_gumtree_uk_direct,
        "Gumtree (AU)": search_gumtree_au_direct,
        "Wallapop (ES)": search_wallapop_direct,
        "Milanuncios (ES)": search_milanuncios_direct,
        "Finn.no (NO)": search_finn_direct,
        "ShopGoodwill (US)": search_shopgoodwill_direct,
        "GovDeals (US)": search_govdeals_direct,
        "LiveAuctioneers": search_liveauctioneers_direct,
        "Yad2 (IL)": search_yad2_direct,
        "Avito (RU)": search_avito_direct,
        "Taobao (CN)": search_taobao_direct,
        "Bland.is (IS)": search_bland_direct,
        "DoneDeal (IE)": search_donedeal_direct,
        "DBA.dk (DK)": search_dba_direct,
        "Osta.ee (EE)": search_osta_direct,
        "SS.com (LV)": search_sscom_direct,
        "Skelbiu (LT)": search_skelbiu_direct,
        "Kufar (BY)": search_kufar_direct,
        "Bazoš (CZ)": search_bazos_cz_direct,
        "Bazoš (SK)": search_bazos_sk_direct,
        "Jófogás (HU)": search_jofogas_direct,
        "Njuškalo (HR)": search_njuskalo_direct,
        "KupujemProdajem (RS)": search_kupujem_direct,
        "Maltapark (MT)": search_maltapark_direct,
        "Sahibinden (TR)": search_sahibinden_direct,
        "Bazaraki (CY)": search_bazaraki_direct,
    }
    # OLX needs the origin to pick its country domain; give it a canonical
    # label so origin-native ordering can reference it regardless of domain.
    olx_label = "OLX (" + _OLX_DOMAINS.get(origin, "www.olx.pl").replace(
        "www.olx.", "") + ")"

    def _run(name):
        if name.startswith("OLX"):
            return search_olx_direct(lock_name, status_cb, deep=deep,
                                     origin=origin)
        fn = by_name.get(name)
        return fn(lock_name, status_cb, deep=deep) if fn else []

    native = _ORIGIN_NATIVE.get(origin, [])
    # normalize any origin-native OLX entry to this run's actual label
    native = [olx_label if n.startswith("OLX") else n for n in native]
    order = native + [n for n in by_name if n not in native]
    if olx_label not in order:
        order.append(olx_label)   # always probe OLX (defaults to .pl)
    # Every probe is an independent network fetch, so run them CONCURRENTLY
    # instead of one-after-another: wall-clock drops from the SUM of ~45
    # sites' fetch times to roughly the slowest wave. 24 workers ≈ two waves
    # for ~45 sites; each site gets exactly ONE request and they're all
    # different domains, so higher concurrency costs nothing in politeness.
    # Results are still merged in the same origin-native-first order, de-duped
    # by url, and any single site failing degrades to [] without affecting
    # the others.
    results_by_name = {}
    with concurrent.futures.ThreadPoolExecutor(max_workers=24) as pool:
        futures = {name: pool.submit(_run, name) for name in order}
        for name in order:
            try:
                results_by_name[name] = futures[name].result() or []
            except Exception:
                results_by_name[name] = []
    out, seen = [], set()
    for name in order:
        for it in results_by_name[name]:
            u = it.get("url")
            if u and u not in seen:
                seen.add(u)
                out.append(it)
    return out


# ---------------------------------------------------------- currency → USD
# Optional "Show USD estimate" feature: parse a listing's price + currency and
# convert to US dollars using live rates (open.er-api.com, free, no key),
# cached for 12 hours. If rates can't be fetched, estimates simply don't show.
RATES_URL = "https://open.er-api.com/v6/latest/USD"
_rates_cache = {"rates": None, "at": None}

_CUR_CODES = {"USD", "EUR", "GBP", "JPY", "SEK", "DKK", "NOK", "CHF", "PLN",
              "CZK", "AUD", "CAD", "NZD", "ILS", "RUB", "CNY", "HKD", "HUF",
              "RON", "BRL", "INR", "MXN", "SGD", "ZAR", "UAH", "ISK", "BGN",
              "RSD", "TRY", "BYN", "KZT", "UZS"}
_CUR_SYMBOLS = [  # longest first so "US $" wins over "$"
    ("US $", "USD"), ("AU $", "AUD"), ("CA $", "CAD"), ("C $", "CAD"),
    ("A$", "AUD"), ("NZ$", "NZD"), ("HK$", "HKD"), ("R$", "BRL"),
    ("zł", "PLN"), ("Kč", "CZK"), ("грн", "UAH"), ("₴", "UAH"),
    ("€", "EUR"), ("£", "GBP"), ("¥", "JPY"),
    ("₪", "ILS"), ("₽", "RUB"), ("₹", "INR"), ("$", "USD"),
]

def _parse_price(price, currency=""):
    """Extract (amount, currency_code) from listing price/currency strings.
    Handles '€ 10.00', 'US $40', '1.234,56 EUR', '1,234.56', '10,50'.
    Returns (None, code) when no number found; code may be None if unknown
    (e.g. bare 'kr', which is ambiguous across SEK/DKK/NOK)."""
    text = f"{currency or ''} {price or ''}".strip()
    cur = None
    up = " " + text.upper() + " "
    for code in _CUR_CODES:
        if re.search(r"(?<![A-Z])" + code + r"(?![A-Z])", up):
            cur = code
            break
    if not cur:
        for sym, code in _CUR_SYMBOLS:
            if sym in text:
                cur = code
                break
    m = re.search(r"\d[\d\s\u00a0.,]*", str(price or "")) or \
        re.search(r"\d[\d\s\u00a0.,]*", text)
    if not m:
        return None, cur
    num = m.group(0).replace(" ", "").replace("\u00a0", "").rstrip(".,")
    if "," in num and "." in num:
        dec = "," if num.rfind(",") > num.rfind(".") else "."
        other = "." if dec == "," else ","
        num = num.replace(other, "").replace(dec, ".")
    elif "," in num:
        tail = num.rsplit(",", 1)[-1]
        num = num.replace(",", ".") if len(tail) == 2 else num.replace(",", "")
    elif "." in num:
        parts = num.split(".")
        if len(parts[-1]) == 3 and all(len(p) <= 3 for p in parts):
            num = num.replace(".", "")   # 1.234 / 1.234.567 -> thousands
    try:
        return float(num), cur
    except ValueError:
        return None, cur

def _usd_rates(force=False):
    """Live 'units per 1 USD' rate table, cached 12h. Returns dict or None."""
    now = datetime.datetime.now()
    c = _rates_cache
    if not force and c["rates"] and c["at"] and \
            (now - c["at"]).total_seconds() < 12 * 3600:
        return c["rates"]
    try:
        r = requests.get(RATES_URL, timeout=20,
                         headers={"User-Agent": _BROWSER_UA})
        if r.status_code == 200:
            rates = (r.json() or {}).get("rates") or {}
            if rates.get("EUR"):
                c.update(rates=rates, at=now)
                return rates
    except Exception:
        pass
    return c["rates"]

def _usd_estimate(price, currency, rates):
    """USD value for a listing price, or None if unparseable/unconvertible."""
    amt, cur = _parse_price(price, currency)
    if amt is None or not cur:
        return None
    if cur == "USD":
        return amt
    try:
        rate = float((rates or {}).get(cur) or 0)
        return amt / rate if rate > 0 else None
    except (TypeError, ValueError, ZeroDivisionError):
        return None

# LPU Lock Bazaar (lpulocks.com) — community marketplace
BAZAAR_URL = "https://lpulocks.com/#/lockbazaar"
BAZAAR_DATA_URLS = [
    # Confirmed live feed (captured from the site's Network tab). This is a
    # dedicated data subdomain that serves plain JSON, so a normal HTTP client
    # can fetch it directly. This is the one the bazaar itself loads.
    "https://data.lpulocks.com/lockbazaar/lockBazaarData.json",
    # older/fallback guesses, kept in case the path ever changes
    "https://lpulocks.com/data/lockBazaarData.json",
    "https://raw.githubusercontent.com/Lockpickers-United/lock-trackers/main/src/data/lockBazaarData.json",
    "https://raw.githubusercontent.com/Lockpickers-United/lock-trackers/main/public/data/lockBazaarData.json",
]

# LPU profile import (Firebase backend of lpubelts.com; config is public
# in the open-source repo). Tried in order at import time.
FIREBASE_CONFIG_URLS = [
    "https://raw.githubusercontent.com/Lockpickers-United/lpu-belt-explorer/main/src/auth/firebase.js",
    "https://raw.githubusercontent.com/Lockpickers-United/lpu-belt-explorer/main/src/app/firebase.js",
    "https://raw.githubusercontent.com/Lockpickers-United/lpu-belt-explorer/main/src/firebase.js",
    "https://raw.githubusercontent.com/Lockpickers-United/lpu-belt-explorer/main/src/auth/firebaseConfig.js",
]
# Confirmed from the live site (Firestore Listen channel + identitytoolkit call).
# Firebase web API keys are public by design (they identify the project; the
# real protection is Firestore security rules), so this is safe to embed and
# saves us from having to scrape it out of the site bundle every time.
LPU_PROJECT_ID = "lpu-belt-explorer"
LPU_WEB_API_KEY = "AIzaSyDGGErdOp0lpzUumA60xJO7BlQr027y9Vo"
FIREBASE_FALLBACK_PROJECTS = ["lpu-belt-explorer", "lpubelts", "lpu-belts"]
FIRESTORE_DOC_PATHS = ["lockCollections/{uid}", "collections/{uid}", "profiles/{uid}"]

# ---------------------------------------------------------------- storage

def ensure_dirs():
    os.makedirs(APP_DIR, exist_ok=True)
    os.makedirs(IMG_DIR, exist_ok=True)

def db():
    """Open a connection with the concurrency pragmas set. Schema creation and
    migration happen once via init_db() (called at startup), so this stays
    lightweight for the many times it's called during normal use."""
    conn = sqlite3.connect(DB_PATH, timeout=30)
    # WAL + busy timeout let the background sync/import threads and the UI
    # touch the database concurrently without "database is locked" errors.
    try:
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA busy_timeout=30000")
        conn.execute("PRAGMA synchronous=NORMAL")
    except sqlite3.Error:
        pass
    return conn

def init_db():
    """Create tables if missing and run one-time migrations. Safe to call more
    than once, but intended to run just once at startup."""
    conn = db()
    try:
        conn.execute("""CREATE TABLE IF NOT EXISTS listings(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            lock_name TEXT, title TEXT, price TEXT, currency TEXT,
            condition TEXT, site TEXT, url TEXT UNIQUE, location TEXT,
            shipping TEXT, notes TEXT, found_at TEXT)""")
        # Append-only history of every listing ever found. Unlike `listings`
        # (the CURRENT results table, wiped at the start of each search), this
        # never gets cleared by searches — it keeps first-seen rows for future
        # features, pruned only by size.
        conn.execute("""CREATE TABLE IF NOT EXISTS listing_history(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            lock_name TEXT, title TEXT, price TEXT, currency TEXT,
            condition TEXT, site TEXT, url TEXT UNIQUE, location TEXT,
            shipping TEXT, notes TEXT, found_at TEXT)""")
        # 4.5.9: listings carry the seller's username where the source knows
        # it (currently only LPU Lock Bazaar rows) — used by the Excel export.
        for _tbl in ("listings", "listing_history"):
            _lcols = [r[1] for r in conn.execute(f"PRAGMA table_info({_tbl})")]
            if "seller" not in _lcols:
                conn.execute(f"ALTER TABLE {_tbl} ADD COLUMN seller TEXT")
        # 4.7.3: listings carry the listing's own image URL where the source
        # exposes it (Facebook, eBay) — used for result thumbnails on hover.
        for _tbl in ("listings", "listing_history"):
            _lcols = [r[1] for r in conn.execute(f"PRAGMA table_info({_tbl})")]
            if "image_url" not in _lcols:
                conn.execute(f"ALTER TABLE {_tbl} ADD COLUMN image_url TEXT")
        conn.execute("""CREATE TABLE IF NOT EXISTS searches(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            lock_name TEXT, condition TEXT, exclude_pickup INTEGER,
            results INTEGER, started_at TEXT, finished_at TEXT, status TEXT)""")
        conn.execute("""CREATE TABLE IF NOT EXISTS locks(
            id TEXT PRIMARY KEY, name TEXT, belt TEXT,
            image_url TEXT, page_url TEXT, source TEXT, belt_full TEXT)""")
        conn.execute("""CREATE TABLE IF NOT EXISTS my_collection(
            lock_id TEXT PRIMARY KEY, status TEXT)""")
        # --- migration: older versions declared locks.name UNIQUE, which breaks
        # the LPU sync (the catalog legitimately repeats a name across belts).
        # If we detect that old schema, rebuild the table without the constraint.
        try:
            ddl = conn.execute(
                "SELECT sql FROM sqlite_master WHERE type='table' AND name='locks'"
            ).fetchone()
            if ddl and "UNIQUE" in ddl[0].upper():
                conn.executescript("""
                    PRAGMA foreign_keys=off;
                    BEGIN;
                    CREATE TABLE locks_new(
                        id TEXT PRIMARY KEY, name TEXT, belt TEXT,
                        image_url TEXT, page_url TEXT, source TEXT);
                    INSERT OR IGNORE INTO locks_new
                        SELECT id,name,belt,image_url,page_url,source FROM locks;
                    DROP TABLE locks;
                    ALTER TABLE locks_new RENAME TO locks;
                    COMMIT;
                    PRAGMA foreign_keys=on;
                """)
        except sqlite3.Error:
            pass
        # Add belt_full (full belt incl. Black sub-tier like "Black 3") if the
        # column doesn't exist yet. `belt` stays canonical ("Black") for
        # filters/colors; belt_full is populated on the next catalog sync.
        try:
            cols = [r[1] for r in conn.execute("PRAGMA table_info(locks)")]
            if "belt_full" not in cols:
                conn.execute("ALTER TABLE locks ADD COLUMN belt_full TEXT")
            # How many LPU collectors own this lock (from the belt explorer's
            # collectionStatsById dataset) — the "rarity" number. Filled on
            # catalog sync; NULL until then.
            if "owner_count" not in cols:
                conn.execute("ALTER TABLE locks ADD COLUMN owner_count INTEGER")
        except sqlite3.Error:
            pass
        conn.commit()
    finally:
        conn.close()

def log(msg):
    ensure_dirs()
    stamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{stamp}] {msg}"
    with open(LOG_PATH, "a", encoding="utf-8") as f:
        f.write(line + "\n")
    return line

def _insert_listing_row(conn, lock_name, it):
    """Insert one found listing into BOTH the current-results table
    (`listings`, wiped per search) and the append-only `listing_history`
    (never wiped by searches; keeps the first-seen row per url). Returns True
    if the row was NEW in `listings` (i.e. not an already-present url)."""
    ship = str(it.get("shipping", "unknown")).lower()
    vals = (lock_name, it.get("title", ""), str(it.get("price", "")),
            it.get("currency", ""), str(it.get("condition", "")).lower(),
            it.get("site", ""), it.get("url", ""), it.get("location", ""),
            ship, it.get("notes", ""), str(it.get("seller", "") or ""),
            str(it.get("image", "") or ""),
            datetime.datetime.now().isoformat(timespec="seconds"))
    before = conn.total_changes
    conn.execute(
        "INSERT OR IGNORE INTO listings(lock_name,title,price,currency,"
        "condition,site,url,location,shipping,notes,seller,image_url,found_at)"
        " VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)", vals)
    is_new = conn.total_changes > before
    try:
        conn.execute(
            "INSERT OR IGNORE INTO listing_history(lock_name,title,price,"
            "currency,condition,site,url,location,shipping,notes,seller,"
            "image_url,found_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)", vals)
    except sqlite3.Error:
        pass   # history is best-effort; never break the live results for it
    return is_new

def _prune_listing_history(conn, keep=20000):
    """Cap the append-only history at `keep` newest rows (best-effort)."""
    try:
        conn.execute(
            "DELETE FROM listing_history WHERE id NOT IN"
            " (SELECT id FROM listing_history ORDER BY id DESC LIMIT ?)",
            (keep,))
    except sqlite3.Error:
        pass

CRASH_LOG_PATH = os.path.join(APP_DIR, "crash_log.txt")

def write_crash_log(exc_type, exc_value, exc_tb, where=""):
    """Append a full, copy-pasteable crash report to ~/.lockhunter/crash_log.txt
    and return its path. Best-effort: never raises. This matters because the
    built .exe runs with no console, so an unhandled error would otherwise
    vanish silently with nothing for the user to send back."""
    try:
        ensure_dirs()
        with open(CRASH_LOG_PATH, "a", encoding="utf-8") as f:
            f.write("=" * 64 + "\n")
            f.write("Lock Hunter crash report\n")
            f.write("Time:     "
                    + datetime.datetime.now().isoformat(timespec="seconds") + "\n")
            f.write(f"Version:  {VERSION}\n")
            try:
                f.write(f"Platform: {platform.platform()}\n")
                f.write(f"Python:   {platform.python_version()}\n")
            except Exception:
                pass
            if where:
                f.write(f"Where:    {where}\n")
            f.write("-" * 64 + "\n")
            traceback.print_exception(exc_type, exc_value, exc_tb, file=f)
            f.write("\n")
        return CRASH_LOG_PATH
    except Exception:
        return None

def load_cfg():
    try:
        with open(CFG_PATH, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

def save_cfg(cfg):
    ensure_dirs()
    with open(CFG_PATH, "w", encoding="utf-8") as f:
        json.dump(cfg, f)

# ---------------------------------------------------------------- LPU sync

CANONICAL_BELTS = ["White", "Yellow", "Orange", "Green", "Blue", "Purple",
                   "Brown", "Red", "Black", "Unranked"]

# Rank for sorting by belt colour: White, Yellow, Orange, Green, Blue, Purple,
# Brown, Red, Black, then Unranked/unknown last.
BELT_RANK = {b: i for i, b in enumerate(CANONICAL_BELTS)}

def _belt_rank(belt):
    """Sort rank for a belt label. Handles Black sub-tiers ('Black 3' -> Black)
    and unknown/blank belts (sorted last)."""
    if not belt:
        return 99
    first = belt.split()[0]            # "Black 3" -> "Black"
    return BELT_RANK.get(first, BELT_RANK.get(belt, 99))

def _canonical_belt(raw):
    """Map a raw belt header from lpubelts.com to one of the ten canonical
    belt colors. 'Black 1'..'Black 5' all fold into 'Black'; Project / Tier /
    Dan / Hall of Fame tiers (which sit past Black) fold into 'Unranked'."""
    raw = (raw or "").strip()
    low = raw.lower()
    if low.startswith("black"):
        return "Black"
    for b in CANONICAL_BELTS:
        if low == b.lower():
            return b
    # Project/Tier/Dan/Dan Points/Hall of Fame and anything else -> Unranked
    return "Unranked"

def _disp_belt(belt, belt_full):
    """Belt label for DISPLAYING a lock: the specific Black sub-tier
    ('Black 3') when we have it, otherwise the plain canonical belt. Belt
    FILTERS still use the canonical belt ('Black' matches Black 1..5); this is
    only for what the user sees when locks are listed."""
    return belt_full if (belt == "Black" and belt_full) else (belt or "")

def _rarity_stars(owner_count):
    """Rarity as stars, over 'how many LPU collectors own it' (user-defined
    bands): 5 stars = Extremely Rare (0-5 owners), 4 = Very Rare (6-15),
    3 = Rare (16-25), 2 = Common (26-35), 1 star = Very Common (36+).
    Locks absent from the stats dataset count as 0 owners."""
    n = 0 if owner_count is None else int(owner_count)
    if n <= 5:
        stars = 5
    elif n <= 15:
        stars = 4
    elif n <= 25:
        stars = 3
    elif n <= 35:
        stars = 2
    else:
        stars = 1
    return "\u2605" * stars


def _lock_id_from_page(url):
    """Pull the LPU lock id (hex) out of a .../locks/<id>.html page URL."""
    m = re.search(r"/locks/([0-9a-fA-F]{6,10})\.html", url or "")
    return m.group(1) if m else None

# The LPU Belt Explorer's own machine-readable dataset (the exact data the
# site is built from): one JSON entry per lock with id, belt (incl. "Black 1"
# sub-tiers), structured make/model pairs, and photo URLs. Richer and far less
# brittle than parsing the rendered all-locks HTML page.
LPU_DATASET_URLS = [
    "https://raw.githubusercontent.com/Lockpickers-United/lpu-belt-explorer/main/src/data/data.json",
    "https://raw.githubusercontent.com/Lockpickers-United/lpu-belt-explorer/master/src/data/data.json",
]
# {lock id: number of LPU collectors who own it} — the "rarity" numbers.
LPU_OWNERCOUNT_URLS = [
    "https://raw.githubusercontent.com/Lockpickers-United/lpu-belt-explorer/main/src/data/collectionStatsById.json",
    "https://raw.githubusercontent.com/Lockpickers-United/lpu-belt-explorer/master/src/data/collectionStatsById.json",
]

def _lpu_slug(text):
    """Rebuild lpubelts.com's ?name= URL slug from a lock's first make+model.
    Verified against the live all-locks page: ASCII letters/digits are kept,
    spaces and '/' become '_', and EVERYTHING else (hyphens, '+', quotes,
    parens, accents like u-umlaut) is dropped. E.g. 'Mul-T-Lock MT5+' ->
    'MulTLock_MT5', 'ABUS 45/30' -> 'ABUS_45_30', 'American Lock 1100 /
    A1100' -> 'American_Lock_1100___A1100'."""
    out = []
    for ch in text or "":
        if ch.isascii() and ch.isalnum():
            out.append(ch)
        elif ch in (" ", "/"):
            out.append("_")
        # anything else: dropped
    return "".join(out)

def _upsert_catalog_rows(rows):
    """Write catalog rows into the locks table. Each row is
    (id, name, belt, belt_full, page_url, image_url, owner_count); image_url
    and owner_count may be None, in which case any existing value is KEPT
    (COALESCE) so the HTML fallback can't wipe dataset-provided extras."""
    conn = db()
    n = 0
    try:
        conn.execute("BEGIN")
        for lid, name, belt, belt_full, page, img, cnt in rows:
            conn.execute(
                "INSERT INTO locks(id,name,belt,belt_full,image_url,page_url,"
                "source,owner_count) VALUES(?,?,?,?,?,?,?,?)"
                " ON CONFLICT(id) DO UPDATE SET"
                " name=excluded.name, belt=excluded.belt,"
                " belt_full=excluded.belt_full, page_url=excluded.page_url,"
                " image_url=COALESCE(excluded.image_url, locks.image_url),"
                " owner_count=COALESCE(excluded.owner_count, locks.owner_count)",
                (lid, name, belt, belt_full, img, page, "lpu", cnt))
            n += 1
        conn.commit()
    finally:
        conn.close()
    return n

def _sync_from_lpu_dataset(status_cb):
    """PRIMARY catalog source: the belt explorer's data.json. Returns
    (count, source_url); raises on any failure so the caller can fall back to
    the all-locks HTML page."""
    status_cb("Downloading the LPU catalog (belt-explorer dataset)...")
    data, used = None, None
    last_err = None
    for u in LPU_DATASET_URLS:
        try:
            r = requests.get(u, timeout=60,
                             headers={"User-Agent": "LockHunter/1.0"})
            if r.status_code == 200:
                data = r.json()
                used = u
                break
            last_err = f"HTTP {r.status_code}"
        except Exception as ex:
            last_err = str(ex)
    if not isinstance(data, list) or len(data) < 100:
        raise RuntimeError(last_err or "dataset empty/unexpected")
    # rarity numbers are a nice-to-have: never fail the sync over them
    counts = {}
    for u in LPU_OWNERCOUNT_URLS:
        try:
            r = requests.get(u, timeout=30,
                             headers={"User-Agent": "LockHunter/1.0"})
            if r.status_code == 200 and isinstance(r.json(), dict):
                counts = {str(k): int(v) for k, v in r.json().items()
                          if isinstance(v, (int, float))}
                break
        except Exception:
            pass
    rows = []
    for e in data:
        if not isinstance(e, dict):
            continue
        lid = str(e.get("id") or "").strip()
        mms = e.get("makeModels") or []
        parts = []
        for mm in mms:
            if isinstance(mm, dict):
                t = f'{(mm.get("make") or "").strip()} ' \
                    f'{(mm.get("model") or "").strip()}'.strip()
                if t:
                    parts.append(re.sub(r"\s+", " ", t))
        name = " / ".join(parts)
        if not lid or not name:
            continue
        raw_belt = str(e.get("belt") or "").strip()
        belt = _canonical_belt(raw_belt)
        belt_full = raw_belt if belt == "Black" else belt
        page = (f"https://lpubelts.com/locks/{lid}.html"
                f"?name={_lpu_slug(parts[0])}")
        img = None
        media = e.get("media") or []
        if media and isinstance(media[0], dict):
            img = (media[0].get("thumbnailUrl")
                   or media[0].get("fullSizeUrl") or None)
            if img and not str(img).startswith("http"):
                img = None
        rows.append((lid, name, belt, belt_full, page, img, counts.get(lid)))
    if len(rows) < 100:
        raise RuntimeError("dataset parsed but yielded too few locks")
    n = _upsert_catalog_rows(rows)
    status_cb(f"Catalog dataset loaded: {n} locks "
              f"({sum(1 for r in rows if r[6] is not None)} with collector "
              f"counts).")
    return n, used

def sync_lpu_catalog(status_cb):
    """Refresh the lock catalog. Tries the belt explorer's own JSON dataset
    first (richer: exact Black sub-tiers, photos, collector counts); if that
    can't be fetched or parsed, falls back to scraping the all-locks HTML
    page exactly as before. Returns (count, source_url)."""
    try:
        return _sync_from_lpu_dataset(status_cb)
    except Exception as ex:
        status_cb(f"Belt-explorer dataset unavailable ({ex}) — falling back "
                  f"to the all-locks page...")
        return _sync_from_alllocks_html(status_cb)

def _sync_from_alllocks_html(status_cb):
    """FALLBACK catalog source: parse the LPU Belt Explorer 'all locks' page —
    a stable, public HTML index listing every lock grouped by belt, with name
    + page URL. Used when the richer belt-explorer dataset (see
    sync_lpu_catalog) can't be fetched. Belt page URLs also yield the lock id,
    and each lock's photo is loaded lazily later from its own page.
    Returns (count, source_url)."""
    url = "https://lpubelts.com/locks/all-locks.html"
    status_cb("Downloading LPU catalog (all-locks index)...")
    r = requests.get(url, timeout=60, headers={"User-Agent": "LockHunter/1.0"})
    if r.status_code != 200:
        raise RuntimeError(f"LPU sync failed: all-locks.html -> HTTP {r.status_code}")
    # Decode the raw bytes as UTF-8 ourselves: if the server omits a charset
    # header, the HTTP library falls back to Latin-1 and mangles accents
    # (é -> Ã©). errors="replace" keeps parsing alive on any stray byte.
    text = r.content.decode("utf-8", errors="replace")
    # Current belt as we walk down the page; headers look like "## Purple Belt"
    # or in raw HTML <h2>Purple Belt</h2>. Lock lines are links to
    # /locks/<id>.html?name=<Name>.
    belt = ""
    belt_full = ""
    rows = []
    # Real belt section headers are exactly like "## White Belt", "## Black 5
    # Belt", "## Unranked Belt" — the belt word is the LAST word on the line.
    # We must NOT match the page title "# LPU Belt Explorer - All Locks"
    # (which contains the word "Belt" mid-line). So we anchor on end-of-line
    # and restrict to the known belt vocabulary.
    belt_words = (r"White|Yellow|Orange|Green|Blue|Purple|Brown|Red|Black|"
                  r"Unranked|Project|Tier|Dan|Dan Points|Hall of Fame")
    header_re = re.compile(
        r"^\s*#{1,6}\s*((?:" + belt_words + r")(?:\s+\d+)?(?:\s+Points\s+\d+)?)"
        r"\s+Belt\s*$"
        r"|<h[1-6][^>]*>\s*((?:" + belt_words + r")(?:\s+\d+)?)\s+Belt\s*</h[1-6]>")
    link_re = re.compile(
        r'https?://lpubelts\.com/locks/([0-9a-fA-F]{6,10})\.html\?name=([^\s"\')<]+)')
    for line in re.split(r"[\r\n]+", text):
        hm = header_re.search(line)
        if hm:
            raw_belt = (hm.group(1) or hm.group(2) or "").strip()
            belt = _canonical_belt(raw_belt)
            # keep the full belt (incl. "Black 1".."Black 5") for display;
            # for non-black belts this equals the canonical belt.
            belt_full = raw_belt if belt == "Black" else belt
            continue
        for lm in link_re.finditer(line):
            lid, raw_name = lm.group(1), lm.group(2)
            # Prefer the human link text: markdown [Name](url) first, then
            # HTML anchor text >Name</a>. Only fall back to the URL slug —
            # percent-DEcoded, so S%C3%A9mag comes back as Sémag, not junk.
            disp = re.search(r'\[([^\]]+)\]\(\s*' + re.escape(lm.group(0)), line)
            if not disp:
                disp = re.search(re.escape(lm.group(0)) +
                                 r'[^>]*>\s*([^<]+?)\s*</a>', line)
            if disp:
                name = disp.group(1)
            else:
                name = urllib.parse.unquote(raw_name).replace("_", " ")
            import html as _htmlmod
            name = _htmlmod.unescape(name)   # &eacute; / &#233; -> é
            name = re.sub(r"\s+", " ", name).strip()
            page = lm.group(0)
            rows.append((lid, name, belt, belt_full, page))
    if not rows:
        raise RuntimeError("LPU sync failed: no locks parsed from all-locks.html")
    # image_url / owner_count stay untouched (COALESCE) — the HTML page has
    # neither, and we must not wipe values a dataset sync already stored.
    n = _upsert_catalog_rows(
        [(lid, name, belt, belt_full, page, None, None)
         for lid, name, belt, belt_full, page in rows])
    return n, url

def fetch_lock_image(page_url, status_cb=lambda s: None):
    """Pull a lock's photo URL from its own LPU page (og:image / Flickr),
    cache it into the locks table, and return it. Called lazily on select."""
    try:
        r = requests.get(page_url, timeout=30, headers={"User-Agent": "LockHunter/1.0"})
        if r.status_code != 200:
            return None
        html = r.text
        m = (re.search(r'og:image["\']?\s*content=["\']([^"\']+)', html)
             or re.search(r'content=["\']([^"\']+)["\'][^>]*og:image', html)
             or re.search(r'(https://live\.staticflickr\.com/[^\s"\')<]+)', html))
        if not m:
            return None
        img = m.group(1)
        if "LPU-f691d3fe" in img or "/assets/" in img:
            return None  # site logo, not a lock photo
        conn = db()
        try:
            conn.execute("UPDATE locks SET image_url=? WHERE page_url=?", (img, page_url))
            conn.commit()
        finally:
            conn.close()
        return img
    except Exception as ex:
        status_cb(f"Image fetch failed: {ex}")
        return None

# ---------------------------------------------------------------- Lock Bazaar

_bazaar_cache = {"data": None, "at": None, "url": None}

def _discover_bazaar_feed():
    """Find the Lock Bazaar data file URL by scraping lpulocks.com's built JS
    for a bazaar/*.json (or similar) reference. Returns a list of candidate
    URLs. Uses a browser User-Agent since the site blocks non-browser agents."""
    found = []
    ua = {"User-Agent": _BROWSER_UA,
          "Accept": "text/html,application/xhtml+xml,*/*;q=0.8"}
    try:
        html = requests.get("https://lpulocks.com", headers=ua, timeout=30).text
        # Vite build references its JS as /assets/*.js and often src=/src/*.jsx
        js = re.findall(r'(?:src|href)=[\'"]([^\'"]+\.js)[\'"]', html)
        js += re.findall(r'[\'"](/assets/[^\'"]+\.js)[\'"]', html)
        seen = set()
        for ju in js[:20]:
            full = ju if ju.startswith("http") else "https://lpulocks.com/" + ju.lstrip("/")
            if full in seen:
                continue
            seen.add(full)
            try:
                txt = requests.get(full, headers=ua, timeout=30).text
            except Exception:
                continue
            # any .json path that looks like bazaar/listing data
            for m in re.finditer(
                    r'["\']([^"\']*(?:bazaar|listing|marketplace|forsale|data)[^"\']*\.json)["\']',
                    txt, re.I):
                path = m.group(1)
                url = path if path.startswith("http") else "https://lpulocks.com/" + path.lstrip("/")
                if url not in found:
                    found.append(url)
    except Exception:
        pass
    return found

def _bazaar_dataset(status_cb):
    """Download (and briefly cache) the Lock Bazaar dataset. Returns list|None.
    Tries the known URLs first, then any URL discovered from the live site."""
    now = datetime.datetime.now()
    if _bazaar_cache["data"] is not None and _bazaar_cache["at"] and (
            (now - _bazaar_cache["at"]).total_seconds() < 600):
        return _bazaar_cache["data"]
    candidates = list(BAZAAR_DATA_URLS) + _discover_bazaar_feed()
    ua = {"User-Agent": _BROWSER_UA, "Accept": "application/json,*/*;q=0.8"}
    for url in candidates:
        try:
            status_cb(f"Checking Lock Bazaar feed: {url}")
            r = requests.get(url, headers=ua, timeout=45)
            if r.status_code != 200:
                continue
            data = r.json()
            if isinstance(data, dict):
                # Field names VERIFIED against the lock-trackers generator
                # (lpulocks-node/src/getLockbazaarListings.js):
                #   validListings = FLAT per-listing rows carrying
                #     name (= seller username), sellerId (= LPU profile id),
                #     make, model, belt, price, url, available, listingType…
                #   allEntries    = per-LOCK groups whose seller rows sit in a
                #     nested `listings` list keyed sellerName.
                # Prefer validListings — it is exactly the flat shape every
                # matcher here expects. (The old order looked for a
                # non-existent `listings` key and then flattened allEntries
                # WITHOUT descending into its nested seller rows, which is
                # why ★ for-sale sellers never showed up.)
                if isinstance(data.get("validListings"), list) and data["validListings"]:
                    data = data["validListings"]
                elif isinstance(data.get("listings"), list) and data["listings"]:
                    data = data["listings"]
                elif isinstance(data.get("allEntries"), list) and data["allEntries"]:
                    data = _flatten_bazaar_entries(data["allEntries"])
                else:
                    for k in ("entries", "data", "locks", "lots", "items"):
                        if isinstance(data.get(k), list):
                            data = data[k]
                            break
            if isinstance(data, list):
                # non-lock listings (tools, gift certificates) must not be
                # token-matched against lock names
                data = [e for e in data
                        if not (isinstance(e, dict) and
                                str(e.get("listingType") or "").strip()
                                in ("Tools", "Gift Certificate"))]
            if isinstance(data, list) and data:
                _bazaar_cache.update(data=data, at=now, url=url)
                try:
                    k0 = sorted((data[0] or {}).keys())[:8]
                    log(f"Bazaar feed loaded: {len(data)} listings from "
                        f"{url} (row keys: {', '.join(map(str, k0))}...)")
                except Exception:
                    pass
                return data
        except Exception:
            continue
    return None

def _flatten_bazaar_entries(entries):
    """Flatten `allEntries` (per-LOCK groups) into flat per-seller rows.
    VERIFIED shape: each entry has `belt`, a `makeModels` list, and a nested
    `listings` list whose rows carry the seller as `sellerName` plus
    price/url/condition. Older/simpler shapes (seller at entry level) are
    kept as a fallback."""
    out = []
    for e in entries:
        if not isinstance(e, dict):
            continue
        belt = e.get("belt", "")
        mms = e.get("makeModels") if isinstance(e.get("makeModels"), list) else []
        nested = e.get("listings")
        if isinstance(nested, list) and nested:
            for L in nested:
                if not isinstance(L, dict):
                    continue
                base = {"belt": L.get("belt") or belt,
                        "name": (L.get("sellerName") or L.get("name") or ""),
                        "sellerId": L.get("sellerId", ""),
                        "price": L.get("price", ""),
                        "url": L.get("url", ""),
                        "condition": L.get("condition", ""),
                        "country": L.get("country", ""),
                        "available": L.get("available", ""),
                        "listingType": L.get("listingType", ""),
                        "rowNum": L.get("rowNum", "")}
                if L.get("make") or L.get("model"):
                    out.append(dict(base, make=L.get("make", ""),
                                    model=L.get("model", "")))
                elif mms:
                    for mm in mms:
                        if isinstance(mm, dict):
                            out.append(dict(base, make=mm.get("make", ""),
                                            model=mm.get("model", "")))
                else:
                    out.append(dict(base, make=e.get("make", ""),
                                    model=e.get("model", "")))
            continue
        if mms:
            for mm in mms:
                if isinstance(mm, dict):
                    out.append({"make": mm.get("make", ""),
                                "model": mm.get("model", ""),
                                "belt": belt,
                                "url": e.get("url", ""),
                                "price": e.get("price", ""),
                                "condition": e.get("condition", ""),
                                "country": e.get("country", ""),
                                "sellerId": e.get("sellerId", ""),
                                "name": (e.get("sellerName")
                                         or e.get("name", ""))})
        else:
            out.append(e)
    return out

def search_bazaar_direct(lock_name, status_cb, belt=""):
    """Match a lock against the live LPU Lock Bazaar feed. Returns a list of
    listing dicts (in the same shape as the AI results), or None if the feed
    is unreachable.

    Production feed rows (validListings) carry the lock as sheetMake +
    sheetModel plus the canonical lpubeltsName, the seller as sellerName, the
    belt as lockBelt, and the LPU lock id inside `id` - all read through the
    _bz_* helpers.
    """
    data = _bazaar_dataset(status_cb)
    if data is None:
        return None
    # tokens from the searched lock name (drop 1-2 char noise)
    tokens = _lock_tokens(lock_name)
    belt_l = (belt or "").strip().lower()
    out = []
    for e in data:
        if not isinstance(e, dict):
            continue
        make = _bz_make(e)
        model = _bz_model(e)
        lock_text = _fold(f"{make} {model} {_bz_lockname(e)}")
        # every search token must appear in the make+model text
        if tokens and not _match_tokens(lock_text, tokens):
            continue
        # optional belt filter (e.g. "Brown", "Black 2" -> match on "black")
        if belt_l:
            entry_belt = _bz_belt(e).lower()
            base_belt = belt_l.split()[0] if belt_l else belt_l
            if base_belt and base_belt not in entry_belt:
                continue
        title = _bz_lockname(e) or f"{make} {model}".strip() or lock_name
        price = str(e.get("price") or "").strip()
        seller = _bz_seller(e)   # validListings: `name`; nested rows: `sellerName`
        ships = e.get("shipsTo")
        ships_txt = ", ".join(ships) if isinstance(ships, list) else str(ships or "")
        cond = str(e.get("condition") or "").strip().lower()
        cond = "new" if "new" in cond else "used"
        fmt = str(e.get("format") or "").strip()
        note_bits = [b for b in (
            f"seller: {seller}" if seller else "",
            f"format: {fmt}" if fmt else "",
            str(e.get("notes") or "").strip(),
        ) if b]
        out.append({
            "title": title,
            "price": price,
            "currency": "",   # price string already includes the symbol
            "condition": cond,
            "site": "LPU Lock Bazaar",
            "url": _bz_url(e),
            "location": str(e.get("country") or ""),
            "shipping": "yes" if ships_txt else "unknown",
            "notes": "  ·  ".join(note_bits),
            "seller": seller,
        })
    return out

def bazaar_entries_for_lock(lock_name, dbelt, data):
    """All Lock Bazaar entries that match a CATALOG lock. Uses the same
    every-token rule as search_bazaar_direct, but runs it over each
    slash-separated component of multi-name catalog entries (so a bazaar
    'Fichet 450' listing matches the catalog's 'Fichet 450 / Fichet 484').
    Narrowed by the lock's canonical belt word when the entry states one."""
    base_belt = (dbelt or "").split()[0].lower()
    comp_tokens = []
    for comp in _split_catalog_names(lock_name):
        toks = _lock_tokens(comp)
        if toks:
            comp_tokens.append(toks)
    out = []
    for e in (data or []):
        if not isinstance(e, dict):
            continue
        lock_text = _fold(f"{_bz_make(e)} {_bz_model(e)} {_bz_lockname(e)}")
        if not lock_text.strip():
            continue
        if not any(_match_tokens(lock_text, toks) for toks in comp_tokens):
            continue
        if base_belt:
            eb = _bz_belt(e).lower()
            if eb and base_belt not in eb:
                continue
        out.append(e)
    return out

def _bz_seller(e):
    """Seller username of a Bazaar row. The LIVE feed's validListings rows
    carry it as `sellerName` (verified from the production data file); older
    shapes used `name`."""
    return str(e.get("sellerName") or e.get("name") or "").strip()

def _bz_make(e):
    return str(e.get("make") or e.get("sheetMake") or "").strip()

def _bz_model(e):
    return str(e.get("model") or e.get("sheetModel") or "").strip()

def _bz_belt(e):
    return str(e.get("belt") or e.get("lockBelt") or "").strip()

def _bz_lockname(e):
    """The canonical LPU catalog name the listing was matched to
    (`lpubeltsName` in the production feed)."""
    return str(e.get("lpubeltsName") or "").strip()

def _bz_lock_id(e):
    """The LPU lock id embedded in a listing's `id` ('cba41cb7-2' ->
    'cba41cb7'). '' when absent."""
    s = str(e.get("id") or "").strip()
    return s.split("-")[0] if s else ""

def _bz_price(e):
    return str(e.get("price") or "").strip()

def _bz_url(e):
    """A clickable link for the listing: its own share url when present, else
    the Bazaar deep-link for its lock (format verified from the site source:
    /#/lockbazaar?id=<lockId>&name=<name>)."""
    u = str(e.get("url") or "").strip()
    if u:
        return u
    lid = _bz_lock_id(e)
    if not lid:
        return "https://lpulocks.com/#/lockbazaar"
    nm = urllib.parse.quote(_bz_lockname(e)
                            or f"{_bz_make(e)} {_bz_model(e)}".strip())
    return f"https://lpulocks.com/#/lockbazaar?id={lid}&name={nm}"

def bazaar_entries_for_lock_id(lock_id, data):
    """All Bazaar listings for an exact LPU lock id — the precise match,
    since production rows embed the lock id they were validated against."""
    if not lock_id:
        return []
    return [e for e in (data or [])
            if isinstance(e, dict) and _bz_lock_id(e) == lock_id]

def _bz_profile_url(e, display=None):
    """LPU profile link for a Bazaar row's seller, built from the feed's
    sellerId (their Firestore/LPU profile id). '' when absent."""
    sid = str(e.get("sellerId") or "").strip()
    if not sid:
        return ""
    safe = (display or _bz_seller(e)).replace(" ", "_")
    return f"https://lpubelts.com/#/profile/{sid}?name={safe}&collection=Own"

_BZ_SNAP_VER = 2   # bump when _bazaar_entry_key's format changes, so an old
                   # snapshot is re-baselined silently instead of mis-diffed.

def _bazaar_entry_key(e):
    """Stable identity for a Bazaar listing across feed refreshes.

    Preference order:
      1. the share `url` when present (already unique + stable), else
      2. a normalized CONTENT signature: seller + canonical lock + price,
         with the price reduced to its digits so '€15' and '€ 15' don't
         diverge.

    Deliberately does NOT use `rowNum` (or the trailing index on `id`): those
    encode the listing's POSITION in the aggregated feed, which shifts
    whenever any other seller's listing is added or removed above it. Keying
    on position made unchanged listings reappear as 'new' on every launch —
    identity must come from what the listing IS, not where it happens to sit
    in this refresh."""
    u = str(e.get("url") or "").strip()
    if u:
        return "u:" + u
    seller = _bz_seller(e).lower()
    lock = (_bz_lockname(e) or f"{_bz_make(e)} {_bz_model(e)}").strip().lower()
    lock = re.sub(r"\s+", " ", lock)
    price_digits = re.sub(r"[^0-9]", "", _bz_price(e))
    return "c:" + "|".join([seller, lock, price_digits])

# ------------------------------------------------------- LPU profile import

def _discover_firebase_from_site(status_cb):
    """Scrape lpubelts.com's JS bundles for the live Firebase apiKey+projectId.
    This is the config the real web app uses, so its Firestore reads match
    exactly what a browser is allowed to do. Also collects any Firestore
    collection names referenced in the bundle (collection("...") calls), which
    tells us where per-user lists actually live."""
    out = []
    colls = []
    try:
        html = requests.get("https://lpubelts.com",
                            headers={"User-Agent": "LockHunter/1.0"}, timeout=30).text
        js = re.findall(r'src=[\'"]([^\'"]+\.js)', html)
        # also check common Vite bundle locations
        for ju in js[:12]:
            full = ju if ju.startswith("http") else "https://lpubelts.com/" + ju.lstrip("/")
            try:
                txt = requests.get(full, headers={"User-Agent": "LockHunter/1.0"},
                                   timeout=30).text
            except Exception:
                continue
            for m in re.finditer(
                    r'apiKey["\']?\s*[:=]\s*["\']([^"\']+)["\'][^}]{0,400}?'
                    r'projectId["\']?\s*[:=]\s*["\']([^"\']+)["\']', txt, re.S):
                out.append((m.group(2), m.group(1)))
            for m in re.finditer(
                    r'projectId["\']?\s*[:=]\s*["\']([^"\']+)["\'][^}]{0,400}?'
                    r'apiKey["\']?\s*[:=]\s*["\']([^"\']+)["\']', txt, re.S):
                out.append((m.group(1), m.group(2)))
            # collection("name") / collection(db,"name")
            for m in re.finditer(r'collection\(\s*[\w$.]+\s*,\s*["\']([A-Za-z0-9_]+)["\']', txt):
                colls.append(m.group(1))
            for m in re.finditer(r'collection\(\s*["\']([A-Za-z0-9_]+)["\']', txt):
                colls.append(m.group(1))
    except Exception as ex:
        status_cb(f"Firebase discovery failed: {ex}")
    # de-dupe collection names, keep order
    seen = set(); ordered = []
    for c in colls:
        if c not in seen:
            seen.add(c); ordered.append(c)
    _discover_firebase_from_site.collections = ordered
    return out
_discover_firebase_from_site.collections = []

def _firebase_params(status_cb):
    """Yield (projectId, apiKey|None) candidates. Live-site config first
    (has the API key Firestore requires), then repo config, then guesses.
    The LPU project id is confirmed to be 'lpu-belt-explorer', so we also
    cross-pair that id with any API key we discover, in case the minified
    bundle paired the key with a different id."""
    seen = set()
    keys_found = []
    def _emit(pid, key):
        tag = (pid, key)
        if tag not in seen:
            seen.add(tag)
            return True
        return False

    # Confirmed live credentials first — most reliable.
    if _emit(LPU_PROJECT_ID, LPU_WEB_API_KEY):
        status_cb(f"Using LPU Firebase project: {LPU_PROJECT_ID}")
        keys_found.append(LPU_WEB_API_KEY)
        yield LPU_PROJECT_ID, LPU_WEB_API_KEY

    for pid, key in _discover_firebase_from_site(status_cb):
        if key and key not in keys_found:
            keys_found.append(key)
        if _emit(pid, key):
            status_cb(f"Found live Firebase config: {pid}")
            yield pid, key
    # cross-pair the confirmed project id with every discovered key
    for key in keys_found:
        if _emit("lpu-belt-explorer", key):
            yield "lpu-belt-explorer", key
    for url in FIREBASE_CONFIG_URLS:
        try:
            r = requests.get(url, timeout=30)
            if r.status_code != 200:
                continue
            pid = re.search(r"projectId['\"]?\s*[:=]\s*['\"]([^'\"]+)", r.text)
            key = re.search(r"apiKey['\"]?\s*[:=]\s*['\"]([^'\"]+)", r.text)
            if pid:
                if _emit(pid.group(1), key.group(1) if key else None):
                    yield pid.group(1), key.group(1) if key else None
                # also cross-pair confirmed project with this key
                if key and _emit("lpu-belt-explorer", key.group(1)):
                    yield "lpu-belt-explorer", key.group(1)
        except Exception:
            continue
    for pid in FIREBASE_FALLBACK_PROJECTS:
        if not any(pid == p for p, _ in seen):
            yield pid, None

def _fs_value(v):
    """Decode a Firestore REST value object into plain Python."""
    if not isinstance(v, dict):
        return v
    for k in ("stringValue", "integerValue", "doubleValue", "booleanValue"):
        if k in v:
            return v[k]
    if "arrayValue" in v:
        return [_fs_value(x) for x in (v["arrayValue"].get("values") or [])]
    if "mapValue" in v:
        return {k2: _fs_value(v2) for k2, v2 in
                (v["mapValue"].get("fields") or {}).items()}
    return None

def _collect_ids(obj, own, wish, path=""):
    """Walk decoded fields; route arrays to own/wish by field-name heuristic."""
    if isinstance(obj, dict):
        for k, v in obj.items():
            _collect_ids(v, own, wish, k.lower())
    elif isinstance(obj, list):
        ids = [str(x) for x in obj if isinstance(x, (str, int))]
        if "wish" in path or "want" in path:
            wish.update(ids)
        elif "own" in path or path in ("collection", "have"):
            own.update(ids)

def extract_profile_uid(profile_url):
    m = re.search(r"profile/([A-Za-z0-9_-]{10,})", profile_url)
    return m.group(1) if m else None

def _anon_sign_in(key, status_cb=lambda s: None):
    """Get an anonymous Firebase ID token via Identity Toolkit. LPU's Firestore
    rules require a signed-in user (even anonymous), which is why raw API-key
    reads get 403. The website does the same anonymous sign-in under the hood.
    Returns an ID token string, or None."""
    url = ("https://identitytoolkit.googleapis.com/v1/accounts:signUp?key=" + key)
    try:
        r = requests.post(url, json={"returnSecureToken": True}, timeout=25)
        if r.status_code == 200:
            return r.json().get("idToken")
        status_cb(f"Anonymous sign-in returned HTTP {r.status_code}")
    except Exception as ex:
        status_cb(f"Anonymous sign-in failed: {ex}")
    return None

def _fs_headers(token):
    return {"Authorization": "Bearer " + token} if token else {}


def _fs_runquery(pid, key, collection, field, op, value, status_cb,
                 group=False, token=None):
    """Run a Firestore structured query against a collection, filtering where
    <field> <op> <value>. When group=True, run it as a collection-group query
    (matches that collection id at any nesting depth). Returns the list of
    decoded document field-dicts, [] if the query ran but matched nothing, or
    None if the request itself failed."""
    url = ("https://firestore.googleapis.com/v1/projects/" + pid +
           "/databases/(default)/documents:runQuery?key=" + key)
    body = {
        "structuredQuery": {
            "from": [{"collectionId": collection, "allDescendants": bool(group)}],
            "where": {
                "fieldFilter": {
                    "field": {"fieldPath": field},
                    "op": op,
                    "value": value,
                }
            },
            "limit": 5000,
        }
    }
    try:
        r = requests.post(url, json=body, headers=_fs_headers(token), timeout=30)
        if r.status_code != 200:
            return None
        out = []
        for row in r.json():
            doc = row.get("document")
            if doc and doc.get("fields"):
                out.append({k: _fs_value(v) for k, v in doc["fields"].items()})
        return out
    except Exception:
        return None

def _fs_list_collection_ids(pid, key, parent="", status_cb=lambda s: None, token=None):
    """Ask Firestore for the collection ids under a document/root. Lets us
    discover the real collection name instead of guessing. parent is "" for
    root, or "<coll>/<docid>" for a document's sub-collections."""
    base = ("https://firestore.googleapis.com/v1/projects/" + pid +
            "/databases/(default)/documents")
    url = base + (("/" + parent) if parent else "") + ":listCollectionIds?key=" + key
    try:
        r = requests.post(url, json={}, headers=_fs_headers(token), timeout=25)
        if r.status_code != 200:
            return []
        return r.json().get("collectionIds", []) or []
    except Exception:
        return []

def _fs_get_doc(pid, key, path, token=None):
    """Read a single Firestore document by path; return decoded fields or None."""
    url = ("https://firestore.googleapis.com/v1/projects/" + pid +
           "/databases/(default)/documents/" + path + "?key=" + key)
    try:
        r = requests.get(url, headers=_fs_headers(token), timeout=25)
        if r.status_code != 200:
            return None, r.status_code
        doc = r.json()
        return {k: _fs_value(v) for k, v in (doc.get("fields") or {}).items()}, 200
    except Exception:
        return None, 0

def _extract_lock_ids(docs):
    """From a list of decoded Firestore docs, pull out every lock id we can
    find (values that look like the 6-10 char hex lock ids used by LPU)."""
    ids = set()
    hexid = re.compile(r"^[0-9a-fA-F]{6,10}$")
    def walk(o):
        if isinstance(o, dict):
            for v in o.values():
                walk(v)
        elif isinstance(o, list):
            for v in o:
                walk(v)
        elif isinstance(o, str) and hexid.match(o):
            ids.add(o)
    walk(docs)
    return ids

def _entry_ids_from_doc(d):
    """Pull lock ids from a decoded list document. LPU list docs typically hold
    an 'entries' array (each item an object with an 'id', or a bare id string)."""
    ids = set()
    for key in ("entries", "locks", "items", "ids", "list", "entryIds"):
        v = d.get(key)
        if isinstance(v, list):
            for it in v:
                if isinstance(it, str):
                    ids.add(it)
                elif isinstance(it, dict):
                    for idk in ("id", "entryId", "lockId"):
                        if it.get(idk):
                            ids.add(str(it[idk])); break
    # fall back to any hex-looking ids anywhere in the doc
    if not ids:
        ids = _extract_lock_ids(d)
    return ids

def _list_kind(d, fallback=""):
    """Classify a list doc as 'own' or 'wish' from its name/type fields."""
    for key in ("name", "listName", "type", "collection", "collectionType", "title"):
        val = str(d.get(key) or "").lower()
        if "wish" in val or "want" in val:
            return "wish"
        if "own" in val or "have" in val or "collect" in val:
            return "own"
    fl = fallback.lower()
    if "wish" in fl:
        return "wish"
    if "own" in fl:
        return "own"
    return ""

def import_lpu_profile(profile_url, status_cb):
    """Fetch a public LPU profile's Own + Wishlist lists from Firestore.
    Returns (own_ids, wish_ids, uid).

    LPU profiles are shared publicly at .../#/profile/<uid>?collection=Any, so
    the underlying list documents are world-readable through Firestore's REST
    API using the site's own web API key. We discover that config (and the
    collection names) from the live JS bundle, then try, in order:
      1. collection-group / collection queries filtered by the owner uid
      2. direct document reads at <collection>/<uid> and suffixed variants
    keeping whatever yields lock ids."""
    uid = extract_profile_uid(profile_url)
    if not uid:
        raise ValueError("Could not find a profile ID in that URL. Expected "
                         ".../#/profile/<id>")

    # The confirmed LPU layout (from the live Firestore stream): a single
    # document at  lockcollections/<uid>  whose fields include:
    #   own        -> array of lock-id strings
    #   wishlist   -> array of lock-id strings
    #   safelocksOwn -> array of safe-lock ids (also treated as owned)
    # We read that document directly. Everything else below is a fallback for
    # robustness in case the schema ever changes.
    primary_collections = ["lockcollections", "lockCollections"]
    known = primary_collections + ["entryLists", "entrylists", "lists",
                                   "userLists", "collections", "entries",
                                   "profiles", "users"]
    _discover_firebase_from_site(status_cb)  # populates .collections (fallback)
    discovered = list(getattr(_discover_firebase_from_site, "collections", []) or [])
    collections = known + [c for c in discovered if c not in known]
    uid_fields = ["createdBy", "owner", "ownerId", "userId", "uid", "user", "author"]

    saw_403 = False
    reached_firestore = False
    tried = []
    for pid, key in _firebase_params(status_cb):
        if not key:
            continue  # Firestore REST needs the api key
        status_cb(f"Reading your LPU collection ({pid})…")
        # LPU's rules require a signed-in user; get an anonymous token so our
        # REST reads are treated the same way the website's reads are.
        token = _anon_sign_in(key, status_cb)
        own, wish = set(), set()

        # PRIMARY: read lockcollections/<uid> directly and split own/wishlist.
        for coll in primary_collections:
            fields, code = _fs_get_doc(pid, key, f"{coll}/{uid}", token=token)
            if code == 403:
                saw_403 = True
            if code == 200 and fields:
                reached_firestore = True
                for own_field in ("own", "safelocksOwn"):
                    v = fields.get(own_field)
                    if isinstance(v, list):
                        own.update(str(x) for x in v if isinstance(x, (str, int)))
                w = fields.get("wishlist")
                if isinstance(w, list):
                    wish.update(str(x) for x in w if isinstance(x, (str, int)))
                if own or wish:
                    break
        if own or wish:
            save_msg = f"Loaded profile from {pid}: {len(own)} own, {len(wish)} wish"
            status_cb(save_msg)
            return own, wish, uid

        # FALLBACK A: discover root collections and per-user subcollections.
        root_ids = _fs_list_collection_ids(pid, key, "", status_cb, token=token)
        if root_ids:
            reached_firestore = True
            for c in root_ids:
                if c not in collections:
                    collections.append(c)
            status_cb("LPU collections: " + ", ".join(root_ids[:8]))

        # FALLBACK B: collection queries filtered by the owner uid.
        for grp in (False, True):
            for coll in collections:
                for f in uid_fields:
                    docs = _fs_runquery(pid, key, coll, f, "EQUAL",
                                        {"stringValue": uid}, status_cb,
                                        group=grp, token=token)
                    if docs is None:
                        continue
                    reached_firestore = True
                    tried.append(f"{coll}.{f}={len(docs)}")
                    for d in docs:
                        # doc may itself hold own/wishlist arrays
                        for own_field in ("own", "safelocksOwn"):
                            v = d.get(own_field)
                            if isinstance(v, list):
                                own.update(str(x) for x in v if isinstance(x, (str, int)))
                        w = d.get("wishlist")
                        if isinstance(w, list):
                            wish.update(str(x) for x in w if isinstance(x, (str, int)))
                        # or be a per-list doc
                        kind = _list_kind(d)
                        ids = _entry_ids_from_doc(d)
                        if ids:
                            (wish if kind == "wish" else own).update(ids) if kind else _collect_ids(d, own, wish)
                    if own or wish:
                        break
                if own or wish:
                    break
            if own or wish:
                break

        # FALLBACK C: direct reads at <coll>/<uid> with Own/Wishlist suffixes.
        if not (own or wish):
            for coll in collections:
                for suffix in ("", "_Own", "_Wishlist", "-Own", "-Wishlist",
                               "/Own", "/Wishlist"):
                    fields, code = _fs_get_doc(pid, key, f"{coll}/{uid}{suffix}",
                                               token=token)
                    if code == 403:
                        saw_403 = True; continue
                    if code != 200 or not fields:
                        continue
                    reached_firestore = True
                    for own_field in ("own", "safelocksOwn"):
                        v = fields.get(own_field)
                        if isinstance(v, list):
                            own.update(str(x) for x in v if isinstance(x, (str, int)))
                    w = fields.get("wishlist")
                    if isinstance(w, list):
                        wish.update(str(x) for x in w if isinstance(x, (str, int)))
                    kind = _list_kind(fields, fallback=suffix)
                    ids = _entry_ids_from_doc(fields)
                    if ids:
                        (wish if kind == "wish" else own).update(ids) if kind else _collect_ids(fields, own, wish)

        if own or wish:
            status_cb(f"Loaded profile from {pid}: {len(own)} own, {len(wish)} wish")
            return own, wish, uid

    # Nothing found — build an honest, specific error.
    diag = ("; probed: " + ", ".join(tried[:6])) if tried else ""
    if saw_403:
        raise RuntimeError(
            "LPU's database refused the read (403) even after anonymous "
            "sign-in. Try again in a moment; if it keeps happening, let me "
            "know." + diag)
    if reached_firestore:
        raise RuntimeError(
            "connected to LPU's database but your collection document had no "
            "owned or wishlist locks. Double-check the profile ID in the "
            "link." + diag)
    raise RuntimeError(
        "could not reach the profile data (couldn't sign in to or read LPU's "
        "live database). Check your internet connection and try again." + diag)

def save_profile(uid, own, wish):
    conn = db()
    try:
        conn.execute("BEGIN")
        conn.execute("DELETE FROM my_collection")
        for i in own:
            conn.execute("INSERT OR REPLACE INTO my_collection(lock_id,status)"
                         " VALUES(?, 'own')", (i,))
        for i in wish:
            # wishlist wins the label if a lock is somehow in both
            conn.execute("INSERT OR REPLACE INTO my_collection(lock_id,status)"
                         " VALUES(?, 'wishlist')", (i,))
        conn.commit()
    finally:
        conn.close()

# ------------------------------------------------- LPU leaderboard / owners
# The public LPU leaderboard is a precomputed stats file (per-user OWNED counts,
# display names, privacy flags) — NOT the actual owned lock ids. To find who
# owns a specific lock we take the top N users from this file, then read each
# one's lockcollections/<uid> document (the same public per-doc read the
# profile import and Compare tab use) to get their owned lock-id arrays.
LEADERBOARD_URL = "https://explore.lpubelts.com/data/leaderboardData.json"









def build_help_report():
    """Compose the Help email. Returns (subject, body, full_report_path).
    The BODY stays compact — mailto links get truncated by mail clients
    around ~1800 characters — and NEVER contains secrets (the API key is
    reported only as present/absent). The COMPLETE diagnostics (full recent
    log + crash log) are written to a file the user can attach."""
    diag = [f"Lock Hunter v{VERSION}"]
    try:
        diag.append(f"System: {platform.platform()} | "
                    f"Python {platform.python_version()}")
    except Exception:
        pass
    try:
        conn = db()
        n_locks = conn.execute("SELECT COUNT(*) FROM locks").fetchone()[0]
        n_tier = conn.execute(
            "SELECT COUNT(*) FROM locks WHERE belt='Black' AND belt_full"
            " LIKE 'Black %'").fetchone()[0]
        own = conn.execute("SELECT COUNT(*) FROM my_collection WHERE"
                           " status='own'").fetchone()[0]
        wsh = conn.execute("SELECT COUNT(*) FROM my_collection WHERE"
                           " status='wishlist'").fetchone()[0]
        conn.close()
        diag.append(f"Catalog: {n_locks} locks ({n_tier} Black-tiered) | "
                    f"Owned {own} - Wishlist {wsh}")
    except Exception as ex:
        diag.append(f"Catalog: unavailable ({ex})")
    try:
        cfg = load_cfg()
        diag.append("API key saved: %s" % ("yes" if cfg.get("api_key")
                                           else "no"))
    except Exception:
        pass

    def _tail(path, n):
        try:
            with open(path, "r", encoding="utf-8", errors="replace") as f:
                return [ln.rstrip()[:110] for ln in f.readlines()[-n:]]
        except Exception:
            return []

    log_tail = _tail(LOG_PATH, 12)
    crash_tail = _tail(CRASH_LOG_PATH, 5)
    # complete report on disk (attachable) — mailto can't attach files
    full_path = os.path.join(APP_DIR, "help_report.txt")
    try:
        ensure_dirs()
        with open(full_path, "w", encoding="utf-8") as f:
            f.write("\n".join(diag) + "\n\n=== recent activity log ===\n")
            f.write("\n".join(_tail(LOG_PATH, 200)) or "(empty)")
            f.write("\n\n=== crash log (tail) ===\n")
            f.write("\n".join(_tail(CRASH_LOG_PATH, 60)) or "(none)")
            f.write("\n")
    except Exception:
        full_path = LOG_PATH   # fall back to pointing at the raw log
    def _compose(logs):
        parts = ["Hi Ferf,", "", "(Please describe the problem here.)",
                 "__REPORT_BANNER__", "", "----- diagnostics -----"] + diag
        if logs:
            parts += ["Recent activity:"] + logs
        if crash_tail:
            parts += ["Last crash:"] + crash_tail
        parts += ["----- end -----", "Full report saved at:", full_path]
        return "\n".join(parts)

    body = _compose(log_tail)
    # hard cap for mailto: drop oldest log lines first
    while len(body) > 1700 and log_tail:
        log_tail.pop(0)
        body = _compose(log_tail)
    return "Lockhunter help", body[:1900], full_path

# ---------------------------------------------------------------- Claude

def build_prompt(lock_name, condition, exclude_pickup, include_bazaar=False,
                 deep=False, auction=False, sales=None):
    # 'sales' picks the VENUE type:
    #   "Auction"  -> auction & member-to-member secondhand marketplaces only
    #   "Website"  -> retail website sales only (shops, dealers, manufacturers)
    #   "Both"     -> search both venue types in the same pass
    # 'condition' is the ITEM condition: New / Used / Both — independent of
    # venue, so e.g. New + Auction finds new-old-stock on eBay. The legacy
    # boolean 'auction' maps to sales="Auction" for backward compatibility.
    if sales is None:
        sales = "Auction" if auction else "Both"
    if sales not in ("Auction", "Website", "Both"):
        sales = "Both"

    venue_rules = {
        "Auction": (
            "Search ONLY auction and secondhand marketplaces such as: "
            + SECONDHAND_SITES + ". Include BOTH auction-format listings AND "
            "fixed-price (Buy It Now) secondhand listings from private "
            "sellers. Do NOT return retail shops, manufacturer stores, or "
            "price-comparison sites."),
        "Website": (
            "Search ONLY retailers selling through their own website "
            "(manufacturer stores, locksmith shops, hardware retailers, "
            "specialist lock dealers). Exclude auction sites and "
            "member-to-member secondhand marketplaces."),
        "Both": (
            "Search BOTH venue types in this one pass: (a) auction and "
            "secondhand marketplaces such as: " + SECONDHAND_SITES + ", AND "
            "(b) retail websites (manufacturer stores, locksmith shops, "
            "hardware retailers). Cover both — do not skip either kind."),
    }
    cond_rules = {
        "New": ("Only include items in NEW condition — including "
                "new-old-stock (NOS) and new-in-box listings on secondhand "
                "sites. Mark each result's condition as 'new'."),
        "Used": ("Only include USED / vintage / pre-owned items. Mark each "
                 "result's condition as 'used'."),
        "Both": ("Include both new and used items and label each result's "
                 "condition accurately."),
    }
    cond_text = (venue_rules[sales] + "\n"
                 + cond_rules.get(condition, cond_rules["Both"]))
    pickup_rule = (
        "For each listing, determine whether it can be SHIPPED or is "
        "pickup/meetup only. Set the 'shipping' field to 'yes', 'no', or "
        "'unknown'."
    )
    if exclude_pickup:
        pickup_rule += " Prefer listings that ship; still report shipping status honestly."
    bazaar_rule = ""
    if include_bazaar:
        bazaar_rule = (
            "\nADDITIONALLY: check the LPU Lock Bazaar, the Lockpickers United "
            "community marketplace at https://lpulocks.com/#/lockbazaar (searchable "
            "listing browser for the #lock-bazaar Discord channel). Search for this "
            "lock there (e.g. web search: site or query 'lpulocks lock bazaar "
            f"{lock_name}'). Bazaar listings are member-sold used locks; report any "
            "matches with site='LPU Lock Bazaar' and condition='used'."
        )

    # Accuracy rules applied to every search.
    accuracy_rules = (
        "ACCURACY REQUIREMENTS (critical):\n"
        f"1. EXACT MODEL MATCH. The listing must be for the SAME lock: "
        f"\"{lock_name}\". Match brand AND model/variant precisely. Do NOT "
        "return a different model, a similar-looking lock, a different size, "
        "a key blank, parts, a keyed-alike variant, or an accessory. If the "
        "title's model number/name differs, discard it.\n"
        "2. LIVE LISTINGS ONLY — NO ENDED AUCTIONS. Only include listings that "
        "are currently available to buy or bid on RIGHT NOW. This is critical "
        "for eBay and other auction sites: EXCLUDE any listing that has ENDED, "
        "SOLD, or EXPIRED. Signs a listing is over and must be excluded: it "
        "shows 'Sold', 'Ended', 'This listing was ended', 'bidding has ended', "
        "a past end date, a URL like ebay.com/itm that returns a sold/ended "
        "state, or a completed/sold search result. If you cannot confirm the "
        "listing is still active, DO NOT include it. Also exclude "
        "search-result/category pages, forum threads, blog posts, and review "
        "pages — each url must be a single, currently-active listing page.\n"
        "3. REAL, VERIFIED DATA. Only report a listing you actually saw in "
        "search results with a real URL. Never invent, guess, or extrapolate "
        "a URL, price, or seller. If you are not confident the URL is a real "
        "live listing for this exact lock, leave it out.\n"
        "4. PRICE + CURRENCY as shown on the listing (e.g. price '129.99', "
        "currency 'USD'/'EUR'/'GBP'). For an active auction use the current "
        "bid and note it (e.g. notes 'auction, current bid'). If a price isn't "
        "visible, use '' and explain in notes.\n"
        "5. Set 'condition' to 'new' or 'used' based on the listing itself, "
        "not on the site type.\n"
        "6. De-duplicate: one entry per unique listing URL."
    )

    if deep:
        depth_rules = (
            "SEARCH DEPTH: EXTENDED / THOROUGH. Do many searches (aim for "
            "12-20) and be exhaustive. Try: the exact model; common brand+model "
            "spelling variants and abbreviations; the manufacturer's own part "
            "number; and localized terms for the relevant markets in their own "
            "language. Check multiple marketplaces rather than stopping at the "
            "first hit. Prioritize precision over volume — a wrong-model match "
            "is worse than fewer results."
        )
    else:
        depth_rules = (
            "SEARCH DEPTH: STANDARD. Do a focused set of searches (about "
            "5-9) across the most likely marketplaces for this lock, leading "
            "with its country of origin. Favor precise matches over quantity."
        )

    # Origin-first strategy: a lock is most likely to appear secondhand in its
    # country of origin, listed in that country's language. Steer the search
    # there first before falling back to global/English marketplaces.
    origin = _origin_for_lock(lock_name)
    if origin:
        market = COUNTRY_MARKETS.get(origin, "")
        origin_rule = (
            "ORIGIN-FIRST SEARCH STRATEGY (very important):\n"
            f"- This lock's brand originates from / is primarily sold in: {origin}.\n"
            f"- Secondhand examples are most likely to appear there, listed in "
            f"the local language. Search THAT country's marketplaces FIRST and "
            f"HARDEST, using native-language search terms.\n"
        )
        if market:
            origin_rule += f"- Prioritize: {market}.\n"
        origin_rule += (
            "- Translate the generic word 'lock'/'cylinder'/'padlock' into the "
            "local language and combine it with the brand+model (the brand name "
            "usually stays in Latin letters). Run separate queries in the local "
            "language AND in English.\n"
            "- THEN broaden to international sites (eBay.com/.co.uk, Catawiki, "
            "Delcampe, Facebook Marketplace) and the LPU Lock Bazaar.\n"
            "- A US-based collector is the buyer, so for each result note in "
            "'shipping' whether the seller ships internationally when you can "
            "tell, and put the seller's country in 'location'.\n"
        )
    else:
        origin_rule = (
            "SEARCH STRATEGY: Determine the lock brand's country of origin, then "
            "search that country's marketplaces in its own language first "
            "(translate 'lock'/'padlock'/'cylinder' into the local language and "
            "pair it with the brand+model), before broadening to international "
            "and English-language sites and the LPU Lock Bazaar.\n"
        )

    wants_secondhand = sales in ("Auction", "Both")

    # Direct marketplace fetching — the decisive step for rare locks. Their
    # listings live on secondhand/auction venues and are short-lived, so
    # search engines rarely index them in time; a plain web search misses
    # them. Hand the model ready-made on-site search URLs to OPEN directly.
    market_block = ""
    if wants_secondhand:
        origin_urls = _market_urls(lock_name, origin, deep=deep)
        market_block = (
            "FETCH THE MARKETPLACES DIRECTLY (critical for rare locks):\n"
            "Rare collectible locks are usually sold ONLY on secondhand/"
            "auction marketplaces, and those listings are short-lived and "
            "poorly indexed — plain web search misses them. Do NOT rely on "
            "web_search alone:\n"
            "- Use the web_fetch tool to OPEN each of these marketplace "
            "search URLs and read the live results on the page itself:\n"
            + "".join(f"    {u}\n" for u in origin_urls) +
            "- From those results pages, open the matching ITEM pages "
            "(web_fetch again) and cite the item page URL, never the search "
            "URL.\n"
            "- If a search URL errors or redirects, adapt the path or use "
            "the site's own search from its homepage.\n"
            "- THEN supplement with general web_search queries for venues "
            "not listed above.\n")

    # Explicit per-site secondhand query technique. Naming sites in a list isn't
    # enough — instruct concrete `site:` searches against the highest-yield
    # secondhand marketplaces, plus query-construction tactics that surface used
    # locks (this is what made the earlier sourcing sheet effective).
    secondhand_technique = ""
    if wants_secondhand:
        secondhand_technique = (
            "SECONDHAND QUERY TECHNIQUE (do this, don't just name sites):\n"
            "- Run EXPLICIT per-site searches with the site: operator against the "
            "top secondhand marketplaces, one query per site, e.g.:\n"
            f"    site:ebay.com \"{lock_name}\"\n"
            f"    site:ebay.co.uk \"{lock_name}\"\n"
            f"    site:catawiki.com {lock_name}\n"
            f"    site:delcampe.net {lock_name}\n"
            f"    site:etsy.com {lock_name} vintage\n"
            f"    site:shopgoodwill.com {lock_name}\n"
            f"    site:facebook.com/marketplace {lock_name}\n"
            "  and the origin-country sites named above (e.g. site:tori.fi, "
            "site:subito.it, site:kleinanzeigen.de, site:tradera.com, "
            "site:leboncoin.fr, site:marktplaats.nl, site:avito.ru, "
            "site:auctions.yahoo.co.jp) using the local-language term.\n"
            "- Also try queries WITHOUT site: but WITH secondhand cue words: "
            "'for sale', 'used', 'vintage', 'NOS', 'lot', 'collection', "
            "'estate', plus the local-language equivalents.\n"
            "- Vary the model text: full name, brand+number only, with and "
            "without spaces/hyphens, and any common abbreviation or part number.\n"
            "- Prefer individual listing pages over search/category pages. When a "
            "marketplace search page shows a matching item, open it and cite the "
            "item's own URL, not the search URL.\n"
        )

    return f"""You are a meticulous sourcing agent for a serious lock collector.
Find CURRENT, live, for-sale listings for this EXACT lock: "{lock_name}".

{origin_rule}
{market_block}
{secondhand_technique}
{cond_text}
{pickup_rule}{bazaar_rule}

{depth_rules}

{accuracy_rules}

Use the web_search tool (multiple queries as needed). Read the actual listing
pages you cite; do not rely on snippets alone when deciding if the model
matches and the listing is live.

Respond with ONLY a JSON array, no markdown fences, no commentary. Each
element must be exactly:
{{"title": str, "price": str, "currency": str,
"condition": "new"|"used", "site": str, "url": str, "location": str,
"shipping": "yes"|"no"|"unknown", "notes": str}}
Order the array best-match-first. Only include listings you actually found at
real, live URLs for this exact lock. If nothing qualifies, return []."""

def call_claude(api_key, prompt, status_cb, deep=False):
    headers = {
        "x-api-key": api_key,
        "anthropic-version": API_VERSION,
        "content-type": "application/json",
        # enables the web_fetch tool: the model can OPEN marketplace search
        # pages directly instead of relying on search-engine indexing.
        "anthropic-beta": "web-fetch-2025-09-10",
    }
    body = {
        "model": MODEL,
        # Extended search gets more room to reason and more web searches.
        "max_tokens": 8000 if deep else 4000,
        "tools": [{"type": "web_search_20250305", "name": "web_search",
                   "max_uses": 20 if deep else 8},
                  {"type": "web_fetch_20250910", "name": "web_fetch",
                   "max_uses": 20 if deep else 10}],
        "messages": [{"role": "user", "content": prompt}],
    }
    status_cb("Calling Claude API (extended web search)..." if deep
              else "Calling Claude API (live web search)...")
    r = requests.post(API_URL, headers=headers, json=body,
                      timeout=600 if deep else 300)
    if r.status_code == 400 and "web_fetch" in (r.text or ""):
        # API key/tier without the fetch tool — retry search-only so the
        # search still completes rather than failing outright.
        status_cb("web_fetch unavailable on this API key — using search only.")
        headers.pop("anthropic-beta", None)
        body["tools"] = [t for t in body["tools"]
                         if t.get("name") != "web_fetch"]
        r = requests.post(API_URL, headers=headers, json=body,
                          timeout=600 if deep else 300)
    if r.status_code != 200:
        raise RuntimeError(f"API error {r.status_code}: {r.text[:300]}")
    data = r.json()
    text = "\n".join(
        blk.get("text", "") for blk in data.get("content", [])
        if blk.get("type") == "text"
    )
    return _parse_listings_json(text)

def _parse_listings_json(text):
    """Extract the JSON array of listings from the model's text response,
    tolerant of markdown fences and surrounding prose."""
    cleaned = text.replace("```json", "").replace("```", "")
    # Prefer the LAST top-level [...] block (the final answer). Balance brackets
    # while IGNORING any '[' or ']' that appear inside JSON string values, so a
    # bracket inside a title/notes field can't throw off the count.
    starts = [i for i, c in enumerate(cleaned) if c == "["]
    for start in reversed(starts):
        depth = 0
        in_str = False
        esc = False
        for j in range(start, len(cleaned)):
            c = cleaned[j]
            if in_str:
                if esc:
                    esc = False
                elif c == "\\":
                    esc = True
                elif c == '"':
                    in_str = False
                continue
            if c == '"':
                in_str = True
            elif c == "[":
                depth += 1
            elif c == "]":
                depth -= 1
                if depth == 0:
                    chunk = cleaned[start:j + 1]
                    try:
                        out = json.loads(chunk)
                        if isinstance(out, list):
                            return out
                    except json.JSONDecodeError:
                        break  # try an earlier '['
    return []

# Signs (in page text) that an auction/listing has ended or is otherwise gone.
_ENDED_MARKERS = (
    "this listing was ended", "this listing has ended", "bidding has ended",
    "this listing sold", "the listing sold", "item sold", "sold for",
    "listing ended", "auction has ended", "auction ended", "no longer available",
    "this item is out of stock", "out of stock", "currently out of stock",
    "ended on", "sale has ended", "this offer has ended",
    "this listing is no longer available", "we looked everywhere",
    "this item is not available", "esta subasta ha finalizado",
    "objet n'est plus disponible",
    # origin-country languages (Finnish, Swedish, Danish, German, Italian,
    # Dutch, Japanese, Russian, Spanish) — common "sold/ended/removed" phrases
    "myyty", "päättynyt",                       # Finnish: sold / ended
    "såld", "avslutad", "auktionen har avslutats",  # Swedish
    "solgt", "afsluttet",                       # Danish/Norwegian
    # German: NOT bare "verkauft" — live multi-quantity eBay.de listings show
    # a sold-counter like "126 verkauft"; only the full ended phrases are safe.
    "wurde verkauft", "bereits verkauft", "beendet", "nicht mehr verfügbar",
    "venduto", "asta terminata", "annuncio non", "scaduto",  # Italian
    "verkocht", "verlopen", "niet meer beschikbaar",  # Dutch
    "落札", "終了", "売り切れ",                    # Japanese: won/ended/sold-out
    "продано", "продан", "снято с продажи", "объявление снято",  # Russian
    # Spanish: NOT bare "vendido" — live eBay.es listings show "34 vendidos";
    # "se ha vendido" / "vendió" (accented, as in "Se vendió el 15 jul.") only
    # appear once a listing has actually ended.
    "se ha vendido", "vendió", "finalizada", "anuncio no disponible",
)
# Words that indicate a page is an error / access wall rather than a listing.
_DEAD_MARKERS = (
    "page not found", "404 not found", "403 forbidden", "access denied",
    "the page you requested", "sorry, we can", "this page isn",
    "no results found", "your access to this site has been limited",
)

# Purchase/bid affordances that appear on LIVE eBay item pages (per language).
# Checked FIRST for eBay URLs: any one of these proves the listing is live, so
# the ended/sold markers above can never falsely kill it — live multi-quantity
# listings legitimately contain sold-counter text like "126 verkauft" /
# "34 vendidos" that would otherwise read as an ended state.
_EBAY_LIVE_SIGNALS = (
    "add to cart", "buy it now", "place bid", "add to watchlist",
    "time left", "watchlisted", "add to basket", "make offer",
    # German
    "sofort-kaufen", "sofort kaufen", "gebot abgeben",
    "in den warenkorb", "auf die beobachtungsliste",
    "preis vorschlagen", "restzeit",
    # French
    "achat immédiat", "achat immediat", "enchérir",
    "ajouter au panier", "ajouter à la liste", "temps restant",
    "faire une offre",
    # Italian
    "compralo subito", "fai un'offerta", "aggiungi al carrello",
    "tempo rimasto",
    # Spanish
    "cómpralo ya", "comprar ahora", "pujar", "añadir a la cesta",
    "tiempo restante", "hacer una oferta",
    # Dutch (eBay.nl also shows a "verkocht" sold-counter on live listings)
    "nu kopen", "doe een bod", "toevoegen aan winkelwagen",
    "resterende tijd",
    # Polish
    "kup teraz",
)

def _verify_listing(it, status_cb):
    """Decide whether to KEEP a listing. Policy: innocent until proven dead.
    Only EVIDENCE of death drops a listing: HTTP 404/410, an explicit
    ended/sold state, or a genuine not-found page. Bot walls, 403s, consent
    pages and other automated-fetch blocks are NOT death — marketplaces serve
    those to programs while the listing is perfectly alive in a browser — so
    those are kept and tagged 'could not verify' instead of dropped."""
    url = str(it.get("url") or "").strip()
    if not url.startswith("http"):
        return False  # no usable URL
    is_ebay = "ebay." in url.lower()

    def _tag_unverified():
        note = str(it.get("notes") or "").strip()
        tag = "could not verify (site blocked automated check)"
        if tag not in note:
            it["notes"] = (note + "; " + tag).strip("; ").strip()

    code, _vbody = _scrape_get(url, lambda _s: None, label="verify",
                               timeout=25, attempts=1)
    if code == 0:
        status_cb(f"  (couldn't verify {url[:60]}; keeping)")
        _tag_unverified()
        return True  # network hiccup — don't punish the listing
    if code in (404, 410):
        status_cb(f"  dropped (gone, HTTP {code}): {url[:70]}")
        return False
    body = (_vbody or "").lower()
    # LIVE-SIGNAL FIRST for eBay: a purchase/bid affordance on the page proves
    # the listing is live RIGHT NOW, so it must be kept before the ended/sold
    # markers get a say — live multi-quantity listings carry sold-counters
    # ("126 verkauft", "34 vendidos") that would otherwise match as ended.
    if is_ebay and body and any(s in body for s in _EBAY_LIVE_SIGNALS):
        return True
    # Explicit ended / sold state -> genuinely dead, drop (any status code).
    if body and any(m in body for m in _ENDED_MARKERS):
        status_cb(f"  dropped (ended/sold): {url[:70]}")
        return False
    # Bot walls / access blocks: the LISTING may be fine; the FETCH was
    # refused. Keep, tagged.
    _WALL_MARKERS = (
        "pardon our interruption", "checking your browser",
        "verify you are a human", "are you a human", "unusual traffic",
        "access denied", "403 forbidden",
        "your access to this site has been limited", "captcha",
        "please enable javascript", "attention required",
    )
    if code != 200 or (body and any(m in body for m in _WALL_MARKERS)):
        status_cb(f"  kept unverified (blocked: HTTP {code}): {url[:70]}")
        _tag_unverified()
        return True
    # Genuine not-found pages disguised as 200.
    _GONE_MARKERS = ("page not found", "404 not found",
                     "the page you requested", "this page isn",
                     "no results found")
    if body and any(m in body for m in _GONE_MARKERS):
        status_cb(f"  dropped (not-found page): {url[:70]}")
        return False
    if is_ebay and body:
        # No purchase/bid affordance was found (the live-signal check above
        # would have returned already). Absence is NOT death — modern eBay
        # pages are often JS-rendered shells — so absence only tags.
        status_cb(f"  kept unverified (no live signal): {url[:70]}")
        _tag_unverified()
    return True

def verify_listings(items, status_cb, max_workers=6):
    """Verify a list of listing dicts in parallel; return only the live ones.
    LPU Lock Bazaar entries are trusted (not fetched) since they come from the
    community feed, not an external marketplace."""
    keep = []
    to_check = []
    for it in items:
        if (str(it.get("site") or "").strip().lower() == "lpu lock bazaar"
                or it.get("preverified")):
            # Bazaar entries and direct-probe results parsed straight off a
            # marketplace's own LIVE search page are already proof-of-life;
            # re-fetching them just invites bot walls.
            keep.append(it)
        else:
            to_check.append(it)
    if not to_check:
        return keep
    status_cb(f"Verifying {len(to_check)} listing link(s)…")
    _cf = concurrent.futures
    results = {}
    with _cf.ThreadPoolExecutor(max_workers=max_workers) as ex:
        futs = {ex.submit(_verify_listing, it, status_cb): i
                for i, it in enumerate(to_check)}
        for fut in _cf.as_completed(futs):
            i = futs[fut]
            try:
                results[i] = fut.result()
            except Exception:
                results[i] = True  # keep on unexpected error
    for i, it in enumerate(to_check):
        if results.get(i, True):
            keep.append(it)
    return keep

# ---------------------------------------------------------------- GUI

class FlatButton(tk.Frame):
    """A modern flat button drawn with a tk.Label so we fully control the
    look (rounded feel via padding, hover/press states, disabled styling).
    ttk buttons can't be styled this cleanly across platforms."""
    def __init__(self, master, text, command=None, kind="primary", **kw):
        # Resolve the palette. Normally the enclosing toplevel is the main
        # LockHunter window (which carries the C_* colors). But when a button
        # is placed inside a secondary Toplevel dialog, that toplevel does NOT
        # have the C_* attributes — so fall back to the LockHunter class, which
        # defines them. This is what makes buttons render inside dialogs.
        colors = master.winfo_toplevel()
        if not hasattr(colors, "C_ACCENT"):
            colors = LockHunter
        self._top = colors
        self._c = {
            "primary": (colors.C_ACCENT, "#20140a", colors.C_ACCENT2),
            "ghost":   (colors.C_PANEL2, colors.C_TEXT, colors.C_LINE),
            "subtle":  (colors.C_PANEL2, colors.C_TEXT, colors.C_LINE),
            # outline: transparent-ish body with amber text — used for the
            # action buttons that toggle enabled/disabled, so the enabled
            # state reads clearly as "active" (amber) vs greyed when off.
            "outline": (colors.C_PANEL2, colors.C_ACCENT, colors.C_LINE),
        }[kind]
        bg, fg, hover = self._c
        super().__init__(master, bg=bg, highlightthickness=1,
                         highlightbackground=colors.C_LINE,
                         highlightcolor=colors.C_LINE, bd=0, **kw)
        self._bg, self._fg, self._hover = bg, fg, hover
        self._kind = kind
        self._enabled = True
        self.command = command
        self.lbl = tk.Label(self, text=text, bg=bg, fg=fg,
                            font=("Segoe UI Semibold", 10), padx=16, pady=7,
                            cursor="hand2")
        self.lbl.pack(fill="both", expand=True)
        for w in (self, self.lbl):
            w.bind("<Enter>", self._on_enter)
            w.bind("<Leave>", self._on_leave)
            w.bind("<Button-1>", self._on_press)
            w.bind("<ButtonRelease-1>", self._on_release)

    def _on_enter(self, _=None):
        if self._enabled:
            self.configure(bg=self._hover); self.lbl.configure(bg=self._hover)
    def _on_leave(self, _=None):
        if self._enabled:
            self.configure(bg=self._bg); self.lbl.configure(bg=self._bg)
    def _on_press(self, _=None):
        if self._enabled:
            self.lbl.configure(padx=16, pady=8)
    def _on_release(self, _=None):
        if self._enabled and self.command:
            self.lbl.configure(padx=16, pady=7)
            self.command()

    def set_enabled(self, on):
        self._enabled = bool(on)
        top = self._top
        if self._enabled:
            self.configure(bg=self._bg, highlightbackground=top.C_ACCENT
                           if self._kind == "outline" else top.C_LINE)
            self.lbl.configure(bg=self._bg, fg=self._fg, cursor="hand2")
        else:
            # dimmer than the panel so it clearly reads as inactive
            self.configure(bg=top.C_PANEL, highlightbackground=top.C_LINE)
            self.lbl.configure(bg=top.C_PANEL, fg=top.C_MUTE, cursor="arrow")

    def set_highlight(self, color):
        """Enable the button and colour its text/border with `color`
        (e.g. light blue when a Lock Bazaar match is found)."""
        self._enabled = True
        self._fg = color
        self.configure(bg=self._bg, highlightbackground=color)
        self.lbl.configure(bg=self._bg, fg=color, cursor="hand2")
    # ttk-compatible shims so existing enable/disable calls keep working
    def config(self, **kw):
        if "state" in kw:
            self.set_enabled(kw.pop("state") != "disabled")
        if "text" in kw:
            self.lbl.configure(text=kw.pop("text"))
        if kw:
            super().configure(**kw)
    configure = config
    def __setitem__(self, k, v): self.config(**{k: v})
    def __getitem__(self, k):
        if k == "state":
            return "normal" if self._enabled else "disabled"
        return super().__getitem__(k)


class LockHunter(tk.Tk):
    # LPU Belt Explorer palette (deep slate + warm belt-buckle amber)
    C_BG = "#161a1e"      # app background (deeper for contrast with cards)
    C_PANEL = "#1e2429"   # cards / raised surfaces
    C_PANEL2 = "#272e35"  # card header / hover
    C_FIELD = "#0f1317"   # inputs / table body
    C_TEXT = "#eceef0"    # primary text
    C_MUTE = "#8b939b"    # secondary text
    C_ACCENT = "#e0a231"  # belt-buckle amber
    C_ACCENT2 = "#f0b445" # accent hover (brighter)
    C_LINE = "#333b43"    # borders / separators
    C_SEL = "#2f3841"     # selection
    C_BAZAAR = "#5cc8ff"  # light blue — Lock Bazaar "found" highlight

    def report_callback_exception(self, exc_type, exc_value, exc_tb):
        """Tk calls this when an exception escapes an event callback (a button
        click, a menu item, etc.). Instead of failing silently — the built .exe
        has no console — write a copy-pasteable crash log and tell the user
        where it is, then keep the app running."""
        path = write_crash_log(exc_type, exc_value, exc_tb, where="event callback")
        try:
            msg = "Something went wrong during that action.\n\n"
            if path:
                msg += ("A crash log was saved so it can be sent for a fix:\n"
                        f"{path}\n\n"
                        "Please open that file, copy everything in it, and send "
                        "it over.")
            else:
                msg += "Sorry — the details couldn't be saved this time."
            messagebox.showerror("Lock Hunter — unexpected error", msg)
        except Exception:
            pass

    def __init__(self):
        super().__init__()
        # On Windows, give the app its own taskbar identity so the taskbar
        # shows OUR icon (not the generic Python/host icon). Must be set before
        # the window is shown. Harmless / ignored on other platforms.
        try:
            if sys.platform.startswith("win"):
                import ctypes
                ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(
                    "LockHunter.App")
        except Exception:
            pass
        self.title(f"Lock Hunter v{VERSION} — AI lock sourcing")
        # Pick a window size that shows the whole UI, but never taller/wider
        # than the screen (leave room for the taskbar + title bar), then centre
        # it. This stops the bottom of the window (status bar, buttons) from
        # being pushed off the bottom of smaller screens, e.g. 1366x768 laptops.
        try:
            sw = self.winfo_screenwidth()
            sh = self.winfo_screenheight()
        except Exception:
            sw, sh = 1280, 800
        want_w, want_h = 1200, 820
        w = min(want_w, max(900, sw - 60))
        h = min(want_h, max(560, sh - 96))   # keep clear of taskbar/title bar
        x = max(0, (sw - w) // 2)
        y = max(0, (sh - h - 48) // 2)        # bias slightly above dead-centre
        self.geometry(f"{w}x{h}+{x}+{y}")
        # Allow shrinking on small screens; keep a floor so it stays usable.
        self.minsize(900, min(560, max(480, sh - 120)))
        self._set_window_icon()
        ensure_dirs()
        init_db()          # create tables + run migrations once
        self.cfg = load_cfg()
        self.q = queue.Queue()
        self._thumb_ref = None
        # One-time cleanup: remove the old 83-lock "sheet" seed that earlier
        # versions inserted, so the picker shows only real LPU catalog locks.
        conn = db()
        conn.execute("DELETE FROM locks WHERE source='sheet'")
        conn.commit()
        conn.close()
        self._apply_theme()
        self._build_ui()
        self._reload_lock_names()
        self._refresh_table()
        self.after(150, self._poll)
        self.after(200, self._reapply_window_icon)   # re-assert icon once mapped
        log("App started")
        conn = db()
        lpu_count = conn.execute(
            "SELECT COUNT(*) FROM locks WHERE source='lpu'").fetchone()[0]
        conn.close()
        # First run (or after a failed/empty sync): automatically download the
        # LPU catalog, then walk the user through API key + profile setup.
        if not lpu_count:
            self.status.set("First run — downloading the LPU lock catalog…")
            self.after(300, self._start_sync)
            self.after(500, self._first_run_setup)
        else:
            self.after(400, self._first_run_setup)
            # Returning user (catalog exists): auto-refresh the catalog and
            # profile, staggered so startup stays snappy.
            self._auto_synced = True
            self.after(1500, self._start_sync)
            self.after(2600, self._auto_profile_refresh)
        # quiet background chores, staggered so startup stays snappy:
        self.after(4000, self._auto_update_check)      # once/day, silent
        self.after(6000, self._start_bazaar_watch)     # "new on the Bazaar?"

    def _first_run_setup(self):
        # The Claude API key is ONLY needed for the optional AI web search, so
        # it isn't requested at startup. It's prompted when the user enables AI
        # web search, or via the "Enter/Change API key" link. Here we only set
        # up the profile.
        self._prompt_profile(force=False)
        self._maybe_offer_tier_sync()

    def _maybe_offer_tier_sync(self):
        """If the catalog has Black-belt locks but none carry a sub-tier
        (Black 1..5) — i.e. it was last synced by an older version — offer once
        to update it so the sub-tiers show up. Declining sets a flag so we don't
        nag; a successful update makes the check stop firing on its own."""
        if getattr(self, "_auto_synced", False):
            return   # a full catalog sync is already running automatically
        if self.cfg.get("tier_sync_offered"):
            return
        try:
            conn = db()
            n_black = conn.execute(
                "SELECT COUNT(*) FROM locks WHERE belt='Black'").fetchone()[0]
            n_tiered = conn.execute(
                "SELECT COUNT(*) FROM locks WHERE belt='Black' AND belt_full "
                "IS NOT NULL AND belt_full <> '' AND belt_full <> 'Black'"
            ).fetchone()[0]
            conn.close()
        except Exception:
            return
        if n_black and not n_tiered:
            if messagebox.askyesno(
                    "Update lock catalog?",
                    "Your lock catalog is a little out of date: Black-belt "
                    "locks aren't split into their sub-tiers (Black 1\u20135) "
                    "yet, so they all just show as \u201cBlack\u201d.\n\n"
                    "Update the catalog now to add the sub-tiers? It takes "
                    "about 10\u201320 seconds."):
                self._start_sync()
            else:
                self.cfg["tier_sync_offered"] = True
                save_cfg(self.cfg)

    def _ask_text(self, title, message, initial="", mask=False):
        """A self-contained modal text prompt. Built explicitly (not via
        tkinter.simpledialog) so nothing leaks between dialogs and so we fully
        control masking. Returns the entered string, or None if cancelled.
        When mask=True the entry shows dots (for secrets like the API key)."""
        win = tk.Toplevel(self)
        win.title(title)
        win.configure(bg=self.C_PANEL)
        win.transient(self)
        win.resizable(False, False)
        try:
            win.grab_set()          # modal
        except tk.TclError:
            pass

        tk.Label(win, text=message, bg=self.C_PANEL, fg=self.C_TEXT,
                 justify="left", font=("Segoe UI", 10), wraplength=520
                 ).pack(anchor="w", padx=16, pady=(16, 10))

        var = tk.StringVar(value=initial or "")
        entry = tk.Entry(win, textvariable=var, width=64,
                         bg=self.C_FIELD, fg=self.C_TEXT,
                         insertbackground=self.C_ACCENT, insertwidth=2,
                         relief="flat", font=("Consolas", 10),
                         show="•" if mask else "")
        entry.pack(fill="x", padx=16)
        entry.focus_set()

        result = {"value": None}

        def ok(_=None):
            result["value"] = var.get()
            win.destroy()

        def cancel(_=None):
            result["value"] = None
            win.destroy()

        # a "Show" toggle for the masked (API key) case, so the user can verify
        # what they typed without the value ever leaking into another dialog
        if mask:
            show_var = tk.BooleanVar(value=False)
            def toggle():
                entry.config(show="" if show_var.get() else "•")
            tk.Checkbutton(win, text="Show characters", variable=show_var,
                           command=toggle, bg=self.C_PANEL, fg=self.C_MUTE,
                           selectcolor=self.C_FIELD, activebackground=self.C_PANEL,
                           activeforeground=self.C_TEXT, font=("Segoe UI", 9)
                           ).pack(anchor="w", padx=14, pady=(6, 0))

        # Button row. Use FlatButton for the themed look, but fall back to
        # plain tk.Button if anything goes wrong building them — the dialog must
        # ALWAYS have working OK / Cancel buttons.
        btns = tk.Frame(win, bg=self.C_PANEL)
        btns.pack(fill="x", padx=16, pady=(10, 16))
        try:
            FlatButton(btns, "OK", command=ok, kind="primary"
                       ).pack(side="right", padx=(8, 0))
            FlatButton(btns, "Cancel", command=cancel, kind="ghost"
                       ).pack(side="right")
        except Exception:
            tk.Button(btns, text="OK", command=ok, width=10,
                      bg=self.C_ACCENT, fg="#20140a", relief="flat",
                      font=("Segoe UI Semibold", 10), cursor="hand2"
                      ).pack(side="right", padx=(8, 0))
            tk.Button(btns, text="Cancel", command=cancel, width=10,
                      bg=self.C_PANEL2, fg=self.C_TEXT, relief="flat",
                      font=("Segoe UI Semibold", 10), cursor="hand2"
                      ).pack(side="right")

        win.bind("<Return>", ok)
        win.bind("<Escape>", cancel)
        # size + center over the parent, with a sane minimum so the buttons are
        # always on screen even if content is short.
        win.update_idletasks()
        try:
            need_w = max(560, win.winfo_reqwidth())
            need_h = max(230, win.winfo_reqheight())
            px, py = self.winfo_rootx(), self.winfo_rooty()
            pw, ph = self.winfo_width(), self.winfo_height()
            x = px + max(0, (pw - need_w) // 2)
            y = py + max(0, (ph - need_h) // 3)
            win.geometry(f"{need_w}x{need_h}+{x}+{y}")
        except Exception:
            pass
        self.wait_window(win)
        return result["value"]

    def _prompt_api_key(self, force=False):
        """Ask for the Claude API key. If force=False, skips when one is saved.
        The value is stored in config + key_var. Masked with dots; a 'Show
        characters' toggle lets the user verify it.

        When a key is already saved (startup re-prompt), the field is pre-filled
        (masked) so the user can just press OK to keep it, change it, or clear
        it. Cancelling keeps whatever was already saved."""
        if not force and self.cfg.get("api_key"):
            return
        current = self.cfg.get("api_key", "")
        if current:
            msg = ("Confirm your Claude API key (needed to run live searches).\n\n"
                   "Your saved key is filled in below (hidden). Press OK to keep "
                   "it, type a new one to change it, or clear the box and press "
                   "OK to remove it. Cancel keeps your current key.\n\n"
                   "Keys start with  sk-ant-  and come from console.anthropic.com "
                   "(Settings → API keys). Your key is stored locally only.")
        else:
            msg = ("Enter your Claude API key so Lock Hunter can run live searches.\n\n"
                   "It starts with  sk-ant-  and comes from console.anthropic.com "
                   "(Settings → API keys). Click Cancel to open the setup guide.\n\n"
                   "Your key is stored locally on this computer only.")
        ans = self._ask_text("Claude API key", msg, initial=current, mask=True)
        if ans is None:
            # Cancel: keep whatever was already saved. Only open the setup guide
            # when there's no key at all to keep.
            if not self.cfg.get("api_key"):
                self._show_api_key_help()
            return
        ans = ans.strip()
        if ans:
            self.key_var.set(ans)
            self.cfg["api_key"] = ans
            save_cfg(self.cfg)
            if hasattr(self, "status"):
                self.status.set("API key saved.")
            log("API key saved")
        else:
            # Box submitted empty -> user chose to remove the saved key.
            if self.cfg.get("api_key"):
                self.cfg.pop("api_key", None)
                save_cfg(self.cfg)
                self.key_var.set("")
                if hasattr(self, "status"):
                    self.status.set("API key removed.")
                log("API key removed by user")

    def _prompt_profile(self, force=False):
        """Ask for the LPU profile URL. ANY profile link is accepted — the
        collection view it was copied from (Own / Wishlist / Any / none)
        doesn't matter, because the import only uses the profile ID and
        always reads both lists. The saved link is normalized to the
        ?collection=Any form purely for consistency. If force=False, skips
        when one is already saved. On success, runs the import."""
        if not force and self.cfg.get("profile_url"):
            return
        base_msg = (
            "Enter your LPU profile URL to load the locks you own and want.\n\n"
            "On lpubelts.com, open your profile and copy the address from the "
            "browser bar — ANY collection view (Own, Wishlist, Any) is fine. "
            "It looks like:\n\n"
            "   https://lpubelts.com/#/profile/<your-id>\n\n"
            "Leave this blank to skip; you can update it later from the Locks "
            "tab.")
        msg = base_msg
        while True:
            # profile URL is not a secret -> shown as plain text, and always
            # starts EMPTY so nothing (e.g. the API key) can carry over.
            ans = self._ask_text("LPU profile", msg, initial="", mask=False)
            if ans is None or not ans.strip():
                return  # cancelled or left blank -> skip
            url = ans.strip()
            if extract_profile_uid(url):
                break
            msg = ("That doesn't look like an LPU profile link — it needs the "
                   "profile ID in it.\n\n" + base_msg)
        # Normalize the saved link to the ?collection=Any form (the import
        # ignores the collection= part; this only keeps the link consistent).
        url = re.sub(r"[?&]collection=[^&]*", "", url)
        url = url + ("&collection=Any" if "?" in url else "?collection=Any")
        self.profile_var.set(url)
        self.cfg["profile_url"] = url
        save_cfg(self.cfg)
        self._start_profile_import()

    def _set_window_icon(self):
        """Set the taskbar / title-bar icon. Uses the .ico on Windows and a
        PNG via iconphoto as a cross-platform fallback. Kept on self so the
        PhotoImage isn't garbage-collected. Records what it found so the
        problem is diagnosable from the log if an icon doesn't appear."""
        # First, keep the STABLE icon copies (the ones desktop shortcuts point
        # their icon at) in sync with what THIS build ships — so an icon
        # change reaches every existing shortcut the moment the new build
        # runs once, without re-running any installer.
        _refresh_stable_icons()
        ico = resource_path("assets", "lpu_icon.ico")
        png = resource_path("assets", "lpu_icon_256.png")
        self._icon_paths = (ico, png)
        found = []
        # .ico for the Windows title bar / taskbar
        try:
            if os.path.exists(ico):
                self.iconbitmap(default=ico)
                found.append("ico")
        except Exception:
            pass
        # PNG via iconphoto (works on all platforms; on Windows this also feeds
        # the taskbar icon and is more reliable than iconbitmap alone).
        try:
            if os.path.exists(png):
                if HAVE_PIL:
                    self._app_icon = ImageTk.PhotoImage(Image.open(png))
                else:
                    self._app_icon = tk.PhotoImage(file=png)
                self.iconphoto(True, self._app_icon)
                found.append("png")
        except Exception:
            pass
        try:
            if found:
                log(f"Window icon set from: {', '.join(found)}")
            else:
                log(f"Window icon NOT set — assets missing. Looked for: {ico}")
        except Exception:
            pass

    def _reapply_window_icon(self):
        """Re-assert the icon after the window is mapped. Some Windows/Tk
        combinations drop the title-bar icon set during __init__, so we set it
        again once the window actually exists on screen."""
        try:
            ico = resource_path("assets", "lpu_icon.ico")
            if os.path.exists(ico):
                self.iconbitmap(default=ico)
            if getattr(self, "_app_icon", None) is not None:
                self.iconphoto(True, self._app_icon)
        except Exception:
            pass

    def _load_header_logo(self):
        """Return a small PhotoImage of the LPU logo for the header, or None.
        Prefers PIL (clean resize + flatten onto the panel color); falls back
        to Tk's PhotoImage subsample if PIL isn't available."""
        png = resource_path("assets", "lpu_icon_256.png")
        if not os.path.exists(png):
            return None
        target = 36
        if HAVE_PIL:
            try:
                im = Image.open(png).convert("RGBA")
                # flatten onto the header panel color so the black art blends
                bg = Image.new("RGBA", im.size, self._hex_to_rgba(self.C_PANEL2))
                im = Image.alpha_composite(bg, im).convert("RGB")
                im = im.resize((target, target), Image.LANCZOS)
                return ImageTk.PhotoImage(im)
            except Exception:
                pass
        try:
            img = tk.PhotoImage(file=png)
            factor = max(1, img.width() // target)
            return img.subsample(factor, factor)
        except Exception:
            return None

    @staticmethod
    def _hex_to_rgba(h):
        h = h.lstrip("#")
        return (int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16), 255)

    def _apply_theme(self):
        self.configure(bg=self.C_BG)
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass
        f_body = ("Segoe UI", 10)
        style.configure(".", background=self.C_BG, foreground=self.C_TEXT,
                        fieldbackground=self.C_FIELD, bordercolor=self.C_LINE,
                        lightcolor=self.C_LINE, darkcolor=self.C_LINE,
                        troughcolor=self.C_FIELD, focuscolor=self.C_ACCENT,
                        font=f_body)
        style.configure("TFrame", background=self.C_BG)
        style.configure("Card.TFrame", background=self.C_PANEL)
        style.configure("TLabel", background=self.C_BG, foreground=self.C_TEXT, font=f_body)
        style.configure("Card.TLabel", background=self.C_PANEL, foreground=self.C_TEXT)
        style.configure("CardTitle.TLabel", background=self.C_PANEL,
                        foreground=self.C_ACCENT, font=("Segoe UI Semibold", 10))
        style.configure("Muted.TLabel", background=self.C_PANEL, foreground=self.C_MUTE,
                        font=("Segoe UI", 9))
        style.configure("Field.TLabel", background=self.C_PANEL, foreground=self.C_MUTE,
                        font=("Segoe UI", 9))
        style.configure("TCheckbutton", background=self.C_PANEL, foreground=self.C_TEXT,
                        font=f_body, focuscolor=self.C_PANEL)
        style.map("TCheckbutton",
                  background=[("active", self.C_PANEL)],
                  indicatorcolor=[("selected", self.C_ACCENT), ("!selected", self.C_FIELD)],
                  foreground=[("active", self.C_ACCENT)])
        # inputs: flat, padded, subtle border that turns amber on focus
        style.configure("TEntry", fieldbackground=self.C_FIELD, foreground=self.C_TEXT,
                        insertcolor=self.C_ACCENT, insertwidth=2,
                        bordercolor=self.C_LINE,
                        lightcolor=self.C_LINE, darkcolor=self.C_LINE,
                        padding=6, relief="flat")
        style.map("TEntry", bordercolor=[("focus", self.C_ACCENT)],
                  lightcolor=[("focus", self.C_ACCENT)])
        style.configure("TCombobox", fieldbackground=self.C_FIELD, background=self.C_PANEL2,
                        foreground=self.C_TEXT, arrowcolor=self.C_ACCENT,
                        insertcolor=self.C_ACCENT, insertwidth=2,
                        bordercolor=self.C_LINE, lightcolor=self.C_LINE,
                        darkcolor=self.C_LINE, padding=5, relief="flat")
        style.map("TCombobox", fieldbackground=[("readonly", self.C_FIELD)],
                  foreground=[("readonly", self.C_TEXT)],
                  selectbackground=[("readonly", self.C_FIELD),
                                    ("!focus", self.C_FIELD)],
                  selectforeground=[("readonly", self.C_TEXT),
                                    ("!focus", self.C_TEXT)],
                  bordercolor=[("focus", self.C_ACCENT)],
                  arrowcolor=[("active", self.C_ACCENT2)])
        self.option_add("*TCombobox*Listbox.background", self.C_FIELD)
        self.option_add("*TCombobox*Listbox.foreground", self.C_TEXT)
        self.option_add("*TCombobox*Listbox.selectBackground", self.C_ACCENT)
        self.option_add("*TCombobox*Listbox.selectForeground", "#20140a")
        self.option_add("*TCombobox*Listbox.font", f_body)
        self.option_add("*TCombobox*Listbox.borderWidth", 0)
        # Make the blinking text caret clearly visible on the dark theme in any
        # plain tk widgets too (amber, 2px), matching the ttk styles above.
        self.option_add("*Entry.insertBackground", self.C_ACCENT)
        self.option_add("*Entry.insertWidth", 2)
        self.option_add("*Text.insertBackground", self.C_ACCENT)
        self.option_add("*Text.insertWidth", 2)
        # results table
        style.configure("Treeview", background=self.C_FIELD, fieldbackground=self.C_FIELD,
                        foreground=self.C_TEXT, bordercolor=self.C_LINE,
                        rowheight=26, font=("Segoe UI", 9))
        style.map("Treeview", background=[("selected", self.C_SEL)],
                  foreground=[("selected", self.C_ACCENT)])
        style.configure("Treeview.Heading", background=self.C_PANEL2,
                        foreground=self.C_MUTE, relief="flat",
                        font=("Segoe UI Semibold", 9), padding=6)
        style.map("Treeview.Heading",
                  background=[("active", self.C_LINE)],
                  foreground=[("active", self.C_ACCENT)])
        style.configure("Vertical.TScrollbar", background=self.C_PANEL2,
                        troughcolor=self.C_BG, arrowcolor=self.C_MUTE,
                        bordercolor=self.C_BG, relief="flat")
        style.map("Vertical.TScrollbar", background=[("active", self.C_LINE)])

    # ---- UI construction
    def _card(self, parent, title):
        """A raised panel with a title strip — the building block of the UI."""
        outer = tk.Frame(parent, bg=self.C_LINE, bd=0, highlightthickness=0)
        inner = tk.Frame(outer, bg=self.C_PANEL)
        inner.pack(fill="both", expand=True, padx=1, pady=1)
        if title:
            bar = tk.Frame(inner, bg=self.C_PANEL)
            bar.pack(fill="x", padx=14, pady=(10, 0))
            tk.Frame(bar, bg=self.C_ACCENT, width=3, height=14).pack(side="left", padx=(0, 8))
            tk.Label(bar, text=title, bg=self.C_PANEL, fg=self.C_ACCENT,
                     font=("Segoe UI Semibold", 10)).pack(side="left")
        body = tk.Frame(inner, bg=self.C_PANEL)
        body.pack(fill="both", expand=True, padx=14, pady=12)
        return outer, body

    def _field_label(self, parent, text):
        return tk.Label(parent, text=text, bg=self.C_PANEL, fg=self.C_MUTE,
                        font=("Segoe UI", 9))

    def _build_ui(self):
        # ---- header bar
        header = tk.Frame(self, bg=self.C_PANEL2, height=52)
        header.pack(fill="x"); header.pack_propagate(False)
        # LPU logo at the top-left (falls back to a lock glyph if unavailable)
        logo = self._load_header_logo()
        if logo is not None:
            self._header_logo = logo
            tk.Label(header, image=logo, bg=self.C_PANEL2).pack(side="left", padx=(16, 8))
            tk.Label(header, text="LOCK HUNTER", bg=self.C_PANEL2, fg=self.C_ACCENT,
                     font=("Segoe UI Semibold", 15)).pack(side="left")
        else:
            tk.Label(header, text="🔒  LOCK HUNTER", bg=self.C_PANEL2, fg=self.C_ACCENT,
                     font=("Segoe UI Semibold", 15)).pack(side="left", padx=18)
        self.count_var = tk.StringVar(value="")
        header_right = tk.Frame(header, bg=self.C_PANEL2)
        header_right.pack(side="right", padx=18)
        tk.Label(header_right, textvariable=self.count_var, bg=self.C_PANEL2,
                 fg=self.C_MUTE, font=("Segoe UI", 9),
                 justify="right").pack(side="top", anchor="e")
        # ---- tab bar
        tabbar = tk.Frame(self, bg=self.C_PANEL2)
        tabbar.pack(fill="x")
        self._tabs = {}
        self._tab_name = tk.StringVar(value="Search")
        self._tab_history = []          # for the mouse "back" button
        # Mouse back button → previous tab. Different platforms/mice report the
        # extra buttons differently, so bind the common candidates. (On Windows
        # the wheel is <MouseWheel>, so Button-4/5 don't conflict with scroll.)
        for seq in ("<Button-8>", "<Button-4>"):
            try:
                self.bind_all(seq, self._go_back_tab)
            except tk.TclError:
                pass
        for name in ("Search", "Locks", "Compare"):
            b = tk.Label(tabbar, text=name, bg=self.C_PANEL2, fg=self.C_MUTE,
                         font=("Segoe UI Semibold", 11), padx=22, pady=10,
                         cursor="hand2")
            b.pack(side="left")
            b.bind("<Button-1>", lambda e, n=name: self._show_tab(n))
            self._tabs[name] = b
        tk.Frame(self, bg=self.C_ACCENT, height=2).pack(fill="x")

        # ---- tab body container (pages stack in the same spot)
        self._pages = tk.Frame(self, bg=self.C_BG)
        self._pages.pack(fill="both", expand=True)
        self.page_search = tk.Frame(self._pages, bg=self.C_BG)
        self.page_locks = tk.Frame(self._pages, bg=self.C_BG)
        self.page_compare = tk.Frame(self._pages, bg=self.C_BG)

        self._build_search_tab(self.page_search)
        self._build_locks_tab(self.page_locks)
        self._build_compare_tab(self.page_compare)

        # ---- status bar (with version on the right)
        tk.Frame(self, bg=self.C_LINE, height=1).pack(fill="x")
        statusbar = tk.Frame(self, bg=self.C_PANEL2)
        statusbar.pack(fill="x")
        self.status = tk.StringVar(value="Ready — pick a lock and click Search live. Double-click a result to open it.")
        tk.Label(statusbar, textvariable=self.status, bg=self.C_PANEL2, fg=self.C_MUTE,
                 font=("Segoe UI", 9), anchor="w", padx=14, pady=7).pack(side="left")
        tk.Label(statusbar, text=f"v{VERSION}", bg=self.C_PANEL2, fg=self.C_MUTE,
                 font=("Segoe UI", 9), padx=14, pady=7).pack(side="right")
        upd = tk.Label(statusbar, text="Check for updates", bg=self.C_PANEL2,
                       fg=self.C_ACCENT, font=("Segoe UI", 9, "underline"),
                       cursor="hand2", pady=7)
        upd.pack(side="right")
        upd.bind("<Button-1>", lambda _e: self._check_for_updates())

        # Help: builds a diagnostic report and opens the user's email client
        # with a pre-addressed draft (they add their message; the short log is
        # already in the body, the full report saved to a file to attach).
        hlp = tk.Label(statusbar, text="Help", bg=self.C_PANEL2,
                       fg=self.C_ACCENT, font=("Segoe UI", 9, "underline"),
                       cursor="hand2", pady=7)
        hlp.pack(side="right", padx=(0, 16))
        hlp.bind("<Button-1>", lambda _e: self._open_help_email())

        # Export the current Search results to an Excel file.
        exp = tk.Label(statusbar, text="Export Lock Search", bg=self.C_PANEL2,
                       fg=self.C_ACCENT, font=("Segoe UI", 9, "underline"),
                       cursor="hand2", pady=7)
        exp.pack(side="right", padx=(0, 16))
        exp.bind("<Button-1>", lambda _e: self._export_search_xlsx())

        self._show_tab("Search")


    def _show_tab(self, name, record=True):
        cur = self._tab_name.get()
        if record and cur and cur != name:
            # remember where we came from for the mouse "back" button
            self._tab_history.append(cur)
        self._tab_name.set(name)
        for n, b in self._tabs.items():
            if n == name:
                b.config(fg=self.C_ACCENT, bg=self.C_PANEL)
            else:
                b.config(fg=self.C_MUTE, bg=self.C_PANEL2)
        self.page_search.pack_forget()
        self.page_locks.pack_forget()
        self.page_compare.pack_forget()
        page = {"Search": self.page_search, "Locks": self.page_locks,
                "Compare": self.page_compare}.get(name, self.page_search)
        page.pack(fill="both", expand=True)
        if name == "Locks":
            self._refresh_locks_table()

    def _go_back_tab(self, _e=None):
        """Mouse 'back' button: return to the previously-viewed tab."""
        hist = getattr(self, "_tab_history", None)
        if hist:
            prev = hist.pop()
            self._show_tab(prev, record=False)
        return "break"

    # ---------- COMPARE TAB ----------
    def _build_compare_tab(self, wrap):
        pad = tk.Frame(wrap, bg=self.C_BG)
        pad.pack(fill="both", expand=True, padx=14, pady=12)

        card, body = self._card(pad, "COMPARE COLLECTIONS")
        card.pack(fill="both", expand=True)

        tk.Label(
            body,
            text=("Enter someone else's LPU profile link to see which locks "
                  "they own that are on your wishlist."),
            bg=self.C_PANEL, fg=self.C_MUTE, font=("Segoe UI", 10),
            justify="left", wraplength=780).pack(anchor="w", pady=(0, 10))

        LF = ("Segoe UI", 10)
        ctrl = tk.Frame(body, bg=self.C_PANEL)
        ctrl.pack(fill="x", pady=(0, 10))
        tk.Label(ctrl, text="Their LPU profile URL:", bg=self.C_PANEL,
                 fg=self.C_MUTE, font=LF).pack(side="left", padx=(0, 6))
        self.compare_url_var = tk.StringVar()
        ce = ttk.Entry(ctrl, textvariable=self.compare_url_var, width=50)
        ce.pack(side="left", padx=(0, 10))
        ce.bind("<Return>", lambda *_: self._start_compare())
        self.compare_btn = FlatButton(ctrl, "Compare",
                                      command=self._start_compare, kind="primary")
        self.compare_btn.pack(side="left")
        self.compare_export_btn = FlatButton(
            ctrl, "Export", command=self._compare_export_choose, kind="outline")
        self.compare_export_btn.pack(side="left", padx=(10, 0))
        self.compare_count_var = tk.StringVar(value="")
        tk.Label(ctrl, textvariable=self.compare_count_var, bg=self.C_PANEL,
                 fg=self.C_MUTE, font=LF).pack(side="right", padx=(0, 12))

        self.compare_who_var = tk.StringVar(value="")
        tk.Label(body, textvariable=self.compare_who_var, bg=self.C_PANEL,
                 fg=self.C_TEXT, font=("Segoe UI Semibold", 12),
                 anchor="w").pack(fill="x", pady=(0, 6))

        table = tk.Frame(body, bg=self.C_PANEL)
        table.pack(fill="both", expand=True)
        cols = ("name", "belt", "pad")
        heads = ("Lock", "Belt", "")
        style = ttk.Style(self)
        style.configure("Compare.Treeview", font=("Segoe UI", 11), rowheight=26)
        style.configure("Compare.Treeview.Heading", font=("Segoe UI Semibold", 10))
        self.compare_tree = ttk.Treeview(table, columns=cols, show="headings",
                                         style="Compare.Treeview")
        specs = [("name", 380, "w", False), ("belt", 110, "center", False),
                 ("pad", 20, "w", True)]
        self._compare_sort = ("name", False)   # (column, reverse)
        self._compare_total = 0
        for (c, w, anchor, stretch), h in zip(specs, heads):
            if c in ("name", "belt"):
                self.compare_tree.heading(
                    c, text=h, command=lambda col=c: self._sort_compare(col))
            else:
                self.compare_tree.heading(c, text=h)
            self.compare_tree.column(c, width=w, anchor=anchor, stretch=stretch)
        vs = ttk.Scrollbar(table, orient="vertical", command=self.compare_tree.yview)
        self.compare_tree.configure(yscrollcommand=vs.set)
        self.compare_tree.pack(side="left", fill="both", expand=True)
        vs.pack(side="right", fill="y")
        self.compare_tree.tag_configure("odd", background=self.C_PANEL)
        self.compare_tree.tag_configure("even", background=self.C_FIELD)
        self.compare_tree.tag_configure(
            "grp", background=self.C_PANEL2, foreground=self.C_ACCENT,
            font=("Segoe UI Semibold", 10))
        self.compare_tree.bind("<Double-1>", self._compare_open_page)
        self.compare_tree.bind("<Button-3>", self._compare_context_menu)
        self._compare_pages = {}
        self._compare_names = {}
        self._compare_ids = {}
        self._compare_rows = []      # last results, for Export
        self._compare_who = ""

        tk.Label(
            body,
            text=("Double-click a match to open its LPU page.  "
                  "Tip: import your own profile first (Locks tab → Update "
                  "profile) so your wishlist is loaded."),
            bg=self.C_PANEL, fg=self.C_MUTE,
            font=("Segoe UI", 9)).pack(anchor="w", pady=(6, 0))

    def _start_compare(self):
        url = self.compare_url_var.get().strip()
        if not url:
            messagebox.showinfo(
                "Compare",
                "Paste someone's LPU profile link first.\n\n"
                "It looks like:\n"
                "   https://lpubelts.com/#/profile/<their-id>?collection=Own")
            return
        if "profile/" not in url:
            messagebox.showwarning(
                "Compare",
                "That doesn't look like an LPU profile link.\n\n"
                "It should look like:\n"
                "   https://lpubelts.com/#/profile/<their-id>?collection=Own")
            return
        # need a wishlist to compare against
        conn = db()
        n_wish = conn.execute(
            "SELECT COUNT(*) FROM my_collection WHERE status='wishlist'").fetchone()[0]
        conn.close()
        if not n_wish:
            messagebox.showinfo(
                "Compare",
                "Your wishlist is empty, so there's nothing to compare against "
                "yet.\n\nImport your own LPU profile first: go to the Locks "
                "tab and click \u201cUpdate profile\u201d.")
            return
        self.compare_btn.config(state="disabled")
        self.compare_btn.config(text="Comparing\u2026")
        self.compare_who_var.set("")
        for r in self.compare_tree.get_children():
            self.compare_tree.delete(r)
        self.compare_count_var.set("")
        self.status.set("Reading their LPU profile\u2026 this can take a few seconds.")
        threading.Thread(target=self._compare_worker, args=(url,),
                         daemon=True).start()

    def _compare_worker(self, url):
        try:
            # who is this? (name= in the link) — for a friendly header
            m = re.search(r"[?&]name=([^&]+)", url)
            who = urllib.parse.unquote(m.group(1)) if m else "They"
            # import_lpu_profile returns (own, wish, uid) and does NOT write to
            # my_collection, so this never disturbs the user's own wishlist.
            own, wish, uid = import_lpu_profile(
                url, lambda s: self.q.put(("status", s)))
            if not own and not wish:
                self.q.put(("compare_error",
                            ("Nothing found in that profile",
                             "Couldn't find any locks in that person's Owned "
                             "or Wishlist collections. Double-check the link "
                             "— it should be their profile.")))
                return
            conn = db()
            my_wish = {r[0] for r in conn.execute(
                "SELECT lock_id FROM my_collection WHERE status='wishlist'")}
            my_own = {r[0] for r in conn.execute(
                "SELECT lock_id FROM my_collection WHERE status='own'")}
            they_have = own & my_wish    # their owned locks on MY wishlist
            you_have = my_own & wish     # MY owned locks on THEIR wishlist
            rows = []
            wanted = they_have | you_have
            if wanted:
                qmarks = ",".join("?" * len(wanted))
                info = {r[0]: (r[1], r[2], r[3], r[4]) for r in conn.execute(
                    f"SELECT id, name, belt, belt_full, page_url FROM locks "
                    f"WHERE id IN ({qmarks})", tuple(wanted))}
                for lid in they_have:
                    name, belt, belt_full, page = info.get(
                        lid, (lid, "", "", ""))
                    # show the Black sub-tier ("Black 3") when we have it
                    rows.append((name or lid,
                                 _disp_belt(belt, belt_full) or "\u2014",
                                 page or "", "they"))
                for lid in you_have:
                    name, belt, belt_full, page = info.get(
                        lid, (lid, "", "", ""))
                    rows.append((name or lid,
                                 _disp_belt(belt, belt_full) or "\u2014",
                                 page or "", "you"))
            conn.close()
            rows.sort(key=lambda x: (x[3], (x[0] or "").lower()))
            self.q.put(("compare_results",
                        (who, len(own), len(wish), rows)))
        except Exception as ex:
            self.q.put(("compare_error",
                        ("Compare failed",
                         "Couldn't compare that profile:\n\n" + str(ex))))

    def _populate_compare(self, who, their_total, their_wish_total, rows):
        self._compare_who = who
        self._compare_total = their_total
        self._compare_wish_total = their_wish_total
        self._compare_rows = list(rows)     # keep for Export + re-sort
        self._render_compare()

    def _render_compare(self):
        """Draw the Compare results as TWO groups — locks they own that are on
        your wishlist, then locks you own that are on their wishlist — each
        under a header row, honoring the current header sort within groups."""
        rows = list(getattr(self, "_compare_rows", []))
        scol, srev = getattr(self, "_compare_sort", ("name", False))
        if scol == "belt":
            keyf = lambda r: (_belt_rank(r[1]), r[1] or "",
                              (r[0] or "").lower())
        else:
            keyf = lambda r: (r[0] or "").lower()
        for r in self.compare_tree.get_children():
            self.compare_tree.delete(r)
        self._compare_pages = {}
        self._compare_names = {}
        self._compare_ids = {}
        who = self._compare_who
        they_rows = sorted([r for r in rows if r[3] == "they"],
                           key=keyf, reverse=srev)
        you_rows = sorted([r for r in rows if r[3] == "you"],
                          key=keyf, reverse=srev)
        i = 0
        for header, group in (
                (f"{who} HAS \u2014 on your wishlist ({len(they_rows)})",
                 they_rows),
                (f"YOU HAVE \u2014 on {who}'s wishlist ({len(you_rows)})",
                 you_rows)):
            self.compare_tree.insert("", "end", values=(header, "", ""),
                                     tags=("grp",))
            for name, belt, page, _side in group:
                iid = self.compare_tree.insert(
                    "", "end", values=(f"    {name}", belt, ""),
                    tags=("even" if i % 2 else "odd",))
                self._compare_pages[iid] = page
                self._compare_names[iid] = name
                self._compare_ids[iid] = _lock_id_from_page(page)
                i += 1
        total = getattr(self, "_compare_total", 0)
        n1, n2 = len(they_rows), len(you_rows)
        if n1 or n2:
            self.compare_who_var.set(
                f"Trade check with {who}: they have {n1} lock(s) you want "
                f"\u00b7 you have {n2} they want.")
        else:
            self.compare_who_var.set(
                f"No overlap: none of {who}'s {total} owned locks are on "
                f"your wishlist, and nothing you own is on their wishlist.")
        self.compare_count_var.set(f"{n1} + {n2} match(es)")
        self._update_compare_sort_indicator()

    def _sort_compare(self, col):
        cur_col, rev = getattr(self, "_compare_sort", ("name", False))
        rev = (not rev) if cur_col == col else False
        self._compare_sort = (col, rev)
        self._render_compare()

    def _update_compare_sort_indicator(self):
        col, rev = getattr(self, "_compare_sort", ("name", False))
        arrow = " \u25BC" if rev else " \u25B2"
        if hasattr(self, "compare_tree"):
            self.compare_tree.heading(
                "name", text="Lock" + (arrow if col == "name" else ""))
            self.compare_tree.heading(
                "belt", text="Belt" + (arrow if col == "belt" else ""))

    def _compare_open_page(self, _e):
        sel = self.compare_tree.selection()
        if not sel:
            return
        page = self._compare_pages.get(sel[0])
        if page:
            webbrowser.open(page)

    def _compare_export_choose(self):
        """Ask how to export the current Compare results: copy to clipboard,
        save as CSV, or save as Excel."""
        rows = getattr(self, "_compare_rows", [])
        who = getattr(self, "_compare_who", "") or "them"
        if not rows:
            messagebox.showinfo(
                "Export",
                "No comparison results to export yet. Paste a profile link and "
                "click Compare first.")
            return
        win = tk.Toplevel(self)
        win.title("Export comparison")
        win.configure(bg=self.C_PANEL)
        win.geometry("400x250")
        win.transient(self)
        try:
            win.grab_set()   # modal
        except Exception:
            pass
        tk.Label(win, text="Export the matching locks", bg=self.C_PANEL,
                 fg=self.C_ACCENT, font=("Segoe UI Semibold", 13)).pack(
                     anchor="w", padx=20, pady=(18, 2))
        tk.Label(win,
                 text=(f"{len(rows)} matching lock(s) between you and "
                       f"{who} (both directions). Choose a format:"),
                 bg=self.C_PANEL, fg=self.C_TEXT, font=("Segoe UI", 10),
                 justify="left", wraplength=360, anchor="w").pack(
                     anchor="w", padx=20)
        btns = tk.Frame(win, bg=self.C_PANEL)
        btns.pack(anchor="w", padx=20, pady=(14, 0), fill="x")

        def pick(fn):
            win.destroy()
            fn()
        FlatButton(btns, "Copy to clipboard", kind="primary",
                   command=lambda: pick(self._compare_copy_clipboard)).pack(
                       anchor="w", pady=3)
        FlatButton(btns, "Save as CSV\u2026", kind="outline",
                   command=lambda: pick(self._compare_export_csv)).pack(
                       anchor="w", pady=3)
        FlatButton(btns, "Save as Excel\u2026", kind="outline",
                   command=lambda: pick(self._export_compare_xlsx)).pack(
                       anchor="w", pady=3)
        FlatButton(btns, "Cancel", kind="subtle",
                   command=win.destroy).pack(anchor="w", pady=3)

    def _compare_table_rows(self):
        """(headers, rows) for text export — lock name, belt, and which side
        of the trade has it (no link, per request)."""
        who = getattr(self, "_compare_who", "") or "They"
        headers = ["Lock name", "Belt", "Who has it"]
        out = []
        for name, belt, page, side in getattr(self, "_compare_rows", []):
            out.append([name or "", belt or "",
                        who if side == "they" else "You"])
        return headers, out

    def _compare_copy_clipboard(self):
        """Copy the current Compare results to the clipboard as tab-separated
        text (pastes cleanly into Excel / Google Sheets / any text box)."""
        headers, data = self._compare_table_rows()
        lines = ["\t".join(headers)]
        for row in data:
            lines.append("\t".join(row))
        text = "\n".join(lines)
        try:
            self.clipboard_clear()
            self.clipboard_append(text)
            self.update()   # keep it on the clipboard after focus moves on
        except Exception as ex:
            messagebox.showerror("Copy", f"Couldn't copy to clipboard: {ex}")
            return
        self.status.set(f"Copied {len(data)} lock(s) to the clipboard.")
        messagebox.showinfo(
            "Copied",
            f"Copied {len(data)} matching lock(s) to the clipboard.\n\n"
            "Paste into Excel, Google Sheets, or any text box.")

    def _compare_export_csv(self):
        """Export the current Compare results to a CSV file."""
        import csv
        rows = getattr(self, "_compare_rows", [])
        who = getattr(self, "_compare_who", "") or "them"
        if not rows:
            messagebox.showinfo(
                "Export", "No comparison results to export yet.")
            return
        safe_who = re.sub(r"[^A-Za-z0-9_-]+", "_", who).strip("_") or "profile"
        default_name = ("lock_compare_" + safe_who + "_" +
                        datetime.datetime.now().strftime("%Y%m%d_%H%M") + ".csv")
        path = filedialog.asksaveasfilename(
            title="Export Comparison (CSV)",
            defaultextension=".csv",
            initialfile=default_name,
            filetypes=[("CSV file", "*.csv")])
        if not path:
            return   # cancelled
        headers, data = self._compare_table_rows()
        try:
            # utf-8-sig so Excel shows accented lock names correctly
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f)
                w.writerow(headers)
                w.writerows(data)
        except Exception as ex:
            messagebox.showerror("Export", f"Couldn't write the file: {ex}")
            return
        if messagebox.askyesno(
                "Export complete",
                f"Exported {len(data)} matching lock(s) between you and "
                f"{who} to:\n{path}\n\nOpen the file now?"):
            try:
                os.startfile(path)   # Windows: open in Excel
            except Exception:
                webbrowser.open("file://" + path.replace("\\", "/"))

    def _export_compare_xlsx(self):
        """Export the current Compare results (both trade directions) to an
        Excel file."""
        rows = getattr(self, "_compare_rows", [])
        who = getattr(self, "_compare_who", "") or "them"
        if not rows:
            messagebox.showinfo(
                "Export",
                "No comparison results to export yet. Paste a profile link and "
                "click Compare first.")
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except Exception:
            messagebox.showerror(
                "Export",
                "The Excel export needs the 'openpyxl' library, which isn't "
                "available in this build. Rebuild with the updated "
                "requirements, or let me know and I can switch export to CSV.")
            return
        safe_who = re.sub(r"[^A-Za-z0-9_-]+", "_", who).strip("_") or "profile"
        default_name = ("lock_compare_" + safe_who + "_" +
                        datetime.datetime.now().strftime("%Y%m%d_%H%M") + ".xlsx")
        path = filedialog.asksaveasfilename(
            title="Export Comparison",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel workbook", "*.xlsx")])
        if not path:
            return   # cancelled
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Compare"
            headers = ["Lock name", "Belt", "Who has it"]
            ws.append(headers)
            hdr_font = Font(bold=True, color="FFFFFF")
            hdr_fill = PatternFill("solid", fgColor="C58A2E")
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=c)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(vertical="center")
            for name, belt, page, side in rows:
                r = ws.max_row + 1
                ws.cell(row=r, column=1, value=name or "")
                ws.cell(row=r, column=2, value=belt or "")
                ws.cell(row=r, column=3,
                        value=(who if side == "they" else "You"))
            for i, w in enumerate([40, 14, 22], 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
            ws.freeze_panes = "A2"
            wb.save(path)
        except Exception as ex:
            messagebox.showerror("Export", f"Couldn't write the file: {ex}")
            return
        n = len(rows)
        if messagebox.askyesno(
                "Export complete",
                f"Exported {n} matching lock(s) between you and {who} to:\n"
                f"{path}\n\nOpen the file now?"):
            try:
                os.startfile(path)   # Windows: open in Excel
            except Exception:
                webbrowser.open("file://" + path.replace("\\", "/"))

    # ---------- OWNERS TAB ----------












    def _bazaar_data_fresh(self):
        """The Lock Bazaar dataset if it was downloaded in the last 10 minutes,
        else None (module-level cache filled by _bazaar_dataset)."""
        d, at = _bazaar_cache.get("data"), _bazaar_cache.get("at")
        if d is not None and at and (
                (datetime.datetime.now() - at).total_seconds() < 600):
            return d
        return None

    def _kick_bazaar_refresh(self, q, lock_id):
        """Download the Lock Bazaar feed in the background, then re-run the
        owner search once so the ★ for-sale marks appear. No-op if a fetch is
        already in flight; the re-run never re-triggers a fetch."""
        if getattr(self, "_bazaar_fetching", False):
            return
        self._bazaar_fetching = True

        def w():
            try:
                _bazaar_dataset(lambda s: self.q.put(("status", s)))
            except Exception:
                pass
            finally:
                self._bazaar_fetching = False
                self.q.put(("owners_rerun", (q, lock_id)))
        threading.Thread(target=w, daemon=True).start()



    def _start_bazaar_watch(self):
        """Startup chore: quietly fetch the Lock Bazaar feed, diff it against
        the snapshot from the previous run, and pop a note if anything NEW
        matches the wishlist. First ever run just saves a baseline."""
        threading.Thread(target=self._bazaar_watch_worker, daemon=True).start()

    def _bazaar_watch_worker(self):
        try:
            data = _bazaar_dataset(lambda s: None)   # silent fetch
            if not data:
                return   # feed unreachable — keep the old snapshot untouched
            entries = [e for e in data if isinstance(e, dict)]
            cur_keys = {_bazaar_entry_key(e) for e in entries}
            snap_path = os.path.join(APP_DIR, "bazaar_seen.json")
            old_keys = set()
            try:
                with open(snap_path, "r", encoding="utf-8") as f:
                    snap = json.load(f)
                # Versioned format {"v":N,"keys":[...]}. A bare list (older
                # builds) or a mismatched version means the key format has
                # changed, so treat this as a fresh baseline: leave old_keys
                # empty -> no diff, no popup, just re-save in the new format.
                if isinstance(snap, dict) and snap.get("v") == _BZ_SNAP_VER:
                    old_keys = set(snap.get("keys") or [])
            except Exception:
                old_keys = set()
            lines = []
            if old_keys:                      # not the first ever run
                new_entries = [e for e in entries
                               if _bazaar_entry_key(e) not in old_keys]
                # If MOST of the feed looks "new", the snapshot key format or
                # the feed shape changed — re-baseline silently instead of
                # spamming a huge false alert.
                if len(new_entries) > max(20, len(entries) // 2):
                    new_entries = []
                if new_entries:
                    conn = db()
                    wish = conn.execute(
                        "SELECT l.id, l.name, l.belt, l.belt_full"
                        " FROM my_collection"
                        " m JOIN locks l ON l.id=m.lock_id"
                        " WHERE m.status='wishlist'").fetchall()
                    conn.close()
                    seen = set()
                    for lid, nm, belt, bf in wish:
                        dbelt = _disp_belt(belt, bf)
                        ents = list(bazaar_entries_for_lock_id(lid,
                                                               new_entries))
                        have = {_bazaar_entry_key(x) for x in ents}
                        for x in bazaar_entries_for_lock(nm, dbelt,
                                                         new_entries):
                            if _bazaar_entry_key(x) not in have:
                                ents.append(x)
                        for e in ents:
                            k = (nm, _bazaar_entry_key(e))
                            if k in seen:
                                continue
                            seen.add(k)
                            seller = _bz_seller(e)
                            price = _bz_price(e)
                            bit = nm
                            if seller:
                                bit += f" \u2014 {seller}"
                            if price:
                                bit += f" ({price})"
                            lines.append(bit)
            try:
                with open(snap_path, "w", encoding="utf-8") as f:
                    json.dump({"v": _BZ_SNAP_VER, "keys": sorted(cur_keys)}, f)
            except Exception:
                pass
            if lines:
                self.q.put(("bazaar_news", lines))
        except Exception:
            pass   # a startup nicety must never crash or nag






    def _compare_context_menu(self, evt):
        iid = self.compare_tree.identify_row(evt.y)
        if not iid:
            return
        self.compare_tree.selection_set(iid)
        name = getattr(self, "_compare_names", {}).get(iid)
        if not name:
            return
        menu = tk.Menu(self, tearoff=0, bg=self.C_PANEL2, fg=self.C_TEXT,
                       activebackground=self.C_ACCENT, activeforeground="#20140a",
                       bd=0)
        menu.add_command(label="Open LPU page",
                         command=lambda: self._compare_open_page(None))
        try:
            menu.tk_popup(evt.x_root, evt.y_root)
        finally:
            menu.grab_release()

    # ---------- SEARCH TAB ----------
    def _build_search_tab(self, wrap):
        pad = tk.Frame(wrap, bg=self.C_BG)
        pad.pack(fill="both", expand=True, padx=14, pady=12)

        # Hidden state — the API key and profile URL are collected at startup
        # (and via the Locks tab), so they no longer clutter this screen.
        self.key_var = tk.StringVar(value=self.cfg.get("api_key", ""))
        self.profile_var = tk.StringVar(value=self.cfg.get("profile_url", ""))

        # compact toolbar: catalog + database maintenance, plus a small link to
        # change the saved API key if needed.
        toolbar = tk.Frame(pad, bg=self.C_BG)
        toolbar.pack(fill="x", pady=(0, 10))
        self.sync_btn = FlatButton(toolbar, "Update LPU catalog",
                                   command=self._start_sync, kind="ghost")
        self.sync_btn.pack(side="left")
        FlatButton(toolbar, "Clear all", command=self._clear_all,
                   kind="subtle").pack(side="left", padx=(8, 0))
        key_link = tk.Label(toolbar, text="Enter/Change API key", bg=self.C_BG,
                            fg=self.C_MUTE, font=("Segoe UI", 9, "underline"),
                            cursor="hand2")
        key_link.pack(side="right")
        key_link.bind("<Button-1>", lambda _e: self._prompt_api_key(force=True))

        # search card
        search_card, search = self._card(pad, "SEARCH A LOCK")
        search_card.pack(fill="x")
        self._field_label(search, "Belt").grid(row=0, column=0, sticky="w")
        self._field_label(search, "Lock").grid(row=0, column=1, sticky="w", padx=(8, 0))
        self._field_label(search, "Sales").grid(row=0, column=2, sticky="w", padx=(8, 0))

        self.belt_filter_var = tk.StringVar(value="All belts")
        self.belt_box = ttk.Combobox(search, textvariable=self.belt_filter_var, width=13,
                                     state="readonly", values=["All belts"])
        self.belt_box.grid(row=1, column=0, sticky="w", pady=(2, 0))
        self.belt_box.bind("<<ComboboxSelected>>",
                           lambda e: (self._clear_combo_highlight(e),
                                      self._on_belt_filter_changed()))

        self.lock_var = tk.StringVar()
        self.lock_box = ttk.Combobox(search, textvariable=self.lock_var, width=40)
        self.lock_box.grid(row=1, column=1, sticky="we", pady=(2, 0), padx=(8, 0))
        self.lock_box.bind("<KeyRelease>", self._filter_locks)
        self.lock_box.bind("<<ComboboxSelected>>", lambda *_: self._on_lock_selected())
        # Pressing Enter (or leaving the box) commits the typed/auto-filled name
        # so the preview picture and Lock Bazaar check update immediately.
        self.lock_box.bind("<Return>", self._commit_lock_selection)
        self.lock_box.bind("<FocusOut>", lambda *_: self._on_lock_selected())

        # The item-condition (New / Used / Both) selector was removed in
        # 4.5.4: every search now covers BOTH new and used. The free probes
        # could never honor it anyway (they can't know an item's condition up
        # front), so the selector only ever narrowed the AI search while
        # silently doing nothing for the searches everyone actually uses.

        # Sales = venue type: auction/secondhand marketplaces, retail website
        # sales, or both in one pass.
        self.sales_var = tk.StringVar(value="Both")
        ttk.Combobox(search, textvariable=self.sales_var, width=19, state="readonly",
                     values=["Both", "Auction / secondhand", "Website sales"]).grid(
            row=1, column=2, sticky="w", pady=(2, 0), padx=(8, 0))

        self.search_btn = FlatButton(search, "⚲  Search live",
                                     command=self._start_search, kind="primary")
        self.search_btn.grid(row=1, column=3, sticky="w", padx=(12, 0), pady=(2, 0))
        search.columnconfigure(1, weight=1)

        opts = tk.Frame(search, bg=self.C_PANEL)
        opts.grid(row=2, column=0, columnspan=4, sticky="w", pady=(10, 0))
        self.pickup_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(opts, text="Exclude pickup / meetup-only",
                        variable=self.pickup_var).pack(side="left", padx=(0, 18))
        # Optional AI web search (OFF by default). The normal search is
        # FREE — it queries eBay, ~46 marketplaces, and the LPU Lock Bazaar
        # directly. Ticking this ALSO asks an Anthropic (Claude) AI model to
        # web-search for the lock, which can surface listings on sites the
        # built-in probes don't cover. It is NOT free: it uses YOUR Anthropic
        # API key and each run costs a small amount of API credit (typically a
        # few US cents). You'll be asked for a key the first time you enable it.
        self.ai_var = tk.BooleanVar(value=False)
        self.ai_check = ttk.Checkbutton(
            opts, text="Also use AI web search (uses your paid Anthropic API key)",
            variable=self.ai_var, command=self._toggle_ai_search)
        self.ai_check.pack(side="left", padx=(0, 18))
        # Lock-only filter: drop same-name non-lock listings (e.g. "Sol"
        # sunglasses when searching the "Sol 2500" lock). On by default.
        self.lockonly_var = tk.BooleanVar(value=True)
        self.lockonly_check = ttk.Checkbutton(
            opts, text="Lock results only (filter out same-name non-locks)",
            variable=self.lockonly_var,
            command=self._toggle_lock_filter)
        self.lockonly_check.pack(side="left", padx=(0, 18))
        # Optional Facebook Marketplace search (OFF by default). Reads only
        # public listings (no login, nothing stored), sweeps ~10 country
        # markets for the single lock being searched, and runs one market at a
        # time so it stays polite — so it's slower, and can occasionally find
        # nothing if Facebook changes their page. Single-lock searches only.
        self.fb_var = tk.BooleanVar(value=False)
        self.fb_check = ttk.Checkbutton(
            opts, text="Also search Facebook Marketplace (public listings, slower)",
            variable=self.fb_var, command=self._toggle_fb_search)
        self.fb_check.pack(side="left", padx=(0, 18))
        # LPU Lock Bazaar is always included; kept as hidden state (no checkbox).
        self.bazaar_var = tk.BooleanVar(value=True)
        # Extended search is now ALWAYS ON (no toggle) — deeper pages, more
        # query variants, deeper verification. Kept as always-true hidden state
        # so the rest of the plumbing is unchanged.
        self.deep_var = tk.BooleanVar(value=True)
        # main row: selected lock + results
        mainrow = tk.Frame(pad, bg=self.C_BG)
        mainrow.pack(fill="both", expand=True, pady=(14, 0))

        detail_card, detail = self._card(mainrow, "SELECTED LOCK")
        detail_card.pack(side="left", fill="y", padx=(0, 14))
        # fixed-size image frame so layout doesn't jump when a photo loads
        self.thumb_holder = tk.Frame(detail, bg=self.C_FIELD, width=200, height=200)
        self.thumb_holder.pack(pady=(0, 8))
        self.thumb_holder.pack_propagate(False)
        self.thumb_label = tk.Label(self.thumb_holder, bg=self.C_FIELD, fg=self.C_MUTE,
                                    text="select a lock", font=("Segoe UI", 9))
        self.thumb_label.place(relx=0.5, rely=0.5, anchor="center")
        self.name_var = tk.StringVar(value="")
        tk.Label(detail, textvariable=self.name_var, bg=self.C_PANEL, fg=self.C_TEXT,
                 font=("Segoe UI Semibold", 11), wraplength=190,
                 justify="center").pack(fill="x")
        self.belt_var = tk.StringVar(value="")
        tk.Label(detail, textvariable=self.belt_var, bg=self.C_PANEL, fg=self.C_MUTE,
                 font=("Segoe UI", 9), wraplength=190, justify="center").pack(
            fill="x", pady=(2, 10))
        self.lpu_btn = FlatButton(detail, "Open LPU page", command=self._open_lpu,
                                  kind="outline")
        self.lpu_btn.pack(fill="x", pady=(0, 6)); self.lpu_btn.set_enabled(False)
        self.bazaar_btn = FlatButton(detail, "Open Lock Bazaar",
                                     command=self._open_bazaar_for_lock, kind="outline")
        self.bazaar_btn.pack(fill="x"); self.bazaar_btn.set_enabled(False)

        results_card, results = self._card(mainrow, "RESULTS")
        results_card.pack(side="left", fill="both", expand=True)
        fstrip = tk.Frame(results, bg=self.C_PANEL)
        fstrip.pack(fill="x", pady=(0, 8))
        self._field_label(fstrip, "Filter:").pack(side="left", padx=(0, 6))
        self.filter_var = tk.StringVar()
        fe = ttk.Entry(fstrip, textvariable=self.filter_var, width=22)
        fe.pack(side="left"); fe.bind("<KeyRelease>", lambda *_: self._refresh_table())
        self.dbcond_var = tk.StringVar(value="All")
        ttk.Combobox(fstrip, textvariable=self.dbcond_var, width=6, state="readonly",
                     values=["All", "new", "used"]).pack(side="left", padx=6)
        self.dbship_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(fstrip, text="Hide non-shippable results", variable=self.dbship_var,
                        command=self._refresh_table).pack(side="left", padx=4)
        # Optional USD estimate next to foreign prices (live rates, cached 12h)
        self.usd_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(fstrip, text="Show USD estimate", variable=self.usd_var,
                        command=self._usd_toggled).pack(side="left", padx=4)
        FlatButton(fstrip, "Open log", command=lambda: webbrowser.open(LOG_PATH),
                   kind="subtle").pack(side="right")
        FlatButton(fstrip, "Clear results", command=self._clear_results,
                   kind="subtle").pack(side="right", padx=(0, 8))

        table = tk.Frame(results, bg=self.C_PANEL)
        table.pack(fill="both", expand=True)
        # "Ship" column replaced with "Rarity" in 4.5.8: shipping data is
        # still collected and still drives the "Hide non-shippable results"
        # checkbox — it just isn't shown as a column anymore.
        cols = ("lock", "title", "price", "cond", "site", "loc", "rarity")
        heads = ("Lock", "Title", "Price", "Cond", "Site", "Location", "Rarity")
        self.tree = ttk.Treeview(table, columns=cols, show="headings")
        widths = (150, 330, 95, 52, 105, 105, 64)
        # click a column header to sort: 1st click ascending, 2nd descending,
        # 3rd back to the default order (Bazaar first, newest first)
        self._res_heads = dict(zip(cols, heads))
        self._sort_col = None
        self._sort_rev = False
        for c, h, w in zip(cols, heads, widths):
            self.tree.heading(c, text=h,
                              command=lambda cc=c: self._sort_by(cc))
            self.tree.column(c, width=w, anchor="w")
        vs = ttk.Scrollbar(table, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vs.set)
        self.tree.pack(side="left", fill="both", expand=True)
        vs.pack(side="right", fill="y")
        self.tree.bind("<Double-1>", self._open_url)
        self.tree.bind("<Button-3>", self._results_context_menu)  # right-click
        # Delete/Backspace also removes the selected result row(s).
        self.tree.bind("<Delete>", lambda _e: self._remove_selected_results())
        self.tree.bind("<BackSpace>", lambda _e: self._remove_selected_results())
        # Hover thumbnails: preview the listing's photo (or the LPU catalog
        # image as a fallback) when the cursor rests on a result row.
        self._thumb_win = None
        self._thumb_lbl = None
        self._thumb_img = None
        self._thumb_cache = {}
        self._thumb_row = None
        self._thumb_after = None
        self.tree.bind("<Motion>", self._thumb_motion, add="+")
        self.tree.bind("<Leave>", lambda _e: self._thumb_hide(), add="+")
        self.tree.tag_configure("odd", background=self.C_PANEL)
        self.tree.tag_configure("even", background=self.C_FIELD)
        # "Scan in progress…" overlay, shown centered over the table while a
        # search runs. Created hidden; place()/place_forget() toggles it.
        self._scan_table = table
        self.scan_overlay = tk.Label(table, text="Scan in progress…",
                                     bg=self.C_FIELD, fg=self.C_ACCENT,
                                     font=("Segoe UI Semibold", 14))

    # ---------- LOCKS TAB ----------
    def _build_locks_tab(self, wrap):
        pad = tk.Frame(wrap, bg=self.C_BG)
        pad.pack(fill="both", expand=True, padx=14, pady=12)

        card, body = self._card(pad, "LOCK DATABASE")
        card.pack(fill="both", expand=True)

        # controls: show filter + belt filter + search  (slightly larger fonts
        # on this tab since there's room)
        LF = ("Segoe UI", 10)
        ctrl = tk.Frame(body, bg=self.C_PANEL)
        ctrl.pack(fill="x", pady=(0, 10))
        tk.Label(ctrl, text="Show:", bg=self.C_PANEL, fg=self.C_MUTE,
                 font=LF).pack(side="left", padx=(0, 6))
        self.locks_show_var = tk.StringVar(value="All locks")
        sb = ttk.Combobox(ctrl, textvariable=self.locks_show_var, width=14, state="readonly",
                          values=["All locks", "Locks I own", "My wishlist"])
        sb.pack(side="left", padx=(0, 12))
        sb.bind("<<ComboboxSelected>>",
                lambda *_: (self._refresh_locks_table(),
                            self._sync_hunt_wishlist_button()))
        tk.Label(ctrl, text="Belt:", bg=self.C_PANEL, fg=self.C_MUTE,
                 font=LF).pack(side="left", padx=(0, 6))
        self.locks_belt_var = tk.StringVar(value="All belts")
        self.locks_belt_box = ttk.Combobox(ctrl, textvariable=self.locks_belt_var, width=13,
                                           state="readonly", values=["All belts"])
        self.locks_belt_box.pack(side="left", padx=(0, 12))
        self.locks_belt_box.bind("<<ComboboxSelected>>", lambda *_: self._refresh_locks_table())
        tk.Label(ctrl, text="Find:", bg=self.C_PANEL, fg=self.C_MUTE,
                 font=LF).pack(side="left", padx=(0, 6))
        self.locks_find_var = tk.StringVar()
        fe = ttk.Entry(ctrl, textvariable=self.locks_find_var, width=22)
        fe.pack(side="left")
        fe.bind("<KeyRelease>", lambda *_: self._refresh_locks_table())
        # Update profile lives here now (top-right of the database screen).
        # It reuses the saved profile URL — no need to paste it again.
        self.profile_btn = FlatButton(ctrl, "Update profile",
                                      command=self._start_profile_import, kind="primary")
        self.profile_btn.pack(side="right")
        # Batch "hunt my whole wishlist" — only enabled when the Show filter is
        # "My wishlist". Runs the FREE (no-API-credit) search for every lock on
        # the wishlist (honoring the belt filter). Can take a while and return
        # a lot — that's expected.
        self.hunt_wishlist_btn = FlatButton(
            ctrl, "Search for Wishlist Locks",
            command=self._toggle_wishlist_hunt, kind="outline")
        self.hunt_wishlist_btn.pack(side="right", padx=(0, 10))
        self._set_hunt_wishlist_enabled(False)
        self.locks_count_var = tk.StringVar(value="")
        tk.Label(ctrl, textvariable=self.locks_count_var, bg=self.C_PANEL, fg=self.C_MUTE,
                 font=LF).pack(side="right", padx=(0, 12))

        # table on the left, a lock-image preview on the right (fills the
        # empty space). They sit side by side in this row.
        tablerow = tk.Frame(body, bg=self.C_PANEL)
        tablerow.pack(fill="both", expand=True)

        table = tk.Frame(tablerow, bg=self.C_PANEL)
        table.pack(side="left", fill="both", expand=True)
        # "Mine" is now two columns: Owned and Wishlist, each showing a check
        # in the relevant spot (blank otherwise). "Rarity" shows stars derived
        # from how many LPU members own the lock (from the catalog sync); the
        # column still sorts by the underlying count. A trailing "pad" column
        # absorbs the extra width so the real columns group left.
        cols = ("name", "belt", "coll", "owned", "wishlist", "pad")
        heads = ("Lock", "Belt", "Rarity", "Owned", "Wishlist", "")
        # slightly larger row + heading fonts on this tab (there's space)
        style = ttk.Style(self)
        style.configure("Locks.Treeview", font=("Segoe UI", 11), rowheight=26)
        style.configure("Locks.Treeview.Heading", font=("Segoe UI Semibold", 10))
        self.locks_tree = ttk.Treeview(table, columns=cols, show="headings",
                                       style="Locks.Treeview")
        col_specs = [("name", 320, "w", False), ("belt", 88, "center", False),
                     ("coll", 88, "center", False),
                     ("owned", 66, "center", False), ("wishlist", 76, "center", False),
                     ("pad", 20, "w", True)]
        self._locks_sort = ("name", False)   # (column, reverse)
        for (c, w, anchor, stretch), h in zip(col_specs, heads):
            if c in ("name", "belt", "coll"):
                self.locks_tree.heading(
                    c, text=h, command=lambda col=c: self._sort_locks(col))
            else:
                self.locks_tree.heading(c, text=h)
            self.locks_tree.column(c, width=w, anchor=anchor, stretch=stretch)
        vs = ttk.Scrollbar(table, orient="vertical", command=self.locks_tree.yview)
        self.locks_tree.configure(yscrollcommand=vs.set)
        self.locks_tree.pack(side="left", fill="both", expand=True)
        vs.pack(side="right", fill="y")
        self.locks_tree.tag_configure("odd", background=self.C_PANEL)
        self.locks_tree.tag_configure("even", background=self.C_FIELD)
        self.locks_tree.bind("<Double-1>", self._locks_open_page)
        self.locks_tree.bind("<Button-3>", self._locks_context_menu)  # right-click
        # selecting a row (single click / arrow keys) loads its image
        self.locks_tree.bind("<<TreeviewSelect>>", self._on_locks_row_selected)

        # ---- image preview panel (right side) ----
        preview = tk.Frame(tablerow, bg=self.C_PANEL, width=340)
        preview.pack(side="left", fill="y", padx=(14, 0))
        preview.pack_propagate(False)
        self.locks_thumb_holder = tk.Frame(preview, bg=self.C_FIELD,
                                           width=300, height=300)
        self.locks_thumb_holder.pack(pady=(2, 8))
        self.locks_thumb_holder.pack_propagate(False)
        self.locks_thumb_label = tk.Label(
            self.locks_thumb_holder, bg=self.C_FIELD, fg=self.C_MUTE,
            text="select a lock", font=("Segoe UI", 10))
        self.locks_thumb_label.place(relx=0.5, rely=0.5, anchor="center")
        self.locks_thumb_name = tk.StringVar(value="")
        tk.Label(preview, textvariable=self.locks_thumb_name, bg=self.C_PANEL,
                 fg=self.C_TEXT, font=("Segoe UI Semibold", 12),
                 wraplength=300, justify="center").pack(fill="x")
        self.locks_thumb_belt = tk.StringVar(value="")
        tk.Label(preview, textvariable=self.locks_thumb_belt, bg=self.C_PANEL,
                 fg=self.C_MUTE, font=("Segoe UI", 10),
                 wraplength=300, justify="center").pack(fill="x", pady=(2, 0))
        self._locks_thumb_ref = None
        self._locks_thumb_current = ""   # name of the lock the preview shows

        hint = tk.Label(body, text="Click a lock to preview its image  ·  double-click to open the LPU page  ·  right-click to search this lock.",
                        bg=self.C_PANEL, fg=self.C_MUTE, font=("Segoe UI", 9))
        hint.pack(anchor="w", pady=(6, 0))

    def _refresh_locks_table(self):
        if not hasattr(self, "locks_tree"):
            return
        for r in self.locks_tree.get_children():
            self.locks_tree.delete(r)
        self._locks_pages = {}
        mode = self.locks_show_var.get()
        belt = self.locks_belt_var.get()
        find = self.locks_find_var.get().strip()
        q = ("SELECT locks.name, locks.belt, locks.page_url, locks.image_url,"
             " (SELECT status FROM my_collection m WHERE m.lock_id=locks.id),"
             " locks.id, locks.belt_full, locks.owner_count"
             " FROM locks WHERE 1=1")
        params = []
        if mode == "Locks I own":
            q += " AND EXISTS(SELECT 1 FROM my_collection m WHERE m.lock_id=locks.id AND m.status='own')"
        elif mode == "My wishlist":
            q += " AND EXISTS(SELECT 1 FROM my_collection m WHERE m.lock_id=locks.id AND m.status='wishlist')"
        if belt and belt != "All belts":
            q += " AND locks.belt=?"; params.append(belt)
        q += " ORDER BY locks.name LIMIT 4000"
        conn = db()
        rows = conn.execute(q, params).fetchall()
        conn.close()
        if find:
            ff = _fold(find)
            rows = [r for r in rows if ff in _fold(r[0])]
        rows = rows[:2000]
        # apply the current header sort: rows are
        # (name, belt, page, image, status, id, belt_full, owner_count)
        scol, srev = getattr(self, "_locks_sort", ("name", False))
        if scol == "belt":
            rows.sort(key=lambda r: (_belt_rank(r[1] or ""),
                                     (r[6] or r[1] or "").lower(),
                                     (r[0] or "").lower()),
                      reverse=srev)
        elif scol == "coll":
            # Collectors: ascending = RAREST first; unknown counts sort last
            rows.sort(key=lambda r: (r[7] is None,
                                     r[7] if r[7] is not None else 0,
                                     (r[0] or "").lower()),
                      reverse=srev)
        else:
            rows.sort(key=lambda r: (r[0] or "").lower(), reverse=srev)
        # belt dropdown ALWAYS offers the full canonical belt list
        self.locks_belt_box["values"] = ["All belts"] + self._belt_order()
        self._locks_names = {}
        self._locks_images = {}
        self._locks_belts = {}
        self._locks_ids = {}
        self._locks_owncounts = {}
        for i, (name, belt_v, page, image_url, status, lock_id, belt_full_v,
                own_cnt) in enumerate(rows):
            owned = "\u2714" if status == "own" else ""
            wished = "\u2714" if status == "wishlist" else ""
            disp = _disp_belt(belt_v, belt_full_v)
            cnt = _rarity_stars(own_cnt) if own_cnt is not None else ""
            iid = self.locks_tree.insert(
                "", "end",
                values=(name, disp or "\u2014", cnt, owned, wished, ""),
                tags=("even" if i % 2 else "odd",))
            self._locks_pages[iid] = page
            self._locks_names[iid] = name
            self._locks_images[iid] = image_url
            self._locks_belts[iid] = disp
            self._locks_ids[iid] = lock_id
            self._locks_owncounts[iid] = own_cnt
        self.locks_count_var.set(f"{len(rows)} locks")
        self._update_locks_sort_indicator()

    def _sort_locks(self, col):
        """Header click: sort the lock database. 'Lock' sorts by name (A→Z,
        then Z→A on a second click); 'Belt' sorts by belt colour (White→Black,
        then reversed)."""
        cur_col, rev = getattr(self, "_locks_sort", ("name", False))
        rev = (not rev) if cur_col == col else False
        self._locks_sort = (col, rev)
        self._refresh_locks_table()

    def _update_locks_sort_indicator(self):
        col, rev = getattr(self, "_locks_sort", ("name", False))
        arrow = " \u25BC" if rev else " \u25B2"   # down / up triangle
        if hasattr(self, "locks_tree"):
            self.locks_tree.heading(
                "name", text="Lock" + (arrow if col == "name" else ""))
            self.locks_tree.heading(
                "belt", text="Belt" + (arrow if col == "belt" else ""))
            self.locks_tree.heading(
                "coll", text="Rarity" + (arrow if col == "coll" else ""))

    def _locks_open_page(self, _evt):
        sel = self.locks_tree.selection()
        if sel:
            page = getattr(self, "_locks_pages", {}).get(sel[0])
            if page:
                webbrowser.open(page)

    def _on_locks_row_selected(self, _evt=None):
        if not HAVE_PIL:
            return
        sel = self.locks_tree.selection()
        if not sel:
            return
        iid = sel[0]
        name = getattr(self, "_locks_names", {}).get(iid)
        if not name:
            return
        belt_v = getattr(self, "_locks_belts", {}).get(iid) or ""
        own_cnt = getattr(self, "_locks_owncounts", {}).get(iid)
        img_url = getattr(self, "_locks_images", {}).get(iid)
        page_url = getattr(self, "_locks_pages", {}).get(iid)
        self._locks_thumb_current = name
        self._locks_thumb_ref = None
        self.locks_thumb_name.set(name)
        belt_txt = f"{belt_v} belt" if belt_v else ""
        if own_cnt is not None:
            belt_txt += ("  \u00b7  " if belt_txt else "") \
                        + f"owned by {own_cnt} collector(s)"
        # second line: the rarity as stars (same scale as the Locks tab)
        stars = _rarity_stars(own_cnt) if own_cnt is not None else ""
        if stars:
            belt_txt += f"\nRarity: {stars}"
        self.locks_thumb_belt.set(belt_txt)
        self.locks_thumb_label.config(image="", text="loading…")
        # cached image URL first, else pull from the lock's LPU page
        if img_url:
            threading.Thread(target=self._fetch_locks_thumb,
                             args=(img_url, name), daemon=True).start()
        elif page_url:
            threading.Thread(target=self._lazy_locks_image,
                             args=(page_url, name), daemon=True).start()
        else:
            self.locks_thumb_label.config(image="", text="no image")

    def _lazy_locks_image(self, page_url, name):
        img = fetch_lock_image(page_url, lambda s: self.q.put(("status", s)))
        if img:
            self._fetch_locks_thumb(img, name)
        else:
            self.q.put(("locks_thumb", (name, None)))

    def _fetch_locks_thumb(self, url, name):
        try:
            fn = os.path.join(IMG_DIR, re.sub(r"[^A-Za-z0-9.]", "_", url)[-80:])
            if not os.path.exists(fn):
                r = requests.get(url, timeout=30,
                                 headers={"User-Agent": "LockHunter/1.0"})
                r.raise_for_status()
                with open(fn, "wb") as f:
                    f.write(r.content)
            img = Image.open(fn)
            if img.mode in ("RGBA", "LA", "P"):
                img = img.convert("RGBA")
                bg = Image.new("RGBA", img.size, (15, 19, 23, 255))  # C_FIELD
                img = Image.alpha_composite(bg, img).convert("RGB")
            else:
                img = img.convert("RGB")
            img.thumbnail((296, 296), Image.LANCZOS)
            self.q.put(("locks_thumb", (name, img)))
        except Exception as ex:
            self.q.put(("locks_thumb", (name, None)))
            self.q.put(("status", f"Thumbnail failed: {ex}"))

    def _locks_context_menu(self, evt):
        # select the row under the cursor, then show a right-click menu
        iid = self.locks_tree.identify_row(evt.y)
        if not iid:
            return
        self.locks_tree.selection_set(iid)
        name = getattr(self, "_locks_names", {}).get(iid)
        if not name:
            return
        menu = tk.Menu(self, tearoff=0, bg=self.C_PANEL2, fg=self.C_TEXT,
                       activebackground=self.C_ACCENT, activeforeground="#20140a",
                       bd=0)
        menu.add_command(label=f"Search for “{name}”",
                         command=lambda: self._search_lock_from_locks_tab(name))
        menu.add_command(label="Open LPU page",
                         command=lambda: self._locks_open_page(None))
        try:
            menu.tk_popup(evt.x_root, evt.y_root)
        finally:
            menu.grab_release()

    def _toggle_lock_filter(self):
        """Enable/disable the lock-context result filter (module-level flag the
        probes read)."""
        global _LOCK_CONTEXT_FILTER
        _LOCK_CONTEXT_FILTER = bool(self.lockonly_var.get())

    def _fb_rate_check(self):
        """Rolling-window Facebook rate limit. Returns (allowed, wait_minutes,
        used_this_hour). At most _FB_MAX_PER_HOUR Facebook sweeps per
        _FB_WINDOW_MIN minutes; timestamps live in cfg so the limit survives
        an app restart. Prunes expired entries as a side effect."""
        import time
        now = time.time()
        win = _FB_WINDOW_MIN * 60
        times = [t for t in self.cfg.get("fb_search_times", [])
                 if isinstance(t, (int, float)) and now - t < win]
        if len(times) != len(self.cfg.get("fb_search_times", [])):
            self.cfg["fb_search_times"] = times
            save_cfg(self.cfg)
        if len(times) >= _FB_MAX_PER_HOUR:
            wait = int((min(times) + win - now) // 60) + 1
            return False, max(1, wait), len(times)
        return True, 0, len(times)

    def _fb_rate_record(self):
        """Record that a Facebook sweep is being run now (persisted)."""
        import time
        now = time.time()
        win = _FB_WINDOW_MIN * 60
        times = [t for t in self.cfg.get("fb_search_times", [])
                 if isinstance(t, (int, float)) and now - t < win]
        times.append(now)
        self.cfg["fb_search_times"] = times
        save_cfg(self.cfg)

    def _toggle_fb_search(self):
        """Ticking 'Also search Facebook Marketplace' turns on the optional
        Facebook sweep. It's public-listings only (no login, nothing stored),
        but slower and best-effort, so confirm before enabling. Unticking is
        always fine."""
        if not self.fb_var.get():
            return
        ok = messagebox.askyesno(
            "Also search Facebook Marketplace?",
            "This also searches Facebook Marketplace across about 10 country "
            "markets for the single lock you search.\n\n"
            "It reads only PUBLIC listings — no Facebook login, and nothing "
            "about your account is used or stored.\n\n"
            "Two things to expect: it's SLOWER (Facebook is queried one "
            "market at a time to stay polite, so it adds up to a minute), and "
            "because Facebook has no real search feature for other apps, it "
            "can occasionally return nothing when Facebook changes their "
            "site.\n\nTo stay within Facebook's limits, this is capped at "
            f"{_FB_MAX_PER_HOUR} Facebook searches per hour (your other "
            "searches are unaffected).\n\nTurn on Facebook Marketplace "
            "search?")
        if not ok:
            self.fb_var.set(False)

    def _toggle_ai_search(self):
        """Ticking 'Also use AI web search' turns on the optional paid AI
        pass. It spends the user's own Anthropic API credits, so make the
        cost explicit and confirm before enabling; a Claude API key is also
        required, so prompt for one if none is saved. Unticking is always
        free and needs no confirmation."""
        if not self.ai_var.get():
            return
        ok = messagebox.askyesno(
            "Use AI web search? (costs API credits)",
            "The normal search is free. Turning this ON also asks an "
            "Anthropic (Claude) AI model to web-search for each lock, which "
            "can find listings the built-in sites don't cover.\n\n"
            "This uses YOUR Anthropic API key and costs a small amount of "
            "API credit each run (usually a few US cents). You'll need to "
            "enter an API key if you haven't already.\n\n"
            "Enable AI web search?")
        if not ok:
            self.ai_var.set(False)
            return
        self._prompt_api_key(force=False)

    def _search_lock_from_locks_tab(self, name):
        # jump to the Search tab, load the lock, and kick off a live search
        self._show_tab("Search")
        self.lock_var.set(name)
        self._on_lock_selected()
        self._start_search()

    # ---- batch "hunt my whole wishlist" (free search only) ----------------
    def _toggle_wishlist_hunt(self):
        """The button does double duty: start a hunt, or cancel one in
        progress."""
        if getattr(self, "_wishlist_hunt_running", False):
            self._cancel_wishlist_hunt()
        else:
            self._start_wishlist_hunt()

    def _cancel_wishlist_hunt(self):
        # signal the worker to stop after the current lock; reflect it in the UI
        self._wishlist_hunt_cancel = True
        btn = getattr(self, "hunt_wishlist_btn", None)
        if btn is not None:
            btn.config(text="Cancelling…")
            btn.set_enabled(False)
        self.status.set("Cancelling wishlist hunt after the current lock…")

    def _set_hunt_wishlist_enabled(self, on):
        btn = getattr(self, "hunt_wishlist_btn", None)
        if btn is not None:
            btn.set_enabled(bool(on))

    def _sync_hunt_wishlist_button(self):
        """Restore the button to its idle 'search' state, enabled only while
        the Show filter is 'My wishlist' and no hunt is running. While a hunt
        IS running the button stays a live 'Cancel hunt' control — leave it."""
        if getattr(self, "_wishlist_hunt_running", False):
            return
        btn = getattr(self, "hunt_wishlist_btn", None)
        if btn is not None:
            btn.config(text="Search for Wishlist Locks")
        on = (self.locks_show_var.get() == "My wishlist"
              and not getattr(self, "_wishlist_hunt_running", False))
        self._set_hunt_wishlist_enabled(on)

    def _wishlist_locks(self, belt):
        """Names of all wishlisted locks, optionally limited to one belt."""
        q = ("SELECT DISTINCT l.name FROM locks l"
             " JOIN my_collection m ON m.lock_id = l.id"
             " WHERE m.status='wishlist'")
        params = []
        if belt and belt != "All belts":
            q += " AND l.belt = ?"; params.append(belt)
        q += " ORDER BY l.name COLLATE NOCASE"
        conn = db()
        rows = conn.execute(q, params).fetchall()
        conn.close()
        return [r[0] for r in rows if r[0]]

    def _start_wishlist_hunt(self):
        if self.locks_show_var.get() != "My wishlist":
            return
        if getattr(self, "_wishlist_hunt_running", False):
            return
        belt = self.locks_belt_var.get()
        names = self._wishlist_locks(belt)
        if not names:
            messagebox.showinfo(
                "Lock Hunter",
                "No wishlist locks to search"
                + ("" if belt in ("", "All belts") else f" for the {belt} belt")
                + ". Import your profile or pick a different belt.")
            return
        beltmsg = "" if belt in ("", "All belts") else f" ({belt} belt)"
        # Three-way prompt so the filter choice is explicit for every hunt
        # (independent of the Search-tab checkbox):
        #   Yes    -> lock results only (filter same-name non-locks)
        #   No     -> show everything (no filter)
        #   Cancel -> don't run
        choice = messagebox.askyesnocancel(
            "Hunt my wishlist",
            f"Search all {len(names)} wishlist lock(s){beltmsg} using the "
            f"FREE search (eBay + marketplaces, no API credits).\n\n"
            f"This scans many sites per lock, so it can take a while and may "
            f"return a lot of results. Results appear on the Search tab as "
            f"they come in.\n\n"
            f"Filter to LOCK results only?\n"
            f"   • Yes  = locks only (recommended — hides same-name non-locks)\n"
            f"   • No   = show everything\n"
            f"   • Cancel = don't run")
        if choice is None:
            return   # Cancel
        lock_only = bool(choice)   # Yes -> True, No -> False
        self._wishlist_hunt_running = True
        self._wishlist_hunt_cancel = False
        # turn the button into a clickable "Cancel hunt" control
        btn = getattr(self, "hunt_wishlist_btn", None)
        if btn is not None:
            btn.config(text="Cancel hunt")
            btn.set_enabled(True)
        # show results on the Search tab and clear the table once up front
        self._show_tab("Search")
        try:
            conn = db()
            conn.execute("DELETE FROM listings")
            conn.commit()
            conn.close()
        except sqlite3.Error:
            pass
        self._refresh_table()
        self._show_scan_overlay(True)
        belt_arg = "" if belt in ("", "All belts") else belt
        threading.Thread(target=self._wishlist_hunt_worker,
                         args=(names, belt_arg, lock_only), daemon=True).start()

    def _wishlist_hunt_worker(self, names, belt, lock_only=True):
        """Run the FREE search for each wishlist lock in turn, accumulating
        results into the listings table (no clear between locks). Free search =
        eBay scrape + all marketplace probes + Bazaar; never calls Claude.
        `lock_only` sets the lock-context filter for this whole batch,
        independent of the Search-tab checkbox; the checkbox state is restored
        when the hunt finishes."""
        global _LOCK_CONTEXT_FILTER
        prev_filter = _LOCK_CONTEXT_FILTER
        _LOCK_CONTEXT_FILTER = bool(lock_only)
        total = len(names)
        grand = 0
        try:
            for i, lock_name in enumerate(names, 1):
                if getattr(self, "_wishlist_hunt_cancel", False):
                    self.q.put(("status", log("Wishlist hunt cancelled.")))
                    break
                self.q.put(("status", log(
                    f"[Wishlist {i}/{total}] searching “{lock_name}”…")))
                if hasattr(self, "scan_overlay"):
                    self.q.put(("scanmsg",
                                f"Hunting wishlist… {i}/{total}\n{lock_name}"))
                results = []
                # Expand the catalog name into every query: as written, each
                # slash-separated lock, and a broadened variant of each
                # (pin-count / numeric-code stripped). See _search_terms.
                comps = _search_terms(lock_name)
                if len(comps) > 1:
                    self.q.put(("status", log(
                        f"  ({lock_name}) → {len(comps)} searches: "
                        f"{', '.join(comps)}")))
                cb = lambda s: self.q.put(("status", s))
                # eBay (API if configured + direct scrape) runs in the
                # BACKGROUND while the big marketplace sweep runs below —
                # same overlap as the single-search worker.
                phase_pool = concurrent.futures.ThreadPoolExecutor(
                    max_workers=4)
                try:
                    efuts = []
                    for comp in comps:
                        origin = _origin_for_lock(comp)
                        # eBay API first if configured (still free of Claude
                        # credits)
                        if ebay_api_enabled():
                            efuts.append(("eBay API", phase_pool.submit(
                                search_ebay_api, comp, cb,
                                deep=True, origin=origin)))
                        # eBay scrape
                        efuts.append(("eBay probe", phase_pool.submit(
                            search_ebay_direct, comp, cb,
                            deep=True, origin=origin)))
                    for comp in comps:
                        origin = _origin_for_lock(comp)
                        # all no-key marketplace probes
                        try:
                            results.extend(run_extra_marketplace_probes(
                                comp, cb, origin=origin, deep=True))
                        except Exception as ex:
                            self.q.put(("status", log(
                                f"  marketplace error: {ex}")))
                    # NOTE: Facebook Marketplace is deliberately NOT run in the
                    # wishlist hunt — it would mean ~10 city requests per lock
                    # across the whole wishlist, which Facebook would quickly
                    # rate-limit. Facebook is single-lock-search only.
                    for label, fut in efuts:
                        try:
                            results.extend(fut.result() or [])
                        except Exception as ex:
                            self.q.put(("status", log(
                                f"  {label} error: {ex}")))
                finally:
                    phase_pool.shutdown(wait=False)
                # Bazaar (matches the catalog entry; use the full name once)
                try:
                    bz = search_bazaar_direct(
                        lock_name, lambda s: self.q.put(("status", s)))
                    if bz:
                        results.extend(bz)
                except Exception:
                    pass
                # verify + store (accumulate; do NOT clear the table)
                try:
                    results = verify_listings(
                        results, lambda s: self.q.put(("status", s)))
                except Exception:
                    pass
                stored = 0
                conn = None
                try:
                    conn = db()
                    for it in results:
                        try:
                            if _insert_listing_row(conn, lock_name, it):
                                stored += 1
                        except sqlite3.Error:
                            pass
                    _prune_listing_history(conn)
                    conn.commit()
                except sqlite3.Error:
                    pass
                finally:
                    if conn is not None:
                        conn.close()
                grand += stored
                self.q.put(("status", log(
                    f"[Wishlist {i}/{total}] “{lock_name}” -> {stored} new "
                    f"(total {grand}).")))
                # refresh the visible table as results accumulate
                self.q.put(("refresh_table", None))
            self.q.put(("status", log(
                f"Wishlist hunt done: {grand} listing(s) across {total} "
                f"lock(s).")))
        except Exception as ex:
            self.q.put(("status", log(f"Wishlist hunt FAILED: {ex}")))
        finally:
            _LOCK_CONTEXT_FILTER = prev_filter   # restore checkbox state
            self._wishlist_hunt_running = False
            self.q.put(("wishlist_hunt_done", None))

    # ---- lock names / autocomplete
    def _all_lock_names(self, needle="", belt=None):
        # belt=None -> use the Search tab's own belt filter (default). Pass an
        # explicit belt to filter by that instead.
        if belt is None:
            bf = getattr(self, "belt_filter_var", None)
            belt = bf.get() if bf else "All belts"
        belt_val = belt
        base = "SELECT DISTINCT name FROM locks WHERE 1=1"
        params = []
        if belt_val and belt_val != "All belts":
            base += " AND belt = ?"; params.append(belt_val)
        base += " ORDER BY name LIMIT 2400"
        conn = db()
        rows = conn.execute(base, params).fetchall()
        conn.close()
        names = [r[0] for r in rows]
        if needle:
            # accent-blind: typing 'semag' finds 'Sémag'; prefix hits rank first
            nf = _fold(needle)
            names = [n for n in names if nf in _fold(n)]
            starts = [n for n in names if _fold(n).startswith(nf)]
            rest = [n for n in names if not _fold(n).startswith(nf)]
            names = (starts + rest)[:1200]
        return names

    # The ten belt ranks shown on lpubelts.com, in order. These are ALWAYS
    # offered in the dropdowns (not just belts that happen to be present).
    BELTS = ["White", "Yellow", "Orange", "Green", "Blue", "Purple",
             "Brown", "Red", "Black", "Unranked"]

    def _belt_order(self):
        return list(self.BELTS)

    def _clear_combo_highlight(self, evt):
        # readonly ttk comboboxes keep a blue selection highlight after a pick;
        # move focus off the widget so it doesn't stay highlighted.
        try:
            evt.widget.selection_clear()
            self.focus_set()
        except Exception:
            pass

    def _on_belt_filter_changed(self):
        # Refresh the pick list for the new belt, and if a lock is already
        # selected, re-pull its preview so the picture/page switch to the
        # selected belt's version.
        self._reload_lock_names()
        if self.lock_var.get().strip():
            self._on_lock_selected()

    def _reload_lock_names(self):
        names = self._all_lock_names()
        self.lock_box["values"] = names
        # Default the Lock box to BLANK. Only clear it if the current text is
        # no longer a valid option (e.g. after switching belts); never auto-
        # pick the first lock.
        cur = self.lock_var.get().strip()
        if cur and cur not in names:
            self.lock_var.set("")
            self._on_lock_selected()
        # belt dropdown ALWAYS offers the full canonical belt list
        self.belt_box["values"] = ["All belts"] + self._belt_order()
        conn = db()
        own = conn.execute("SELECT COUNT(*) FROM my_collection WHERE status='own'").fetchone()[0]
        wish = conn.execute("SELECT COUNT(*) FROM my_collection WHERE status='wishlist'").fetchone()[0]
        conn.close()
        # top-right: "Created by: Ferf" with Owned / Wishlist counts beneath it
        if own or wish:
            self.count_var.set(f"Created by: Ferf\nOwned {own}  ·  Wishlist {wish}")
        else:
            self.count_var.set("Created by: Ferf")






    def _commit_lock_selection(self, evt=None):
        # Accept whatever is in the box (including an auto-filled completion),
        # drop any highlighted tail selection, and refresh the preview.
        try:
            self.lock_box.selection_clear()
            self.lock_box.icursor("end")
        except Exception:
            pass
        self._on_lock_selected()
        return "break"  # don't let Return also trigger other default handling

    def _filter_locks(self, evt=None):
        keysym = getattr(evt, "keysym", "") if evt is not None else ""
        # Let navigation / selection keys through untouched
        if keysym in ("Up", "Down", "Return", "Escape", "Left", "Right",
                      "Tab", "Home", "End"):
            return
        typed = self.lock_var.get()
        matches = self._all_lock_names(typed.strip())
        self.lock_box["values"] = matches
        # Inline autofill: when adding characters (not deleting), complete the
        # box to the first match and select the added tail so the next
        # keystroke overwrites it — e.g. type "Mast" -> "Master Lock 2" with
        # "er Lock 2" highlighted.
        if keysym in ("BackSpace", "Delete") or not typed:
            return
        typed_f = _fold(typed)
        first = next((m for m in matches if _fold(m).startswith(typed_f)), None)
        if first and _fold(first) != typed_f:
            try:
                self.lock_var.set(first)
                self.lock_box.icursor(len(typed))         # cursor after typed part
                self.lock_box.selection_range(len(typed), "end")  # highlight the rest
            except Exception:
                pass

    # ---- thumbnail
    def _lock_row(self, name):
        # Some locks share a name across belts (e.g. ASSA 700 exists as Blue,
        # Red, and Black). If a specific belt is selected in the dropdown,
        # prefer that belt's row so we show the correct picture and LPU page.
        belt = getattr(self, "belt_filter_var", None)
        belt_val = belt.get() if belt else "All belts"
        conn = db()
        row = None
        if belt_val and belt_val != "All belts":
            row = conn.execute(
                "SELECT image_url, page_url, belt,"
                " (SELECT status FROM my_collection m WHERE m.lock_id = locks.id)"
                " FROM locks WHERE name=? AND belt=? LIMIT 1",
                (name, belt_val)).fetchone()
        if row is None:
            row = conn.execute(
                "SELECT image_url, page_url, belt,"
                " (SELECT status FROM my_collection m WHERE m.lock_id = locks.id)"
                " FROM locks WHERE name=? LIMIT 1",
                (name,)).fetchone()
        conn.close()
        return row

    def _on_lock_selected(self):
        name = self.lock_var.get().strip()
        row = self._lock_row(name)
        self.name_var.set(name if name else "")
        self.belt_var.set("")
        self.lpu_btn.set_enabled(False)
        self.bazaar_btn.set_enabled(False)
        self._page_url = None
        self._thumb_ref = None
        self.thumb_label.config(image="", text="loading…")
        if not row:
            self.name_var.set("")
            self.thumb_label.config(image="", text="select a lock")
            return
        img_url, page_url, belt, mine = row
        parts = []
        if belt:
            parts.append(f"{belt} belt")
        if mine == "own":
            parts.append("\u2714 Owned")
        elif mine == "wishlist":
            parts.append("\u2605 Wishlist")
        self.belt_var.set("   ·   ".join(parts))
        if page_url:
            self._page_url = page_url
            self.lpu_btn.set_enabled(True)
        # thumbnail: use cached URL, else fetch it from the lock's LPU page
        if img_url and HAVE_PIL:
            threading.Thread(target=self._fetch_thumb, args=(img_url, name), daemon=True).start()
        elif page_url and HAVE_PIL:
            threading.Thread(target=self._lazy_image, args=(page_url, name), daemon=True).start()
        else:
            self.thumb_label.config(image="", text="no image")
        # auto-check the Lock Bazaar for this lock, enable button on a hit
        if self.bazaar_var.get():
            belt = self.belt_filter_var.get()
            belt = "" if belt == "All belts" else belt
            threading.Thread(target=self._bazaar_check, args=(name, belt),
                             daemon=True).start()

    def _lazy_image(self, page_url, name):
        img = fetch_lock_image(page_url, lambda s: self.q.put(("status", s)))
        if img:
            self._fetch_thumb(img, name)
        else:
            self.q.put(("thumb", (name, None)))

    def _bazaar_check(self, name, belt=""):
        # Result states the UI uses:
        #   True  = confirmed entries in the bazaar for this lock -> button on
        #   False = feed WAS read and had zero entries            -> button off
        #   None  = feed couldn't be read                         -> button on
        # We grey the button only when we positively confirmed there are no
        # entries; if we can't read the feed we still enable it (the search
        # deep-link works, and a real lock like "Abloy Classic" must not be
        # greyed just because the feed URL couldn't be reached).
        try:
            hits = search_bazaar_direct(name, lambda s: None, belt=belt)
        except Exception:
            hits = None
        if hits is None:
            state = None            # feed unreachable
        else:
            state = bool(hits)      # verified match / no match
        self.q.put(("bazaar_state", (name, state)))

    def _open_bazaar_for_lock(self):
        name = self.lock_var.get().strip()
        # deep-link into the bazaar's search for this lock, e.g.
        #   https://lpulocks.com/#/lockbazaar?search=assa+700
        url = BAZAAR_URL + "?search=" + urllib.parse.quote_plus(name)
        # if a specific belt is selected, pass it to the bazaar's belt filter
        belt = getattr(self, "belt_filter_var", None)
        belt_val = belt.get() if belt else "All belts"
        if belt_val and belt_val != "All belts":
            url += "&belt=" + urllib.parse.quote_plus(belt_val)
        webbrowser.open(url)

    def _fetch_thumb(self, url, name):
        try:
            fn = os.path.join(IMG_DIR, re.sub(r"[^A-Za-z0-9.]", "_", url)[-80:])
            if not os.path.exists(fn):
                r = requests.get(url, timeout=30, headers={"User-Agent": "LockHunter/1.0"})
                r.raise_for_status()
                with open(fn, "wb") as f:
                    f.write(r.content)
            img = Image.open(fn)
            # flatten transparency onto the panel color so PNGs don't show black
            if img.mode in ("RGBA", "LA", "P"):
                img = img.convert("RGBA")
                bg = Image.new("RGBA", img.size, (15, 19, 23, 255))  # C_FIELD
                img = Image.alpha_composite(bg, img).convert("RGB")
            else:
                img = img.convert("RGB")
            img.thumbnail((196, 196), Image.LANCZOS)
            self.q.put(("thumb", (name, img)))
        except Exception as ex:
            self.q.put(("thumb", (name, None)))
            self.q.put(("status", f"Thumbnail failed: {ex}"))

    def _open_lpu(self):
        if getattr(self, "_page_url", None):
            webbrowser.open(self._page_url)

    # ---- profile import
    def _start_profile_import(self):
        url = self.profile_var.get().strip()
        if not url:
            # No saved profile yet — ask for it (same prompt as startup).
            # The prompt itself starts the import when a link is accepted, so
            # return here either way instead of falling through and kicking
            # off a SECOND import of the same profile.
            self._prompt_profile(force=True)
            return
        if "profile/" not in url:
            messagebox.showwarning(
                "Lock Hunter",
                "That doesn't look like an LPU profile link.\n\n"
                "It should look like:\n"
                "   https://lpubelts.com/#/profile/<your-id>?collection=Any")
            return
        # normalize to the ?collection=Any form (the app reads Own + Wishlist
        # regardless, but this keeps the saved link consistent)
        url = re.sub(r"[?&]collection=[^&]*", "", url)
        url = url + ("&collection=Any" if "?" in url else "?collection=Any")
        self.profile_var.set(url)
        self.cfg["profile_url"] = url
        save_cfg(self.cfg)
        self.profile_btn.config(state="disabled")
        self.profile_btn.config(text="Working…")
        self.status.set("Reading your LPU profile… this can take a few seconds.")
        threading.Thread(target=self._profile_worker, args=(url,), daemon=True).start()

    def _profile_worker(self, url):
        try:
            own, wish, uid = import_lpu_profile(url, lambda s: self.q.put(("status", s)))
            save_profile(uid, own, wish)
            self.cfg["profile_url"] = url
            self.cfg["profile_uid"] = uid
            self.cfg["profile_synced"] = datetime.datetime.now().isoformat(timespec="seconds")
            save_cfg(self.cfg)
            self.q.put(("status", log(
                f"✔ Profile loaded: {len(own)} owned, {len(wish)} wishlist.")))
            self.q.put(("popup_info", (
                "Profile loaded",
                f"Your LPU profile loaded successfully.\n\n"
                f"Owned locks: {len(own)}\nWishlist locks: {len(wish)}\n\n"
                f"Browse them on the Locks tab.")))
        except Exception as ex:
            self.q.put(("status", log(f"Profile import failed: {ex}")))
            self.q.put(("popup_error", (
                "Profile not loaded",
                f"Your LPU profile could not be loaded — {ex}")))
        finally:
            self.q.put(("profile_reset", None))
            self.q.put(("profile_done", None))

    def _auto_profile_refresh(self):
        """Startup chore: silently re-import the user's LPU profile so newly
        wishlisted/owned locks are known. Runs only when a profile link was
        saved by a previous session; never pops dialogs."""
        url = (self.cfg.get("profile_url") or "").strip()
        if not url:
            return
        threading.Thread(target=self._auto_profile_worker, args=(url,),
                         daemon=True).start()

    def _auto_profile_worker(self, url):
        try:
            own, wish, uid = import_lpu_profile(
                url, lambda s: self.q.put(("status", s)))
            save_profile(uid, own, wish)
            self.q.put(("status", log(
                f"Profile refreshed: {len(own)} owned, "
                f"{len(wish)} wishlist.")))
            self.q.put(("profile_done", None))
        except Exception as ex:
            self.q.put(("status", log(f"Profile auto-refresh skipped: {ex}")))

    def _start_sync(self):
        if str(self.sync_btn["state"]) == "disabled":
            return
        self.sync_btn.config(state="disabled")
        self.sync_btn.config(text="Working…")
        self.status.set("Downloading the LPU lock catalog… this can take 10–20 seconds.")
        threading.Thread(target=self._sync_worker, daemon=True).start()

    def _sync_worker(self):
        try:
            n, url = sync_lpu_catalog(lambda s: self.q.put(("status", s)))
            self.cfg["lpu_synced"] = datetime.datetime.now().isoformat(timespec="seconds")
            save_cfg(self.cfg)
            self.q.put(("status", log(
                f"✔ LPU catalog updated: {n} locks loaded. "
                f"Browse them in the Locks tab.")))
        except Exception as ex:
            self.q.put(("status", log(f"LPU sync failed: {ex}")))
        finally:
            self.q.put(("sync_reset", None))
            self.q.put(("sync_done", None))

    # ---- search
    def _check_for_updates(self, auto=False):
        if not auto:
            self.status.set("Checking for updates…")
        threading.Thread(target=self._update_worker, args=(auto,),
                         daemon=True).start()

    def _open_help_email(self):
        """Generate a diagnostic report, COPY the complete report text to the
        clipboard, and open a pre-addressed Gmail draft; the user describes
        the problem and just presses Ctrl+V to paste the full report (a URL
        can't safely carry the whole log, and the app can't type into the
        browser, so the clipboard bridges the gap)."""
        try:
            subject, body, full_path = build_help_report()
        except Exception as ex:
            messagebox.showerror("Help", f"Couldn't build the report:\n{ex}")
            return
        # put the COMPLETE report on the clipboard so the user only pastes
        clip_ok = False
        try:
            with open(full_path, "r", encoding="utf-8",
                      errors="replace") as f:
                full_text = f.read()
            self.clipboard_clear()
            self.clipboard_append(full_text)
            self.update()   # hand the clipboard to Windows right away
            clip_ok = True
        except Exception:
            clip_ok = False
        if clip_ok:
            body = body.replace(
                "__REPORT_BANNER__",
                "\n"
                "============================================================\n"
                "  >>>  THE FULL REPORT IS ALREADY COPIED TO YOUR CLIPBOARD "
                " <<<\n"
                "  >>>  CLICK BELOW THIS LINE AND PRESS  Ctrl+V  TO PASTE IT "
                " <<<\n"
                "============================================================")
        else:
            body = body.replace(
                "__REPORT_BANNER__",
                "\n"
                "============================================================\n"
                "  The full diagnostics are below; the complete report is\n"
                "  also saved on this PC (see the path at the bottom).\n"
                "============================================================")
        # Open a pre-filled GitHub issue (the open-source way to report a
        # problem). The full diagnostics are also saved to disk and copied to
        # the clipboard as a fallback.
        link = (GITHUB_ISSUES_URL + "/new?title=" +
                urllib.parse.quote(subject) + "&body=" +
                urllib.parse.quote(body))
        opened = False
        try:
            opened = webbrowser.open(link)
        except Exception:
            opened = False
        if clip_ok:
            self.status.set(
                "Problem report ready \u2014 the full report is on your "
                "clipboard; paste it into the GitHub issue with Ctrl+V.")
        else:
            self.status.set(f"Problem report ready \u2014 full report saved to "
                            f"{full_path}")
        if not opened:
            extra = ("\n\nThe full report is already COPIED — just press "
                     "Ctrl+V in the email to paste it."
                     if clip_ok else
                     f"\n\nPlease attach this file:\n{full_path}")
            messagebox.showinfo(
                "Help",
                "Couldn't open the issue page automatically.\n\n"
                "Please open an issue at\n" + GITHUB_ISSUES_URL +
                "\nand describe the problem." + extra)

    def _auto_update_check(self):
        """Startup check, at most once per day: quietly looks for a newer
        release and only speaks up if one exists (offline/no-news are
        silent)."""
        today = datetime.date.today().isoformat()
        if self.cfg.get("last_update_check") == today:
            return
        self.cfg["last_update_check"] = today
        save_cfg(self.cfg)
        self._check_for_updates(auto=True)

    def _update_worker(self, auto=False):
        latest, lt, status = find_latest_version(
            lambda s: (None if auto else self.q.put(("status", s))))
        if status == "error":
            if auto:
                return   # silent when the startup check can't reach the server
            # Couldn't reach the server at all — don't claim they're current.
            self.q.put(("popup_info", (
                "Check for updates",
                "Couldn't reach the update server. Please check your internet "
                "connection and try again.")))
            return
        # status is "ok" (found a version) or "none" (reached, nothing found).
        # A newer version -> offer it; otherwise the user has the latest.
        if latest and lt > _version_tuple(VERSION):
            self.q.put(("update_available", latest))
        elif not auto:
            self.q.put(("popup_info", (
                "Up to date",
                f"You have the latest version (v{VERSION}).")))

    def _show_api_key_help(self):
        win = tk.Toplevel(self)
        win.title("Getting a Claude API key")
        win.configure(bg=self.C_PANEL)
        win.geometry("560x420")
        win.transient(self); win.grab_set()
        tk.Label(win, text="How to get your Claude API key",
                 bg=self.C_PANEL, fg=self.C_ACCENT,
                 font=("Segoe UI Semibold", 13)).pack(anchor="w", padx=20, pady=(18, 4))
        steps = (
            "Lock Hunter uses Anthropic's Claude API to run the live web "
            "searches. You need your own API key. Here's how:\n\n"
            "1.  Go to the Claude Developer Platform:\n"
            "        console.anthropic.com\n\n"
            "2.  Sign up or log in (this is a developer account, separate "
            "from a Claude.ai subscription).\n\n"
            "3.  Add a payment method under Billing and buy a little credit. "
            "The API is pay-as-you-go — each lock search costs roughly a few "
            "US cents, so a few dollars lasts a long time.\n\n"
            "4.  Open the API keys page:\n"
            "        console.anthropic.com/settings/keys\n\n"
            "5.  Click 'Create Key', give it a name (e.g. 'Lock Hunter'), "
            "and copy the key. It starts with 'sk-ant-'.\n\n"
            "6.  Paste it into the Claude API key box here and click "
            "'Save key'. It's stored locally on your computer only.\n\n"
            "Keep your key private — anyone with it can spend your credit."
        )
        msg = tk.Label(win, text=steps, bg=self.C_PANEL, fg=self.C_TEXT,
                       font=("Segoe UI", 10), justify="left", wraplength=520,
                       anchor="w")
        msg.pack(anchor="w", padx=20)
        btns = tk.Frame(win, bg=self.C_PANEL)
        btns.pack(anchor="w", padx=20, pady=16)
        FlatButton(btns, "Open the API keys page", kind="primary",
                   command=lambda: webbrowser.open(
                       "https://console.anthropic.com/settings/keys")).pack(side="left")
        FlatButton(btns, "Close", kind="subtle",
                   command=win.destroy).pack(side="left", padx=(8, 0))

    def _clear_all(self):
        # First confirmation
        if not messagebox.askyesno(
                "Clear all?",
                "This will completely reset Lock Hunter:\n\n"
                "•  Delete the local lock database (all locks + search results)\n"
                "•  Delete the activity log\n"
                "•  Remove your saved LPU profile link and imported collection\n"
                "•  Remove your saved Claude API key\n\n"
                "Do you want to continue?"):
            return
        # Second confirmation
        if not messagebox.askyesno(
                "Are you absolutely sure?",
                "This cannot be undone. Everything above will be permanently "
                "erased and you'll need to re-enter your API key and profile.\n\n"
                "Click Yes to clear everything, or No to cancel."):
            return
        errors = []
        # 1) empty the database tables
        try:
            conn = db()
            for tbl in ("locks", "my_collection", "listings", "searches"):
                conn.execute(f"DELETE FROM {tbl}")
            conn.commit()
            conn.close()
        except Exception as ex:
            errors.append(f"database: {ex}")
        # 2) delete the log file
        try:
            if os.path.exists(LOG_PATH):
                os.remove(LOG_PATH)
        except Exception as ex:
            errors.append(f"log: {ex}")
        # 3) clear the saved profile + API key from config
        try:
            for k in ("api_key", "profile_url", "profile_uid", "profile_synced",
                      "lpu_synced"):
                self.cfg.pop(k, None)
            save_cfg(self.cfg)
        except Exception as ex:
            errors.append(f"settings: {ex}")
        # 4) reset in-memory state + UI
        try:
            self.key_var.set("")
            self.profile_var.set("")
            self.lock_var.set("")
            self._reload_lock_names()
            self._refresh_table()
            if hasattr(self, "locks_tree"):
                self._refresh_locks_table()
        except Exception as ex:
            errors.append(f"ui: {ex}")

        if errors:
            messagebox.showerror(
                "Cleared with issues",
                "Some items couldn't be cleared:\n\n" + "\n".join(errors))
        else:
            self.status.set("Cleared everything — database, log, profile, and API key.")
            messagebox.showinfo(
                "Cleared",
                "Everything has been reset. Enter your API key and profile "
                "again, then click 'Update LPU catalog' to reload locks.")

    def _show_scan_overlay(self, on):
        if not hasattr(self, "scan_overlay"):
            return
        try:
            if on:
                self.scan_overlay.place(relx=0.5, rely=0.5, anchor="center")
                self.scan_overlay.lift()
            else:
                self.scan_overlay.place_forget()
        except Exception:
            pass

    def _start_search(self):
        lock_name = self.lock_var.get().strip()
        if not lock_name:
            messagebox.showwarning("Lock Hunter", "Enter or pick a lock name.")
            return
        if self.ai_var.get() and not self.key_var.get().strip():
            messagebox.showwarning(
                "Lock Hunter",
                "Enter your Claude API key first — or untick 'Use AI web "
                "search' for a free eBay + Bazaar search.")
            return
        self.search_btn.config(state="disabled")
        # clear the old results from view and show the scanning message
        self._refresh_table()
        if hasattr(self, "scan_overlay"):
            self.scan_overlay.config(
                text="Extended scan in progress… (this can take a minute)")
        self._show_scan_overlay(True)
        belt = self.belt_filter_var.get()
        belt = "" if belt == "All belts" else belt
        sales_map = {"Auction / secondhand": "Auction",
                     "Website sales": "Website", "Both": "Both"}
        # Condition is always "Both" (the New/Used selector was removed in
        # 4.5.4) — the plumbing still carries it so the AI prompt and the
        # searches log keep their shape.
        # Facebook is opt-in AND rate-limited (see _fb_rate_check). If the
        # user is over the hourly limit, run the search WITHOUT Facebook this
        # time rather than blocking the whole search, and say so.
        use_fb = bool(self.fb_var.get())
        if use_fb:
            allowed, wait, used = self._fb_rate_check()
            if not allowed:
                messagebox.showinfo(
                    "Facebook search limit reached",
                    f"To stay within Facebook's limits, Lock Hunter runs at "
                    f"most {_FB_MAX_PER_HOUR} Facebook Marketplace searches "
                    f"per hour.\n\nYou've used {used}. The next one is "
                    f"available in about {wait} minute(s).\n\nThis search "
                    f"will run WITHOUT Facebook Marketplace — everything else "
                    f"(eBay, marketplaces, Lock Bazaar) runs as normal.")
                use_fb = False
            else:
                self._fb_rate_record()
        args = (self.key_var.get().strip(), lock_name,
                "Both", self.pickup_var.get(), self.bazaar_var.get(),
                self.deep_var.get(), belt,
                sales_map.get(self.sales_var.get(), "Both"),
                self.ai_var.get(), use_fb)
        threading.Thread(target=self._worker, args=args, daemon=True).start()

    def _worker(self, key, lock_name, cond, excl, bazaar, deep=False, belt="",
                sales="Both", use_ai=True, use_fb=False):
        started = datetime.datetime.now().isoformat(timespec="seconds")
        conn = None
        sid = None
        try:
            conn = db()
            # A new search replaces ALL previous results, so clear the whole
            # listings table (not just this lock's rows).
            conn.execute("DELETE FROM listings")
            conn.commit()
            cur = conn.execute(
                "INSERT INTO searches(lock_name,condition,exclude_pickup,results,started_at,status)"
                " VALUES(?,?,?,?,?,?)",
                (lock_name, cond, int(excl), 0, started, "running"))
            sid = cur.lastrowid
            conn.commit()
            self.q.put(("status", log(
                f"Search started: {lock_name} [{cond}] sales={sales} "
                f"excl_pickup={excl} bazaar={bazaar} extended={deep} "
                f"ai={use_ai}")))
            results = []
            # Expand the catalog name into every query to run: the name as
            # written, each slash-separated lock, and a broadened variant of
            # each (pin-count descriptors and numeric codes stripped) — so
            # e.g. "Dejo 7 pin" also searches "Dejo". See _search_terms.
            components = _search_terms(lock_name)
            if len(components) > 1:
                self.q.put(("status", log(
                    f"Expanding “{lock_name}” into {len(components)} "
                    f"searches: {', '.join(components)}")))
            bazaar_direct = None
            if bazaar:
                bazaar_direct = search_bazaar_direct(
                    lock_name, lambda s: self.q.put(("status", s)), belt=belt)
                if bazaar_direct is not None:
                    results.extend(bazaar_direct)
                    self.q.put(("status", log(
                        f"Lock Bazaar direct feed: {len(bazaar_direct)} match(es)")))
                else:
                    self.q.put(("status", log(
                        "Lock Bazaar feed unreachable; Claude will search it instead")))
            # Phase OVERLAP: the eBay lookups (API + direct page probe) and the
            # optional AI search are network work independent of the big
            # marketplace sweep, so they're launched in the background FIRST
            # and collected after the sweep — wall time becomes
            # max(sweep, eBay, AI) instead of their sum. Results are assembled
            # in the exact order of the old sequential flow (and the DB
            # de-dupes by url), so the OUTPUT is unchanged — only sooner.
            cb = lambda s: self.q.put(("status", s))
            comp_meta = [(comp, _origin_for_lock(comp),
                          "" if len(components) == 1 else f"[{comp}] ")
                         for comp in components]
            phase_pool = concurrent.futures.ThreadPoolExecutor(max_workers=8)
            try:
                ebay_futs = []      # (log label, tag, future) in submit order
                for comp, comp_origin, tag in comp_meta:
                    # eBay Browse API (official) first when in scope —
                    # reliable, all marketplaces via one keyset. Merges with
                    # the scrape via dedupe. Inert unless credentials are set.
                    if sales in ("Auction", "Both") and ebay_api_enabled():
                        ebay_futs.append(("eBay API", tag, phase_pool.submit(
                            search_ebay_api, comp, cb,
                            deep=deep, origin=comp_origin)))
                    # Direct eBay probe: eBay's own search pages (origin
                    # domain first) so live listings surface even when search
                    # engines haven't indexed them. Skipped for New-only.
                    if sales in ("Auction", "Both"):
                        ebay_futs.append(("Direct eBay probe", tag,
                                          phase_pool.submit(
                                              search_ebay_direct, comp, cb,
                                              deep=deep, origin=comp_origin)))
                ai_futs = []
                if use_ai:
                    ai_bazaar = bazaar and bazaar_direct is None
                    # give the AI the component names too so it searches each
                    for comp, _o, _t in comp_meta:
                        ai_futs.append(phase_pool.submit(
                            call_claude, key,
                            build_prompt(comp, cond, excl,
                                         include_bazaar=ai_bazaar,
                                         deep=deep, sales=sales),
                            cb, deep=deep))
                        ai_bazaar = False   # only ask about the Bazaar once
                else:
                    self.q.put(("status", log(
                        "AI search skipped — free mode (eBay probe + Bazaar "
                        "only, no API credits used).")))
                # The no-key marketplace sweep runs on THIS thread, fully
                # overlapping the background phases above. Like the eBay
                # probes it only applies to Auction/secondhand scope —
                # "Website sales"-only searches skip it (same as before).
                extra_by_comp = {}
                if sales in ("Auction", "Both"):
                    for comp, comp_origin, tag in comp_meta:
                        # Extra no-key probes: Catawiki, Leboncoin, Kleinanzeigen…
                        try:
                            extra = run_extra_marketplace_probes(
                                comp, cb, origin=comp_origin, deep=deep)
                            if extra:
                                extra_by_comp[comp] = extra
                                self.q.put(("status", log(
                                    f"{tag}Extra marketplace probes: "
                                    f"{len(extra)} candidate(s)")))
                        except Exception as ex:
                            self.q.put(("status", log(
                                f"marketplace probe error: {ex}")))
                # Collect eBay futures (old wording + old per-phase isolation).
                ebay_got = {}
                for label, tag, fut in ebay_futs:
                    try:
                        got = fut.result() or []
                        if got:
                            ebay_got.setdefault((tag, label), []).extend(got)
                            self.q.put(("status", log(
                                f"{tag}{label}: {len(got)} candidate(s)")))
                    except Exception as ex:
                        err = ("eBay API error" if label == "eBay API"
                               else "eBay probe error")
                        self.q.put(("status", log(f"{err}: {ex}")))
                # Assemble in the historical order: per component eBay API,
                # direct eBay, then that component's sweep results…
                for comp, comp_origin, tag in comp_meta:
                    results.extend(ebay_got.get((tag, "eBay API"), []))
                    results.extend(ebay_got.get((tag, "Direct eBay probe"), []))
                    results.extend(extra_by_comp.get(comp, []))
                # …then the AI results, in component order. An AI exception
                # propagates to the outer handler exactly like the old
                # sequential call did (the search is marked as errored).
                for fut in ai_futs:
                    results.extend(fut.result())
            finally:
                phase_pool.shutdown(wait=False)
            # Facebook Marketplace: opt-in, SINGLE-LOCK only. Runs ONCE on the
            # original lock name (never per split-variant), sequential and
            # throttled inside the probe. Its results are marked preverified,
            # so verify_listings won't re-fetch them into a login wall.
            if use_fb:
                try:
                    fb = search_facebook_marketplace(lock_name, cb)
                    if fb:
                        results.extend(fb)
                        self.q.put(("status", log(
                            f"Facebook Marketplace: {len(fb)} candidate(s)")))
                except Exception as ex:
                    self.q.put(("status", log(
                        f"Facebook Marketplace error: {ex}")))
            # Verify each listing URL is live (drops 404/403/ended-eBay etc).
            before = len(results)
            results = verify_listings(
                results, lambda s: self.q.put(("status", s)))
            if before != len(results):
                self.q.put(("status", log(
                    f"Verification removed {before - len(results)} dead/ended "
                    f"listing(s); {len(results)} live.")))
            kept = 0
            for it in results:
                ship = str(it.get("shipping", "unknown")).lower()
                if excl and ship == "no":
                    continue
                try:
                    _insert_listing_row(conn, lock_name, it)
                    kept += 1
                except sqlite3.Error:
                    pass
            _prune_listing_history(conn)
            conn.execute("UPDATE searches SET results=?, finished_at=?, status=? WHERE id=?",
                         (kept, datetime.datetime.now().isoformat(timespec="seconds"), "ok", sid))
            conn.commit()
            self.q.put(("status", log(f"Search done: {lock_name} -> {kept} listing(s) stored")))
        except Exception as ex:
            # Best-effort: mark the search row as errored (if we got that far).
            if conn is not None and sid is not None:
                try:
                    conn.execute("UPDATE searches SET status=? WHERE id=?",
                                 (f"error: {ex}", sid))
                    conn.commit()
                except sqlite3.Error:
                    pass
            self.q.put(("status", log(f"Search FAILED: {ex}")))
        finally:
            if conn is not None:
                try:
                    conn.close()
                except sqlite3.Error:
                    pass
            # ALWAYS re-enable the search button and clear the overlay, even if
            # the DB setup above threw — otherwise the UI would lock up.
            self.q.put(("done", None))

    # ---- event pump
    def _poll(self):
        try:
            while True:
                kind, payload = self.q.get_nowait()
                if kind == "status":
                    self.status.set(payload)
                elif kind == "done":
                    self.search_btn.config(state="normal")
                    self._show_scan_overlay(False)
                    self._refresh_table()
                elif kind == "scanmsg":
                    if hasattr(self, "scan_overlay"):
                        self.scan_overlay.config(text=payload)
                elif kind == "refresh_table":
                    self._refresh_table()
                elif kind == "wishlist_hunt_done":
                    self._show_scan_overlay(False)
                    self._refresh_table()
                    self._sync_hunt_wishlist_button()
                elif kind == "sync_reset":
                    self.sync_btn.config(text="Update LPU catalog")
                elif kind == "profile_reset":
                    self.profile_btn.config(text="Update profile")
                elif kind == "sync_done":
                    self.sync_btn.config(state="normal")
                    self._reload_lock_names()
                    if hasattr(self, "locks_tree"):
                        self._refresh_locks_table()
                elif kind == "profile_done":
                    self.profile_btn.config(state="normal")
                    self._reload_lock_names()
                    if hasattr(self, "locks_tree"):
                        self._refresh_locks_table()
                elif kind == "compare_results":
                    who, their_total, their_wish_total, rows = payload
                    self.compare_btn.config(state="normal")
                    self.compare_btn.config(text="Compare")
                    self._populate_compare(who, their_total,
                                           their_wish_total, rows)
                    self.status.set(
                        f"Compared {who}'s collection \u2014 {len(rows)} "
                        f"match(es) on your wishlist.")
                elif kind == "compare_error":
                    self.compare_btn.config(state="normal")
                    self.compare_btn.config(text="Compare")
                    title, body = payload
                    self.status.set(title)
                    messagebox.showerror(title, body)

                elif kind == "bazaar_news":
                    lines = payload
                    n = len(lines)
                    shown = lines[:8]
                    body = "\n".join("\u2022 " + b for b in shown)
                    if n > len(shown):
                        body += f"\n\u2026and {n - len(shown)} more."
                    self.status.set(
                        f"{n} wishlist lock listing(s) are NEW on the LPU "
                        f"Lock Bazaar.")
                    messagebox.showinfo(
                        "New on the LPU Lock Bazaar",
                        f"Since you last opened Lock Hunter, {n} new Bazaar "
                        f"listing(s) match your wishlist:\n\n{body}\n\n"
                        "Open the lock in Lock Hunter and use the "
                        "\u201cOpen Lock Bazaar\u201d button to see them.")
                elif kind == "popup_info":
                    title, body = payload
                    messagebox.showinfo(title, body)
                elif kind == "update_available":
                    latest = payload
                    self.status.set(f"Update available: v{latest}")
                    if messagebox.askyesno(
                            "Update available",
                            f"A newer version is available: v{latest}\n"
                            f"(You have v{VERSION}.)\n\n"
                            "Open the download page now?"):
                        webbrowser.open(GITHUB_RELEASES_PAGE)
                elif kind == "popup_error":
                    title, body = payload
                    messagebox.showerror(title, body)
                elif kind == "thumb":
                    tname, img = payload
                    # ignore images that arrived late for a lock we left
                    if tname != self.lock_var.get().strip():
                        continue
                    if img is None:
                        self._thumb_ref = None
                        self.thumb_label.config(image="", text="no image")
                    else:
                        self._thumb_ref = ImageTk.PhotoImage(img)
                        self.thumb_label.config(image=self._thumb_ref, text="")
                elif kind == "locks_thumb":
                    tname, img = payload
                    # only apply if it's still the selected lock on the Locks tab
                    if tname != getattr(self, "_locks_thumb_current", ""):
                        continue
                    if img is None:
                        self._locks_thumb_ref = None
                        self.locks_thumb_label.config(image="", text="no image")
                    else:
                        self._locks_thumb_ref = ImageTk.PhotoImage(img)
                        self.locks_thumb_label.config(image=self._locks_thumb_ref,
                                                      text="")
                elif kind == "usd_rates":
                    if payload:
                        self.status.set("Exchange rates loaded.")
                    else:
                        self.status.set("Couldn't fetch exchange rates — "
                                        "USD estimates unavailable.")
                    self._refresh_table()
                elif kind == "bazaar_state":
                    checked_name, state = payload
                    # only act if the user is still on the same lock
                    if checked_name == self.lock_var.get().strip():
                        if state is True:
                            # confirmed entries in the bazaar for this lock ->
                            # clickable, light-blue.
                            self.bazaar_btn.set_highlight(self.C_BAZAAR)
                        elif state is False:
                            # feed read, no entries for this lock -> grey out.
                            self.bazaar_btn.set_enabled(False)
                        else:
                            # state is None: couldn't read the feed at all.
                            # Enable anyway so a momentary network hiccup doesn't
                            # hide a lock that may well be on the bazaar (the
                            # deep-link search still works).
                            self.bazaar_btn.set_highlight(self.C_BAZAAR)
        except queue.Empty:
            pass
        self.after(150, self._poll)

    def _clear_results(self):
        # Wipe all stored search results (the listings table) and refresh.
        if not messagebox.askyesno(
                "Clear results?",
                "Remove all search results currently listed?\n\n"
                "(This clears the results table only — your lock catalog and "
                "owned/wishlist collection are not affected.)"):
            return
        try:
            conn = db()
            conn.execute("DELETE FROM listings")
            conn.commit()
            conn.close()
        except Exception as ex:
            messagebox.showerror("Lock Hunter", f"Could not clear results:\n{ex}")
            return
        self._refresh_table()
        self.status.set("Search results cleared.")

    def _usd_toggled(self):
        """USD checkbox: fetch rates in the background on first enable, then
        redraw. If rates can't be fetched, prices simply show without estimates."""
        if self.usd_var.get() and not _rates_cache["rates"]:
            self.status.set("Fetching exchange rates…")
            def _w():
                ok = bool(_usd_rates())
                self.q.put(("usd_rates", ok))
            threading.Thread(target=_w, daemon=True).start()
            return
        self._refresh_table()

    def _sort_by(self, col):
        """Cycle a column's sort: ascending → descending → default order."""
        if self._sort_col != col:
            self._sort_col, self._sort_rev = col, False
        elif not self._sort_rev:
            self._sort_rev = True
        else:
            self._sort_col, self._sort_rev = None, False
        for c, base in self._res_heads.items():
            mark = ""
            if c == self._sort_col:
                mark = "  ▼" if self._sort_rev else "  ▲"
            self.tree.heading(c, text=base + mark)
        self._refresh_table()

    def _export_search_xlsx(self):
        """Export the current Search results to an Excel file with columns:
        Lock name, Rarity (stars), Title, Price, Price (USD), Site, Seller
        (the LPU Bazaar username, where the source knows it), and a clickable
        hyperlink. Honors the same filter/condition/shipping toggles that
        shape the on-screen table."""
        # Pull the same rows the results table shows (same WHERE clauses).
        qy = ("SELECT lock_name,title,price,currency,site,url,seller"
              " FROM listings WHERE 1=1")
        params = []
        f = self.filter_var.get().strip()
        if f:
            qy += " AND lock_name LIKE ?"; params.append(f"%{f}%")
        if self.dbcond_var.get() != "All":
            qy += " AND condition=?"; params.append(self.dbcond_var.get())
        if self.dbship_var.get():
            qy += " AND shipping='yes'"
        qy += (" ORDER BY (CASE WHEN site='LPU Lock Bazaar' THEN 0 ELSE 1 END),"
               " found_at DESC LIMIT 5000")
        try:
            conn = db()
            rows = list(conn.execute(qy, params))
            # lock name -> rarity stars, same map the results table uses (the
            # rarest variant's count wins when one name exists at several
            # belts; non-catalog searches simply get no stars).
            rarity_by_name = {}
            for nm, cnt in conn.execute(
                    "SELECT name, MIN(owner_count) FROM locks"
                    " WHERE owner_count IS NOT NULL GROUP BY name"):
                if nm:
                    rarity_by_name[_fold(nm)] = cnt
            conn.close()
        except sqlite3.Error as ex:
            messagebox.showerror("Export", f"Couldn't read results: {ex}")
            return
        if not rows:
            messagebox.showinfo(
                "Export", "No search results to export yet. Run a search first.")
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except Exception:
            messagebox.showerror(
                "Export",
                "The Excel export needs the 'openpyxl' library, which isn't "
                "available in this build. Rebuild with the updated "
                "requirements, or let me know and I can switch export to CSV.")
            return

        default_name = "lock_search_" + datetime.datetime.now().strftime(
            "%Y%m%d_%H%M") + ".xlsx"
        path = filedialog.asksaveasfilename(
            title="Export Lock Search",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel workbook", "*.xlsx")])
        if not path:
            return   # user cancelled

        # Use current USD rates if the user has them loaded (or fetch-on-demand
        # isn't done here to keep export instant; USD is filled when available).
        rates = _rates_cache.get("rates")

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Lock Search"
            headers = ["Lock name", "Rarity", "Title", "Price", "Price (USD)",
                       "Site", "Seller", "Listing"]
            ws.append(headers)
            hdr_font = Font(bold=True, color="FFFFFF")
            hdr_fill = PatternFill("solid", fgColor="C58A2E")
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=c)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(vertical="center")
            link_font = Font(color="0563C1", underline="single")

            for rec in rows:
                lock_name, title, price, currency, site, url, seller = rec
                price_disp = " ".join(p for p in (str(price or "").strip(),
                                                  str(currency or "").strip())
                                      if p)
                usd_val = None
                if rates:
                    u = _usd_estimate(price, currency, rates)
                    _amt, cur = _parse_price(price, currency)
                    if u is not None and cur:
                        usd_val = round(u, 2)
                fk = _fold(lock_name or "")
                stars = (_rarity_stars(rarity_by_name[fk])
                         if fk in rarity_by_name else "")
                r = ws.max_row + 1
                ws.cell(row=r, column=1, value=lock_name or "")
                ws.cell(row=r, column=2, value=stars)
                ws.cell(row=r, column=3, value=title or "")
                ws.cell(row=r, column=4, value=price_disp)
                ws.cell(row=r, column=5,
                        value=(usd_val if usd_val is not None else ""))
                if usd_val is not None:
                    ws.cell(row=r, column=5).number_format = '$#,##0.00'
                ws.cell(row=r, column=6, value=site or "")
                ws.cell(row=r, column=7, value=str(seller or ""))
                # Clickable hyperlink: display the title (or "View listing"),
                # link to the URL — a real hyperlink, not raw text.
                link_cell = ws.cell(row=r, column=8)
                u = str(url or "").strip()
                if u.startswith("http"):
                    link_cell.value = "View listing"
                    link_cell.hyperlink = u
                    link_cell.font = link_font
                else:
                    link_cell.value = u

            # Reasonable column widths.
            widths = [26, 10, 46, 16, 13, 16, 18, 18]
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
            ws.freeze_panes = "A2"
            wb.save(path)
        except Exception as ex:
            messagebox.showerror("Export", f"Couldn't write the file: {ex}")
            return

        n = len(rows)
        note = "" if rates else ("\n\n(Tip: tick 'Show USD estimate' before "
                                 "exporting to fill the Price (USD) column.)")
        if messagebox.askyesno(
                "Export complete",
                f"Exported {n} listing(s) to:\n{path}\n\nOpen the file now?"
                + note):
            try:
                os.startfile(path)   # Windows: open in Excel
            except Exception:
                webbrowser.open("file://" + path.replace("\\", "/"))

    def _refresh_table(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        self._row_urls = {}
        self._row_images = {}
        qy = ("SELECT lock_name,title,price,currency,condition,site,"
              "location,shipping,found_at,url,image_url FROM listings WHERE 1=1")
        params = []
        f = self.filter_var.get().strip()
        if f:
            qy += " AND lock_name LIKE ?"; params.append(f"%{f}%")
        if self.dbcond_var.get() != "All":
            qy += " AND condition=?"; params.append(self.dbcond_var.get())
        if self.dbship_var.get():
            qy += " AND shipping='yes'"
        # Default order: LPU Lock Bazaar first, then newest first.
        qy += (" ORDER BY (CASE WHEN site='LPU Lock Bazaar' THEN 0 ELSE 1 END),"
               " found_at DESC LIMIT 500")
        conn = db()
        rows = list(conn.execute(qy, params))
        # Rarity stars per catalog lock NAME for the Rarity column. When the
        # same name exists at several belts, the rarest variant's count is
        # used. Custom (non-catalog) searches simply show no stars.
        rarity_by_name = {}
        for nm, cnt in conn.execute(
                "SELECT name, MIN(owner_count) FROM locks"
                " WHERE owner_count IS NOT NULL GROUP BY name"):
            if nm:
                rarity_by_name[_fold(nm)] = cnt
        # LPU catalog image per folded name — the fallback thumbnail when a
        # listing has no image of its own.
        self._lpu_img_by_name = {}
        for nm, img in conn.execute(
                "SELECT name, image_url FROM locks"
                " WHERE image_url IS NOT NULL AND image_url <> ''"):
            if nm:
                self._lpu_img_by_name[_fold(nm)] = img
        conn.close()

        usd_on = bool(getattr(self, "usd_var", None) and self.usd_var.get())
        rates = _rates_cache["rates"] if usd_on else None

        # Optional column sort (header click). Rows without a value for the
        # chosen column always sink to the bottom, whichever direction.
        # Rarity sorts by the underlying collector count, so ascending shows
        # the RAREST locks first (same convention as the Locks tab).
        if self._sort_col:
            def keyval(rec):
                if self._sort_col == "price":
                    u = _usd_estimate(rec[2], rec[3], rates or _rates_cache["rates"])
                    if u is not None:
                        return u
                    amt, _cur = _parse_price(rec[2], rec[3])
                    return amt
                if self._sort_col == "rarity":
                    return rarity_by_name.get(_fold(rec[0]))
                idx = {"lock": 0, "title": 1, "cond": 4, "site": 5,
                       "loc": 6}[self._sort_col]
                v = (rec[idx] or "").strip()
                return v.lower() if v else None
            valued = [r for r in rows if keyval(r) is not None]
            blank = [r for r in rows if keyval(r) is None]
            valued.sort(key=keyval, reverse=self._sort_rev)
            rows = valued + blank

        for i, rec in enumerate(rows):
            price_disp = " ".join(p for p in (str(rec[2] or "").strip(),
                                              str(rec[3] or "").strip()) if p)
            if usd_on and rates:
                u = _usd_estimate(rec[2], rec[3], rates)
                _amt, cur = _parse_price(rec[2], rec[3])
                if u is not None and cur and cur != "USD":
                    price_disp += f"  (~${u:,.0f})"
            fk = _fold(rec[0] or "")
            rar = _rarity_stars(rarity_by_name[fk]) if fk in rarity_by_name else ""
            vals = (rec[0], rec[1], price_disp, rec[4], rec[5], rec[6],
                    rar)
            stripe = "even" if i % 2 else "odd"
            iid = self.tree.insert("", "end", values=vals, tags=(stripe,))
            self._row_urls[iid] = rec[9]
            self._row_images[iid] = (rec[10] or "", rec[0] or "")

    def _thumb_url_for(self, iid):
        """Image URL for a result row: the listing's own photo if we captured
        one, else the LPU catalog image for that lock, else ''."""
        listing_img, name = getattr(self, "_row_images", {}).get(iid, ("", ""))
        if listing_img:
            return listing_img
        return getattr(self, "_lpu_img_by_name", {}).get(_fold(name), "")

    def _thumb_motion(self, evt):
        iid = self.tree.identify_row(evt.y)
        if iid == self._thumb_row:
            return
        self._thumb_row = iid
        if self._thumb_after:
            try:
                self.after_cancel(self._thumb_after)
            except Exception:
                pass
            self._thumb_after = None
        if not iid:
            self._thumb_hide()
            return
        xr, yr = evt.x_root + 24, evt.y_root + 16
        self._thumb_after = self.after(300, lambda: self._thumb_show(iid, xr, yr))

    def _thumb_show(self, iid, xr, yr):
        if not HAVE_PIL:
            self._thumb_hide()
            return
        if iid != self._thumb_row:
            return
        url = self._thumb_url_for(iid)
        if not url:
            self._thumb_hide()
            return
        self._thumb_ensure_win(xr, yr)
        cached = self._thumb_cache.get(url)
        if cached is not None:
            self._thumb_set(cached)
            return
        if self._thumb_lbl:
            self._thumb_lbl.config(text="loading\u2026", image="")
        threading.Thread(target=self._thumb_fetch, args=(url, iid),
                         daemon=True).start()

    def _thumb_fetch(self, url, iid):
        photo = None
        try:
            import io
            r = requests.get(url, timeout=10,
                             headers={"User-Agent": _BROWSER_UA})
            if r.status_code == 200:
                im = Image.open(io.BytesIO(r.content))
                im.thumbnail((240, 240))
                photo = ImageTk.PhotoImage(im)
        except Exception:
            photo = None

        def apply():
            if photo is not None:
                self._thumb_cache[url] = photo
            if (self._thumb_row == iid and self._thumb_win is not None
                    and self._thumb_win.winfo_exists()):
                if photo is not None:
                    self._thumb_set(photo)
                else:
                    self._thumb_hide()
        try:
            self.after(0, apply)
        except Exception:
            pass

    def _thumb_ensure_win(self, xr, yr):
        if self._thumb_win is None or not self._thumb_win.winfo_exists():
            self._thumb_win = tk.Toplevel(self)
            self._thumb_win.overrideredirect(True)
            try:
                self._thumb_win.attributes("-topmost", True)
            except Exception:
                pass
            self._thumb_lbl = tk.Label(
                self._thumb_win, bg=self.C_PANEL, fg=self.C_MUTE, bd=1,
                relief="solid", text="loading\u2026", padx=2, pady=2)
            self._thumb_lbl.pack()
        try:
            self._thumb_win.geometry(f"+{int(xr)}+{int(yr)}")
            self._thumb_win.deiconify()
        except Exception:
            pass

    def _thumb_set(self, photo):
        if (self._thumb_lbl and self._thumb_win is not None
                and self._thumb_win.winfo_exists()):
            self._thumb_img = photo
            self._thumb_lbl.config(image=photo, text="")

    def _thumb_hide(self):
        if self._thumb_after:
            try:
                self.after_cancel(self._thumb_after)
            except Exception:
                pass
            self._thumb_after = None
        if self._thumb_win is not None and self._thumb_win.winfo_exists():
            try:
                self._thumb_win.withdraw()
            except Exception:
                pass

    def _open_url(self, _evt):
        sel = self.tree.selection()
        if sel:
            url = getattr(self, "_row_urls", {}).get(sel[0])
            if url:
                webbrowser.open(url)

    def _results_context_menu(self, evt):
        """Right-click menu on the search results table: open, copy link, or
        remove the selected row(s) from the current results."""
        iid = self.tree.identify_row(evt.y)
        if not iid:
            return
        # If the clicked row isn't already part of a multi-selection, select
        # just it (file-manager behaviour); otherwise keep the whole
        # selection so a right-click can act on several rows at once.
        if iid not in self.tree.selection():
            self.tree.selection_set(iid)
        n = len(self.tree.selection())
        menu = tk.Menu(self, tearoff=0, bg=self.C_PANEL2, fg=self.C_TEXT,
                       activebackground=self.C_ACCENT, activeforeground="#20140a",
                       bd=0)
        menu.add_command(label="Open listing",
                         command=lambda: self._open_url(None))
        menu.add_command(label="Copy link",
                         command=self._copy_selected_result_link)
        menu.add_separator()
        menu.add_command(
            label=("Remove from results" if n <= 1
                   else f"Remove {n} from results"),
            command=self._remove_selected_results)
        try:
            menu.tk_popup(evt.x_root, evt.y_root)
        finally:
            menu.grab_release()

    def _copy_selected_result_link(self):
        sel = self.tree.selection()
        if not sel:
            return
        url = getattr(self, "_row_urls", {}).get(sel[0])
        if url:
            try:
                self.clipboard_clear()
                self.clipboard_append(url)
                self.status.set("Listing link copied to clipboard.")
            except Exception:
                pass

    def _remove_selected_results(self):
        """Remove the selected row(s) from the current search results. Deletes
        them from the `listings` table (the live results view) by URL, then
        drops the rows from the table. `listing_history` is left intact — this
        prunes what's shown now, not the app's record that the listing was
        ever seen, so a fresh search can surface it again if it's still live."""
        sel = list(self.tree.selection())
        if not sel:
            return
        urls = [self._row_urls.get(iid) for iid in sel]
        urls = [u for u in urls if u]
        if urls:
            try:
                conn = db()
                conn.executemany(
                    "DELETE FROM listings WHERE url=?", [(u,) for u in urls])
                conn.commit()
                conn.close()
            except sqlite3.Error as ex:
                messagebox.showerror(
                    "Remove", f"Couldn't remove the listing(s): {ex}")
                return
        for iid in sel:
            self._row_urls.pop(iid, None)
            try:
                self.tree.delete(iid)
            except Exception:
                pass
        remaining = len(self.tree.get_children())
        removed = len(sel)
        self.status.set(
            f"Removed {removed} result{'' if removed == 1 else 's'} "
            f"({remaining} remaining).")



# ---------------------------------------------------------------------------
# Embedded icon data (zlib-compressed, base64). Lets the window/header icon
# work even when NO external 'assets' folder is present, by writing these
# out to ~/.lockhunter/assets/ on first use. Primary source is still the
# real assets/ folder (nicer for the exe build); this is the safety net.
# ---------------------------------------------------------------------------
_ICO_B64 = (
    "eNqEumN0JdHyPnxi27Zt28nEObE5mdi2bdu2J3Ymnkwysc0TG/+5974f37V++0N319P1"
    "1F67uqtXVe0GAMAA0ABUVMC/QQpAhgAALP9zRfo/2R4WAHD9h7Gy/k9mRwYA9hAAAFHR"
    "/8kWmAAAJj4AEBj4P9maEQDwYgb8f4MUIK/3T18XAIhSUZJBgseH/wciyclKqv07/2dG"
    "VFjof0eSuy8TAAB8SU5STMMz1y5ZxDbVjTCiAAR62tr06d0CRs8EcaMnisVaV42xGmvT"
    "syV1smamAnEyxuYhcYxPDyy1tX+Kme6JciJTL4TWI3OyGZvTF0mEEOt/9XR1P732gZ72"
    "ASXKt7Vz97gDVQnMW7BkzPvCSCs2Xyhg5CMBGzVLdRH92+3qCDHLZmpGD9WSsMOwsFm7"
    "DQJgJujxZJKNAOzKyTMbAS93No4SKrlBxk7jQLZ6eeYI8EAfwxI0DN1I451kPVixKxaM"
    "gxTmSGOzqO9jGUU7IVdowiI32NC5Pu3D+1vM/d7iu8VgRdX0iPHQvyycnT34I3OYVBuo"
    "XBsMzfweOZ+slP/0esd5M+1aKriidglQyzhDdU0byJqWEXWtT60IcsyxkrT0uEcctgkg"
    "3rHlWB39Q7WtZSh3mQ+ajZ3ZGzpT8fXKPFkZsxTPLkCXbGgUGGzN/R3jtmrHg43mNZgZ"
    "ahR38CkppFbaURho/oSxp4tqwj5faEbTQmY3j2iQhOefki0eAyzndCIoY4yKlIgMFRhx"
    "id0uZWmOZHG5DTA/opoce+Y0blNnH5nV17u/USm/PF8KP2Q79RF3JkrXt6I7e6CnrAgz"
    "nqYl5RWDezRvoi0B8t/bUAamNzBXmXcU3aOjDfFWhRnZoiiTWtiqnynjBFkYkDetU9Bg"
    "IqaZeAFO18ncbsi9g3foxHEK3NJQ4yVJnd3Ug8RHVAopXJrI0u/teNc9xhrQgOyP4Id+"
    "AvOwFs4/xmba7FGOMu4XmyhYgX4424rK8w0jKL1o1qGLhnW5CgdYvgJ/zPnxt58ietBi"
    "4RuvfrWwTF9UUpLENApuHsxQOoKMp33iIANp9pOn8qBZ5I81AySHHTk45CDO+ziesOT0"
    "fdzSeR2d6nXMhOturGvABl6Npwu8BklKgpVYWYcZ96rrCQamo7ocpjuDFpGMITIs0lya"
    "8sAdRaasx0UZMgkpax1u9b7sRQw/UHaa/Gvw7AxV/hMPclJKkvXixsH/P9Hwn3gj/W80"
    "uHh73wMAsFz/i4a6DJ9YIiDRuEi/i+WVDejxyt4jA7tZ7ztcuU1hIGZNsLJ4a3SkU8nY"
    "Dok4Qmp00yBMsizGNUpJFNp8Nvl302zdaEpUSC9MiAVUeN3CxBVAmKlCx6X709X+loCH"
    "5ezLy/i37m4/2oDrnG/sLsT940/vhjIAMgqht6GpfsqxjcoTmSB842fjCQRCVZOcnYDP"
    "sRjlvunrRotzTRpEFL3wcNdfFHPbVGP1aAM+T2C/8sjQUM1Ul8KfTmd8Pl8WQbjE7eCN"
    "B8FT/gSYCjylTgcAMC+jdJKQq4/TwkJiNJsQcDlCg7vHWbunX8kIwk9eTYNmW12Yg7N7"
    "bwF0M1QwIB+YwjMpxGr8qj8I4A3eLUDgogL0YpXcRO4QuBm8yMlE+ldx0T0s78BdABrb"
    "WR6HmLYTZQc0/FSeB7j2/YEeVCN41FAouGklBSNS84MZbo8lrBaUOFjFlzravkF1pT8b"
    "qiuxocuOMuTAgWYcf5nCzpTF1GjfzBDn0bAa2I7Yn4PewgFk9QMQ03hlBDPF+GP+NPUD"
    "seTRVNaUsIu8FERh1pyhOpomEKKcACm5RUQSZJGBIXyHCfhVwWJ2/4sGdntDsa0kxKdc"
    "/WqGRxNyQZ1CzJF+hQlvI4bR9qL9HjL9gSKFk/kT0sb+U/gyiIN3Py0ArPGrxqii6x2d"
    "+/1Wy/ViZVYZZ3q/unn/p9O4APLR4acTO3sl+xsQiDhZqhUO7tnDY+wBz/DItfp3dcpj"
    "2VDZ6Puq5/ql/tuFgv7btIJtz6X7q8TUGkGDmZ4Vyz+1lbyhbW+IcPDM+2iqfb8ewZJp"
    "P0iq20lwCJR7SNsf/JrLRE6L5ofRh4z4uMZTISgtFncfF7YC7mPxTDkgq3+Ay9EOKYMW"
    "I028bm29ADcXEco0irZx4SHH1rgIDSlptWT1xqadF9njhpoy30ygatlTq3+Pt3vvI3+Z"
    "cxYyhStukQZzy3nb8s3Z6SVovNI5O0jvjI6lVcjDQ6849ykK9qDR5nG4U6tPjR0RKtxt"
    "H45LtYceHveX+AwulGB2f2LCoQu8J/BHcS2BMi48klY64r5wCpvMkW2ECrtCQ43u54rs"
    "aoG3PADcac+6HWIx3boN6Y+ickkZp+6k658/Pwj+Ie0QyZwIfeiNIg/vJvyHr9UzoSQk"
    "UNkyQb8k7K1pEaHDYcZMq5NX8TrUsZlqHZe61k8FbrE9fF1IHtLumsbAd3yT6teEX8f9"
    "h7AkKAQX9crRyiCxq8EDrEv1N1VNrWl7kmgtsn6hjA/NJ/JNSoUlhsy20Nxo3B4ILqn8"
    "NTI4MDvuU/B9VpwVaYH8ax4FTorl6LZHMed02gy+3V09khCwXd5BnRooFOU87yWsxtqz"
    "pnTCyl4yoKqyBm8HxIYZ2tLIus2fhy2yXFuKhvOf3DMQ9WEfgx4qSsIGfCBMdmqowl3/"
    "ie+rhkh5EStmPUzrrWqmRF2rbsX6lu8aB7SY9NL0CL8gqmTfQHNW/X37KJoGzCNTnzUD"
    "OslFy5H7LhaTwlG8ja+4WKUIeXDjrYpIi6H7Lg5RolnnOqeCqmcTO2Cq9TVaET3M2TXn"
    "9gg8gdHgczJ0JbMuFjKyUqEiPKbu4/nk8wBHuBE6ymD8kDj1LQ51AKMc/+9OpKgI7UEo"
    "T/jUX8/xXnM81VDZhl6VSdfSe3Ty4So7QeCNHPJePbun3ENlh4HdY+v3UOKj8Orw2eWo"
    "HWRoa3CMgpum0BSQ2p2NMX/N3QlHost32kTFjK0EDqPCnaWoDlY1JzcfIv1taA/8jTVa"
    "eFng7rV3t5v2c+cbeTHFJ/awRmngRuw4rSUWwjVjfgyq0OF+byP9dbzHLSzXLFghD8uI"
    "HOYiPny18tssTM2+t3pWMmgqaQWNCneSkUWX2I5yGgcMSzzOSVUSX6c2SBocJ0AQMmxM"
    "kgDoSJz7AIkfaMlnMUlXcAHLxgcC37HEGjBTLRSFjS2dNIsLJfhBRZwdf8d3WBcS3Kb4"
    "owgqxY2AaCk3Zi4Zy5DU8YKgLvEGiA6sb0jzrE1iTfZ+rZumm3OZbILgh/IX2jJ/M8YZ"
    "XgLXq1i7TzMeFjkKUuaroSG7oXEgrAv6eEhEWre5xxSiiLLq1DXX+d18tIOZx77sDdDp"
    "PBgyo5D/a0zsNFmtG1Xw4Ob4a4vfgSqe2Qw35oh7quIs2kc+/+cTIxzTHNPZ3t9hpKus"
    "yEyHMI1jehJzeCILstLqe1wTs95RlcHbA2Mchi/14zT3yeDm8QNrDaxkXy3ZdLaEdmOK"
    "qOZa3OsTUZwcavKrqo6TpvBFhgOpl0rBkrOEA3xCG4o8+k1pbY/v2TMYaMoo7i3tNdOl"
    "6sgj+ZbVhEqCPEv0Gxo1jry/xv4hvzR01oFHp4P23MqStQKcihlJOvNzO2VnNBBSX/Sm"
    "/XXCbqTu3Kl3kP6ejsk4emwe9x7N1TXT5VHYc9xmhtHos0yWU0iTC+IDFfRQfUAGdbah"
    "bAgM27zYk3ZqlMPVUqVYLs9+Ce85fhBhyAVmaarAQW5IO/i+0mnz0FIbu1GO1t79wIJG"
    "WbIRRbw/iSyRHg3FJ+fv4qtUe4ZODD2lGSxYxysE8Pnukx4QJHfn8humBhG5ZXqCyS8Q"
    "pCEdLCOz9mutiV8ADeAcRdFpxetxJF67yJKymyCXdg3iR/C8N44D7/1sPRd0+OXJlhbo"
    "jrkPrJRsbK2WK9c6J9dS0aIBqxocBn38jzoajNpc0IjVdlQJ/bwNlEK7lf0bFLmp5Mgr"
    "5ZbjSZ6BTRZd9DB2+88+KYnjkkeZwgIo7wSR+CyzX5NnL2WogRklocUTcGzCMYLkbkD3"
    "wK+Zk+4GEWfM3S9wZp0ipxCigIf/I9v4TzbP+t9sQwv8JeIf/vK/bGM02yvWaMltXOR4"
    "62nLIStrO9aD4GJlbL45NFsKlXZkQntgwhZpEFiMGXkqLrFIC7nGCSwJRgeW0c2HY8LC"
    "z7uilcXL05owUBqH4JMWX4zGBa4Mkpr5PL7YgTZ9lL88QTwcxHOecJRCJSUvp5tvAk8+"
    "VzpXXccfoXLJe/3Eba8Xw1RyzAiOZJO/UV9FEfvVFgonAsUMVZ4zeLqzvyaPZuKRCBhz"
    "uF9OC6FUAoE0BT51kV2zuC3aQFWkwteaUeEgpUTTbUy/LZ/eK6uvj5U/6mmJyCxCl61q"
    "q0tXXKrOTo54ZJAsC3g4WX5pRtRVRyuDkkmqQKRBg49HBOWNQR0rKdutC0QCblxRBesn"
    "v7ExnG53hj+oGgPFuJhiaV8k0iqIxsFTPd4Nf6k7GgSdNtOIPNAD02jqu6mVJnm6a/Pg"
    "KUn+froFU/tqg5MF0pKf5eY3fxyOGyUfVtQa2Rq8ipAv1FQTZbwOlIrWPOt/NoYeJfc2"
    "NSFtCd09xljlYmXV9v6W/J27wy6CxkaSCCmL1fc3b7jPX59OLolKmHd1p5n0b37X7UjF"
    "HE+wYS0ZLckQNR4pvfAyJDPtoT4Dpcy3n0swFEAwrMOn4kLv1lDS0TsJbCqoSgAnptoA"
    "GDCIbWYUuWZhXj8IknbqwMNfrAaXpKfy7DQpchGeLcI8CPm6XdzAZUBG+pogyiJNFRlD"
    "2YHsJt3+UKitBytUS31Oj1mrXaGQsj+RPJiefDrG2YYQX+fWgBi5cXroEGsXA2s4qskY"
    "361VY0irenGdDpBxD2yagV6wxzhyu1Er39hGFFxR27pR4ehye//EzsuKrdXufUxMTQhk"
    "A1aT29bCB2hzMUaGeaznj+PjxbzOWZ2OLyeOwTHYf/ID1vvujyl1JuMrOhrvfkfwNuC3"
    "tJqv5tF5cIJ/zaucVd9H7iCrhpyovG+Bh8BXjAzF9YYvRV17Q/Mg8hI7eAn7zODo9Noz"
    "1XGWWYZGtwFcWQzGvqtgk7hxNjCKRLc4vAjM0jFoPWn36XLoa2j2CaI/y1VPs0VYwqX4"
    "bAetFiNCWUjoPH8KTjklOQ5HHheLzXlAj71MjzE8L3qdG7uG/rgWbDrEMyRxuH71DKFp"
    "F5Tg8sQl2lU/rtv34DYcNIDY1+c96TIZcbGlFGSday4iRwzzrQsZpZpKEDkPpI6Mzusv"
    "/tFZxmLajeBrv3K3jzh5iT1Ocnfuu08QIJdtGQkb4HLH43Uho9zUWEEQaGC2f9+gN9W2"
    "/I5WLY8+rv80brTbVidSBqR38UbseNDRlAy3CSfatoM8goc5VIJpFbU6ZqFF4HmQsnUL"
    "UoAhnGuDQKhLtKCz8n1H3KMtV6tyBPnleD7//jL3N2JRBGZXT4POD4tkiPe33vpkytTI"
    "lExn9V2oddKb5+Hn2lyy9Ax2yvsiESZhHOBa+p1KvH4K0E8O9oPS4iDi7ZSybEgfgtSS"
    "G88i02hZ+fnH1HMKijTPB70SgI3wsz4/2ojz1945zUrMKoeq0xjkkkMXYY7wV66cV0E+"
    "h/+MlpAXqRqvsIx9xNHLqxy+ZlsEodUTza5kQ3lLycV0B7elO3fFohUZNqg3LLORF8of"
    "oyxF3dyY9ujTYejE/CAFlmml1unsrO3GQkl39lnChcV+6Bq3k0HEQpb125406F4Rewhu"
    "3E9zJWJ9LmcJiDU/3Z0Kus3eg6wG9CFQeIQ8EgRea9I/m00oDx7uL3itqkZD6C12zuyn"
    "ErX6uI4qwU3WBEPI0Z2pjaWepsQz2QO2uWoGOsYIJQy7fq6r5ZvMVBHOuEYSGDMOnFIN"
    "j0+mOyRY9lw6G10efAc2Sld2pLXX7JVFfOvGxVcfIwCzgWKaxUUbfBM9S/vWfb0gsQ6U"
    "4yhTAyfBpJOPbh5YV3rp+LlfrTv9K+p6371wfNq3IdJSYJwd+MrgRdzZvxsVWJhyhh90"
    "Y7mTechGM+2XSDVDc1LH2fWAwDb7QEoe9hWcyIAhlb7P8PVNqyV41rrJVnXkMS//hxxs"
    "2I5ofjvIwfOzEPuPUZOttZdKVpNhssX451EsvWeYDDawEr7Wp1/OWB5GlHRm5/olXObr"
    "Z8yLn534ullt0I+w1XPxMOmUHYhy29VPw6EXoyoOznLJ3+ZsG1GDJMzV8yonAf1ynDET"
    "fgvfS1+UWJqGkpquTfErYaiy2MfhF3Mmu2jwjuaeOMdNxWlTLOAuB2MfVeRtChV9WMuE"
    "tH2J5a1RLM11xGEtskvYpkTQeb80prTV2djpRFQrYQ7LOB3NGRmbusKK8Yp5F+jC4Q6M"
    "CIG87M0acddqZ3jXYLEq7OxZaJCm5785IUToRsHxSeURle4io0b2FgJhKxk25kPaQ+Uv"
    "YkBH7ODzsJOiKobfhDLcZe5he/YguIKcg4ok0NSnNbG11aMmMJXo8djDhmsd5bio0Tz7"
    "Lq2Vfxv6ieqny7iPRwCj0Id9HBJuzwIQIRzfR/0Ydf4CackwuqmIe5blvPijyJj+zmLO"
    "1mkfEZDlB4tRY4o2DaioMJ/KPR8NyZOp4N1PHtvAFYAwUJ+H6qm3iRR0jAVIMX1FP0q/"
    "iguzGg/fQM5Hc/Fiqu0b/WnbnxuTG2p5LeQHmonGgF8Mm6e4wey4qSI56pMQDvDhfNLX"
    "mDoTEFhBXz/47fFsd7E0x8IfljJJONm3ZCFU2QyTBAeAYiEwx7fVNY029bL5QYbR7t/4"
    "I6ZUKlKYmgG886hF+LtuDohKu5kjjfX4d703sXY6A0rj802B4GzX8Qj6IiumnLRWWuTC"
    "q0R3w+X7DhHri7Eq2GRot+/M4GSi6cWzhRjwtLjD2S5Hjack2xDGw3wRYo6VCBiKwkii"
    "2KS4tOZK49ejfeg4rqMbwgtHMoodPOFDVjTMnPCU4EC06mQYtfJG/qZ7PJN6NLXWPxcM"
    "nWZgWY2BgD4oBGx87+5b4tXgMOht+aQqOAGAUta1yIwFDfhmMszZ2UgyFp5108xgN3K4"
    "EuK3tCp4vGiAgP7OsY6xm6lc4COmFA0v5DXemQUoJi9wKtWSQ0TAgIEq+ik/ZTywuKjV"
    "o7d0XloOHso/H2rLDYxMRL+CRPctFT3MiA0z1UovXsEHXXNsKEensFnLRV6FZKBsZ9oQ"
    "iQQoS74bmoLBKWoqJJzT7grQn7D4Vpur11jlaG7LW2++ypnOTjgkOa9UoQOCKL6WpSgx"
    "D/388Y3jQu+nriCZS2XO2NbuGl5FHrHLKx5bfd6lwDif+BErkXbYt5J5RPkALmq2IYo4"
    "Ro9eHc8x01kIA+h8+9KepUd9Qwanue8xGjQKemyTV6qD3+UnyDcQZfz+MttyDWdqsZ7h"
    "1ns4a8ZTQB3pdfZo7tW1WylQlE0FFoDRBGbVqcVhPnoeziZhy6QyZjyk4j6SgeC8Tuig"
    "iH3bJlNYMk4vfyyktJ7QcT3t6fIdn875JFP7GSgaPm9nRhjWLBjA8xYU7A0LO+ooR1HI"
    "pwvI6CJ6UmVDldBPJFMo56tykXmoIwj+bM0oa9rt8ZB+tj46xSaDNVeiHQ1CiCcyM8yg"
    "+Vb3bFkiOMW+wSl4K41CwaRAN7npop9mFMOUwHRuIYN9ss0WHAFXMxfp8qs6xZJgmDwW"
    "0sKO6+izjCvfhl/45YoWmUpMRVUcdWzNsphQWBv9V933OPLsihPY3W2UBqcfVHsYAtJe"
    "zUm/bcw1n4KsrV35+VKzg8YJ6IzW/+ywYLvfPMX8fPnIM2uF1UANpvUD8uyPKdy4bh9o"
    "acCL3RYubbh2MF0dtdZnErgvefUksS8feAUEYB463FiCvH0KlXfKB/lj6hyXepjlwbVA"
    "hGYJKNNN4TtPl+V08wuCUulSj6rziRmHEsKs2/j2ehvDr7pZAjBnG5zaFcqQATv1RfhA"
    "aPjxTdAFM1pfj0+u3tOijhHD3Umz8F8NM2jfwiIvse8xEF0uJxudhTkEOK4pVUquLXUH"
    "TksqFgM+/1SIDf6O/7P4j/aSsrnvSvScZ1oRTpwmxaDsKRrbpSJxMnIRSzlKxn2Vtn1Z"
    "c62F4FOJu5u/S+2ByOPZe0gdLU4hnoGWb8SGE6jegaSG8lLWW9ZFPp5GXp+tfjjZBz9T"
    "+H5az3aFb79g52JDZHwXgktBFnpMAbstNiP1gmXhPtNWOWgj3enSOehSF9swe266nwuf"
    "8ZTu5dkzE7JvZ2hrcTEBPlMYi520Ce9V1M4ZxdtE6GTiEfWwSVWLRoxvd/MQxM/Sxn6U"
    "WSznXPgxt8QBjgJrjF2v3buK61MOfYyWtQRPFB14UwhpbNc2YvVGf++RUSkJDYIAmQfl"
    "bbXQx6PuzEw3reCnj2LbICnieNr2zAgNK+bbtrKJdvCd/PVWpTmRhV1tAm80fjnDiL1O"
    "8U0BIh9ZQeXBlZChLFIkR/6hZyOpGSE0iNZBxzta8/1biHLhGlR3GXucjTfHzbEl79LJ"
    "lTLHd6LVVF0VtUfJ+1RmsG7UJ2UcEqbo/3mbYGht55sBTaKyFNIEielhXFgmz3Fgjw33"
    "79nTaE1DGF+2g4X7TtRYFZpSc/wh60FQ0chrcrrwe9y5umwwRCT4BpkjmIt1udr3chXG"
    "IIyP8Ogey4ZPPTny0SdwBaAl07ClgTjESXU+vi2Wlwrt/BO1VLzuzlooIASJxsAd3euB"
    "kuq7XHG3q1h0HTqDzL1isG741v50qkP2fIxmhmom+qiFqOQo6pTdUsgZ3Nx2ubLmzin+"
    "Wctxxacy5U6eCIVGMaYdljf4du7uEHDY4X7h2b6q6rg2e3QFuRJGovNxGviF6Ng8oP+Y"
    "YlqRkmjU/vMbcx9x6sm8kSuMsbbUViKpyDX9F2fn/qFH1m0LaeAXBHzjD3DsCk6w/6Ni"
    "/M9+j+h/K8Zqa6ctAACT8X8V42aOT+oh0O1Y5C13G+QBqtN7+lt3hjOP8j1bFLPUWh9j"
    "ccWVB1pWWqqEW0nfhpAouBY3LsDclDVOITLwNgKOgcMpztogfLqWd0Ltt6xlFDznD2zK"
    "3aVgZupWbaSZkpqabuKOM2H/49w8kdnehzO9hHK9b5bL/DVXvf13K28+j6tWIL9ZZd78"
    "0M1uJI/NQ7DZvgOebVvB/rgT91SmQi7UPyWbLAWYEazwt6kDX0ihAijvd71zYI0+HZeT"
    "nBml4x5QvMcPWvrxEyQQ3bFN0lyF1ZvjX4+yYGBiA7iQZJUi4oksyjnv7Tifp28lvaKZ"
    "CIFgQ+FIDrRYf5E27Gs/UEWvkTWEyBeq+lFbA2xKN7PQzWqb6Aa7TRp216xIup66L1aS"
    "Gh0KzjWv4DHCAoaxJKDGeeoUty/u2kGLjS6Y+9tocIXV0Tx/qQPR39vmBG83WWagyGZP"
    "rjaPpocj1AWy/LapT3N8t2alKD/KbgPk5SBgx6ipDy6/nHduQF+KHnyyT8Or68zbOaWr"
    "oJ89D19Z1wYLgGJy6bDrowCvg26UFp8OrECmhn7bjMyLvp1cCVlDJPfQzisAeMMCuPrD"
    "GDU35NfXms5H8THJz2YLgKqDvbiqSpHNScDgKENshPZ7DSd15MOsgE2LS4vmYYmjeiLT"
    "+tattJQ/7Gu1tifZrCf5dVQf74ZVOVaFKoUhlCIsubQ4LAuUWwOXWNeato0KUv9zEwXg"
    "D7wexrOxEaNeyvN6TWs/PpJhXVsFQRNILlCBK3RinQXXH0J7gN4c/mHWM2ts27Ziq4YV"
    "ozq6+LYe0tVlHhlQgRtUX6ROCrOGH0aAwFrPTnHdOZRUyItxbgVO1zJ1iMPKIr2EcNIO"
    "2bqGt5tqXXjq+pe8uh1Kk4CXAY2bZgifQT8x8A/vNcxUuBHDQKSW/1GqSUMs5LVogWgQ"
    "7BDdIqIYVuhSs9d5vTq0o8dYBRd8/++hsixHa2r9TLy/ekWZHWxbu5KZ3xAAGrfYUfSD"
    "1K0LsRKGq9vIFzYMKO1px76QRcpKvdK57ADGeegzMGbZXQ/5cAaEIUsESuxt2xvnK9aK"
    "vHi9FuSb/NugCrDMNWfiH0y/JkyGFFgamRUbR8F55+aEfCvK3u0ors2RP6DZi91z4eXv"
    "MXesg9jCuKSfS6p3e6VMFK2CRBrCfWSo+Vhb43k+l59aPoeWf+MTFr8fYNxvu1MAZLa9"
    "uNRxbMscL3AjHkHHJJ/LRLkc7EOHIRaz8uDLVEPd0ekCOZPHFeAUYniJ6hDBfHjj4vyH"
    "lREM7LvnIeEJqWFL2I/HkzTnji+DcH2OY109N0d6q51+ArPa37mb6rCDKhfal4ItMawQ"
    "CTocjhULPhG/7mKNdm9DMDZUEvzq2q+Tmq/jyIaxshwX5+ID2Nl79kakl8WV6d6dRhs9"
    "5m9rqq6LI3KowUF82M93v60nCXhucXgudy+K8Gea2GJh0+TDjBGxgqJRRbTocdGyzcxy"
    "e6vHAvqtUT4IeJzfTjj/0TqdyK+DI+7QTUCFh6/RJtUtNbnf3KVnfzqOwzUE1V93DdlB"
    "pD+lFZMLrWjaZ16/U2r9kFbJcDln4nH6LaWHR0JSoUG/e5vb4P9h/7jZn3C+tg/wJSMw"
    "Hq/P0/+o90gspYuTDgkKAiMOeIf7NJ+rSKS0YKRtUAxVRtF2Nw8GGFlDCgtfyR2m+SQ5"
    "H50rIQbaWxaWOfrM3G4JSyjsNlSA8yh6YlZSh5uluLlf9qQSOHTm30TBNTOWbM523fwp"
    "2N2N+IpSC8jaGOtkeIiX9otW2Lom6bwJV/5Xil+65wT9SCNw+GR8ErkmDUnzp9YXDVQG"
    "MpEwErAHNSjBCHMf0U0wTOlmYhhwQluopRHUpNcIDH15bbp8dFJglFHK10dHNtocP67I"
    "F2kz3o6yfHGhCL0n1Bq2RJtmq6CVkcuzKWr3WmKhCN0n+N5MeagwYqvGUVytNRGhTXA9"
    "5kU5f3Tuu7gpK5h4FOyOVVuoPWDhtWLoCsM5cg7AvCIupyIp7bKc588peUIreljDzvJB"
    "20DkujyutuoyyZeRD8OgOG9/oKjVICUe84m5jcFCA28FKzQJdj7fSMyfd1xS5aV067D1"
    "2m3LNWt2cylGYklCoO1zrt8Tfo7T6atz+xbOOXI6DjX1RKmwGMXSPxTaKoYpa2DUDvDg"
    "B+nREEos1JXD7uSZFeqWqZ2O88mkWwjeTc9VKv9lM83M2sr63LeiCP/1Sw6Z8BlPJ303"
    "t0S3SqVG82bWLUW/cihxoA4j3oETXHMHYym0rkoFJWk5baotNtm16s1Leal5DVb5aDL/"
    "Jrc4sDhUCUkKrnplxLQurjBy3pYbjo2OQ93BPVcJVAyJxImPJAkJHNWNUWQUcys3PGll"
    "hrvAm/Ni19IkMFFemt3HSk9X1BlEXqjjxN6pTjaxGmQtm2Taj1cl6W1mrbgq5qRZPAM6"
    "bLVQFSfb94oS+iwO/7tJB3THRtRKQzQzOeweik4PXP5eW09HU0q+8ed6Y91xkVJYr/VS"
    "fFarNhuhzDC3IZv2iZvwiVOwCq2eTo9+6obLA44xIpFq7LI7OjyGw4SR1uc8XzJdiSQy"
    "EJV/1ExSrjZRO50uGE3quuUy43ekZGiv21ymVrtFuIGUhIHVHYKCanNuMa3Tal+JuVxc"
    "o8PkjfaUOltMRIV0qibzwT2GSq5Glj0Or+90UkTgXa8eNXZ8nED5OlQuYiIFbuiGlIp3"
    "3jeoQsZr7amWhlMvGb7d6uTDfGuSFN8psEYV9eNkgGKYomZnqfE1/Hb2I9HDWR0YuYoD"
    "rFW9fcuWL8E7Z5hlkMGx4HDd8hItNnzovI+KWajlVONH7rj2eq65aE5lVWP5wtFpOWvq"
    "IPnKLvDlb+uxnR9cBiZiTu11sVazxYJg801bf9+JC0Is1OHmYWi/LnCnEKHnw9HFst9g"
    "ZR4ft0DNHTTd+WEmFb1wAskArwM/ip4nKemwgBw8A91697wKK0MgVf9xwYkiuFt5D2cC"
    "BgvOazwxsF3H1EzH5vu9dXS5nJCLQkE6ToEFLE5psn3aGNJ/A7nnJPWT+IxFMn9Om8r3"
    "MTs1Gahbhy5drLIGG1r1eOtFA2WGPmKOmpo4rCLV9TEM5f5RVLKqWcvnX6MQhGNwcFO7"
    "2MjEsQEE4S5Sjekv+CCKqoa6pKwMp1yibm4Qu8WJlkA1rNIiagpeVhHW2I10oC0PheZC"
    "yhTbH/XfNE1/MA2aiMd3Yu+ta6+10jhz1yPZCq3Pf5caPcHKPq4WzBqv4uNhw5Tak2Qi"
    "23eCS+qi4SauxOcVikoUOQfHzqvi4BHi2teOLN4aNxibg+kXwKs2rktMLc5DZkrwiJ64"
    "weBu8fGWtXVt/tk/d6gFiLm1JRbdvZesxC0ru2uZQXgQynqKGVHPlgOih7aLB5AvIrYC"
    "X7oQODEXAHapwkmVewanYScvUDi/z00Cg8WT8c/27fK9XTED71gMpQK4hg9NoEO9SknB"
    "OE9sJyruXmZQ8eGHmec1yyxJxXBPccU7ErfxAPzbpcP6YqEiVVi8LFycXNTkU+gC/UWa"
    "xuS0UkEVdwPHY+utx/wYm6e304BGFdQc2AczpqvPoTm+u7A2wMbHPJj/28+Wk5pSyOVh"
    "2R8V/XzdO6bvmooHbYxDx1pYmZret+ujxtSe6IVFk6NQU6Bl+DvBLjYSnMmyMp3lPR7I"
    "vj+O75WnJwXnVZXBd5XHnX68kFRMpYIPJ2jvZFPxQD3A822Rar7vPaxDzwXnrWU82x1B"
    "NfpMzxqeRsa+VuDUq6XZcNj7LGDeDieQfChBL7+f0GbmT6AdVAzYIxchhhhd1LUr8bPT"
    "lLokLLzlWLfGdV8HdOh3nRRSwt90n6tIEI6SFbKtKL7E68yS9SVg1dBLmQso3oL9SEDo"
    "2EauFsidPmseZYijWwEZCMu4aFQOIR9pAkp0ZefZxnBQnqplHdeGc+1oSFxuddRj0dzE"
    "jVgEPSQOmgFpBG3n6YAXzv3NIgUeSqwwxEA2ktSsQD/K8gTa0v6T9uV04qw3CcLLDCR1"
    "sLM13eF6RRrfBQQJsfV1MMpvaMYhkFNTIBR0rAp20Q/ICHbi1OmBrCiV1GfbAVkUo+4U"
    "3GvmK0gE+MGUDsIlMJO3Qy50BznfoccpGw2Y0J5DUsJtpd4IT61TXo3jWz3M3WnJCcRY"
    "GCqatdDRlzBzk8W/bTwGrmMQhM5BdzFwORghWlTfE3Lfp7fwyVsMvflJi3fWg9NEp+No"
    "NcH78ihc3O54yzkfe0lNc6catCxcTEVDJZSyroMW2I9h5BpujB/5tdXT/r0iWANQWVRp"
    "XmOefe/sy9D30KJYklAFPuKRQN+rpB3Vhl6KfmRANf74CZs/QC1KEFwwzycbv0PYe7EA"
    "zLhJpiWENqGNkIy2gUFbqSX8N6qsjzu/B5ws6HttCdjHh3y2VH/fRsXzhhQZDhjTjwH9"
    "DputytagP7W7Q8FUkIe/0yFaabTvwLT3rAHysAmcRCydA+UjjnKDt20/KxpZJzKdVvnx"
    "Ub0Fdsh39Fg1Qjy+KpynQUhBDl1lw6HMAA6+kVUZ6EzkST2F9LioXSgcz9ZTWjfGT7VW"
    "c3HJKL0JR2LPytNGaU+ZgHZJCrFIbtE8CYmEvf2AZfdaQyibv4EOLGoderkQINPtmAHz"
    "oJpX9X+l+A/+8Lv5X1rMCAOxD7qtkKnd53a3hReMVPr6rYtc0zGlGlgUQb/CUfWXnVlU"
    "+QXEt2LTA6p+ngmdCf6xS/SKSvLEVlkeGJWamlWTJyZvMIFyHoRnn1AqMgoFKR4LLrgu"
    "1u5KLJHnFdponkPcW0xSwCC8nxixv77pdm5g5Oq1ViZYVPEzP91n6Aq6Tdw+u0DJWw/M"
    "rSICbgHLwJh/3KTyfZ8/fB/knRbm6gZF0Nqp3rAXRA7JNbePn6kMU3Rmwq6xsozi37S+"
    "PrWWb4cLsE8V1u5X3ykVcmjyZ557BZYXviO4KzHmF8a+DKQdpnY8VRxU/+OrobD9MIiy"
    "dX7lazCyzwK09D198fVOYh2mfjDNASOov1cms5/34AWVx8Ta622X9qOeQDvRiK+d2vDj"
    "5PNqFCch+GXN4nJJ9HhooO/qndfoazf6WtEfexjmFrZ+G3LP169w0MED3JzF3GZPOz+9"
    "0tfU7c7+wl7u07AamN5qCM3EgKX/diwO2/sRhJUkjSqvM+D+GqpDYiFI66N8ftybBi7j"
    "u5MXQf1svf5nWTG0NlDTV2Lkz0wfvz1x3pGLzZuXR9mAS9dxiky5jp2V2DU+aCONQ8oc"
    "DWQfVpr39s2we7v4WqbntqLmzVO5LBDmP6SmdDPguJ8RXd68TjEEdNx12Q31sQUcSX3a"
    "c8igsC58R6ySesvv7Xy4ze6rXfLFhxkt76Wy9afvKvbaS+Pj0G+yqMIif3cU9hXDY3fS"
    "z83BqCq9Nd3DAZKD9OA+0tRVeLt+S+AkRkvj9AwbrOxSoqc86B/pVK9wor0thd2qa6aU"
    "gWna26S+9opZ23VpE4EN3zhXOdjlpMypDdzBD5BOZMfBjFRtqNZCgmFJ/LySBaizSwci"
    "flD1/eyb5LMeikTgeArCmQ3goUgjf1/9CR/QXi++1lW6VmNPjlaLhrOSTL3R1Tfp1oZO"
    "5zgumZLQIygn5h16OBVNRvsNF+iFqCIV0twssu0kqqYq/LoA+W+lvzU9Ffj2cwhgnh1q"
    "JinBZfhfuq6wVjMHNMrFjfV+O9vVjbS1oa9Hp1ehEjGcLnKNCB6g4chCfjoCSbbXWAcc"
    "UjnW/uOvm98Z1LJUt3taEz7HygKrqwVU0f4GidviD6HTbst0TDGOU0l0xA78wr3lBcuY"
    "6PaGm3NX8kFfQlsH6w/UMHcnGViU5au/PZvXDcB1rJR6xNUh6mbbu+i4fI0LRwKIaZvb"
    "EoMHlPXileMXuyq2ii1NjWj8QzS66U5IJgay9GykhpUlp63x/aYx+lzTYAnvgtqiDs0k"
    "/HoadE3t8P1qdzfuQk65UZTaVredey4E57oHLA5ju3kQwa9njxOctjVyHqJMDEy7GjMd"
    "rY3GHtXGdHR8LU4diZ/Xo6ADnFyzM/Ks5scI6BsS9/LRp3MgjTd19ETsbZ31rrskDrB6"
    "2Uyj6Fq9YXvqM28/g9AdLlka+RRMYSsnVUk8yZ4Ip4WqYDNS93L4UHicnYHcZekcalOJ"
    "XL/BYCdb7Ma3l/U22LZ2xmdy/TFYZHpvFZLPpNBLzvyvb1vePOCNuW09JYrsCBIIzGaB"
    "D8Tmo91D3zokl075v80ActiHZ/vAM9I4mOnLbWWZLCg4hojOc12RZ9E86WnnjvNpdPme"
    "ltYFWTctOy/p8C0Jc0fNhrCczkYoo7z7zw93aYhRV+x2mfuvahFYVJNIpPVYQYEnob9Q"
    "aXqxjOWxEGoVEXPUp1Z1hn7YdFZm3v6VRwv0oc4R/D3JtwcOsxZv5TJngsM/m6Fyos/A"
    "f7q+V90lt7Ao8DSatPuxLH5cVsUMa8Od/+WfYu5R+Tq/mqF0hrXaRdvOpftQyIoFnvGA"
    "8/JYTgbYcr+mfVrKSY0To7Rv0oloUXcX0Ofi4ZaTuP5SWbjpQET50kX2CJv17p8ajjc/"
    "6PILaLH/E1aFuxpDih9M0cfbFbwonsB2Wr/0TK9LL05uPMpqbMUhD6FZqSsBLKZZdqtQ"
    "VrrP4mBFMf6WDA0KSPXyWLq+G1bE5Q6czdr1Q3anZCHZ+KJ6+JSXCrBHzxnRsgw5LqWm"
    "uOorajHyAa6Recd+SH+PeM4BLeoIp3RMG6PyC1KMabONve580d/JsAR8vhQQWjzadmQw"
    "RU0V/dxtk8hSrMe5DyG35GLNieKTy9PY+/tS1YYFsOcXlGugGMQoV4LZXM4FSgC6jMQ7"
    "L69dlP09lJgM5s3JpFAB0HO8AWbZk0vwX6rAORxtqAbbmlwFsMwzXvFDC9MgDFKUWAgY"
    "UeLQ67ikEOUPWuE1u936/dZnIzDlvyYsUL8wFP6PJuN/fiIP/G+TcVjYZAoAYGT9X5Px"
    "sterx0LK7eTLAOhTcDPbcQomVnYKLVlejjdEGYyx4yIry/C74vOb/5iRIefvykTYyX0q"
    "dL/sDByn10rQ09v3WGEpP1XTVFrpRDFqPKkjMdTOTb+32SdEH626bZK2py6J7au+p/w/"
    "EER/I4q751Yrqt3rgDb1wCFnkz/UU2la3qJg3PYbuMoTYONevIpSPNFnSdG12K0otEGZ"
    "FgkPJGgaedGLkO2BgkUnRNIHx1XWW88rMyJEkyQCd9McJTYNvX+Tdk2qSBU+RwVHSLyH"
    "apS5THm8HhqweTdwbcAwoYAaOlcb+TCQaGBwAdXei29WaP7hk8sk9X/9J9qWy0XjcTru"
    "0dJvL8oVqUnMN+EahEmmC2Icw5uwQUJnRLRBvJC/Xo34k4RQWsT0uuVGQr8YCz1BV9IP"
    "JMT13H2T/wzc8RybGidNdC0OWBDvzMc1eHVFvvOYTeNV7vWqA3Mjs49X34nBpJYoInFt"
    "JBG3utchP8+K1us/5TzNoY26gw0wRMd0/9puaNNNRScQHPSSB3tYG/OQu7+AbgVbYc5U"
    "F1doIw1AwgKj12FzG/5v96ZSQ0pwpOqsC27jeX6la8821NHcz5w3lWd3kdsRoksQK53G"
    "PS79mBmfxxFXz78aXN4vLyoEqAM55mI7H9fuE/cNb+QxRrc5ySFmuzRCK4f/BBDzQ1aX"
    "cjcnPfFf/DkIHr/qOaLhI/a4zwmDI5X5DNfqnZW8tVOl8wTlCvafP3mpAK1i2+g55RZ4"
    "hJyPubbX8WhW8gHsq03uwwxGZH8/UZdgH6IFG3aErESXKMYiZc+Wem9+rwfchhA5BXdB"
    "Op6WhU8x37GsPZ9RmHbCnaiDPu4mApCDq8dxwX5sdWWIuB9zdTCX4k7m9jusmf8V1GXQ"
    "9N3HGVsr6VXYvJvccsHVyNO2WXqrUFsXvInSayAi3+mxIJ+9uJj+pe7eprzCK7zWzelH"
    "/DLnhNp0pv/D1lUldcwILFUx58jx4zpB/4rkrfrPN8blmju64EMvli1/sH/rji/htPbf"
    "aRKofyHJ4bZr1rIsJpW4LoM2vgw/d2BXdU1WLrD6eMOjjSRfRDplQ9n+BCUMlYFneA1n"
    "WUk7Zhl174bzZNPmi31+Evb/Fn3htkn2aycJqfTmGtPM5SiATFUYxgnP6N+ACD4LOwOq"
    "WmJLPM1ZBkqa5KAw/RJeX79by1NEPf8zgmcW4HHZ1MAf9883tmvmSBkGRyh8jxcYoysp"
    "SxVZDgFfm8jQlXSI+M9/5L+HUPu2b6//jI5GW3yOAeO9lcUKlODu2QuQVn0/ZpZTezTT"
    "agBYh7833MPLjO7e+WB4jmeyX5Alf2S9g6lICkrH+9y3k2zb/K0TqKB1jH9CNv/agoTO"
    "hJxlgDpAeZmPjx62Srn40hfFsrI1p3JzgGGtXxPE1q7YGvb//OxSXKfYcZOGu8cXzvmY"
    "/oQnx31AxtpqtN/kG4U94DSrRbr5+AKUG03UiYViN+DAFVyCekQgSmYZyM99l/JQg/Ir"
    "4at/bqf0o6MSXTPe9alrYG1+uQFUHsbG/r4cMEJ4d+79cn5+1JtIwUBC+PNyAXo8DwMk"
    "+g5PgSQHsge+zPu9QC+rlcI2cIMbCsaGmLZXHvmFBNOakemp1l1zIZkD8A+NsKVaEleP"
    "LukAS4oNCMXCqkR0VB1ANbgfNDSkZs61wecQ3W4LBXkjuBm8Orwb0oeUqJi4SHiXcPox"
    "rSz1rdYvsvG8S8/4H198U7MeDtVgv2sYazuGdtFqY3ZIKRDxT3hUkgPnXyam3qy/kwaa"
    "NLyPDUrQCR7s7Gp4rYU51YrqkM6+6mi8XgrDnZFrFKhP4AyYW25ojDVr4fLSYSsxw9KM"
    "N3KjQ39baYLnHWL2+GmmWQ8Zik2FOsnPkIs/hzrWiEPsNnkJt7GiciWfVieMqg60PLon"
    "3IfXROZ7rD0eyXWhYgKjxpocGAt2VFL11BHTJoA64OJq1U+PgdwlTT7viv5uiTCLJzp/"
    "sXpBCb1rnmIKO6KaOI8Pn1DhaKD+YiLR1Vv1DcIAT8uT+RRhDg680TzyyTEwzCVm6MVi"
    "t1O20tz/Xd8QotA4DjVqJ4U6ubSQ+JeZQRg5/jiHJGkwJdyCLSQ6GpyV/h5sFEJpNa+I"
    "GTVBMC2dHO7TrOSLveLLcnjzqlgGXGjjgYuuD9ijNTgrZGm8j7b17ozFMJSFSR4So5/u"
    "MPnxZC6VRGzb4plbOxXdTfVSJ9BGFloz8fEppVZf8aHeaMFlCulZjSpWvcDyvfRPJv+j"
    "597MHHpcLKrDejSed7rrCR70XqRzleQm3xk/ThN0pbnAlwnMdiXddrVXOOF9ZpbHihbZ"
    "Gk5TQcsbHJLtB0pNuAcwVKKkcV3GSnENHbOZnDXKsT+kwB7P4gKKZa/zbu44jCB+XKvm"
    "O2aI98KVYfFvtIPwgqJDUf5kSk5DV9jh456g76aII6TMUuXUXrePRs7HHv0SIIeCnUMI"
    "odqO+ZgF0ulRDEn0AQRG31Y9T4XV2zK6ua96hofpmZ6G6LDvRjmas1fjgj09tPorFrxN"
    "pIuY0u/T9gmw7B1wG3m0Ioo72aehRoo8vG8PDcdEDr2BEN4+cjZcpasjnPiysB4UM0f6"
    "HTKPOpx8Spx88dm0kUHiSRKz/lBLcTTXSw1+IV/+Syi8P9V2LuRHVtkIMRCgwMFKITcL"
    "6x0LwxKe8GdZGApy79991PVeB/uosYv94L7QWPMvM6d8EB1qcBkRKPwld3O4Dru55ljy"
    "WDe+VWpPGfmffPQ8LjWY//iRfTTCazrDCDVXPMZ8cYcuFjeja4lblfwttqVae6f/c7XA"
    "cWj3/GWkQmSciUKxTzty1VjoOAOpgQ6qxmxhLjPZLWzfdJfkjcko7/F9s9RqUaIB0nRV"
    "y7gg8uNYKdckTN1099ZwuKmJdQ8iYLIfxEUWclxQLmjUQPdUEY1VdIKhiWUm7H0jleu6"
    "EIWYEG8aK3Yx1FRHlBWEMDujmz2rSkK03+NiOGpsRBJw16Ssooxhf4+2zEUk8w3OoQo2"
    "ZzJcDKMTB5IDGOihhVSBGYqS4GNvS//RDOJNfvyJQqqVPIa5BWR7B4FN7SY1culJUq2s"
    "SgU1KDM4ELrrs+96cQK941SkboiZ3+eNHPfkaBrtzDFt7Cw/Z6BAQltv6TwsMDPHSX4i"
    "JJKQNkzr7XGOC79orosQthOe5LHK5mATyiap9vN5EAX6uC/pZDNdhMc0USKkhnhemQL0"
    "uWKtfg4bGvk+3hIizmtQAZ9BQuH9l1yewe4FPi+viqwtW80zN59NqePqc4GiqphYJqY/"
    "okW8jwOOLKxyJycmcFVbP+qXp/MbNGs0GCwWTJNWcAmjJaLiyJiwqzQnNWqW/t7dkUWT"
    "S0uPea7iVdxEfC2vZ2Zmi0vRRrOTXkueeX1l7tetq13JpoJxlMwfYU9tuxMI/zaT5ep5"
    "26IG97phs7rCUjKH03VdGqjwYTpORRz2D6dI5Zh54UjJn1Pm+5Fln9htGMf2TUnRafIX"
    "zQ1Phl+obl37sIERIzYamn33ZYb+B2iZ3fWroWvom5MrGeqUkOK4T1vMkgzhoM0sf0oc"
    "QYIM0JoTH7dFKFojzblqhOrroaHr9EjvtFpTL1cqs8nwU2HpNr96cOLbbBFNaamBgdaW"
    "QDr74XD3FdOhYwOqaPy7MITRWQIvFXY4Fy0lncmySC4tTpxAvQ1tL/U7tRElGalY7C5I"
    "5MEJ2esmpnL5fJ1sik9ZSyuNpsA3mcX39Gv/SM9eKk8mRgqZUbKnHhP2qPrnrSuHFD7S"
    "CL6S8KnkZ4HIpwSy8AkpdUV5493rw5Y6fZrnxc241xfMF/yD8OJ11oN/fCFCrirmP6fp"
    "FJxyvggg4z5LqXsty06eGFKCt9CopwAHaRfZ7Z4ZPtBo5sT7KC7iSZbT7iLU7b/Thugc"
    "KfHXoKBNHSxyNgi+MSIb9cf/8x7Rbe+Tq4fWyyGRx20fyjsWz5YnR5pz2A8tHAPVkb8a"
    "L9OhxxwY+XVVC1hSGt0/K1MZNjz3Z3pAEAGjRB5bUh5ay+7eQWzhunVjuxF1x2lWvbm/"
    "qk4dU8WPZ91UxDd23ndTA2p8oBU9xyMLp/MJEYxOE07SGqmpn1RAg5aPB57Q31EFrYyI"
    "I1dUUKXZD3yGZr9uU4/z1tWkfY1f4DgmNT4r7uJ/tE12NXhRzVCuXl03P+XLl0339v8s"
    "SDB058NISbm/d9/vQVFMo0V4vc9JhLvQ0Nt+unfnYGMblGHpzjXafLKhT7t1lyQl50wJ"
    "dZn1wkqThzrezOWjN7kSAi8h/nTM2Q9Q9Hqd+RbpJWkiwejrRVpMbmJ3FJHzBFB6nFYR"
    "V+nUSGN4ZarABFdoxbv1gRLm2+38zlhGK686dHf3xiZdufSj536Cgvd+vm3xPO4C07X9"
    "ulpSdeT5mbLg9S2g0LBBlR5L7o+rkkj2zHhShL+lbnZ5qrzUjzZr4QcR4ZmREUpVzBlV"
    "xDi2xXOkv0Q/0zOIvSDbvJlyxjf6CNe8C2LG1P5NeJKWVZqCFqYOP4T1gS11s6p7Sddz"
    "wUalAet+YzSn9nzb6dsQlKDIKyx8dZLidZxx/EQR7BKPwW1fiDru5ZWLHbpptCvwdewj"
    "eDtRolEDzNX76erG7JgVQeyVCnovALFMQC0yhPEejcSNYygtsOSq77c8gbCenzv3zfDQ"
    "aE2QwoVy1dnf3RIs1J5sLWeJyQIjd2beSqp384Dr4n0Qr5DCXLvPb3AyX4z+rRSOW6rM"
    "GZ6/0CJSyk9zpbSA8GHgOTwOR+PL75yNRxak/VzpLv8+G5+XNCIWd+l1mPA3hMyk4FUK"
    "fHuRw7Rp8kxsjagHIC7jHRANUUE24kj+OeQw8ZZO27vqUmvqqBBYdx6AcbDX54dqZ6mA"
    "KSPHg7Pi81fera+pr9MEDpr+2cnlpcXDGZASAVW0SuOgvht0uFo0M+LUUyjUr60C3OO1"
    "/P49shvUl6FWxZkACKWU02nx5vLz6xU4BDcHHQX5BynHwuQZkyooc0DW7CZfRarHDGWG"
    "t61vNurVAX47XjfjPK/nR2pbonO+Ib1X1vu6MiI99jYrIGlj6p/hnpzbND2+QJLC/md7"
    "d+slS73/sgciWEHoafuLNNhWBiCmoXECOkeAbIz5vH26QHs+0Z5MOOOILoIxhbNZojG5"
    "dF8y6rbroA+CthBPQ7nXjqeTZbDVRfOkVRXmTTZuPj5iMBUYkLYZjAkbcu8WkJmYn0eV"
    "djZvW75VE5hDRrKTceT7jYUAVb3kjMoTU8g/s7Va7OtIeVgkWWW+qURPqI4qPG+XpbF5"
    "6f5qjwHGdy5S891TZqmd019iYqbbkdNkgoFK58dQqGndKK5OYi+DfO/WyffkbtjF84dY"
    "2eO4iuim8Anm1B8/za5DD9LIGvQ5s/eG7G2lRLYN23p6cPegu9UxL7hFROUdIwwXidps"
    "9ZTnnxsyv3W1F93PNt9lrLlDd5/n8QtZijci+v+91mfLWrCxRJdhdDRRlqTTm9p8UOoO"
    "gzjo/gqOPhzZQzwHoXeWVfnaL/y2eqzrfN4Cl3GMCrBfcFKgoIwXjaTJ1UBlrlzbnITi"
    "VwJ07XWo7WwC1+GDYXctWEgUKi32jOEIXfqhZhL17hxfPBcmnab3qQL022pz61Ddcb8p"
    "MF2qcNBibLmIjTYyc9fZbhpCbOe7v/dcnaCNHsKvOY/pIEpcXftctrPB2fymmJw7j3dQ"
    "7k6nrWSojQlDoA+HA4WIrzj5uqHeSmDW1F7//THSSi8+EUbgfv0SlbpXW7l6xwoBjyn1"
    "/C452YhiQx0qGNyhQaGps3r1zYymsATSGNbL6TjNzuaJYd4ilKFC/UpZMI3K5uOLbQBj"
    "ev8NYy5Nnc4gK8hQS5uEwFCv/dvr7NRF/ONxUHTQdoN28rQkQpHyoY16fPLE3cQ2HOQf"
    "lee1VLN1Rg/H4skXrG49buVx2AGAerUtI+w4FuKfriOH0Bqb+Df17clbHeHpjFEEAp5V"
    "IQUiBwHucnfO/jKxfDi3gLOun0cnk8W5f+D/UQhihlwjDt1bN/G6LAKTfwU09LSbV0Wq"
    "zFGvaRyuq9NJ4ziXMDoRIlW81R7s62Iq9J2QJeInlSdS5lOq9jg4aJilUt6U0tGqclvG"
    "B6MeQI+a03vPn6eLlKPncmzny7/0RtuVJQEleRvMu+qPdPeKMF5FZZSg4XYDx+AZclWb"
    "F5uhaBE2lu431u41JNeUvKIROdUtbCcDtcnLOTRCeXtkohZ/690g3uvSYw2+Pab/wi3C"
    "1fFmyJeliaw0040643vtzsxs92zI5L77JBLD6Ww77ao45Ss3zlSrrHKW/sX+PSycPeZi"
    "PSRekgiJtWryl9TUGm3u3q//x8c5RWfWNe06tp6oY9u20UnHttNxx7Zt27Zt27aTjp/Y"
    "yd/vN8Y+3Yerxlw15qxVdd11UHM1bxCu4qPpYTNdAfzYmTiONdkVDqzNrA2t722ZvjSj"
    "TuafECyBw89L9jYmoxU7VLoanekOV4yDkdMegFIoWeSDrejBLV99f61F5ShD/8s78UDI"
    "ka67w47DIaMwWDZ0pOgAt0/N7flFf0RIfARe3mDlyJQCmMSK/aRYdR+1Z+tMn5W+lyaH"
    "N0VA7qcBI3R1+oj8NjyqbS8iusdkySCGLQTb8PvtVU+4Qpn1foEgFbAh9eH2Wl623l2M"
    "GiNVHuL15FVz+37Hxpo8AVOcq9i8veSEZ/D7bG69T8edu93A3HQ7GcInmQqZwgYbNBeb"
    "Cark3G5Ka1DQu0hEXdUiShgdozHzjyl/FuwIm6rpnR0PaL4xQFPoBU4q/rcO97Hb5Cpj"
    "OXAkVkbRiipeN4cnDjqCQBIPuXhC+nBPTp6tyl6hhPdQ1kI7f+r4G4e2KjrY9GkEEmBj"
    "xTXS3XI8qgYp9prpmvL3POStaDTjI0Up9k2T/p3znFU+9UZhMb3+JPaM3y2UPWeDY1RR"
    "AofBx8q/3GMOOgU+pU/Y1ooLDjuZU4WbDXNFI/+X2VjnlV2jornHcd0ak7nMpPaxvz6H"
    "lzYs4k+DDijKqtasOyXidDuxFmXJK6/IKEp8Rx0706F4aFzLR/Fi6Mu68AF8xNUDE1gm"
    "c4RQ8ZCS1ZggB14jpRK4UAkRYVOlnNHuu4shZXPRL1xYre/bC4SSPGJqR0KbAoJUjwnd"
    "CQsJTcvpY0dQcrv5d1STLSta5x4iOw28mBRnMB5TyKpIcFZhYg6gUwHfIV99U6mJW/CX"
    "5V+N7t3TyPVyaIzB8x6m5+PzAilJZM8Rdw7e4+D4uBfg61ubo2p3ocs2a4xQMrSESEa6"
    "L/NA4gzzeTP9jf6eEnbtXTudjVWHI28uIv7Sj2HPganTnd76IRumIFFyeZ7ZfMUgVKQF"
    "AjxsOBGXknx+80jwI+avCn/wKuLpEmNIn4sksWTLzSfLlLdihk1da2coP1PtRVMG8VN0"
    "lXiXpha1Xb+E4srmsyCSIVCfTJoWdAFpRSQmkeU5QOhcAazQWs+MTc3lIzYs9zWP8pQh"
    "FSLSovmOrQO49a/ElSjwJuN1yFH8CgUq7SXqx/G9RlpnzSGcDPbQKa7XFGgBK+mmb0yt"
    "Xb7QO83SamfnjOpWAAfvds3fY6VVV0VyI4T+pcw7R005n2vvORjxRPE3lvRvt3Ns+4jV"
    "9lI9yLIMBW2z1IL84upzhHd9BqOG3WHIQdFW28MZ3NY690k2EFrkMrutcui66za9hRJb"
    "W+krycMd2Etzb8fYvcNfpo9mB03mYMQIxR26QXa/LiZJuPyIX/LVNQd5O408WMRO0ddV"
    "NNB+Vf6pIB+q01dsqWoRl2JxXJhGqQqsLFEOySrt+gusY/Bgdl0LFNFh23phnLX7jM0p"
    "oaVpENCD7BMl1X659kZ2fOOGdji3Uk9k+7jDD5fOOqxZaa2ZUAyO045vfMd6WcwBmy2R"
    "Eja38nY9h0UPeLZIxuipbxhSHLaD8Pja4DIgieyA7Te8DZGaKjnfOk3Kw+gtpj1A5lhd"
    "O3eyVwtAVTk3iQBRLVZo4XMbheH8+IsN1UGTZRb7O44Bws3Cnv07Y2cUUV1wcAI8smp3"
    "LjM80ARhBRW08+vKaLW1nHgpx7emj9B5dJN7LaWxh/cYxD3uNHsZTAKDTrS8RaT+Xn9o"
    "5v25KGSD/353hBpPta4rH+f+rB4KvFMJNRgbPw8xeYM8x9N2IySSLzQDA8HyhzIbCNYS"
    "AxI7bdu6QD8/p7XxleI2HQOy3X49GOglD0F92ePyHnK67VP0WzC9X3HNTUFSH7djRu0W"
    "g/i6bssfVSBPokT7UC4O+ON40Hy1bBV7P1xaa8E5hhVioXbmEiVaBHkjEeiNQ1sWSEod"
    "E6/ubjbSyFObBsllEgieAsZ+bikqQ9kInipbJzroK1+HGNWU/u+BkrbW/dcdlQY3NObn"
    "C8ZVOIW+98vL7GPA9oRhRwJ4Ea4fPU8/aUcyFm4AXWP7tNIPdvX+H8RgmJ099+IgENIG"
    "EiQG1Be4V5YS4n0K56+B+RZ66tF/OVhrtokGk5ZXgCc1bc164oAezmw728XmpL6PV9ha"
    "Hfbkuz81RZe/P+oOzT5AJrvnRosVKmC8AwGHpLDNZePI295V7LQgbku2xaYZZWhvUMZS"
    "GMSyKeUwpmqRFRbT3mZWXZWPiFJ5x83JaSEzaYHfVl3jjTNorfU96y+8viKw+Tv0M47D"
    "GsbCrMCxNaCnLjMzsY0u6su408SPUdyHKWkIANgyI3XR+h34WAO3H2BwnvIJ+AK+iL6k"
    "QLkv940OAkmnhhh+j+dtJC95o1WzuZ6oi1fHhdnQ+DhNGnBzTe3gw26qFQgZB5KLbWW3"
    "C6a5Iq27K34G4u884De2dbHiZxNwWrufXADLq+L3YzREjxZhxlQ0YT9BvCP0/nUW82ty"
    "DwSMzSrmUUKV6/VSwKbTJqqNKIldB7+igEz2+KjxXcNXxWSEuOlP7uDCSWfhYzNEytfB"
    "WlNBBIW2FW2/BxWTVVhxDaM0EIImXdRsBf7YEf9RmPaqdeFutWrmK2eCTkKBVyrvLaqY"
    "oTfSFHGGqzXXdNtxVPYwiVb4V7m8J0fWwiD26ZNyLsKYLfTzZD3RP6x0Vaq6jKQaQf7s"
    "aqnyc3qS56mGYDBdSNnr3zhIerQngoXu6stQEw5rMWyBMS/FReYqc8nkDirCERgMhkBO"
    "15d2oWaSgGiC5jbjGstYoYhAaeEymQhdXL2KWXCRSpvfe3FneRRjKcQiwIPvJlliq6ts"
    "+XmbTey39fMZHmLRO8Pqu4e/Yhq2j3E1Cibu/dNTnIsLM0EGNAoiR4Yf8TEJoJMpvRDv"
    "cKu33z/U6lSyt82ChwX7zMkPmJSBs6jWN4Vq8aw2o5VqUKBsLz/blE5MpwfbyPP3IsXH"
    "wVXTEmWRvvGgsvfv2ZhXZjHnrSX6AY7Mhj8t8+uXvx5pF1kLG2adVapAmEvPVTRzCkh/"
    "MRJjOH33TxFialgBf+gMn/7KmxRwa0IEWAupAfBw4zXPDnI8Haa+W/FJTorUWRgMBdhW"
    "t328c1lP/471/I1X+yuZWYVdvR7rOVa+iJaTnktX6BjG4EXhCuJb08pnvjiuA7q5RVD8"
    "PH6FgaGcX/zIlDm71OJsWQsKx0NKTdmEBu6BRyR+TzsaoLUa8OfB6ZHCDb2eEpuaEnDV"
    "MgccfYC2v+HDYT9X1OdhYpfwZ4uujrVmQmJUYFmpFiu2lOTDG2PujadDfyA4P14VE7pS"
    "DW/uWXDwXSDFYukYpqcpUFAm5egNbjgDfi2SNgnOIIDS+fJfnHNXGWjk4jeG7kF0tOeY"
    "QottjuKamyUJkdcl4FpS35whtlhWMJkHN04CWpeUKibnObTul9Kj8jj6yec2SbtuZVUJ"
    "PPJg4axWRxEQYH41ykKj/Fbrq/oYNrGjB8UhmGGAXAXWwa8PIwbjWv6K+85e+3KiotXw"
    "b4BxXcGPJYsRNRKVHQzSV415rWLbQAN3TKp8OqERWlDp6Nx3laMHiyPDcfGFbjrH7G/x"
    "J4Z2igRWSOElmcNeP0qgjmYjSrQHqqrx/ARhrhv2rKksxtOkUkIrLplgHg5LnACsylI6"
    "pR9jhM0hOYCQN9ajpdmEzN7df3PPbDRosSDVbaE2p6DMuWtyn6QZsV0UNdtZ05tUWn3n"
    "Xf35q1ENSu5W/WVSsU5jFEf4e21JNc1x8HTBY/pLJVd8xUs7+kOejyS/f8q1ab335G3i"
    "sp0RgFdX3jnHY44WYDOBXJJhnduYhozfybWAZjPRUZ3jYDZoL8JspJT8qzZLOdz9Vz/Q"
    "/CDVTYEtycucjGcSbW18dRdxD7Iroajx4osRScSHqKvIgKsUlTV92IGWZGtFsr06R0lX"
    "Ct053tgFIimdhXk3yxzFbJOfnCrQ9oZaaXIrPTK7RyKMbut9EyyDCkpQnnDtOaMsdOn+"
    "7wqM8QAvnctnUBzR21MOSaCaVjCFkb+tHFvaBJcVZavfWdqidOBNh3RQJUUmXxI+/Gwi"
    "+s6S6lHYGsuV2fXIuqojUgXk7GNsUvtzOdjx2r5O6VTTDOcNP+SFm0vm/ojERJJxNjOM"
    "y3zUQBNyTi+nM0Y60LAxCYJABk9dluMMtd/Eapu6y7lJKg0M0qBRuYdjpyjBvRIV7Ieb"
    "rX54x3mCE6iVQYPhywUhItE504tdghkhFwJPp/upfbkSVYnX9ak3HzuOYhRf5DJk4IW6"
    "u7x+eH8VRR7q2vfqCO4fFx7Op13iOPYLub/mue7rRRUwgnxloT/ESjCQqSLqzs7sa2qL"
    "00OZzLx4oa2S5jOO5eV3R0Z3gXQ7sMrTT1sD2RCgEfFKVLxfZyU+X5e8VlTDvPKQmaDD"
    "JFGex1EOQ61Il/n5/d466C/W59zP5fYlbIyhyy6/TZCcPp7syNOCRGfuX+d5fF2FkrUo"
    "OumA8f1wtl4efKI+X5HZgfdz2g5eT28+OTjRbS/Lme7jE1Aa2++lwRjAgUVjgH6bB0Zu"
    "p49fgG5XBm2QaI4JeCqCmlaoIF//w/uHeTbx6Zltuk/2cGqVZqt9qzUZ9Me9P/+fKTow"
    "SKgou9cWHMJNjmk4UumKMhPEqGUEb5PRM6tSbiOT7IDtTi8i+4HCTbitB7fg0p/wDyk6"
    "tOxWzCig2t0Ot0LByMhJ6ynjXwwAG9uirtuz220P/T4bUj++/FuT5DTb6rP8nGrJnCVg"
    "SSllW0/KXOtvEoUFGYgin5tV/JQ/csca1D8JWoOKksPipFw9uE+0oV1aY1ZFDb5tfUwD"
    "LkoSZyCWXEAaTua2R0Gr+jhKve78aL6La4fUCMWENYMoUUcvfHlnwpQjpGXw0QYMY7tj"
    "v0Ai6Ljhp1B0jj/VXr7gPdwU8c5pPQyCAQKcWqei2iVYdeknvZlCK3geHzYXAmOtl1l9"
    "8B5cBSg0YbT7BoPLyrI7T1MQ4meSDpOOUHeu1lQfhxk2esr0TBn1XxUwvTYlXfuth2wq"
    "vGooDhJ12rUxj7//zLgLNJkkkHkSg0dxHW9Z+k8il7Fw3KjMf205YN1PHEOrcLdqGQxN"
    "U43Lqv5bGKF4qu3s73Lq5fLe7Y/0l3kAsTlJOrRUezSXFFZ1aHdZ++CgX0r6dX1T6vMa"
    "gpjuhaNWh09jX7ebvOXVLraesmqxeHBxS2Gz28GX4zI+U3QwGTjwQ8LKy6KKYCsLh8pn"
    "ifYHokqbncE3RcBHd4v78YA/3D8D1rQl1Uk7mj5O5f3wPM6ntErUpj+/i2sdqtSor9sk"
    "mWDanxavbRGD0FQktZcxA42kmE6qZ5ajr3fFWh0Wjf2FfjjjdJGsUSkfpvWjQe2SkyzE"
    "LTSiylUbyXXW1iTT48P8P3sX0E8OltymmZmG/t4/X+7Mo3u3WJI14TtvTJn91RAFay9b"
    "xGNhAAKxd63QTrc1PSyF3HcYUFV2yvT7yDkEhScWwX7PKG6BLM2jP0nXWpX13EdZD7fU"
    "5M9mh/yJ8dzfDr9Y9LUXRXsxfrK+rlnEo1TY9tcph2QnXELFJL7z3pE6frsZMN5c07Ky"
    "wL9+qrVV59w9mY2k30lzhm5Cil9AzU+zOOKP2EBLAMGjeGYh5WPtlekNmTLmYVt4jtl1"
    "IWt4ymJZaZi1Fw04XCLEj9nd1cRCnh/V3fR9gzHiLdoreUHbz1rYVeVqh3ScFQhPVeD/"
    "ODf+9mFrvJdoKNzyFDmbFAIIJylqOo+FEGIZzyg/xvCTwRH1/pAvxxn8oebqqKceb3h6"
    "uhW3r/h12oKTS5BecpKcIyPA4JoatMhhJ0Tmdd+nF8MQjHfieQ5r6+MRRX9GYoV5BXLH"
    "YnuoRx0ONAPH2OFUGCFxbCiCqt6HTkT8Olqvu39mugPtCmPe6dyMDvsK4B8F0UI5h4Fh"
    "v3F/XzT2snfKACWtFq+yMJgmJKgCfkbhRrnt8li14dUbQFAs37/VkEikGoUozSqcc3eg"
    "//eYeJ6quqkOozebQUO64+hBfuIJ/Wx/axCsYkE+gCVCdxRZuNj1ecZsiZ1/+Yqrws7Z"
    "eh2L97UbPas8+0QH4DLPJvL5noPoteH62oygIY+PQyFmukw+lAKo9+TF0lKee3lsoj+z"
    "45GjtYntzBoJq2pnuMxcow+YaZpOGqkTU65YqJ4jxdBp7WzqmyjLx0sG7U9N3eO82Zbd"
    "/cikfD34y6arGZ7KR9Nhpq4njCFuG+kx/VB9aLE6OJI2qOKdE9g22oeGWS8fxBgl4yjP"
    "v59Y2bGugMC7g5L0E3RnRfmgyzLeNmb1ZUIj+x2HOkQBizkhXcqBF3K7Z0lhErE7btKG"
    "AKvhPhaHJZsPmCAyxPs2oTOHlsA94sWFncx4BGfV1hgP7M9HAcFlyp/pz1F5OQnQUL1B"
    "cAh0f9kPxYXZTbSj8qmd+gXb6TYHx7mKqvd8RJWLnV5CtSBFrm4OGfPllDcQi3R9Cl/p"
    "yVWpJeS/r5I+cYCgJpKp78EeOuzwmmV9+ik85r6VATWUq0kNfF3e/gAGf3aZ5ckoDuA5"
    "4pyziCgGzcbI3HdQSrS3p9QNuJzI6OPuiSPAISMmZbqKg5B8Ve509oKsG0vticRZ2oVy"
    "dN/l58P6cK3vHWCoylH/SFuxmyGCbCnu5/Oysg2X5Se/tiTT4V8/Kl2UJTAvivq931gg"
    "RLcQIagBW7IuQjeD0tOoSacP6yGlnGzaQu8RVZePZa/WPGZVIR5DK44OQdHdQL+sZhLf"
    "AGMCsKlgf11bofbhWS7ahOH/yaN26vuapzeEdfsl8bztqs8aID09V5i63Xk3PPqVWf+k"
    "88gAWdB9HOdMpRYESqvDub2XUprR1Vw6lPugzl2FgiR1FMR0MFBAqB3FXY53L6C+vz9k"
    "f+80cPUYpM+ulPNLN5JcO7h8qRVJeyokg0bZggv/iB1FrhaAFE2qpxW1phHBXY/nqNG/"
    "YBwcgtzga5nCRx2hwDwgOp30X/Aeb3GcaxE0AkPFS/uuTOsqkMR6hdOmKeDVO8BXCDwz"
    "xGvUl/ktka4RqoJkDplvNE3oba/o71t2WEIx6KiFOcdok2gNoVPFtJCSmWv//OKXJ4uh"
    "RBuNc+RbceqqqBzC7Yys7npnIP24FwsklF/hLjLQxNfUQ8FGihnc9WS9KMrP/6PNYyVD"
    "y8ozJio7qv6P1N5VI40z81YyF/hPiAqUIziP1yf5o9cCyelaJak4bVO1HrfMu+wcb7o3"
    "SqGXuvYgcbZY0nW7lF1mxc1dNZ1jmZ4TePox5Vj5J8kd51kZbu2UDJjEM29xCNb7Wwdb"
    "G53DVU1akqyw4dXhITT41FO/npiw85Mb63015V9ONKIrKXARimNs0M5JmyMErh5lc/0r"
    "0NA0b7ZubHwOanxDWAu9LKHmwJF8zrSMCzv9SXhX8p/gjlpclAOn6OGzNBiAuR2RWZrU"
    "dT0oimz5KXKDtLBNMCJmj6DwqzuP8Hy53yOSkk70OMQpBjP8C5Kv606JvunmdlbV403I"
    "WUVxyQ67Zhze8Evh9Wrd38Nzh67ZKd5cvP70YQHJjDwnjFHR3zPvZJBna8MJ2QF4xHqB"
    "MJLo/ziMWBy2SAMpRz1aRStWVRciYrrJGxCQXsCL+HHkrQ0mjZyqCyObtL90r8B6nKI1"
    "MFuauaUov37AxywMcVTlsyvKF4etRYLqPmqh7eN4Gs3dc3oFxATo0K/0aHl3yKMKI0Gh"
    "wHb2nAT+qcEylP3N1YVlL0Z8UGc0BG88CWeFVReiRIF6QkrjisbJYruzsnTvduBWF38n"
    "xHrc/3PCVnnWCTOPxmxLqJGvxxtHRdn4kMp5HEM7t+101oFv8IluB6TjLq5QreBJDcCt"
    "sdzebU1wDXf3O1In0NayzWq13IhMx2zVJMjHyg7OV0Dh+Pu6luUc7Pytan+SVPljsIW6"
    "qq9MuaSrtGaPXYNcHLY5Tqbk88pE2wcWRaPX9eu6oPJjRlkefyuIVntRXwjdhiWTsNN1"
    "rNihZwDd6A5rRADe3Y/KrD94Kgg3DMFbjSp8TmUty8NnW9KN4rF6I7id4lvmaftkSVm9"
    "mEgkF9H7Acdi+cheysX4xRF0iXSk4k2R8a2kNOdNHLEm6/BCp4/SoTiQtXC/9Lpr+f1B"
    "VpgYEw2pt/G5kebHvAqr2Ovc3WAqc5NYn1cY67QABR7XgmbvLAPG5sVNc1MJgZiCrN3b"
    "+2Ds7r2Qwacj5z3Sx9dqiZ1p02O45zfG6ayERQQWeEdoYUp2oJeuSxHEzNd1e32nK9S5"
    "WJSwbMDi7u5HpmeNP2uhzvVsSDaEDxjgj89dzk3hF5/qNs2IBvBMQGjHuhkf0LeTo26W"
    "IT+MVfmXw+stl0PQffYZ/K5Wl+m8K3JM/ydEb/spDqVHh0f3dhqVtPour3+hXN8HCfzE"
    "V1xQBRJwRCyexzf09gw7M7freZRXVIDuvLtW1jAR3+cjBxrOZhN7p5MWxVisav7P3Ez3"
    "0HUw3pnxungUgxVYlAlmv2P0JzQehCwh4LC7c4AieDH3c9r7GTK91/hocIHPm97KABgs"
    "0TsQpsMwkhVEU6GpWYWivGPWFXpvSEE+BbwKhD4pCvr8StxcS2jxXX8kCFCIlxqKGHdI"
    "6+rdidRreLz1rJ3O0Ux6bFH5DfuHRTh97Tey23ygrcSjt9ERljAwX6TE9EA4rT3MOdd/"
    "RTw6/vjZYyC4axe/ljWpg1znd7uEcHq9MMRQGEwCtimVoPpLKH1v+zvv/q2sS0t2L60G"
    "g4MIh7Kf+TK8vkeMxE6BXQpAk9/xOKRhy8O9E+3uHjZRIK9zMo0/O7bXDWU8Dv2r9hx/"
    "Ahzmm2KglUPQ/ka2MScnOnXYeXdmybRN2E/VKlazfCa3ZAK7D0pVj2wAbm6i50R2KlC6"
    "bvYN2F3ALvxTOBeD/dm9TXceN0Qd8+bgaTPxsv0KWIf+8g16snsNBqYuSf3/HzT+72cH"
    "oCD/GzTWcSgzBQHRBvy/QWPXLF115+dPXc9c7dLKSdkWOtznGGbVkrUSaKYCUT/K/CNy"
    "IaIlkBhYehRJKmo+f6E+ZH8YGIoHG561kTZJhrbxkRa68YSSbY8nvRvC96z0eD7VDAdT"
    "J5/c2foRx5utzlkBq+ruN9QG8PU2+8Lug497/lS989GeuqfLUjXeP15NHr4VYXixp5IY"
    "uqENSmIpoPxC2E0gRj9+P3IjAcrALcoWUNxA5vHDtgJXN1FsIO/fDaNtem7ObZdGvpZH"
    "onELaWu4dgeL4f6C0X8JxwXrgaTqGK11qK51rRf38lw1wfNYinvGFMSKcUuLzYJqBksO"
    "p/XtVsDhPClhsDSVciNBhkGcAIRRUIS7JZY5hVrhHZIrvsmNfB3cJD7faGvaSOS+2WAC"
    "afGwiTaCUUEGSmRN6RZreT/lPQqNDkUUcwoqYllH8Az/BFWERJMbuqmmegWQTabGz07F"
    "G3/FMdVaEskJDfmFhRGN+MEI2HgawFfs2gGbeWu+dmwlBAwPR8IgMSHup9mQjT4OCd6O"
    "RstcGSFybKp0A7GWeFX4K4YCpNbWIDVtY/BYxQ1MVZELYOKxIrECYzf1gnPDU2c9Dxh3"
    "fR6R30DXmsYb3BvNQCH6P5hV2n5B1xkOhqB6X+RG30kB3JvSqEw8aiN+EDcfJir79LMX"
    "43GbhgE6VAOGz8XDopr9+jvdcM8Frbt93k6sfOwY65Bjp05pTrFS0wCjoyPR8/bNd5Ho"
    "O+ZDOWARzwxZD/5n92gCo0ohUKcYNCNazJg3G8/g30ZNr4ytOt7DNcAWWcwTCZQp8sGr"
    "wutawYsAMxzjHZMoftHlwhlI+e8yj1zS0XwI1WFs0niNSNcvpNuvT9+X7ecv/q6u8YGt"
    "TEhHMhAGZhQR99Ec5/31n1OezA7xeW6nXgedk9iCEjF5GG71KxFYXY6CqoS9jy1WPi+V"
    "wDsqxmnIEyEQiMfjy8ZZhkvDYW3EN456xeHONWhDCorWc8cjGKxB+IgKTo6GnXeNOWTM"
    "C/5Um+2OdvZUiv4/OQMDJXVzj2aC+0vdDR8SHBE71Qr29pvBN3ClPjMf3A4QUfuQwkMe"
    "P3jnJAsGNrWT/QkJHU/Q5AOnGYArl4RrjY/+IUOb7ftZo714cISAeIIbo9S1l3ZHRomY"
    "dKFlbO26Cp9a9ee3zdqvL1/UF0k/TVZYU2HwtiGQnMbZmBj+9wFfVGofHDaTr63Gf4Hx"
    "FxryWHXJ614Ry922mwkmNA16ZUxnn/40SF8Euw67Z9BRs7znRsPIOdis/liajTdeyYNz"
    "Q8Oojt3c/aMOx9hzcbq7JTrgCf1+Bs3wK0D2YCXbZ35+mGOXk8Mjwrb18H+vcBCDGXl3"
    "4uPi8pNXtPraPaTH6kIX+fgHvj+Av1WSGVt8YstzzWjcI9Y/uRGOMos41cFPxJOOQigY"
    "oYnH8zxDm20p37Tdhbbi/Xnam4Lswzr5QbvLOkJwlT2Z/u/zvPqh4lqR3ny2tETmeq3e"
    "+DY9GgT1Tcn4STMw/9m5jBG4r78IO9IeOZn6d9YukbCuyGxmU13iiB7vKYtYwZzpTOQ1"
    "vDR/qDywgn/amdSjlf9obFy46L3lglXvbvDPvbvYGEFEdHP8/TUHwP3SWt37xoahfehy"
    "V8aMh5C9+9l647fcRsNwy0SoPgOHh5HN/elYfZ+CMwWNxs3Wcr2VgOFms/pDaa3MM2mJ"
    "MVgU+iNbhJXgnz/74QIxm9z/W4tSLxBjbHy1Hk+tNzISnSKbaYoekKsFQjbV+zVF2TX1"
    "M3i1rS07919S/IYUOTrkalHcpx5FECM0ujdttYASMYjgDJYqMpbv/fub5JNdQRYHeJzs"
    "9W8tdMRBVDTJaSZj56er4NulEyl4IiyIv/TvwZBddK/U9/DmJ75TI8oE1+3peNJAIEmA"
    "yLTH/Zh2zeuMc1Cgh1Lp3hDla4ksFMNjT+OnKa74KfePUf/XxZqR12T91+Rs+MId4yGn"
    "kQOBzv9ICpumMkmoD0J2+2U847VlMZLW6r4q9/Zw9js3mVQiBjIwwKAr8nDsp5maztUN"
    "SO4LOMVQ0QwfQCCqGSN6U05MZubr33bxAqJ1N8Oj9Xwv13/4GN+fTAuaJgzk2y8Fzi0d"
    "prES+tbXz6c6N1VxSsTgRYLmN8bZM/TkG9Igd/6jMRuWMW3U8Ne5f59pr83hSENYl3bA"
    "pGDt5WS8MbPwkNsPXlwMOer2N1xQ5MNMsPcvljV4y7eryKhwOYeZ1y6rf4yLRCE5yL4f"
    "GcuM4T29vTBsPE2A5kTFvgBH4bB9kshy1RPPOJGt0BO/X/yXDOJwhefZ0XC6M807iF1h"
    "9k2viuFxWV/DLqes0eEx6qdsgp+pHAtQjtwpFGRpYASXjqu5r0GObQ8L0Baghd+N7o2h"
    "0lE8eqI8LpyjCWIbDdRyDL98Uo3HhRVUja2LgDG3OygQQmXkK4+l+rd2zA45zzFb971s"
    "xH1+Irh/E+yaE3ePN/pQxczZa8DkeU+iRgWTpBkfjT2hIQj7flFS+0V417P/9NqaEfWV"
    "RkcMLk3Z4zbZfxkCd5JG+N5W6FkMUmKATCFQ8Vbi+PDYC6buHSzDLruQ+07UIw3kxDVW"
    "K6+gZHzRVEIBu+/NB9fTxuCbkxrLo9o39K0q8fqNNZGw6SuNkJ+JzBb0srRZ036w0jQ3"
    "W2dfCMnE4uZfP0+jmsv6jRoIGoKeowE3JVjGlJYGao31+r2qFkyZPODpCKahhFJ7HhU9"
    "Yb4v/081VEEQKtWo/Yjkf5BGTGQhkkzVMwGDlD+5j4GNZkuANB6QxV1bmtDnE7G6JiKS"
    "gp3CA/02m+0RByUaupDpzO2Nxlwt1f368yyidqC+FIoIqltpn338XMaeEydEooEomZPU"
    "S+YWfl2RgUr2D+PYelq4WOi1f/rPDoJfc6tA6tBWQdP+1Epm6EeSuPkElzQGazAuMozL"
    "TLNjC6bRbJkcYtE9IqgkA+EBLAxxOBqLV4QZcaes0HFa+IXLy2m7c0kCcsO9KeXAcXS9"
    "i+eOs7YQZADUG15jn25wJwPqOg65Z9IEZGFkfbuLaMTgSWPCVRxEaaiecoKJKe0264mv"
    "5AkA3tr/6lti8c+Vnq4oO+2aKnOnIEEcLkcOkIIwQyYrC09EWI79GPtHP9QQN+cyu423"
    "RQYNi66/UMb0ee5//+W/KGh+s28sjmsH6EqwPsNRxL89MsujqGSzr4CUYgO4pZYDj9wJ"
    "uR3wWFBAlxHjC2hFIRQcQOvEmJ6EM44yfrmNxrx21rsY1pAxAGAK+s98/wQ14I0TXhkF"
    "Xbq+MqeCd0GVWMfLJeXgSP1FZCoidlQAIS0EM7yca+dTqtMg7kcmaqAA8xCBsouhM6Yr"
    "pPXvNyaSq3jLCmnYSYAYuWKeIxVpFSr1X2bpVtkjAT401G0P1iEP5O3SOueV5xe1lQtT"
    "oftJBozEYf44FjFUFBrAA+hiwyrhvJrETBl8J/JDbcDJgiWqQXpE2SX5PwLwTA2ag5PB"
    "KhGrWkRg64yxmov3G0c8586ekyB2dslJA8g/aC+f6rZpkd3aVvt1Ca7DDDNDhWsRrObY"
    "IiCE0QFNwTpEv1HDZxizAQXDWc+NoIZJFkZggy0wUzv0djJkWaj9vyA1bPepcGmeK8bH"
    "b9jMoWL62qgPi78Fs0urgpQF6NFiIGiE9CFGSUl7ragRKEsGdJFsXIGhWC1D+YNzeqqW"
    "4diKUhAiNL8dDOMgqcqSAvNmZP2C/S8nH+iDPZRqNs32kgKAQJIsACxyuEE7oAPVs0t8"
    "DobPHwDtvMvQi/bJO20i3FCyBMulllcWw4PpCamGGXQPt+8tCtfJt4U1iMiss65kE5W8"
    "jQOhIp5HVIj0U7aXavRDlC3BWM73T88nvF/TUYusgQc8nul+V4CdHGCGGiumKd+ilHJt"
    "u3wb+T0DBj78U9VLmutKOtzWQHKfAFq9NVm/JWm3Kfldp+3G0ihHaMfvd34cvJ+7mMcG"
    "L9iip0G1fyhmMCV6cChGcDNHVhwiIo6wVi7/ynuOc9QyHt2+zwTEvnNvKO3mBDLORO3m"
    "RDfDlrrqT160Ywhu3zHIPj4ikx0KhcW57NOoACClyf5H0lni6gYElIYMgiCGkj9R0HsV"
    "GX7otTMwMeV0E7GtohqBgkIk3DtYMqZdlAHkBKU8Y1ey6/mCTkcFlYEG1gwzmEYYzZz5"
    "hwyjMR7IOLJPBp1JQnJpT74m29IoXCCJ/z5un7P1MS7isaesSSYuBnocfEC72D8/INSH"
    "yJSWX+HxwKkI4EOCXRJihR2/4g1yzhl4YSWEekEZ1MgjgSLh1+O+a49KBs++MsbinwAY"
    "S4go2FxrmZdsiNnBsWe8IAPJYH9iJk8x4A+vxrNXS/yWIR4jAiEGmBOwdpA08aIofWm+"
    "E2R6mNlqmB9JQccUTN6Jivp9Mswr+6ZiYWJzXCFcIW34WE0KQvDiXvl+tS57PDy3+Nkv"
    "RuHn0auk9nz1KArihWEb4aVGWrJTQBQXNxhl3JGmTcnO4CZUUvIIIhoiRbGqCHV+qtHf"
    "lu+pMAMGsQPIuBOnFxOq1sMg+AmZUI3RVKdf4INrCpoeEavsWddm6EGp6kGUKaJZPhqG"
    "N2lz6zJMT0lRfgUw5FZYlmzrizZojw8HqwQxvAdUWpVqaM/0lNe6PzttsXuUxKOFI8Vl"
    "GseZfe95ZDqjcQYDRXpJwp5fRPyvGD3Ze1REtOiHPGW80wZy+MU8XdCutmUNV+Wm+F2J"
    "fbtwn67zWzXyGnUKzvjoVY1+Ln+cwXXUK8aXQKHtrY8T+JTme9XU6KyrwUFEE8m6E7zj"
    "OUzhzS/ZjjMTjIR0awUpfy9JvjRWAIscJPvxeYRG5OUH+4D9gTsgUEw6ixGv1t9EgVZI"
    "p2ul+mP2sW4QwTU2Xu1RGzD2SLdIWfCWHc95EU2bi18+Fq35GhBW8yL+bToUevVw7JjN"
    "HF4SjcKdvbnavq/SNp9frmc7PoynLurcPejz0cgt3qjjktBxchJYUdnKdZywsXo6AAY6"
    "+N0/VLk11KutyRWwAkFdkVf9xBMmp/5CP1E1nyoaVmBqb1D0VJjpOImpS/UlHI9KA5PO"
    "6IWXo2HcX1qHJpKt21gwfdYFJQuVjj9KdaZYc75UYnS6E9rmhbIHlyxhIkgKaXrG/v4S"
    "TDSafegm7YEI5vqpItE7iCXl7Z3lczv8wH8+r/P1aAus3X1DWtYN5CvoDFysdrABxDSL"
    "1EIQz6X5M0lCgkZdH9/TDgST738vJ88u8T17FEv/zBWXm2SVqAGdQvF3E732FN4bxthX"
    "xQ9BjGe6lJCvPmP60oh1w/k6V3+xln79Jf3qLt2+OEhQirBqyXdQxep2PKk3ZUGZI1Q4"
    "X8TN5Ec2wx2VzntHSKow4hM6wlzdD7ry6HR/IuiL/YP/bt/rjjX3sMGq5rhymaZbjWul"
    "rzCCeMtdiRgmEFMYq6hkWqWAenuyG6cEs7tG2U1pHGjLPquOKyUaODflCGOKxP32cH/P"
    "n0mxQJZDjGyGGdTqhMOR8pss07aD4DPpFh5iwBpce2PiWXtCPnNiRQP7QWpYy7w1TMyt"
    "35HBdLEkMhwO0rKZZwSGJVtZF6MZU7vS/oHcdmjFsqz58B0ka1C1csv9LNR+meZgq4kN"
    "WnXYU75Zi9ym+WUhtXci9euddqjCPF9ym7sjyk4HHaI/cWIAVpFKDUfOgdC0wzWZLBJY"
    "KnsMPphjpSVVCkupZYMpQ6jIwFIKLU20UtGn74yjp8n+kJZXEquqTinqIZIMS7mOhLxr"
    "49azreIzj/B6N3zUjKE4ZFctt9sgoPecWnsSH+uGNTjIvvHIs2Jq5RIbZlBdrFdrHLUv"
    "r38nEsP/pCHwFu5rcpPTXVWmYs49mpJy15lyl46TpY2So4NJoM1Sxa5g1gRj0dAp/dyR"
    "v29LoIsqx6glwH75+mw4q5Bgngzvf0bMCVoipFXWKCGqcZNgM9gOaagBfj8bdLu49cyb"
    "7eAu+iol2z59Nmbl/StNuqrY1zzNeJVwWnv3A4qjSjDE3184ihkrkf6EEN4YoWKir48o"
    "1mPIvUqzO+LZG0i6ZT/KwcPLS8Urkayg8TBYchdY0kCslKOImaOHyYEe1AIVeq4SQORt"
    "zQTIsD1ny9JH3eHfe7XWPkuZPx/EQyPtXEoC1RCRrCgHniMNmkzSmWEEdf9CnVWZqg16"
    "TiZn+9vGrLbDvNfu1noIHphjLXcmI+WtfC9SZWJnHpYox+P/3PWejdd4TQUFBV2/Iid/"
    "rG9QtNHXt0c0itcY89z9np3ldCCKhYczoMGb8LFEnT93VqPxuhd3uu7JZ3vzvIUOfSK7"
    "GFyqpEC/7+3JnmaSq2/q7hyK7W7zWpQ5Hg/vB2sAS/mhRfuhnXts24kAKWLOeOXDClAa"
    "YYR7wGljV9tZuXH/shkttczhJ/BIwPlq8k4FW8EJCQnhOi5Ce9107Uk+9vZeOneu0vX6"
    "/afapE3N7PXO5OOKd3dinDKnb24OZSUDwLuTeBkfVELyrXjP59p6SA+eI0gWkMeFKByX"
    "ASmKo7Ih3YGQe9lY1cxiTX+lIM4A8iRa4XGtuKaB+mWVvj9uki94qyUvLnMGu2q1R1Da"
    "1bh7n7Ud86/WZRGHcngqRlCb/7vlM/P2OAqMae0Oz7m2wZk2LVH9KjeIzt8xX9hcHOTd"
    "+TJdIXRka9YyzvWUCjbZ0M4rBr5mnK534t9gAbcukvRYK4hl37yDzJZqgUMJwGGT/l+d"
    "Eo2KPNnidHrnGeBFYhxVwAqkwVIP+lBNw9jcnfIb2PMu8C/ntcYz0FnKUP7Hf1k2u/vN"
    "hgbQquaYTKE/tLb1AadVZ9B/mMolwQCrVGM4/c8y+eZh0bqBzGKmJGwAs6GovL/IPttI"
    "CPTyGltcqWtVDn5Tss4TnN8RWLZ6jURqlL8Sj1E8eo/vQ717gO8TfIBIJgxodaeRLyc9"
    "gnCC8+rM/y+WC6CzNrUP6Yq8znwDGAC28ULfT/2eDMSLWEO3tpBs8pPnX2FIpOb5Y78R"
    "DdhhNw0hbj6LbrJhq9gRkFfX6gz5a72nKwcuhZftvXHerSA8u+5y+vPEvTsBXhTlmkxn"
    "rqg5+H3k2yG+fznExm8cEg3QTXdRE0XvIxRCVDQZJnqfsfUUYFuMM8pYAAvKvat9cqUf"
    "zr2HyZ7GNuaeHmLUYQdrvksnPakwCKsh7juOjP8xYC1s8MYgnsWuWKektBBcvWZOwxFr"
    "ZPgQjgQyKtKihhvhn2VxXrhchyD8lpCuqhkBAeHurwR/1in/qeY/L4nogbfP4rCeP8IX"
    "aCdk609VPY31dm/aZqm7F+egFhD6fZBna3EEtq00rQcbW7DzOul0HB4scCk3B8ljw/aE"
    "BYqJvJpGHZTQuz4WuR+lgxK/J51nBZHcr6OSc8SMZEJNlDq/oxblj7tvfFZhFgmiWnc9"
    "Y8MUqtUsWje3PC6sdbwfb16OH7W5Xns6X8y5TWHWJdqKBV0LumoA9SZXPAT257+yvGfW"
    "/3C58AeL+GtVh1QQJGkm7GSGATlvPDhX9XzBl6VMtMAoPld8kCV+DuDyImbSmBa5mfP0"
    "pNvAli/Xnrluzal53Hn5i2qDiEGY7RpHDEh4kqqRu38kMvt6nLvC/3EF+hOPpl7kgBFv"
    "caIFKKkWUB3c7LpSjxK6bkmSTPsPeMoXK/W3p+l/n0j1PGbfOA6mVCMgIgLWYfTpgCXX"
    "YLUXTG4Bg4gBUtZ+sUieAmxzI00bQwwljY2JyvLc4ZT8tyIzvM/572TveJc4tj3HmDt3"
    "T9J2M9OSe9FBpv6RkyGtzporGyj16TSZTBRpFp1Zdxugo/KZUrlRT5p86PZjcH5Jr4xU"
    "8IQ3MWUZ8ydZaIZHRegv9ov318KzKy5ApdJa+R8WVR/Bs55ItS+NqSxMTE592D5n41rV"
    "au1fIfQ6T8GMDo/sqwfZpUrUoXOZUCdWNHztdH8qsGJiB563UT05Vvl3LanIYTPCYbhO"
    "y1dxsay43qdY3W7hcyapS7P4ChDxykoOsko7KxPlfc5ZvfalM8+uzwfazyJnANZC+22S"
    "bUWTdsLjTHnh1KzrrmgbWasKeps3Lzaimi4w8KDcb0ui8iAa/t0EOWoAgQU0RSL0Mhlx"
    "ruEj22/3sft54lDKo7tRpnbmDsIogNYVR6V69cEgoQfhtq7ukL2XEMJxd6kCTLdz78uu"
    "/n4vp2+yR6GwRDPcQJlHdZQTLO4dMXQ+b/s+KKHc8swQaVIPZxX7h0ObWqES9f2WmWDk"
    "qcCpEcjG1ROoBXj5yaIgUx/u15v15VZzpIhKWvoiM2xpiWZg7RVTV8nkAdqVAY2wHsQG"
    "le3ls1f39Rl7+JNLmP0g8Q/Iuj4C6B/qCIQX4WFpBPeucAgebtMEo7vIHXONMXErPI4F"
    "AK1pTHnPiUSNZgfyq6fAQaNF/+ccpfIhIJDpxn0X+jV21ofAL8YS44Z/51XRF3lff9mg"
    "Bx3mmfXv4Oy2xTEw7blgL3vUkxAiYsDX6+m+6/mTsECtUpEGA3lMmIqkadNarVZrfDHn"
    "jSxjGrzNF9FTn6Lq+yaUr3YdjTWPx4ReiEedmOTFO0cV8MRTF4p1D54W/Cw6s2sCa/m9"
    "SDPwj22CdlrhHAAjEZ23kJa7gG3m1FV+ftjqkvzudB0Skcblg7HCmCP7WG+ER4/PvDTf"
    "5xz+rgvb3VEMwW/y7e1tooB7nwX2InzPhYZjp4e1376HXdCAjo7O1/FZn35eRm9Gz6fL"
    "YhUa/DqUslVcvq96XO67sUGLFi3S4xHWdqWfNxgvHqLuO+CFSD3jSx6D5tAnyM8aOWrs"
    "pK8E8LKl9+74LDh9YdP6qF/1nFRhzIm2fpTofz/yztznr9NwjfEVMR5BVolH5v40u328"
    "xhoOR2A0bbKqf6omi+T37e2h3PjMQrf7TEY/OLkeZmXG42Sj9sv0rHJLyLtyrur3Jrv2"
    "3gT+q1FFqikFbzMVeI12I8hY3u/5Ji36KG6mXBEaGA9Im5jkHzcxr8KLi5yIskQwsPuE"
    "nnQ/kkVX9vJViF0lYfzYlp7TCoC3O2EBY0WUvSznZ71VJ+7DBKN5Mz/Ty1FGmSc9PN2n"
    "Cgl/VS7zEZ3If5+QEr6b2aw1BF9fz2BzzyNhznB3dHa2Xb/RfxxKfENiyZpnwIoFMB4x"
    "x/qg8/PzD9UfGqC6KDN++w3QPnMxVLp7TB0FMCxDkPjHPv3kK2HrZgIq9SfIQ0T4CavK"
    "he8degPwKbBt+ftDioIW6IH1xKOZb4PYgLdZ/CmtPq18vVrHt7NSu6N4Cby58tIKlUrV"
    "nn81c1Iq+M69va3LxZkMJvie6sjMv62rc9uYe1s8rV/8T9o1QLCMzZu1qnjPJeUEr9Bx"
    "onlWSt7t/S5EJEbyetsFTwoDY+Sw8XUFrnlg2OYN9w67AaU//U0vhauM63WktcCZhGST"
    "2ccDdZnFy+RhgxaIgfnCo+/QuiwNUrdoCbNO+b9IooF1oTmi+Sj7n1MCrNH5+r6fmD1N"
    "r5VOFDM9wSIbkH8jM3sHELVblcqTzLoIbviDy/oGZ/gp50mQeupTKX+/qULuCP59EbrD"
    "o0QDSxHCpUX0quI0ig4OYbCHMoTNBBFRgJ0E+17DtTcS/inlPio68WHi/4zAJcAXnWj8"
    "8wIQdpjW0yMgx/v472DpP97DYGaGvb7gvzFQHaqym70+wf0rQo7ytKJn9Z/Es+5ZOSyi"
    "mNkyEP/F4febh2e4MSftPxFkb49F4v7Xc1LJi8Ig0GRpoHyVcxaQFcrKpNKkFIFISgtm"
    "wqU+ifG1/KZovlAK+IviduN3iCZqQ/ZTQPuSHDXLd9SaKwWDNmWWC3yGvPSq1seRXP37"
    "WqsC3uutVLLZr9LHFbH39SqZVU8dxTImmVatwXCwvkmreu6vE2YcTcc22DfPkdusH246"
    "C9u8o+ynOG5kiHQnacY4bmoFikYCWB8KwUGuq0aOPvNlTtCrshOF5XcvEq9HqR0ePgxE"
    "/zAzZY4dX7KDyYgcb6LbdO/n1c2+8wnllY6IGHDmvvrG69sPNSPLUYXGRM5t7BY6C4Ud"
    "sHC6ZhyA8rKXV2gVEStP+vOf5n91DUYVYTOK07W/NwRNtYymItheNyDI3wL55kz20ERj"
    "fPEciIlD1k18YChN2nVO0oxjRl7Ts78spr6RW3dGHrVblAT61t51+v8pqM7B24Xt3sqm"
    "Taa9jdAcGN9FLHeGQ3vUcyPuPBYthADP0cuaAgpYYvUcyptIVzuteZx0c+NcAipICJOw"
    "mS7EzU14WELdtlgVJlrYwTcitITiKR6zIJgcfrJUsLMEAffhVC/bte+xai98ML/iU4AK"
    "RwQWFs7BRKWPRD1XYxcIgNimVdN1KT4/bfyi6mLrzPEqSaHPWdNa5u3GzYqkCxPWK4/S"
    "KBsDh445hHiqTeOv5FCVLlzZQGcHaUz4T8/HWGjoB+wXn8H7GR8J6tfumf4OiP7RmJ4d"
    "PY93D1M90FSVfV3fT0QNq3k2jzh/fvniYg/yhpSLml2Rao5GgRASkUs8b+/uh6eHPeNQ"
    "osaeGRUKVfQddpT77JBS14QtJNG8fGOStqLNAtho+hef2YPm3xAZ5EsR37kVx+Mmg3qC"
    "9RMpJ0/fPcVd9vJRiwbP08mi/GvfOcm06gpZBbSzX5L8V/84fDxj7Rr99YlX/eiXZ5g9"
    "LBLPe7M1stPtGZiXse+NCSeuVq3jH3lIBHJiVJ5RJOj5csHPvNpEWYtfFzRQT5mMyHjB"
    "hCAjoh4JkBYBKRLlnaGHqb1pWaKI/5OjANs62a9/Fb5+LPrXloHeEtToVGNlVcI+xKKx"
    "QoFK7ON2x7N8+Tws3X1k20tIFUD6rwNuO915DyuAyfZPUlKNGAbr9u4ecrzclbTYA0W2"
    "3UK798O9muMjUPLtw+0gmaEPSUeRFoehe5MREg5XpCnQWUCbpyZuduy6F/2V+RPuBs+8"
    "sfjqTGbDZfqMB1KFPzsbf40qIx8EiIVW7zgqlzFN8XNB0uWJgVSkctloNCZbLF7crG/U"
    "ItxO3Yzy/LPoasD00IEqu2ALhXdnfpfA4o1nJ/GXoxQPeVoqBiwb5/1C8aVupB44U9BR"
    "tOBFU5Z2Ft/6KFqIopdX3fTx43i2UCPiMgokzDR3CSMIEdsyHhYWUcyhjCscxwjKqPUb"
    "NBRTwsqyLHJFBHuIVHvWb/AZfhHQeDv3wfvbfreFMSEe172tAqBZdgqxLs75fUFPqAo7"
    "byVRVwNrxkN+Ef++z9yDFx3s87QZkSxGARVPHlYrDFuG7Acf80h/H7d847oYkl/DkKSS"
    "Jtf7RQbMWT5wnng0NsOnqPrdjQmCGDgRE8aacGsXmeGvcXShLpJ2Qr89L/n5LFbh+2tP"
    "ANp3kzDxZ/pJ/IxBQLMsbvlKC0wea5W9hiTV+IcEYlP9dnCLeWI6GD/FnA5b/zsmGqa1"
    "KAr4g+fxOsbkqzKZJVX+++Q7lJFNuHNuu+O9Nz4Vpkdv6p03zwDdQObGxDuUSGtp+DUx"
    "PvbzWsLxQUSxvgT0GMOIlaIliMhf884VuNrzpqa0FyfnoBiBTCtm/ZpRZZeIcXtzEv71"
    "Xh3XPzW8HiWSqJaFXR2LDSsw9LpdzWCRSiBBLPpu7qHHbZLmf184D2CJe6zlXscMZOzn"
    "BUrw7oPBgPv+adb1SqZegooZRE5eZF1dizI4LjM8mNsdyieC1VRdbaQZCwBloZexPOPM"
    "zHc335jh6HlwcAEWFOW0oy5J/ADGip9ZWjpdIgbMNelYx7M8dJPVMlj3rkKpMLsWXD06"
    "jq/oaXiLIpYsPu3XrzaDi+yg2xCsIGpY+7O8lwXstIXu0e14ZAuL35cF+9QYk3mr9vi+"
    "UpJXV88TgkhkjUDMOwHV9dbTJ1Dvsb+nvx/T6fPfVfZyoTb4ujQK4sxj3gnZ8v3D7o4i"
    "9h0HB5rp5Z3jLdIj5cCBVIwXTU7Y6UAx5FJBr3xMR93JRX/TvfyiIS8in7B677uGmGHf"
    "XIK/IDUW1PplwTTz4X9HAVBVXV2DMyKbba+jBQAQi4XZdvHJxImUfFg5X3mYsy1M2lyc"
    "CGpGfw3rmFcJKYlvAXRHzyq0t1g6w4UFNNqUUMzSa8F4drqS5oK+SB667MQgc3SqcGB5"
    "flonpbsdt4MMFDkXxE148jdWzgeRvVAqQG/NMP0XqApmHRWEEI4LdzkRX88f7Iw8Mtbk"
    "DW3tnO/yzl/TQXjcJQGkU9IfUtP1xdVoG8/C8AoMVNpyNGm/3hV+I0aaZDepuLuQG7Un"
    "w/LXwKEDBHjxJrxIHR4FR/1JnSluHIC5oszBv53W3cL8j9XcXNgLijvMWeRYGEGCmMZu"
    "/44BlyryOYRrXZBVIGusPHv0XJ2ODHvxGkEksLAc3Lftt8n0vvKKaeQrsWJiEKbFE/DT"
    "3EEN/DpCIydCCklYDTeKwcQBPVx4jZvUwe56ZjHFkWBAP+UeLj57B4WNkK8xybayBrcp"
    "v6laZi6NvfLlsCeICtWMfmftzfvqjkJLMteP4zR2t7xBmN0KT5tY54FrYREsza+liaOY"
    "czriRbA+lowXRw9hiqSK8tEnX8sYz+2/f9rwuE3B+1shnnlmv5e2ggD+S5reWv12LgkE"
    "WY8rhoopmuTT1J5awwrMc3ij77UQGNahgbXLxlIhCCFGSBy8Nt+z4xufxRodz+VC432r"
    "lsyMZ2E/vWTerYosJIEWGJOzGQyirkwCBVU1TNPNAlMwp6n0zoL2w/qp/raIvDbWnjzR"
    "ce0+6I3Cs+8/HMwA0i8Cyqz3Ky2nESHGyqH4rfdOQE3q01SEUGzD1lUZiUKWTRQxP0Nk"
    "ZEZspEbTgQxf0b+IwMcM+Kha+N6n8Bpo4zdfeiaeB6bD1CngDpKX+snrK7VbcO3qAazE"
    "VKli5SDp4KBUiGEt6XZpg5uJ7ZAMGHnTu8ef9ZAV6UM7BZy3zvzqjwqMLF9N5y4YooKo"
    "Dsrgs4ffiMgTByrmo1J4gwijcDO4Kvn4tHLBTHm8OsrEVzW0bhQxTYx8eArsA7WZOdXW"
    "s+NoZJ9FTe9tDcdWzVyDrkmKoEw04sH3UikZhLmqSShv7zbd0wQ/KD5/x3P+OUg1bweT"
    "61CAAmxgbSVBMHJVYZajG7910PFnBziuRzFp05r/+WeoQkJTHJQjBhFxvIKRj0+Ty9KK"
    "8keYFtiGBtRt0u7Wes7QaVkeZslAlbkf75NPGsYbmqGmzBQQjf1r967KQtXpQsUH5iDA"
    "iGTAlXwW7NBZNR5IUWu+DJGV9oDX7xoYvv7JXzw0DdDS0eXgKREP8La+Dpm9PXthc6pc"
    "VpwZIkLNiUb7P9DffCcVqV7GuTXrh/4vjDRzy3Qy7Fdi9VyqUT+DdUNHpxIZcBmL56aH"
    "zo8ajld1ReOKt9R/tTYSzisa/ns4ppRR7TSTAkAJQuBhJUNEUPX3FwAe6BD4TdNzUzt7"
    "RT/GQBxDo6Vc+NCIv3jWah52wrVEA78GY1G1TDsKqxAVhGrJZJf5U4igvoVqM/z/NNZb"
    "YA9VuR3pv5I3XlRirV3Jv/5Pg3t3K91VyZQQSJ2BN7U9W1qeh5bMuBWAnhCdOq9k+PJw"
    "dh9b5gRA2np5mRHOKVoaE6T+pg6MnzQcg17py5CENQaaLaiW4H2vqPKSJ5XACK6hOO0E"
    "GKP6T480vAJOs59fqoDIDGb7CWDxFMFDT9XzF88t88TzCrz932p896o2vR3ZEgIW2OOV"
    "HzcdDy+ZcZ+gARqqnFOy9NnQt3Jmo1LmB5lLBEpTKk8sGo6MQsfV8RYZEdoz/6zpKU5S"
    "pDHfMBKYf1kXXPbODl7yJyXaDaU17EEVwSHqQH1y6bQu0WAti2b8StY43ftK90HUITic"
    "U5r9nvVrDZe9q4OXXVRk77BiM0SpqRnws4ZnQMc/RdPo1FGRcEIhu7UBmUsESlt+n1sy"
    "dI2Jo47FSNlvW7m17ZPUzIW/DxGIHVRKwqffUuW8MyOaezzRWG1/4ZeZXSRMElloNZTI"
    "wkdeWyV28Pnvt+jtFFwGkjzCcFHhD23PHbFy4gRZqjGwTIRzS4afNUJnoawhcyZAGto7"
    "vWDGzbRK31MW4bfNmH7v6cmI+i9AvaV84M/LnHdmRGOvJ0pF7MLT7aKClRDuVJQP/kWZ"
    "O7Y7frU1pquSDXMgQun3ym+bjjOLBYbHMVWFEJU6o2Ay2yswQ4rVqP3fa+HEgqU5ife/"
    "oXBla3Su70LDGhisKxedU+CFFxVpDSg2dREvdLbHIrxUE42qrZSL8MFXVOjtMLRddkKE"
    "SqDB1L7fHwZoeuWkgqE36WORkaWPIHMCoK2wITKssONP8FGC2rLHKb9veiomG/a/91As"
    "wGueUUCMoupDiE89Qfbn14wv9VijNGueE48XXvbEAkN1xWSAizxQMcLvm549zk+Y7hsD"
    "K6xwbGRoZ9APYBZa0o+9jAaJeUrBjvRj33/DFCgI/CH27HJKlIV1Cww3lHOPj3jYCZa4"
    "5rFZzv9cZJcI+Iby0gsKHL5MaLUTuljgdUUKO51yW6wUxknPSE2ALhFOKUjQaDOwn2Ov"
    "TGkAnuBdPbEgE47RTh0wN7Qdg6qZCK0YgVYMjz3DYKtK7JVRL39+Hexl8MQtz+GHwRPO"
    "jhiuL3xUQBmdC3B9y1GSiVO6RIJJW5zkPQuFTAkAR3AAHhvJhH3VDCFNeEtTMzMfwzno"
    "qginrzcQK5JK2ByzBu/Dlj7mVEsUkZmKT1XY0va0mKQ4SJVjI6GawTZ1mYkCjDgAjXBU"
    "ZCe0l9IOrDfFnlIG4v9CSFetlGDdKoH2GBt1oRe3hGAFpKWccCR0V4VWHH63kFucaqM3"
    "tZWhCcp+U7/WUQWhS2CPZqtZbbYEgMK6SOgy449iUsKCdzrlntiHKsF5X+l+SNZQsNBd"
    "1bDwBV/UEoQCHlb1CMUCNNuEI3eB9zoS+GPs2eOUNePMqhQCSXSLsK5g2dF04/oLFgqZ"
    "EgAtVY5JVKWhcRwUSnAA3hHrSLulrGykCEQGVDU7i1pK0KBuFyyZiALAqB9gWOHOWEcy"
    "V/dfngM6BdZHwi8aSkdW4phAlJX2WQo4lCMiCc6ScRIrlOAkvC12DHmlJyPlvyMmf878"
    "c4tkoxXdx5m9kAiOQLg9djxeLIPj0G3q3D48CjSehXWnyIwGEJgb1tiJMwAhbOa2WHHo"
    "DApO5glZocqliAzvbazKtnjiozQ1A9ZYk7mGtZkSAGURVluZsOjNAE1VtjvFZi21UkGT"
    "ioQsPeClgn34P2MbHAlsj5WGjt//KPi3go+gnJHEtRSZEABpwkQJWGOFmAeHAJW0xBLu"
    "jz1R1k5/AHQ0kS2Ly1vMUDK5p6lj+j7vR/oE7v/40wjXWiuUgCbZ8V9lJg9ACf3WVlqD"
    "m2Bn0lLh+1xGO61mbkH7QcF7HcmzFyPhSnouer8IfBgZW19wBArbnU5a8usUVlhDJWND"
    "azKhAUDYoB4jk/ZOswL9ToOjJYOnwYgTMGNECmFpxkCpYtFYaTVC7T2EXn7FskUiQ9xw"
    "eAeSmaMhQYZ9AFZgwIfqwK5Ixj3AlNAhqMfAA5M5ueYZczoXYLr3aAiZXd3ChLH9NOSy"
    "1ylOQ93AdD9/zmlmhDgzJgCSegQFimVDfTDm5t/tZtvNg+ze3qA+FGIolU7LsjVl1h3f"
    "xfpTeqh0RbQafrQVeRbuZ7Kc9jnCdOhLCPTrFPZ6ZV3S62ncHBaBnjERrulqAnO5/XOi"
    "AVjCzU00xnt/hO4pSqcRChISgsZ3Agr9XsOQkBmkVUbJd8x9yDALnDIGiTFaKBhu/M1u"
    "fvOD+3lgex1RwUSSTIWA4UFhx911brxqN8vXVHjYBas58Zw+2i1/SPozZkq/ELL99jrF"
    "jCMx0yhAQaBDhKbqjOoC5pJ+Z10ACKFTb9kI3SLTqt1OPadrrBAxcRmwADVVOoxMuwuw"
    "Kgyo0vBK11x7YLOmAQDGCFf8zz1cfcVOrBXKlYg2DqNQTGYot2wMRSgQsXdnk299fhv3"
    "313j0U9bi3cZ8QvMkwlwIPSbdrGq6+QKUwSsiYRuE67pMPRc0++stwRrKDyjI+J5nQWO"
    "KxhKTM+Bq0BBxs+kgnRQqPLocsQVa+20Pk8IHtc/tD1fHGrzzVpMaS5PM5VsCAABvBKV"
    "LFd87R6u/PFOKtUgZtvecYws53izml6pALBX69zi7+cOfYCoYIgKwpU/3IGqcv4zDqPd"
    "9IhZ4PzVeYoDHgz9lkWoTRIKbCm8s7fEW3uKmaHfWdMAwmRU5U29Jf6qp4TXoNbP5FEp"
    "k9tFQUiEAozpQoAjyhGPr0T8Y3+TD+1tUpmzssxs+He9C86+G6/czdVX7KLaEaEe2jjO"
    "NkdytjkKAJest1vKHGX7uNrfxdX+bgpYqp0R11yxizVHVjjxnD6adYfJjGNw9qX4bNDv"
    "VGUgSuhlWc0Q/c6KALBAv8KfVAv8VU+JQa/j9vKfCtPZFmVmthmE5CGj8Fc9Jba0PJfX"
    "YnrmojQzI55qa6A+2OY3P9wR2pIptIhZLys42xxFmzg5YcKOO2IEONscxW6tcbvuoqQR"
    "xsJvf7CDdQ/ppFg0C1t+PYd7O5/0eyC1YnNJv7Mi0z1QBp7fURjJg7HMfK7kdDHTz02d"
    "Oj5ZY5k5Oqsz0DzDO09UFG67oZ8HtjeJihavSoThBLN6JI9+bBq1jEwpUB5iVhNhwt8U"
    "Lbu2N7n9hgGiouDdwt8fOvtP7lCm34MWAGmWU7cRji8aWhM08lxoGKDlleOLhm4zNw0a"
    "J4tUzdeFgI+Vu/4QxtEEp5RSIKJXKnjGr6EQBIfSKxUKRGiajalw161D+CQ0s9D3N9sK"
    "wKFOv5lJBJovLPUogBFoNWIeuL+JjSSUJ88AgagSfUAVGwm772/SasTJ5836kqeHjJhX"
    "C43Zvv1ZEXYRIRPqlpanmLFUxxQeKBrhDy3PwDjDRmYFM5qIM0cX4GJPYzgO2XwaTvc2"
    "MXu1jkEYz7WlKBZhj9Zo40a0BDFQH4pxsc/UPc4mDmX6PWgBoMmHNID/Gm5jGE1amCv1"
    "bqaf6ZI1mWSNDeaomcxCM8YETBLUXM9N/n7SqX1jhcCIuo9ws99BjB9HvczAPc2BAJhN"
    "+p3u92WJfqPZ2M8Y6BL4Zq3NR/oNb+gp4UnCKDP4fGVq50bqFJkuREIzBgN8pL/JN2tt"
    "ukSIZ5GO5uhgOjCoDwk/HRHDgw5sWFuRiDv0Aa72dz0oDGiTHb3a38Ud+gDFMT4A9VDp"
    "jLBGwv9kIC1wtuXBbNHvdDz8qXNxuphr+p01TdgrlBA+tKfF1qbneZ0Fji/OPBGoPEm+"
    "iRBSLhs6fsfg/T8vTaS4peXGJFLI3HWUVT+z42AO4BWKJcPy1QV23NOgUDSoBt9/hOF3"
    "/m52a42HmFX0ShWAvVrjZr+DO/UBIsyIdiAiOOdZtrJIsSS0254Fq2af6VE7Q8wG/ZZF"
    "Ju33J4RswXaG6HfWMwHLAl8bbvGdWptuI0wngcwSbLBndBT40PIyDR2/r1q3Eb4x3OaN"
    "DzSmlQqcJMQx4EOzhjlPBc4A1CvGwFHHVrnxmsEHvR5huF13sc3tppic/C0cDk8R++D9"
    "EThqQwVjwmdLJsow5+4pHgj9pll+H1le5ikdBQb8g7MBFSgJvH13g8uH29NKBZ4P+p11"
    "X5gSyno9YWjCdFZsgEGvbI/9hKqNENS9qgjDXqnPgA7T7ME5LwbyfsEbghiEdtOx/iFV"
    "lq8qsXdXi6gw6r1XGFHx28mOGASb/C6FCLRbnhWri6w/oUq76YLDyC+Qiyw9Euf46w+E"
    "fiHQZsWEfR7v0afhxu0uOPGmY+6mmEv6nZOQZ+pAiWZwFUUYUmWybs8epccKFQlaQ2Ga"
    "n506Ug4NKC5WKp2Gc8/vxTl9UEHLiIo/4hLc1ykIQQB4rzx0Yx+VDoOLF9i2mUfMhH4L"
    "BFosmzDTYrwx9ansaisMOaUkMiPemEv6ndM8gOmSS5p2OeBDKXA0jggVwib0mtCAIS0a"
    "yg5JJqfFDE6NuYIRaNeVE07rYMc9vVz1072Uq9NP5RWB2rDnoY/p4YTTOmjXfWjAspD3"
    "NeIDSBcy9yrWTOjXEmhzol4fQqDtAR1tG5YF2s1MIpAllGG2USoyvuEVugYZuo1hKE4G"
    "cGYIqppcC7uOdPvitvKoC/tQVa75ZT9R0UzL+dRuKuc8uodHXbiMuO1H2G0h72sO0wAO"
    "Gk6h2xp6zCTt7CQ4/wb81A7A+URmBIARqHllp1P6klTH/RE8raGm+u5Yx+SwZwXZUJNT"
    "JvE+CKWNT17GijUFfvKN3ZMfnhoI9YJnreDkszppt3ywaSWbjJcFBM1UWRsJpUkchhbY"
    "ESs1n612dplIex4b8tgeK9E4jJ36f0oCa4xM2DVoQbHQCTL7XZIk78Qtx/EnVymWzKQh"
    "pDSEePzJZeKW2+czMnVlCKlqv8YEATCe/zf4E4TtzmeqIzBkRADA2J7/fsKhj44Qa10T"
    "TT48JMe+EJRWwzOdugBVpdXQzOlWWUXq3V8TGcoTeOqVZHaAC519skS3mTEBhNDye7vz"
    "k0pIAdYVZMEnw44L78Mxmo2EuVGYmZeropqdGdwwJgyYpY0djRYcHU08p0oTx992pyNj"
    "xLNSbzDriUAHA4Pyx9jT0vFLMkPzUDg2MnQIyXiwrNxBNlYxMWZCcgfSDmM+kY29TiNT"
    "HQbWFwwtJqDbJGr1x9iNybPMBua0LfhMoEARuL3tJ5z8G2KpyjGRpUqYIJwJFSb1/anP"
    "RBjwQVBmZjurMpLWnBWM7PE4TQkWEI5QR3BMwdL242taYYKwcnvLh9mAGdrXzIj51FFy"
    "ZztkSo2XDJRK3D4rHF4ws1rQMyvYnzCzdh1q9zEPiBUOjwy9SWrveA7ANNN1W+wpZMy7"
    "kikBYCVs1F2xn7CoIp21flLBhl5p87zOybHQnJFLgPlE6rg+sWjplPGz9ZSQyrut7RlM"
    "QoDZuYOMaNApgqoEt7U9jyhHDI+jK3lC2vDJRTutnu3zipFQ1UIvZD8ozNwEIFv3kT3+"
    "B4J9f3LRUhQJI+v2e10J9Hpb7BjWcHhlxQEIGRMA6YSgG1tuQnpNpe6pJUvXfBT4TBea"
    "/CcXAHODsevJyLqCNiqcVppcG1WFG1uetipmzlrSHxiijOwlEDa0KMJ1Tceg6ogjcOxB"
    "L0Ab2BAZVhjhPq+TDhSdD2RXSQ04GAMgK8iaQZDS4Vpj2BAZ2oxv/1tCVeF1TUcxObCy"
    "tK+Z0gBSe+kPbc9Or6w1Qnu/96SZV33WcEY54vahFqVpjlmac2Q6CjCT92dQkxnh+GzY"
    "fQLUvXJW1dBrzYRNaiLgPq/cOolfayGRKR+aEjas33u2Nh3lCRo0hoxAeFh56hFh84qF"
    "9o1NevnpzblL+4At+HonuTIASa5zyxHlCRyAHigZYUvT0e90pLQ3S8iUBgBBItUUrm3G"
    "PK2jMO6GGUJbsHNLET1z1OP/wJAUee5vtywwvIeoAN3dhtqww9oHuwTSgp+OTkMUgV+o"
    "xh8TYWS92djYmFD/f24pojGB/Z+aANc2Y2qqE6YKLyQylQkIJF5T+E3TMThB++M0I3B9"
    "wXBc0fD7pqO64N5V3feEytC2qodSWTjh1BJ/vHuYYjFpApKsUQSMgVbLs35DhVJJaDQ0"
    "Q7MAE4zs78JubnpInVSyHFMwtHR8VToihLV/3YiTFnYZIooEWXvEKKHib2vLcXfsKU7A"
    "2DGhddNjyhGtCeyvecHYZzpSraZMWs0231WBorSanpNPLXH6WSVqNU/c1pFJW3EbhoaU"
    "Y48rcvrZRdotj8jCr3vfCsCxP4+z9/OI0ANQeUw5GpnSsz9CuBruij03td2kpcILicyk"
    "AqdQQuXUHuf5bSPmpGKJ2jgqVpoW/LhKgU/2NydsxDAvGLFPxxBolvY1KZhQr2x8QoUV"
    "Ky03bWkxOBhEa0eHYf2GiNPPLlMogE/biGXwHrLgDwjpv8LjqoUJy9I9oXL1t42YPS40"
    "9MxSbVWKzPkAAERDIdqvGzEv6i5NWBhUVzilZDmxaLNjBqSmQJaQyqSkSvGMs4uceEqB"
    "uB1eiCKhVBbabcXFZI/5ITN+lVT9P7NoObloqU9SuNZW5Vf1GK+BprOITAoAD1RE+FU9"
    "5t7Ys9yO3wDEAX1GeFwl4reNmI6FTg3MYvhsfyg0G+GEL5WSX6mO/E6S92QSC3jqp0jV"
    "/8dVgwN6zwQtwIsC98bB/q9kdNwYZNAHAKMb+MfY88t6THWC7Kn0YVzUUWCZHd8Wm19k"
    "lXP2RXrCe6d4F3wBCy07p4+F3eMYWGaFizqKNCfwPY0cYI1wgBVZ6FVPjExqADBa+fej"
    "WptndxbH3ejUDDi5GPHwcsR3pjlwYW6gyUwAv9B2yAEg41rLWPi0TnT+JZYFBlR5UqXA"
    "SUU7YfgvTVb7Ya2doRD1+MisAPBA1QQp+kfnWWWF9jhmQOptfXpnke/W2gtKx0FDzWKw"
    "Z/Fjgf1+I2sQ4GkdRYrCSN+K/d+Taq+/asRUM6z+Q0ZNAAgbWQDuiT0/Gm5TmcAMMMCw"
    "V86vFNgQWRq6gDc1togmv+boWhgREJLP4LiC5bHViGE//unvCdOrflRvc0/sF7xOZSpk"
    "LhFoLNJ0y/8dbvG8ruKE6lYbWGOFizsLvG9PnYrMyfDvKZA6ABeROr3YMCIA5l8XEISm"
    "ei7uLLHamnGdfxA0gppX/neohREy1v7jwcisBgDBB1A1cFUz5veticN8wRegPKuryOGR"
    "WZDEIK8QuzFttBf+qFxalyqCJ3Y67/H01Nl8WGR4ZmdxwtCfByoGft9yXNWMqU5QI5Al"
    "ZFoAwOjk4P8dahGJjKsBCtDwcGzB8pSOIsOT1GbPBUSg5Qw7hwxW0orA/Jrty4jSXxdi"
    "L/MatTCEnn5P6yiyvmBpTtD7TzW0tfvfoRaDE2gIWUPmBUBqU3271uKutqdkGFepkqTz"
    "6gu6iqywhvY8nRIKRAYGGoYb7ytiIo/zLDizLLXLq2KNsu2BiFpremO7ZwtthRVWeH5X"
    "MdDVONyvQMnAttjz7VorA0lp04NRWGjlbtIr9fLf0fZ8fbg5qTOw5pXTShF/0lFgSEcb"
    "i871JQJtBz+9tYy67Nt9ixGqgHiuvafIUNNgzfw8WwMMqfKMziKnlaJx09Jh9KD6+lCL"
    "O9qeYkKnC80/U12ZqwUYD0roFPSlwRbP7ypPWFgRBjDCy7tLXD7cou7nZwyT89BRVH5w"
    "U5W7dkUc1hvTigWT1fzPRYhIhEZD+PEtFYo28QPM8famDuaV1vCy7vLIROpx1wfscsp/"
    "DzYpTmCqZhGZNwEgca4AW1qO79VadE6Q7JNqASeXIp7dUWRonuwwVSga5Z69Ef/5my4K"
    "BYf3Gl5YaBG/BC7noFxw/Or2Mldtq1AtKPPRrsACQ17Z1FkMef8ThP4c0GmE79ZabGk5"
    "KiwO9R8WiQBIYRQ+19+c1MEiEuK1f95T5qjI0GB+HIIe6Cx5/uNXvVy3rURnyQVfAEnN"
    "Lfl1QJd6jCjOKZ/4aR8tJ8g8PNC06cxRkeHPuss0dOJ0aUuo+/9cfzPY1IsIi0YAeKDD"
    "CL9txnxruE3XJFpAXZUNRcsresvUJ/DYzjZUoWCVvXXLW76+ilpLiIzHu9HX82v6V3r6"
    "tz1UKzGf+VUvP7y5SlfJzcvpL4Q081f2ljm2aCYM/aWn/zeHW1zZjOnIeObf/lg0AiCF"
    "UfhUf4OBSbSAVCK/qKvImeWJs7ZmG84LXWXPz2+v8sb/WU1kFWOU2M8z9yyBy6vSjqG7"
    "GnPFTR28+7srKBdgPjI80uzSs8oRL+ouTapxpmHqTy/C0x9gnnyps3N5lKqFq5ttvj7U"
    "CoU/42x6OrK51xre0FdGJM3Qn/vLeaW34vjCVd28/surEZRqwRN70AwwVmYvRn92XlFV"
    "ujvaXHFzB3/+xbU02kGj0vS9c3gpoTT6jX1legwTNv1wCl1G+PpQk6ubbaoWfAb4ZCbX"
    "otMA0OBx/WR/nZ1u4lbLqWR+crXIc7pK9M9jYobz0FP2/OeVvbzgPw/n9p0R3ZU2kQk5"
    "As6D10Dk+RUu50h8JkpXyVGOPP/+sz5e9Pkj2FOzlKL5yQC0QL9XnttV4onVIgM+jKzb"
    "H0poYb/LeT7Z3wxVdfOwvtmGrN1w9qJbtgX2eOVdyyv8dV91wrxsD5RE2B57nnHvANtj"
    "R2keJ7PYJEFobVfMqx/1AM89o581PWHSgXOC83nGgADGKJEBjFJvGn51Zwf//PNl/OCW"
    "TioFJTLzw/zp1Kk1keXyw7pZHZkJJ/6kzWg+uLvOe3bX6MvKbIoZYlEKgLTeuscIlx/e"
    "zbrCxLXZ6YO6bLDJX+4YomueRzNZA81YqLeFY5e3uPCEQc5bV+O4FS2WVR0Fu+i2f1YR"
    "e+ivW+7aW+Daeyr8+NYOrrq7SssJXaVEoZ6nLUqn+H5idScv6Cqxd5KDpSzCnW3HM+4d"
    "CD3/M9r0cyosSgEAQQvY65U/7S7xT6s6J1XxPdAhwl/uGOK/Bpsjo5znCyJhiGSjLdTa"
    "hnLk6Sx5CnYx2mCzCwXaLuzLcNNQjILPxAjz2ug1ImiVz+8q8fFVnZPWk3ig2wiv2zHE"
    "5wbmn55mE4tWAEDaGRg+v7aLx1ULE0YGUnttj1Oeee8Ad7TdhCnFc7peASOKV8F50IyM"
    "uVpoiChWgimgOv/dc9OY/zEFy1cP66bPGto6frzBEZj/h7U2L7lvMJPjvmaCzHYEmg6C"
    "t1953+4aZ5W7KEmICuz/4ARoKqyJDO9ZXuXF24fwjHaani+ogkuY3grZbRW7AFBCGHW+"
    "IZDQgvCeFVXWRGbygwTY6z3v3V0jRiku8soPk4FIxAFfXqED4ep6zL/uadAxyameRgUu"
    "6Cjyxr4Kg04XtBJqoaNuWbsW6jkYhUGnvLGvwhOqxUnzSzwh6edf9jT5fT2mAxmtSVik"
    "16I3QT1hSMP/3dvg1/WJMwRh1Mnzur4KF3cGJ8+iVoFyHBRSP9IzO4u8rq/MwCQJY44Q"
    "8/9VPeb/7q3TuQAm5FxA1h57li70Ig4WFmFIlbNKEV86rJti0oll3KYNBH/AgFc23TPA"
    "1pajUwTHot+GHDNASjMnFy3/fXg3XSbY/RMO+ZTQFeg59wxwTcMlBWmLn2YWvQYA4FA6"
    "BX7TaPMPu2uTdmJN/QHLjeGjqzpZYYWG+qWxETmmheD086wwwkdXdbLchDZyE3r9NdT6"
    "f+iBOr9txHQalgTzA8iaJaAByJh/G6p8ek0XT+ucOI4LQUPoMcL3h1u87L5BIEj5hXAO"
    "5ph7pM/UkIQXBT6zposLOoqThpCdQq8VvjHU4uXbByjLqNNvKdDIkjj4Up8GhAf9zl01"
    "bp8i1Jfafxd2FHn/qk7qaM78SxgJz+OBBsr7V3ZwUUdx0kPCAxUj3NZ2/M2u4ZGDZiy9"
    "LXYsCQGQIs3Q2tZ2vHHHMG0N8eWJHlZEyPt+SXeJdy2vMux1pBV5jqWF9LkOe+Udy6u8"
    "uLs8qRM4tfvbqrxxxzB3tR3lJeL4G4slJQBgTKLGcIv3PlCbNDQIYQP6vfK6vipvWlZl"
    "IBcCSw7p8+z3yhuXVXl9X3VSjz+MZo++94EaP6q1FnDk3NxiyQkACA+vNxI+ubfO5/sb"
    "LDOTDw4VQnjwrcurvKavyt5cCCwZpM9xr1de01vhbcurDPrJuwrEwDIjXDbQ5P/urdO7"
    "yJp8zARLUgAoIbmkLMI7dg3zg1qb3kmEQEoMw165ZHmV1yUnxNjXciw+pM9uwCuv7atw"
    "6YoOalM815hQPPbDWpu/2TVEKXH6LRWbf38sSQEA4YFFEso7X3P/IDc0HV0yuRBQQhTh"
    "0hVV3rysypDXEedRjsWFlLAHvfKmZVXevaKDuk7+PGNCUtkNzZjX3D9Iwy/eKr/pIoQB"
    "l+IdJhxtJXR2PbFk+dJhPaywob/bRJ7fdCu6jPBPe+r83a5hKiJE81ydluPAYSSUizdV"
    "eceKDl7XV2FoipPfARURdjnP8+4dYGszDsk+qcRYis9eQNasPyMGWQxTjA4YEcEGfEy1"
    "wGfW9lA1YYrQROpPqvJ1G+Gy/gZv2zlErFBaBLPeDnVYQqKXFfjAyk5e1FNmcArm94Th"
    "M3UPL71vgJ/WWpOajEsIzig0F3oVc40Y6DXCFbU2r9o+QFtDOvBk2YJpdOAlPWU+e1g3"
    "vVYYzGsHMo2IoO31WOEza7t5SU+Z/ikcup5AC7HCX2wf4IpDh/lRtGUEGUx2ZykqOSNI"
    "nTvfHW7xmvtD5t9kQgDCabLbK0+oFvniYT2cUorYnSSOLAW/wFKJdAijbeJOKFm+eHgP"
    "F3QUR57VRPAEPxHAa+4f5FvDLfoODebX5Ok3jVd2SiCDJS0AYFQIXD7Y5C+3D6JTaAIw"
    "2inmlHLE5sN7RqoI07TSxYhUgMXJlTLQYkQ6/3GvV57VVeIrh/dyaimifwptzRNq+0Xh"
    "tfcP8rXB5qHC/AECAnuNCNuXxjkwPTiCObB5sMmrtw8SKyPVgxMhVS27rPB/13bzzhUd"
    "tDUMIIkW2dalCTFtDY7OLiO0lRFVeTEhIkRtYuBvllf55JoueqxMORLOEZ55DLz6/kE2"
    "L/K2XgeC0I9GH4iM6O2IgKoeKoLAERI9vjbUpH6f8sk1XXQYoa4Tn4SW0H7MEbLJzihF"
    "vGPXEFsbjt6kb/RiSBapq/LMrhLP7y7zkFI4I29uxvzXQIP/HWpSnmj+VYaQal57vPKQ"
    "ouW9Kzu5oCM081Am12YcUJYwQ/JV2wf5znBrykSxJQgFES/sNN6ZWxd6NQuBNNvre8Mt"
    "XnrfILucp3MKQkgJr98rj+8o8tXDe3lh4mVuThJazALSltdvWd7Bf6zt5gkdRVZZYZUV"
    "ntBR5D/WdvOW5R0TdlfOCoKXXxn0yvN7SvzPEb08PinqgcnNspjQ0We3U15638ChyvyA"
    "+nDoy93GWK5T53wSClzyfoCxSH0CP621eN49A9zUjKflAU6HR/RZ4ROru/jn1V0cFln2"
    "eD8SQcgSLDCgylM6S/z1spAKO5CYAW0NmXJpwszTOksMZFCYmeTa65XVkeETa7r4xKou"
    "llkZaeM1VXpvrxFubMY8555+flZrH6LMDyRbJbDNON/6PUK/HCLq//5ICWNLM+a59wzw"
    "k1qL5Yk9OJk0tIRcgiFVXthT5vIjenlpT4WGBvUyK5GCtAS2jPDC7jJew/+n60sdgJ7Q"
    "+OIF3WXKyEhp9EIjXV9NlXqyvsuP6OOF3WWGNTyDyYSVEtT+5Ub4Sa3N8+4Z4MZmTM+h"
    "y/yAWPUOrN4igKxef8ZPjNjHeHWe7B1gc4ax9d0RQl2VsgjvWdnBi3sqDHvFMfWGpHZl"
    "UYRvD7X48O5hrmq0qYhQTrLJFN3n++YLhiTF1Ri+f2Qfq6Lg9NufudNWaffHyoV372HQ"
    "eyLm168xuj/hOLICDa/UVTmrXOCvl3XwpI4ibYLzbyotJRV0nUb4/ECDv9kxREOVigjx"
    "Aj2PhYeogCjaECdnGUAV/WWIp+hi8GPNGsYWecQoJYE2yuvuH+StO4cAEmKZHKk2MOiV"
    "J3UW+eoRvXxgVScrI+EB53DoyHy5+Sa2tDBKUIpT5LUrIdtR0NCtd57WOPb7AawoDmW3"
    "c6ywwt+t6OR/Du/hKZ1FhlRpTYP5Y0IxmAi8becQr9s+QDt5xnHyTUu5yGdiBOmvcN92"
    "t/POMNPQ6RWKfyuIPQR3ZASecOp0GuFfdte4uRnzD6u72FCw7E3qxydSi1N1eiCJP/9/"
    "vR08saPMv++tsXmgwf2xp8uEmgJPQnjztdczpfT55gwZzb6ME5/ECmt4cW+VV/VVObYQ"
    "Mej9tAa8KmF/+5JOPm+4f5AfDbfotaGqb76HjmQP6sTYCI1/xx//WA/abS3+uXrdjkhq"
    "Mh6yUA1E0meEHw+3eNbde7l8sEmPEew0agFSe3qP9xwWGT6wspPLj+jlFb0VSiLscUrs"
    "wS6FnuyzACHsRayw1wUt5c96KvzvEb38/apR52qqzk8GRygE6jHC/w6FZ/eTJLvPK/M2"
    "YzDTSIS7IFcBGDZtsjt3bh1S0e8aYxQOLTNgIqRNQ7fHjv9zXz9v2zFE0zMyd2AyWkod"
    "Vy1V9njluGLEh1d38bUjevmLvirLrGGv15GQm5VsONzmC+n+hNBk8Oz3GOEVfVUuP7KP"
    "D6/u4rhiyOZL1f3J9id19HUZIVZ4+44hXnbvXu6LHT2HWILPlBCxqg6j7hcAETt2JCEB"
    "8xX1+qcg+cGUwMFIQ4h/3lPjt/U2713VycMrBYa90mbyUykl9LoqNYUTSxF/v6qTV/ZV"
    "+Opgky/117m15UL7qcQ8UF26KlhqQsUwUp57fNFycXeZZ3aVOK4Y7ZOVOJ1QpCOk9PYY"
    "4Vf1Nu/cOcSv6216THAk5sy/DzwiRr2/O4o7roJR01WPOOK8SqvQ2GJEjtHgDMwFwRhY"
    "YDiJEryir8pr+6r0JDHo6cb+U9u/LMFBtcMpPx5u8q2hJj+ptXggUYErEsyN1BF3MJqr"
    "kIxSt8JP1i2jx4ZTcrwoQCTQ75Tzt+2elZHXKXFJ0kuhoUpLQwLWI6pFnt5V4rHVImsj"
    "Qz15faZ72W2Evc7zL3vqfHJPjWHVZNBLjgdDY2MK1vn253bcce1L2bTJRoDCJvvHP26u"
    "rz7mzMtEzDtVxyORQxuOMBzCAf/wwDA/Gm7y9uUdXNBZGqkLmMxJCKOE3UqIvUPgud1l"
    "ntVV5oZmzHeGmvyo1mJLM2ZPIgyKIkSMxvOzbsaOZeCYkLXX8oFRTy0VeEJHkad2ljix"
    "FFGS0IZtT+Jgnc6Jnzr5KiIUBX4w3OQ9u4b5XT2my0rO/JNCjKoX48z/ArBjx0jitwH8"
    "ymPP3CBerxehTN4Na1ykBD6sikV4bneZ1yyrcEIxYigxC6YSBClSr7Qkp35Jgvp7QzPm"
    "h8Mtrqi1uLXp2Js4wcoiFCRhlGlqCHOpAYw94dGkujARbgB9Vji+GPGYapHHdxQ5uRTR"
    "kxQf1VXxGpx2094rQq5CpxhuasV8fHeIsMToSPfnrAvIBYRHxOD9/R0mesjtt1/dT/Lo"
    "Elxi4FK/+ujT/5+x0Qu8j91S7xR0MEjLUAe8sjYy/HlvhZf2VlmZNA6JmVmJrScwtJUw"
    "jKJAEAZ3tR2/rrf5Ra3FNY2Ync4zmBS9lAQikRGHmhlHKAgh3fdABEBB9v2clNlVg0bk"
    "gFiVZvJ5XUZYZQ1nlgs8tFLgvGqBDQVLt5GQvJPsy0zTpV2yl91G2OE8n91b51N769wX"
    "e7oTW3+p+k1mD+qMiax38b/df+e1r075fUzJ9NbgDBT5J+/j5wEml6cTIyW4Xgv9znPp"
    "riG+NtTgVb1VntZVpseEslQ3zVPOwMibhn1IxIkEjitGnFaKeElPhT3Oc2vLcV2zzfXN"
    "mK3NmHtiRy3J5W+rEiX9Cy3Bj5Ay2kxUubQPrk/u02lInnEaGN6K0GGCun1kMeLEUsSp"
    "pYjTSgU2FC191lAWaGgSCXGKJGuZrlBMT3xLSNUe8Mp/9gc7/7pmTIcIvRacLs2WlnMA"
    "471TteZz4X+3pvUAY5FoAetOu9xE9uneu1wLmAbSE62mSqxwVjniFX1VntJZpjsRBO0Z"
    "qLspUsJOVeWI4BMoidBEGXTKgPfc2Xbc1nLc1nbcFzvua3vuix0DPmTU1T10GPjF0Svo"
    "nUID2OuUR965i2EPFRMESY8V1kaWNdawtmA5MjIcW7SsL0Qst4ZOK5QQGhqEUAwjgm+m"
    "nuSU8SNCQtagV7451ODTe+v8tt6mIMEXk6v7M4E6Mdaod7+6/47rHs0YJXHcpimx9X8X"
    "eXlyEhLMfQFTII1DVyRw+LXNmFffN8DDKnVe3FPhos4SK62hpkrTM3IaToV009M04pgw"
    "qmrQK0aCPbwqMhweWc6vhvc0NDjeGqr0O2Wv9+x1nrqG6MNkmXA+ec9H1/RQEei1hl5j"
    "6LVB6JREKAsIQpsg1GJVhr0yqDoi4ARm3CglZehS4g/Z7ZQvDtT5j711rqy3kSTBJ93r"
    "HDNA0hZRRT/BqHyNYXzGNoBfffRplxkbvdC72CG5FjATpMw97IMafVIpYlN3hWd1lVhX"
    "sDhCxeB0zYOJoGP+TbPc0lM3+AQS/0DyHTU/tbosQNWMps06RtXstGJQ2XfdB7P+VLup"
    "mhDtuLvtuHyoyX/119najBFCjgTkdv6BQb2IFa/+5m47fMatt97aSl+AiQWArjzy5PXG"
    "Rr8HOhidrJxjBkijAbUk/n10wfKkzhJP7SpxVqVIp4TX2sqMPOJTQcf5WZm+/e0Yn7ln"
    "a21p5KOYqPM1hWsbbb420OBbQ022tR2FRBOAnPEPDurFREade/n92679DGyysHlEiZro"
    "mVrArTr6tHcZE12qeUTgoJAKgqaGU7jTCmeXCzyzu8zjO4ocXrAUGRUUOovCIAvYn+kr"
    "Saz+7rbjR8MtvjHY4Df1NgNeqSbh0FTjyHEwUC/GGu/dlor2n7Nt27Z0BMDIGTERjQkg"
    "RxxxXqlt678Ra05V73yeJnxwEAJjOx1Vx9dGhodVQpz8MR1FjihYChKGVLQTpyIy/dyC"
    "LGCEeZOwZkGEikAbuLft+WW9xY+Gmvyi1uKPsR9R862Mmhg5ZgPqRKz1vv3MHdtu+BrJ"
    "wT72HZPRlAXc6nWnn4+RH6qqCpg8RfDAkOTKjGze2KzAepICe0TB8shqiKE/rFLkmCSG"
    "7gnJNS5JtiE5TQ/WBj9YjOeDsBIcgOXE/zDolbvbnl/XW/yi3uJXtTb3xI5Yg/pfTBbv"
    "x3yWkAuBg4aqMzay3re/s+PO65+URvj2f9sUtBPshRVHnfoPNiq8UX0cg+TDcWYRqVag"
    "GoRBLUkp7rOGE0qWR1SKnFUpcHIxYmVk6DSCIbw3jc07EhU7+cDUEz/2Ow4E+zNhyuip"
    "E9CSMDyjjDzklftjz5ZmzO8bMb+pt9jajHnA+ZF06qIEAZaf9nOG9FE1jeWh99927ZYD"
    "FACBltauPbvsSu1fiNgzVFNTIH90swsZEQZo6EzUTBJpqhJi8ccWIk4vR5xSKnBsKeLo"
    "gqXLCFUTQnReNcnOC8JhbFbgTGK56XvHZv9Z0gQjsCI0VUcSkO5sO25txdzcjLm+EXNT"
    "q81eF14vilAyQUjsy/Q5/cwNBNBYTBRp7N65465r38M4qv/Yd08FA/hVx5xyGmp+jtCR"
    "PLvcHzCHGMt8XkdzAJpeKSSZeF1GOLoYcWzRcmTBstpaVlrhsIJlTWRHUoULySndmkbW"
    "nBCSjeIxcf62wi7nubft2Ok898aOO9uO21uOO1oxgz7kArRUKRkZKWAaa9Pn7D5fUCfG"
    "Wu/9r1ZW4/O3bt0aM8kjmOahEEyBlUef+qfG2M+qd+lE4dwfME9IBYKRsfH5cNK3FLwG"
    "wVBKTtxOMfRYoccaOkVYWzD87cquEQ/8eJmAae+Cv905yH1tz5Aq/c7T7zQkMSX9+Jpe"
    "kUSVj5KUYztmbTnDLxi8iKCqw0btw7Zvu+ZGkgN8oj+YLgMLbDJBCJzyEWMKfxWEQD4s"
    "d6Ew1gGYOgRT5kvz931S0tn00G2Fq9avnFYq8Dm372TAKSWTpjnLPglGaQnZ2MahOcMv"
    "ODTt9+e9/9Odd173uf1j/uNhumq8wmYPm+zOO2/4a+/9N8TYCPTQba2+wEhP2ZFiHR1N"
    "pzWENOGSCFUJpkLaIWcqCCHltstIEpMfU4I85vvcmO/LmT8TcMZEkXf+o2OYf8pUipnY"
    "8YkQQFvKi9W7KxMhkKdmZwxjHX8jFX0z+Hs35u/GflaOrELjcPK7bx+1ovCmMcw/5WOb"
    "qQqvgOnfdu3eZRtOfVbU9j8RY49V72JEopxKMoqZcnDO8YsDAmhQ+9W73xfj+EVXX31d"
    "G66edirFgXjyPWyyu2+9/o8+bj/Jq94hYiM0L9LKkWNeoRqLsVa9v9Vgn/HHP27dzWiv"
    "mmnhAEN5mx1ssrvuufEPXt2TFX+niLG5TyBHjnlDLGIjr/5O3+Yp2++8ZltQ/WdWQnEQ"
    "sfzNDrAPbLvhJnXuCer9zbljMEeO+YDGYkyk6v+gwkW77r3ulul4/MfDwYbxHGB33r3l"
    "tr71J11YcPYrYuw5ecrwYkfuBMgukiw/535vWvEzdmy/cRtwQMwPs5PN5wC75/atdzUx"
    "F3jnviYmipLf51SUI8fsIInzR5Gq/3Y9bj9h+wjzH7j/bbbSeR1JdGDnXdc/S73/oIjY"
    "MGswDxPmyHFwUIeIhBRf97Edd/Y9ffDemx4g8O9B8dds5vOPTBPase26t3jvXgLsEWNt"
    "IgRybSBHjpkhze6zoHu99y/bue3618EVjilSfKeL2S7oSRdkd9615fOxjx+l3v9cTJTW"
    "DeTawALBEyoLJ8sGDF2LQh/DvMpjoaEOEDGRxftf0NaNO7dd/9nE0w+z1DBpLir6lMQv"
    "sPuurVt3rCg+znv3boWWiLG6Xxm45tecX1ZgwHlubsaUkHE7A3uFEsItzZgB57EZWPeh"
    "dDH6s1fwYiKrUFON33X/tmXn77jnhusYdfaN8wQPDHNZ0hvUlKuvbu/cdv0lTuQxXvUK"
    "EWtERBR1Oos3kmN8KOEwrytc1l8bKepJN1+Tn9PJQpf116hr3pVnvqHgQZ2IMSLGeHXf"
    "AvfIHXfe8HdwRcws2PvjYT4UvVBQlix+5bpTXwryTjFmPepR7+Ok7XiudM4hDKHU960r"
    "unjzik48Qd2HUDRkgA/uGuL9uwZHBm/kmA+oB0HEmDBG2V2j8L6dd13/5eQNaXLPnMjj"
    "+WS6EadFz1Gn9hWEV6O81hi7GvWoalqmvn9HqxyziIYqT+8q84LeKieUQqrGTc2YL+yt"
    "8fXBBmXJt34eoKAOJBITlHD1ep3Cx3atLHyOq69uM8oHcyqLF+Bpj2YsLT/yzMNE2v8H"
    "eLkx9mgUVD2EbMK0Ea7kCulsIZQED3pPWYRuG4hvwHkaqnQZk+xyvtezjHRDg1ksYkQs"
    "eIciPxXRf9vR4b7C1q3J0I4Dy+o7ECyUuB9pMALQt359j3XVi0XlxaCPFmNtohWQCAMZ"
    "05I8P6IOEqlOGSdkGckcGZiHLpKd1UR1l0gkIWFV1Pv7VPRy1H9h1903/mz0zzZGSYhv"
    "3iTwQjPTPoIAYPkRJz/UGPkTVX26iJwUYqAwKhDQ1G5iX3Nhoe9l0SHdsPy8PyDovv+m"
    "jdHDAJ2U4QXw3gN6t8IPBflWHBW/t+f2q/tHP2r69fuzjawwjcBGm3g7E2yMVq7beYp6"
    "OV/EPELVnwOyVoyUJZGkkDS51JyEcywAJO2yNDqkQb1HoYbq/SJc5VV+J8gvCnHx6vvu"
    "u7o2+sebLGxWFngAUlYEQIrkRN9o9hUGsHbt2dXYNI7DmDMQPc7DsQYOB12lyHLCYNsK"
    "0x+BlyPHgcAptARtoNIP7FZhhyB3qfq7RMwtVvwN0ipv25fhAbCwUeZbzZ8M/z92fvbj"
    "Bk+jHwAAAABJRU5ErkJggqUgqJ4="
)
_PNG256_B64 = (
    "eNodWgNwXU8Xf8GL1dho+OLGZmOjsdEkje3kRU0b29aLbTuNbbux1ejL/7t3d/bO7Nk9"
    "Z3+HO3N/KytKoSIRIgEAAFQZ6a+qAAAU4L+OAPcx6DkVf/8Y4EzkpL4C3v97czog8QAA"
    "OUDmq6i6R8ZZpmeGnkTDY5er/cX6bN139GGzdAxo7eXe7wgyCOHlgWGfr7G5lbOUJEiL"
    "cz4rmT8h5uU0h4edSpMGS6AaaQXHhIjcj5lNmtGkNrRBkr1PHoao2m3Yfb07mryPn5/m"
    "4Geck50vKvhcj+dK1AIxc9+CnUNbCzzrsb8HxZh8DoSjaiAXzwF+qiWUwbhcYirLOz0C"
    "qIogVxOBGpH2dNTiT/RUdTvvzyROAFZVvxYw3JCn4E3Wrm5orVXTKUMEsHhgimt06UJ6"
    "MQO5KPZPVGsxkt/GfnJnr35vkvbdxgeyet/bLN4zXyyr7iIZMcf1uI0WKO9BITLCPHEj"
    "Rtv7PId8c0N2IW9kAF/RZggSNGMOf+tWbttXoUMKMypJWPNWp3s7HAjHg5ziDRU9R6Xz"
    "vosp/EqnyjCKR0EKybQRHXnuIL0afTDw96vSF8PWhlorjRTebriJWUMaozvvwEGC/2zZ"
    "hy0s+63/wkqnvRny2MSCuWsx9s4a2AZGHGiFiKmIccj2KBBt+m1XYTvC5W4hbf21Qj5W"
    "S2uHp2Chu+3ADmPbUgA92tRkhADVAOwWEy7uf9LsyfvHRXKUH1ToRmKag+1t9bxG1h3K"
    "tP8Q4XyFisjkdU+IWTHEyFD9dcKADQPdc8JGXbVS/zJwF9IaknSBHJX+5cIkoFGM0y/k"
    "sxRVXTBzUlS24/3+x2RWFbJClk8qrcCPvXUER7ytyUVXnDCoKLweqR+h0fP352u3Tp0s"
    "WckU3Qmgm6BP6NEU2jHRJyDssL5PcUMwoEqdlusQ6Y3K1BPgoJKwRQx2ybsEAs5UvjI2"
    "cQ0j8NDW4N/LRLseU3L8kHw2u85s/aGo/FgoK9KH2k5CRjlpycIjajUAHZGG73isbJoU"
    "2EgSgzYiWM5cnXe9C3rMcfOMIn1jy66tRSz/Cd7LPGg6okmkmMLPlRxuZgMlnyXAP7fr"
    "+g7T8Esqt5oEZJnA7Kyxm+5vP3HwEsEIHo5nuekXPKdVjjTMdYAVY5NUsvQn4rXW0FC7"
    "IACBSZWeSVRZWdKJPo7SSw6k9drNDd1aiAi9TYIJGJxI8e/EeIGFp/qkQocd+qXk+5KG"
    "/+qEAou5Uxw1UFnYVrSDZkyiGEE4pHc4Q0fQ5/RfY2b6RoOX2MpGu/va5lspdkZKug1S"
    "SB6Cy3FrRud5RTKXNi/PWQrFue17648XJIlhDC6Uq8uxVPZ1DbRzdwRWt4M1FUxSXPrF"
    "TjEV2Yi/PZCmM5sXqVaY5wZA8B4EOvApXi5XGRvecbd59zAz7b64843VblAbmWqn5QWp"
    "UJku7E/pj2ku06OZwuFONJ5aoMOyeevXNV3Wmg88v/A0KmgYrm5CSakDTLkbfLETOzRz"
    "6eA73eduPfue/+WfIHQhUr3rE+oS0WOHcag3Fz5Rbo1Qvl6xIhut1dcxf3Wy6RlQpcMe"
    "EP3FMKiQ6t/P5ZsCHSO2pu7x+2PzT9MFdqQ1+mQcyKx+x2tHRYP8A7EyH5J+6purN01P"
    "6qw+6CZDv4ZK6W1XrbgD+Su699G+Bg4/d755gzzoaDwbJoc/SxDknBR5Yl2WwPi/5DRM"
    "ZgmuNhSKb8KhgBrNwnYcK8jeKB3MqN/Ti/csaj4jF9DRgMSdI/lYc//CC4GZzirrVCK7"
    "ynnX1913Qze6Mu1AgXoqec0us83H3WRvk4h0YwxhzFF5xT4z2X4eFvv4EiDK9yAYxvdX"
    "JBEwzyhpAPaAyXVsnlAtj6TKjDw+u+/E0fN77n++9KCV8GR4r3W0lgil/IZAUbnx/k4I"
    "Zt1Pvf+d/4F2WMLh2SRZMZR0gCoAaQAsikLI7zuMskQGb+T0s+KaLiNEB171HsE4C6a9"
    "IjLL8541Z9Au4AOsq5hnt2s3RPZp8m+3m4CeFfekEVh8hAwevxdeA1Q1/hU7mQ+iaUcC"
    "HEEpa4SSxnsNOrjPRW4ZNOoHZCEmIp+DPTPg5laYfonl0glhOkbAqS/9910dIsqcAE0n"
    "twZdCqv+4/mNoXzyduwKiU/eMnnKWCw0gWq+X4RXtRBbBHEXl1Ux9KrVKAGtjxM/t89B"
    "Xxlinr2yPxEWNJmAoCO/ZCXdSFTdFzwOZ8NqFJ+bID78ucQFiqJtX7lACp5QL7CEOWXn"
    "/PV92gvuue0QDICzZtXCkhekMhRB6jlIMOuAYoivFqDWzobOR3ACOX1VZhqUF1Q7sofP"
    "Q+TRwcvYCcBaCzRdzzPTOB6gKpeX2VDstpYZTitAYnKil4Xi+QUGzwoGzgoEapOKKqx0"
    "pMsFCdK0Z8UwykgbyEiblA7ywhUCbAaRM/Zwxm4QDjzkIMmTSc3B/rTULIPU/tEM5kbc"
    "slBL/dh4ejuLKdwAQl/Dfqp2jGQBAnbt1FrdtIDfqgCJArHu0FfC5EdY0/DtDgYKriIq"
    "pcHt1BSZ5NJYCXEQ+DOrHp7ecVD0Iyc9nKW9fWJKsMVuXGgeTfq8uliPClIOWN4FsRIN"
    "YeI6LE9+kSTsE4oxnDj+s6zDvJPcoURU5k/SI8lol3xX91wpSmDSrQPKVmHzXpPKcGtC"
    "aV6IGAgMzaoRygLHeeAh0cPO9PVTfWiZo0LGL57SLPV18jhOVyrrkxjSC9GhqK439qTK"
    "lu0sQ7Jb+08XPR3u6OcdFc26uqzc3LF6XT+PJ9yggVlb8bf0gYZf5blA4eQXzeonk2RS"
    "R4EsQNibDugKAJMwsFypJ6zALHN1qJjXmffq5ebSo8Nqchx4dBeLzxXYX/0J/nWDNPpm"
    "oTfE5vV500r6jnUeFUvvqcENlnWSgWc0v46OokXYpFYBFNieW9t9GJADnoTjrLxEFPC0"
    "e1xF0Ypj1Y4DN8XHFVQ6b/AiN9/mIV8++z51aP/06DBQbLtVNTnMGjJVEhonmewAD3Hw"
    "9FDyzH80VMrbl3YMhe4Wgx696IlskzhDCYW80ocRAhVh10avSO08xOYpiKOtSD5XtCi/"
    "tWoc3k6cfz+6WzCnPP4knvA7gSQaWda8ax0JtJb5SePZywtel0nfoALZi0ZVH3HVoE+V"
    "UR6D/0rNfMILx2fPoiF1lZ+8OdgiSRHz2/zGeGo3ZuK2nFT+J0e2j7oRQcK/i6z03I4k"
    "9Jy2Q6frTzkVcnEAYGNCgfv+WYGrOfaUIs+pzWPXlMP55m7IFzal3WBzmeGHVoW0hHxW"
    "hq+v2IUAth6gHsHco5jI6PKakqwWMbdnu9T1RX/s7j4DPf3j8d2i09qeTMt6xA4d+LXc"
    "wx3n4eXWr7W1Ff05HVDUCjx85ph4QHa6MyoAC0vKn0en/TA3+UkHjIFCtes5wyHcdeIw"
    "X5Ufpvzy0eI7vm8OoRJxJaTbv/CJw6GPHPk9g4qgAUsNZalkL3+Z3c5+nN+k6KAyP1jr"
    "kU4Vj6jkVir1KGTCESVwGqWyadupasE6FBpeDtELQga7OKmfRCE9vBVw6F//2b+YDkUc"
    "NOraT+juNgXlioc6rLp3Kbz+VfElrRsKfqakQdZRfckZNbNGI31DZPYo1jkMAxdhrut+"
    "BYE5UGa23ugz30un6rie3PHjLtwI9doXt2pIZQTGr+7ypa7O0jtleIgnzn+i5bHp9TBt"
    "tsZO50cfPzTzXGqN7cUJgjoXzWvryPaXyAygNhGs1jQLD+e/6FYWa86hswsGds0184BG"
    "k7qGgslcUvXqF7dKSK1AMoNTueJPCwrcGEjAulYIFmJIoM9zl73R35geq2gFKKVBeea1"
    "+J3Q9VddmVfHMKlLh8gJjqZa4L8PtxZqnmPxIB81oAK+UNO0WCqFDZkskKjh4FDYdV5Z"
    "q/8cZZHTNTdP+YA3zXwgKeFysXj4yVv15bNJ3B8TLUMIImn5QgtfcamnYTG3M4LBln8v"
    "4ebHucPivDgpWsyUwsjcCVwslK04FXbLDFsWVFlSmYpYfLdiEk1rzf9O5b4canYGJ4mm"
    "mMn3wnyKH30o0ooLwXWAMkT88UKe8VyuwJuniUrgW7YuxhNCj3pe5uSYxu3c0byYsLZh"
    "7qTUJGqB2zFvnmzQn1qH3CZTZ39HL2Dr/U+QRkc46jxekSkFRTwUZZBVrdyggyv/ho9K"
    "YHW2jXTLo6k7WzTJ9C+qitAypQcvxW8ZTrlqso4f0q8VbsjbsMeuj+YN/GWWDe9nX54m"
    "0YIStRXFuDkiDldYfNBijQsbe+9J+LVvX+QEbg+lPKzX4XA9nruGUBWYBBpcgN6rK8WV"
    "pUV5n995TWkV0v7U7l/MLCwkXbWkq8ZBTGmCXHuCUbrmMfRZhyi14mJc6EgA4OLs61yo"
    "d/qiFqBwyVzJmANQKRTUUjG6VehxXzW17+Q4hG5vuvfrtNf72eSoFufFdvv9+u3ZGu3f"
    "a2vOoEWgFFrfs2v/VuefHoMm0+jUfiCZUPTRbNHet04m3UanPoXNHtba7If0i/nmFjHK"
    "jG13a+7tf1HqC0htGfiXuN3y0Vud9Hg4BBp5TXFxLI0QD+4hdQ28DH9mSKaTkx0eL7ho"
    "XLf5OevZ4/jIwtLKSvD04l1RdkNJ9thFNJJGejp95FlKv0MtIbS2rTBRM7l9sPRBk72x"
    "M6u0dKT9j7a2tgwHrul6K6q2oWk+hSczM+8KE35WTx5lNhgoFPmg6tRdxUg8hdt9RNpD"
    "ksg8ozUAbyw7HKQSbyhJ/T3KdHhkhPXhtYv/at/4W7ZVmsC/R7Vnk0tktTJX/UNKT634"
    "6mwtxsGGRf/MysrKjMp8xB68QMLscVYAHfM2EQnIYPx7WK78VBkcujyyUaWmg1QiA+Wk"
    "Q3IwER9K1i+Z622p1p385C9Wqudtde5Sn1/+TmRXe1yV/YHmay4AeV50zNtodcXuWkx0"
    "n7lBhEvI1uy4xSzL9VrouVuo2hNjyhV7jgxONvyS/1bAR4KGi6DSEtSriMHiYPZ7zjf5"
    "oicuwsuWxpdIIgaGPp/9lcn8kxam8JO/p3ttm+8UJwsW9rIWI5UH6+11PsA6RT4o+6hh"
    "aclo0gE5f0N5gdg/2pXjP/cKSYdlSEUbBigymWEtvIBlak4iB+TEDo4A2jP2Z3arF3sa"
    "quyNeDM5UrxUxhRhvOih+FnOz9N0m21nQzYbtZxKqtoWNwcVPhf3ZFiRg06tUZMn5m/t"
    "7DdUODZfqLLX0ncaUMa8mjDW9EucRwAo2PjKSEZpUpfxpWODL0pD8pW2cvuc3xoqT8Yy"
    "2fkh1ilFw+vCLdHjuo17EkU2lj9WetcfXk6rXd9T2S7K2W921JaU1HvvzuAmb7WSR5lg"
    "3TVaeOzAIWjo+EDlqj7BRC/JdIcXPhMHOYhoZ2goe5FCYWsrPcT6vzxS/sfxNAM9qUpr"
    "OF5J+OMEP6ytU/n2lwRk5uRUZTtnjS/Tf7m2CZFvbfzAy79OGCzHnvK6qUxyL8GN8y7V"
    "Jxi/Ecb48mjVpILhqSh4RzXqRPy7qIhJ4K1/FhjLlH79x1TpQACF9y4h/oY5+auTfVNP"
    "x/pCokgDOFvz5PnevP59Hbc1oQO4BTwvcWNfydNI7ibRh0M3U7ZiE110/S6fnjTL6fe6"
    "ummKbVnKQoTSt6XZ9SquuMjjd+DiUsHaRvtCY+JgkRxQ7/Ei3bFIy4qhB21eBbCrcPlC"
    "HeS4nqkP5x+GW+nIySFPjnFp4fu2esFoerfi6plAJ/wRin7uM2h+9/Wjwy8w7m0WkHVh"
    "RZdAPT+ufdOHeHJ0qqm8ydSN6kDpEUzxDFm8fvJf0V8aRHdY2zN14ns/ssgaj6d6OjyC"
    "vHNDbmgPKS6M/QWwOR894ZLWn/wio+FHWn8nVEGnHE03C8GgvqX/6qlid8A3H+3/eaWd"
    "3Wb7jKZRZRD/JpKudff2w+0+/p2FBBWLZk8ljfJm9FePff6CBXynxrZJGsQmXcaTHiya"
    "tjVEMHb3zpSUmk+HNxhRMvlvbdHAn9LQBBsWRC1nMGow+OUu/4ZOYXNtNdHEeCwN0wHK"
    "/ecod0fKAUDKzCA8pmPXr4zMxqp37znzu4pZGUoooqZZ9tlhpjmDrX/X0Qj/aLmtiBdV"
    "aVkZAlAcmKwi29xAOZCOaC2bo4i3hutGQXzpd6M/oljly5bjWCrxcDaYsPevEXv4M/Uv"
    "/DjxxuwsjwyFdMIP9sZLzTWxX9GMO8RD0YbzOXzjoTun7aYsv7Q6J2nuhgklqSSzmRcC"
    "BltgF7WJH9zT4qoaIWyyq4kMrpycH6iYjzZmUJ6HBRjj26w3Cp7cWFuXfuz57pIWd7Ja"
    "IeVpQDZj5ZD8LR6ULULeBCbTRrnhLNw7iMLWVf2uYl5ePCY07u7s7KxiPl54utLCcHJY"
    "Xo6C922BCRRv704Vi2t2g9DMVL7t1M0/oQkicr1PzZnwCAfl+X364slp8LLBESPeO954"
    "HyUgq6q1ws9XJ8SdjP/bRnvN9PrEzhpqE9UuUil0hFYohxG7Wyt5S5gxanJDPkVAK/io"
    "k3LbydzQcecq67+leYLLVRguJ/cclc9tqBs1f7HbaYrVDGlOgLIGUJZil2fVDf2k+ph2"
    "R9L7GcO0ul9/Wg2hFcT8FrqOHOvRpHDxLi2AenH3jkNjf/W+R3o0SCSqAMjohUd+f6ej"
    "anKKNTRtHRCJXOWLO+UosK09CFu/aF3Ldk5G3Lyo58JOqbV3N23bdK0sLGAC59A0c5mN"
    "3xris2sl9PxzGOjeGyCNP2RciK8KdoGmtoNNqId+9FGGLARqr3+PmxbDGOg+kMqUudfB"
    "R5o7Cvy4qohXm+F00yVv7UJZd7yjM/kph0CznbAKEUwRaPei9cR9Axgi6ehWvcquqRd1"
    "fTbFHURJ0kqh/5A4bsC4GAskwzLIo+/QGGLeKQNqqWM3b2uHT7Sp0T2HeswQPpXXXkn0"
    "CDC0iGC2Seb+yMWx4EWNv0lFRm/4X5b09iy73uQql5jeu70mPhQXyrQXOhXAov63cTRd"
    "AdRC0T1t10Ns6BVvHQ16SlSAEg673k7WX3iRuunQGUxP5AuIUwO8f3LRXnk7CqjE/rT/"
    "qLYnnRxsqPUHJxexx0LWQTKnALvbIXmL0f/fhj+EwnPGpBc+5DNJtBSqm5YWl8BkFXt5"
    "/3dYWdkPj69P2j8MON+puNJMEJYBPblS+QbnIDd/1tLREBmmAyNB4Wn/feWGUCEjZyyp"
    "fbn0oP9RPS69WaJ3vIiQGPNJmQ1pqxc8qrq0T/a5pP1iL1/5o3TWcqCH5+wJZYIr5g2I"
    "UwEka3mUeCsMs3/kaexfkHw4W1w9uWQ4WQt8ZUwko1Jd1ole7FHDfLirz3/y8VNupYfz"
    "p2/v92GT6VIVJPWfENfoGDU4foj1UsZiT/hl6ZmNasYofKHbPPThids5zoniVyBfoYbK"
    "dOePk4ij9ya2h3oEjj1ANAHvv4h+9ETnbZ1+oX6vr240C9gMaUK43gsbPaozt3gstz16"
    "1ervPpqMc+76e3m4qtQifIilEvRk9b0AGlrx5rh9btcmO2GXCezCQmiUTuG5aiLEHfJZ"
    "0yCayKaBldqjMaTVmjtKAyAKBfGe7eqH5N3vP6mhCFBVTwWyX4oLjcfd/aCNyqriWYfo"
    "jty8i40tf1/wgwvJQ4llUe0hXKslQT+zxD1kKjJEEm7neI6UCBsV7nmBFFJSM8epY5hS"
    "nZnAkXZwdO7npQzjzZ5eH2AFzvq25p9A8nG6CTZu4UaOCwqGd+4V3b6YZ4QTVV2yW18T"
    "5UjSw1NI7xizkN6rgjSdbjz5ghTIhL3ZFKMTBx65plPYtLM5UmO+Q3Jm29NUAcs5iD2M"
    "yKN+2dG83OeThX1Mmk9P/XosYls+F0kwl8egyrY/Q6EWno//XlU2DE7WLtqKaTugedYP"
    "xYxsDyn4HcrRjeLktGQfn9BDvbTLxAKFN4j34m8EfjmtaTR2m99i9IhzA7UdAvnb0JVh"
    "3+xyb0rmjfKCHrBcys0gOtJPDeMORU8XM9tvVXA8T5YdTYua1GU7rnLnbRuTvLS8aog2"
    "sjJuOimoPAO9ygvMxESIHnkh40Jv87tjf7p0hfZ3C/Ib4lbGh0osof1eeMmIGiFaXeOe"
    "aAn+A6jMclglXoGau9wDaz/5YIhL0vS6a7Jud/35uR7lhxJq+DTHHFc3vJPUEL2ohd/D"
    "yJJhIPPJWviBis7WPJxKvyhxumGbktyO7JG0CN03Ht/TydRM38QSHPuAZMKznuIB/rTy"
    "M5HQgWlGgaOxCuOpVj0QXYe0crmUr5kJl5q9S8aEdTRHjaMqlfsMeN6XzafJu90zttJc"
    "BTO6ebMCE+V4UepyaowR78m0mRZvk9klGX4hjA04DoM1E0TwjE0YP4UGckeDSEoAUY2s"
    "b4bnIjfv9cOEjWW3bCyHGXsCwOqxQGhLh7/qObvQAQJ57ptdLuVOOUDW9qSNDVkVVppI"
    "FwOl1pFAaPxdzOYQB71anSEQuIaZqUL5E+IqCc+7MB1jAjQdK+WPs6Qth0aCuOICw6EG"
    "XYRj/cA7GYt59A9il9xEAFEczYRXZAdsB6YUkoak+uN/xPKg5RiO2X+VhHqBd2hymHtx"
    "aGU1d7jqMysFAbN75wO/2RLJFF/uTxjpwCrM2zzLJBZQKE2shq3DAbv+IqnvMCmSyGoB"
    "lMT4yFsCQCNjqJhEVMJTkZA3NL1GbRUAGgq8S4YsoZWonJ17MbQyil753PAntDXBJRVc"
    "sC7CD6biAhLc+gomOmSjbpMBepKCuwkHXlBfQD62SOFCOkKB8Mr4BghPKg+7G0xtIsun"
    "INOcTVHOnKOKdeqo/M8rOT4jjtOKUFgkfXuvhGsICQ5sLJ22cVHg6EpbsWr69d5W71O5"
    "ixp372SKae+cElL+h017YaD5jttaYQNEEsGCaHGbKrgiicCbrnXRlRtWYhsojs7F4ZNw"
    "t+nlJJ5T+7MKxxXdXRutacXmwb5yoDi+s3whqmoS1xlNhKKzfGKl6Aiv8WmUKmDeywt1"
    "6nxqW3hLGkheuBd56F5nIlNWuMv/QOPxNqmb6qYTEX+AKsozoXCqV+6iOI/7vSVMRpQN"
    "2vcP5KYECX3Iy0VO18uiN9Uv010KjomMLLGa3vgytBKWBY/PP+dmRdie8cf3AWtvtH+2"
    "GVy/4L78+vEJkKMgqU2ru2TZ6OMwdu1/Rx8Rf4jOnFWNlhtz9/A+70URpNCHpO9/Ywfl"
    "Q326nVZNu5k8ngYGY3Xw0Xq9Ciylbm1M9ClsvWwzok49+i2KRhqS8ygQyXW4OYj4ZiKo"
    "0Iqh0I51jSmQfS1jetdkEI++VhEGKUC/5Kk1G7jkzG+2E3qQXR9FeXmGOiks9YUyFhBO"
    "Ej5rjvopRyTs3R26k1FteNhOGUe+ixw36AxaB33IPZz3PSKh/tUTjR5oQr6w2U4PVvrG"
    "87B/Bslh4yO3GnybufVbNEbrcsxjPqMk9pLX1I0bL2uz8c8Kkg3A/3XHXSUOYPgtcFYU"
    "w8ZnlSu1mzbR57cqfHARTUky29wY/iVe3DN2dBil/1+6JnAr43iTGxi56D7FKDKq+AEa"
    "6UhsVo65d3R+VoVyQi/WQdvbgSvCpUEv702675UBhSbmii0hmXKefiiLgH38dH30lGOw"
    "6Wd9mdVsCL44OEDL/6YU8Db1TRrd24+uOMKNO4Ij0qwPatbHt+YNYQgh+HVSDRbIinFS"
    "lTR5Ucvz/dJIFLLstNszARtuKgl1UugijxiMRrZum1DBkeYLD2vLprtMh/7uTX5arrcW"
    "8s8PpubtQBBtKysKdnQXTjMEGUs6FvQMA/Lb1I4QIddeOrHoztk24LsnFUCJkX4ztzOB"
    "z5THCCr+1AH1zds8+K42RMGXLlS0TyDKq1GMberYXvQzlOG5zJWT0kE46eGz/pr+ov9M"
    "00AN9mLDYC1ORTVPceQcN+xydenzL/9Vs8ZaROvcHsAdVuB/y5HjbV9fPTOz2V9eH0oW"
    "mwYZz+j1GIkbcHPGxRhdE0PuoO+3MzTYinqUmlcULHThtOmoQ4HLIWV6EVl/h2QmD9RL"
    "GcvzXFZMuKQaaBhYIDMqmGSB6xM1MGFFPcPOpWzx7wh37gtraRuRTCSX8vA8usQyiSEh"
    "QopqALQ6wAyZ2m71vDr8dvUIbxFF91x8b7nHzWUOOYM8P6/8K2zXbtZoiZb20eJ2tYvb"
    "h7ORFPUn1G6WiySmkZM7WGqQHyK6zqR+IzcaIEoFBZTQhvI6uuEo+tDQYs7A4lhhgykS"
    "gVMsMiArTmOHNMuhbeZMVJlDoC0PJ8OlTB0DAdtaH7JSgfsLotFGoWuH4HsniFvCq5yq"
    "zvaLvNQ+DGwoCAA+A8X3eL0TGhXIVfcY85cDorgST2gfGUloA/5dCUkJh2PVxF17t8HE"
    "L/OpQsk/3FcBMlIMSeiyJdWYJ1W+ltv82+53uZ9QA+AazPBEo/HMwFbrggVUPuYEBrWX"
    "ZmdGDigGisgDRFqmzCiMd742/ZzdgyMGqWrwvPFm0BoFAYFTCWyOaadRssho1Ru9Bieo"
    "dqioAUccavNW12bhs9l9xjtWYqwbgdBayawktG27PwHpcR1ZM2+wN8g9CPrVnPvVTUf1"
    "Vemc+9fEA9t0+ltKdRAJVdk0Yh7nfR5gzNJD3mDbszss6eH88nH9l8RZb8WyhFFn7N+7"
    "7Ufp029R41Y8ilCAcoU0K5xujuiJuk+hvNCjhdv/JzkyQD7oEm+qdy+A1vrEzcOydf6Q"
    "4qIW0qjMjArRy8qXHqs2gmKM5p98gXfgvVJjCYdwJxGudVg3UKJ5AWipgeNAiHBofAGy"
    "VaAjbuOrQRIrMfWDH7EV1Aezr1Swz55VreBdX2YDt6mV49oj3tF93rDuYSHNFZKKNoWj"
    "P1OuDktOW6/4mBsNRsACuU5ll4K68O/GDWkY96OpjTy/1HyPTrZqgwNgggBQofQAMDdT"
    "RKpi3eV1ePg6U/676rf4DXvMyjnb5LoEju4jXLdjz7apkr9OInXvpqQtjL8LHy6jXawi"
    "nsjpp6AGzr8SWI3Upx5xFxILT5u2ZRxhBfLkR7gj/9TWgrn8oTTjpv1MH9UA2FDCr9ep"
    "Tj8oTGWio/8vSJjrBaUjSj10NMzF5DIXjdEBgVsoB8AZ+9n5Hynlzpd+MIp1ucIi0FV3"
    "LMvEMwAAlsB//wJQtZAvvcsJrV7sJJMMxIKY0EozfK9y615QSPNu3kZgSUg6QEUMvYRY"
    "NF4pbH1BchCqrD5Exuqpp/aZ+N5h+1lHayi0vURL+gponpV8dslFS5ev6jlBhgdCwWDT"
    "KGfnMUffbwaHwkFYX4kPFxgH3rsIFXjFNjodyoGASydGiWjQVeGVQTaJy42IlJv39j7O"
    "vDMa/kTN1PNQwV4i40w4HtbQPwAgD9dwa4BCaaVlKc/oYckjBzt9OnirV/nf1fuOL0lK"
    "vGM9NEALu4GLyCEPSzwIx3wYJJS5jdv+PunK755e88rjhQDo6AMXJ8DO3F6SlVxe/hw/"
    "SFKLD6S/Hi3E9Xq3U7su7Rv3p4MCTpnK4krE6cUdFGbvu6uUId8vETgVYossq0GQJ+ox"
    "Av7ZcMjkwZYGmcIEoDFGu0LcLz2Dt3Ypos9qS1U39rFwSNdwXxSi3Qdsk+dlHdvRAV7J"
    "8FUNGPKOEkvpXOxyFEdRiDbuMYmeR8c1F8//OHr2cY7yQYvUqBq8PI+HIrJiUAAB455c"
    "1RJE5mpWIQZsfj6OVSLpluT1zPtvy3JBpHLP74Iz+z6Hm36yklACODJaOs/+rkvW6aa7"
    "ryTUGJK6O+ZSbYJ1gnYUzfQZnQ8XXNSTX2WloaQduqWpfI/YqKNdCx+AxPszzax7DRKD"
    "EmUwaB5q+/pJNrfLTofh4qRy6f4vDfyiySCSy6JyHKlQaLKRCz6v5s+n1PooC3gMHS8V"
    "/KJiuA/bBtoXFvZNkgrVHIP+hX/CraGx2ya8grduPnCkujyc8JjeHw2JI4AGcBDFJTNS"
    "QZMczDxKS1z5Ql0lbG0kl51pnpjrJNlcz9ZDHsuEWdA98X/kPhjOZRpBKBcKtCAF2IBu"
    "Fn4vRtASi4bXrN2iHgKpX1sgagWzjT2TuYY86Gk6nDPVN0IAY4GzUyvTz/Sbmqzf/BH4"
    "jo5A1QnJyLZ3C4i83dPQOohHP5tu+0RjhdT58lf5s2P1qrGG4n6uf6invcVV16HrN4NS"
    "Jheq8DyZheQ/c5sBf3E4++wbdH/JL8ewPmqFwMd+n3jXqpq4iaU6kzCIqaHNq/OsCn3C"
    "avuU1/zALGVgf3tzB/lpMrKj42K6zrdZhRxjt9ETc2nINJ23WOeTozmH0jZvR8FWOJ4w"
    "5+1CKPDqT6iPINnGqVtkvtDkobW4Y2wnDpus4YXTB2rf2DtcGQrFksNtukLemugRldTF"
    "fxbGktEK03D9TM0NuzSptGqS1P4CeVrPyKYpkc71MFb4VfYmkDFuO0IbY/WlcA/d/pvT"
    "XbbUFAGDH2cH655xUkhBnSPiQ8MDtmher3KcOACsqIY5BcvO6FOA9WKXdayIwXBdT94d"
    "hcYdzVj1FTo+9HZrOpEzJmoxLdDn6cpYoDGv9u93YZEGmmZ7B4SukJr3oEzBdmymzO5s"
    "8U99g5OMy4QE2Jenf2Thjo68f4XEwnLXfw0j/7b5uE5zQ9jKe/iXOYpgRVx/qvuEvsIq"
    "3wdLBsnNtXV5+ej7cBE5S2VVjlVVx2+DC0LsnQwjLwI3IppDm5P6QtNhPOoKAIAZzfNJ"
    "VL97psbQ+velHmKIFX3YpKcoujsKkg7ffXqFShZouubdT3YBvjobQomYpLMkgbC1sAIu"
    "lGcOCxhWTh5ssn5JivItox3oxbR/UYZD6k6Iu6iXZOKFZeDoeZ8DZsYKK2/TuUC5LwAR"
    "dfEVB6l9LJQS4rgyH3I11ZOLsC5FVlB/AxSvIWRNkQ5Wlt3piivq6sb8hhXXzlkDhXsP"
    "jUOPkJCYW9WcPSbw60tpx9A1GmN21bVjUbwbturSyFpO6KR6IfoDUCue9VEfUemZJAyf"
    "y06SV4sNEOF+IsCXh1kFRBb+vJ2/yr+0N2WbjjOQu1Fxs7htLHciuYMp1JrfN63ray67"
    "rAX4uCnBuTdJMJEQYTtcxMQUfsnOEyiZAwoTqRp7Ti+ReOebzz9C3Soj/oR+GBrxCNtt"
    "Z5Zsm7WIKlngLPJ0xZWMS1i0+FbQxwQrTES8w1PV8KLyHkbd0iFP7A+v6raeDWtnd1Y0"
    "sNN6uutr16mJJzKOjEIM6+CtHxPPxBOZy711OJgkvczpRycOBEz9kH6rKyQq81tBFjfo"
    "1LG1fdNU+eeTT9P9hCv4JgNq6WxJF5T0lLRJKoQGeEBoB0i8y0PfDNT2D5G/ySIm8C9q"
    "PLe0zqQNkW3dd6gk6aJwMEiMaEABXmrYJ3pR7i3aOlcmvxmfJIqr6Pjwne62oiS8zK47"
    "kjL3MAqnKi8gAbxeoUbZprY2y1HpqJV/6uHcO4pXNHBSTlmL+kHmVyJUQWr3SOQU23Jt"
    "rZa1PJssYCJrmJfka8QE1EX1p5aahdZ8d7mRvdG2tvCsbytoIwZQgPr0OBYud4uE2qM0"
    "emHDMpU6mMKInMXt5N9s9uzU2avzAxtsrv9IwkhLSf9tJnQz97TM2KUwRlLzNvVaWzlL"
    "gQmE8KRaOAyiaL7Ai9grK3oEb4UGBfpwBG3rfZUVaiqLbzA/veUPboarX2WSqNEiVXI8"
    "q2TeIajfR6Zvf/HTCikcq0ay+IhapMQaPU8UJ9rBu5iV68sDYVlQs49QrRw5fXagsGGi"
    "GqpFbvfYsRyp7HPnm6lL2WOmr3j0p+YRuofbxsaKt9qGp4kJcSwXx3vjWdda28lQ+jAj"
    "0EuTo9OwhpvrmbGY+CfTnmZfqZ1tZFlLpSc/TFJKUp2NRqHxL4yM7LD/2fnB/BLY+gWg"
    "71XstTTQDdtGDeIiioe7oP2oEDqaGXUeDCh4uPcGIrcqPGoZxcXlPul3a1pDr+nh9aeL"
    "1iY3B5+4udYs9Uqy7/8KRiVUbbdoIne7cuLosXB3xRzuxsToPix7fbHSopZKtPSqWYbi"
    "7iIi/X6UQpgUVbLt3tc3xjz9gtCnR9hFRHiQRu8vyc71Re26UD8TpuEQXgTh/m/uSm4f"
    "OPwMd5DZrUB+41TGSWMqXDhow7RYbjcPiWw1Hu5Ejg7S7f9hVObKv2KZSZowmkM+auPD"
    "F+WLylH7hBis05eYZPifP9nZnqDuCWkWT493zjC9FgLJBzZvHc4Q9GWfPNvMa+IPEoY3"
    "KREVkZtLRsiFwZoIE2ieJe4njGcyvEKCt6fgdF24kH8vz1d+3BcDahzUmYLwBmXWa1w8"
    "/Q3zbxHyfuCigh6Iu0ttZXCCW27Mpe54fKkHGf5WaXWTiY+NFK29hTFmFtn19dofWeYq"
    "ureDxNRLK2THS42omnggBcsk8beBFm0e0+QhNR8TKVpdWRvkC7FcigVEMMVuUWcUNu0q"
    "UwzWJUJzBXM/EGe5At6DM23rXvATIclSXzt7UZJIGik7PTU74QO2zOeTiLu/Pm1c/Yzh"
    "1o/o+yXnPJtWBPjbHRqf0vDLoaRBuMtMEWrS0s25hem6OmzpNGuTR1Do/vylaeknYsEk"
    "uAIJlVe6gkUpwZw9iz+gBQTz6MsGs6HPV4j2B9f1S6/I5EcrUUd19AC9c1lSKU0rjDH5"
    "j9YW6ESYcMIEqj05KB2vdlsj1dE3gaR+C7i70xNZM6U+aBjzlLBkqN5PqgqsJ3RjHyjf"
    "jWcykv2d6KQk3zZfasIvbrFa0wXfnAqla6/E7+3Y2ran7Qs4RlZenbWST0aXHlxXR89k"
    "WDIw21E8ahVG40k96Kgsc70l5yMyLy9UEvqTWQb698KXQZFxgg9tEoZq6gnjJs9OP+CJ"
    "jVjAIP7Dmjz4ABfi4abOSYb+fMlrY7kF+5K2hZXoh0vybaDCleKnWWoFL/P2c8Wnzuq6"
    "OptZj32Y8bl3qBfD0/6YmtlJsZHkU6zhJw4PvbUvBYG/CSb4yWTbKJ6cLdD9z39N6/TL"
    "tn7XyyzcjdpY0jGSeYvm078clhX+4FQ89kXi+dgjzJ7DvaYnAE0rtKauxfHJYStnC/TF"
    "bKwetVvL24weVa97fh6qsgxNawVcIBynWkuVtFdSycMiuYMPJPXml4CyFHOonRzYNyaz"
    "AAomJVs8SOj1Btmq+E2Tjo448k/FXD6vR44Ko8IfQsV024NXTzFK1aWndZcNjqNuXkUp"
    "oiaddJC8htrqYEi/wwvYKMF51DJM1MXypI/Y3Px+jbj5m0F06+S4lNykn7YZQd1SWLjq"
    "NEgLBPivvAvDl2ysxKpJNa/l/YaKIaciQ9jqRWIxLEP8LqA2RXaCCco6U7w0oZoYFKCH"
    "4WcEX9llPlGgHwa3y1/uVelwQuT/s5fUSlJaeZ0ETP40ItqIlcMHfp/X/ph/Vv5iDhJX"
    "+/T60GKXIWkAzW//nElxiUeiEoYZpmGrJLHaCeRmMkV16sRBP6GLp2BpP4ob0ksHye6+"
    "agGWlDC5R1vvqhay0W19FHWqtZo2xyfU53dfLVcSLq9BKjV2jS8i18CD+Fi7TBZNeB7h"
    "PXEWcP5mTDVF6k5mFvZFP2Z5BGQpjY0JwqDLbhfzf52ZHPEoIEi2cBQ9rsk/tKRsO7P3"
    "VOTjkKf8ibmUd0cXtg2v1U991qb8UnRup9fg9N4V/Vj4MyLzHx4VGUOW1v6AVvSE67wo"
    "Lqwg0AoFHpOd+iQeznoHxEWtaqNAYKOUaiNHGKqNvEwPjqHBP1ogOJrBP5ohOJxO/Xtm"
    "e/oGrWSVplpLF7cFsVUhxPqhKCy1YWmn9+Wkmyzq8e5nVJpfxNy4u4qGiOKH6flU8nUv"
    "Pyet4KYzq838SHHBj3ObNXPrXoFOG5iG/ZYmWchcPMs52nZ4EE9VXcw/wbKaV66JKiyM"
    "ILTWbjXQG+R/HERhWiQE65O3JYYKOHaJjA7x9SmdqWqkVBL/Kpv193Sb+0zFrfSb95OZ"
    "hPNu2BEx5t91CDS+urBn8U9rVGVty2E4JJy0DxMgS/72LyVFKtqDt5ordmc6ijTr+Hh4"
    "eLqEh03s16g7J7P/bfXmM7+V9OtcFbQ7nnwp1Jh6dnp0cEeG7jbl32ocj4G81QSBVZLD"
    "mih8BnhFZj9+rdm1ehM2hxSZef/xdMFVzoKSM1ujHwGwRJcfSXQOHAqkQHcW36jVhHNE"
    "rBrCZDwOu/Ghqiy6DgPH7LTrpycLe/d3FfJ7Ti2VQBDjk1XQegJ3F70oYSYRPHgEt4at"
    "HWIz60/DzT2o44wcGUWjf9wOiF8+ng/ikB09T+LiqwrfhWJ2pgqXYB9hdE8b4EHgBTL+"
    "urOvH9Kofy5DaFMYrs2ieo+i9qROM7IVKMy7O6UoRT5LyYwc5XpwBTcQcno3WjvMbmiC"
    "ta2iv2HD6nfDRqen+nHPtjPTeOlH3DXbTH6JSYiL/I19ggk3ugvaiKJZipHH8TGO74UO"
    "JHLjrmzPfxdnZHssNh3KDXcXCXZOlo9FX+vWig5HNMAzXfOyr57oZNPC8kQWYyssWP6L"
    "yrxCZjaXwqGFpzSdIv3Fgf2En+Bq5Vxn0SaJ+ezhxkxdHC5a6BwL4UrZLhAUFkBUPMdZ"
    "6iLaUKlWiHxO9MNW8qFVn6SeKI6rL2c9pp6oSJZIX+fU0WowgiR9YzXuvDui7VdPPmQt"
    "YOtkebiwOr5hF16Xz27n5zow6auKSIGIwdra7fmsBtrE0ni7980IEZVARoseUkjelvR5"
    "7cojRVuJ4JFiszDMC3+4Mw6SxLaHnXvmc5G84p8feZ5DUYUqwZ5QUynjT0ghny/aQSZ6"
    "SRNH7nTa8AUXdnrzAj6ne8sg8Xd4VubVEzriX2m9DEywJyETwPFxBsrUrDS14uUohQtb"
    "x5YxdlGbP/t8H3tYNx6JZS+2O5prZK1mmCq42+g1fAubOyLfnLzT+Zim5DJJq3sOv5WT"
    "MlgoaOToaCqzuy1oIawM6Y7w0VMzpv2BogzdfqRIPbna6augyO+1b9sgc3Iwzw+Td6KO"
    "Js1brkyH1CtXxPKFZXEQ3dNge3Fubq7O+6ZN83O/u8r6L7FYe6pfbW6LnhO2+m2NDGR4"
    "jP1eai1tVvsVqtcq5Cxi9NhI9wgglqUOz/1F9szjk+hAJh775evz/NYkPsXPb9rF3nRs"
    "lCYUvh5NJdeXvKggDzlyRcuN2h/G//LZKZv0mWen1jdWPp1ZvuV3/XDA+8WuzI+2PFIC"
    "hU65/up/7e4VJ7C8rEb1G9vqL+4v+4Y5TyE87Q6Bw6u3WLpi6KXQCLdTP0E93lMQaM9l"
    "q/jzmpoj7ArqcCiDzyvpzWsSc8NkFmVMNBEzs6WtzwOv2zljirYGChAg4i7lLFdv+mAf"
    "mBi0KfUrVcS15rzibyHIags3cPdVaoNsK6Dj09vMJ4Za9TBxrSAMzUYbv4d7rdXzCbHN"
    "sSxiNXaL1hMdTE3kHlvpclvN/D9EFnAdevpYRmFN3NlTIx5xm+eR15PSa9+H4TgukPoz"
    "25rkX2bfIvnDAypkDuilSDTPBtoHv/rIdFvym6vQ2Yhe/8S+vl9gumDq/1niDDFLnfsZ"
    "wWcX9budlAf98E+hMjagBmbPVNWyy369/ZpEl7rm5kAqz+22Jp4jXg5DD4E4GsZcna7W"
    "iPbPYQi2/3KH68xmuNXBQ2g+kebZnxLK4O9ZLDGHyDwSYpHrbCDsJIx6ck8J6Sd0caz3"
    "tTrP+R9tEnyYURqAUDbLvQJEy0or6WVyIFbIHh0FPqMxa8cXuG9uo8V09rBUYDvEM+lW"
    "DBfkflK1EVdLpwrWBlevmQavxUpzJiuWA8Cb8ENFbZnAFkcE4OORkVD8WilmFPg/oqGh"
    "mQ=="
)
_EMBEDDED_ASSETS = {
    "lpu_icon.ico": _ICO_B64,
    "lpu_icon_256.png": _PNG256_B64,
}


if __name__ == "__main__":
    try:
        LockHunter().mainloop()
    except SystemExit:
        raise
    except BaseException:
        # Fatal startup/mainloop error. Save a copy-pasteable crash log and try
        # to show it; fall back to printing (which any console/installer log
        # will capture).
        _path = write_crash_log(*sys.exc_info(), where="startup / mainloop")
        try:
            import tkinter as _tk
            from tkinter import messagebox as _mb
            _root = _tk.Tk()
            _root.withdraw()
            _m = "Lock Hunter hit a fatal error and had to close.\n\n"
            if _path:
                _m += (f"A crash log was saved here:\n{_path}\n\n"
                       "Please open it, copy everything, and send it over so it "
                       "can be fixed.")
            _mb.showerror("Lock Hunter — startup error", _m)
            _root.destroy()
        except Exception:
            traceback.print_exc()
        sys.exit(1)
